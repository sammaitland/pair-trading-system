"""
Unified daily data capture module.

Combines three data capture tasks: closing prices and market caps (from IBKR),
analyst estimates archive (from Alpha Vantage), and earnings calendar
(from IBKR Wall Street Horizons). Includes staleness detection and backfill.

STATUS: live
"""

import sys
import os
import asyncio
import logging
import time
import sqlite3
import argparse
import requests
import xml.etree.ElementTree as ET
import warnings
from datetime import datetime, timedelta
from io import StringIO
from pathlib import Path

import pandas as pd
import numpy as np
import nest_asyncio

# Enable nested event loops for Jupyter
nest_asyncio.apply()
warnings.filterwarnings('ignore')

from src.shared import config
from src.shared import config_helper as ch

from ib_insync import IB, Stock, StopOrder, WshEventData
import json

# Setup logging
ch.setup_logging()
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Alpha Vantage API
ALPHA_VANTAGE_API_KEY = config.alpha_vantage_api_key()
AV_API_DELAY = 0.8  # seconds between calls (~75/min)

# Base ETFs for analyst universe
BASE_ETFS = ['VGT', 'VHT', 'VFH', 'VIS', 'VCR', 'VOX']

# Archive paths
ARCHIVE_PATH = config.analyst_archive_dir()
ANALYST_DB_FILE = os.path.join(ARCHIVE_PATH, 'analyst_data.db')

# How often to refresh analyst universe (days)
UNIVERSE_REFRESH_DAYS = 7

# Ensure archive directory exists
os.makedirs(ARCHIVE_PATH, exist_ok=True)


# ============================================================================
# STOP LOSS PROTECTION IMPORTS (conditional)
# ============================================================================

try:
    from src.execution.stop_loss_protection import (
        get_short_leg_info,
        calculate_stop_price,
        get_existing_stop_orders,
        cancel_stop_order,
        place_stop_order,
        update_stop_losses,
        get_stop_order_tag,
        parse_stop_order_tag
    )
    STOP_LOSS_AVAILABLE = True
except ImportError:
    STOP_LOSS_AVAILABLE = False
    logger.warning("stop_loss_protection module not available")


# ============================================================================
# TRADING DAY UTILITIES
# ============================================================================

def get_last_trading_day(reference_date=None):
    """
    Get the last trading day before or on reference_date.
    Accounts for weekends (not holidays).
    """
    from datetime import date
    
    if reference_date is None:
        reference_date = date.today()
    elif isinstance(reference_date, datetime):
        reference_date = reference_date.date()
    
    weekday = reference_date.weekday()
    if weekday == 5:  # Saturday
        return reference_date - timedelta(days=1)
    elif weekday == 6:  # Sunday
        return reference_date - timedelta(days=2)
    
    return reference_date


def get_previous_trading_day(reference_date=None):
    """Get the trading day before reference_date."""
    from datetime import date
    
    if reference_date is None:
        reference_date = date.today()
    elif isinstance(reference_date, datetime):
        reference_date = reference_date.date()
    
    prev_day = reference_date - timedelta(days=1)
    
    weekday = prev_day.weekday()
    if weekday == 5:  # Saturday
        return prev_day - timedelta(days=1)
    elif weekday == 6:  # Sunday
        return prev_day - timedelta(days=2)
    
    return prev_day


def count_trading_days_between(start_date, end_date):
    """Count approximate trading days between two dates (weekdays only)."""
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    if isinstance(end_date, datetime):
        end_date = end_date.date()
    
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    
    trading_days = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:
            trading_days += 1
        current += timedelta(days=1)
    
    return trading_days


# ============================================================================
# STALENESS DETECTION
# ============================================================================

def check_closing_prices_staleness(closing_prices_file=None, warn=True):
    """
    Check if closing prices file is stale.
    
    Parameters:
    -----------
    closing_prices_file : str, optional
        Path to closing prices file
    warn : bool
        If True, print warnings for stale data
    
    Returns:
    --------
    dict : {
        'is_stale': bool,
        'capture_date': date or None,
        'expected_date': date,
        'trading_days_stale': int,
        'message': str
    }
    """
    from datetime import date
    
    if closing_prices_file is None:
        closing_prices_file = config.closing_prices_file()
    
    result = {
        'is_stale': False,
        'capture_date': None,
        'expected_date': None,
        'days_stale': 0,
        'trading_days_stale': 0,
        'message': ''
    }
    
    try:
        if not os.path.exists(closing_prices_file):
            result['is_stale'] = True
            result['message'] = 'Closing prices file not found'
            if warn:
                print(f"⚠️  WARNING: {result['message']}")
            return result
        
        df = pd.read_excel(closing_prices_file, sheet_name='Closing_Prices')
        
        if df.empty:
            result['is_stale'] = True
            result['message'] = 'Closing prices file is empty'
            return result
        
        # Get capture date from Capture_Date or Timestamp column
        capture_date = None
        
        if 'Capture_Date' in df.columns and pd.notna(df['Capture_Date'].iloc[0]):
            capture_date = pd.to_datetime(df['Capture_Date'].iloc[0]).date()
        elif 'Timestamp' in df.columns and pd.notna(df['Timestamp'].iloc[0]):
            capture_date = pd.to_datetime(df['Timestamp'].iloc[0]).date()
        else:
            # Fall back to file modification time
            file_mtime = os.path.getmtime(closing_prices_file)
            capture_date = datetime.fromtimestamp(file_mtime).date()
        
        result['capture_date'] = capture_date
        
        # Determine expected date
        today = date.today()
        now = datetime.now()
        
        # If before 4pm, expect previous trading day's data
        # If after 4pm, expect today's data
        if now.hour < 16:
            expected_date = get_previous_trading_day(today)
        else:
            expected_date = get_last_trading_day(today)
        
        result['expected_date'] = expected_date
        
        # Calculate staleness
        if capture_date < expected_date:
            result['is_stale'] = True
            result['days_stale'] = (expected_date - capture_date).days
            result['trading_days_stale'] = count_trading_days_between(
                capture_date, expected_date
            ) - 1  # -1 because we don't count the capture date itself
            
            result['message'] = (
                f"Closing prices are {result['trading_days_stale']} trading day(s) stale! "
                f"Data from {capture_date}, expected {expected_date}"
            )
            
            if warn:
                print(f"\n{'='*80}")
                print(f"⚠️  WARNING: STALE CLOSING PRICES DETECTED")
                print(f"{'='*80}")
                print(f"  Capture date: {capture_date}")
                print(f"  Expected date: {expected_date}")
                print(f"  Trading days stale: {result['trading_days_stale']}")
                print(f"\n  IMPACT: Daily returns will show {result['trading_days_stale']+1} days combined!")
                print(f"  ACTION: Run backfill_closing_prices() or capture fresh data")
                print(f"{'='*80}\n")
        else:
            result['message'] = f"Closing prices are current ({capture_date})"
    
    except Exception as e:
        result['is_stale'] = True
        result['message'] = f"Error checking staleness: {e}"
    
    return result


# ============================================================================
# PART 1: CLOSING PRICES & MARKET CAPS
# ============================================================================

def get_active_tickers(portfolio_file=None):
    """Get all unique tickers from active portfolio positions"""
    if portfolio_file is None:
        portfolio_file = config.portfolio_file()
    
    try:
        portfolio_df = pd.read_excel(portfolio_file, sheet_name='Portfolio')
        
        if portfolio_df.empty:
            logger.warning("Portfolio is empty")
            return set()
        
        tickers = set()
        tickers.update(portfolio_df['Co1'].dropna().unique())
        tickers.update(portfolio_df['Co2'].dropna().unique())
        
        tickers = {t for t in tickers if t and pd.notna(t) and isinstance(t, str)}
        
        logger.info(f"Found {len(tickers)} unique tickers in portfolio")
        return tickers
        
    except Exception as e:
        logger.error(f"Error reading portfolio: {e}")
        return set()


def get_universe_tickers(parameters_file=None):
    """Get all tickers from the parameters file (full universe)"""
    if parameters_file is None:
        parameters_file = config.parameters_file()
    
    try:
        tickers_df = pd.read_excel(parameters_file, sheet_name='Tickers')
        tickers = set(tickers_df['Ticker'].dropna().unique())
        logger.info(f"Found {len(tickers)} tickers in universe")
        return tickers
    except Exception as e:
        logger.error(f"Error reading universe tickers: {e}")
        return set()


def get_cumulative_version_tickers():
    """
    Build a cumulative set of tickers across all V9.x version parameter files
    found in V9_BASE_DIR, plus the current active version.
    """
    all_tickers = set()
    versions_found = []

    # Scan V9_BASE_DIR for version directories (V9.1, V9.2, V9.3, etc.)
    try:
        for entry in os.listdir(config.v9_base_dir()):
            if entry.startswith('V9.') and os.path.isdir(os.path.join(config.v9_base_dir(), entry)):
                params_file = Config.get_parameters_file(entry)
                if os.path.exists(params_file):
                    tickers = get_universe_tickers(params_file)
                    if tickers:
                        versions_found.append(entry)
                        all_tickers.update(tickers)
    except OSError as e:
        logger.warning(f"Error scanning V9_BASE_DIR: {e}")

    logger.info(f"Cumulative universe: {len(all_tickers)} tickers from {len(versions_found)} versions ({', '.join(sorted(versions_found))})")
    return all_tickers, versions_found


def get_megacap_tickers():
    """Get megacap tickers from VGT_MEGACAP_ADJUSTMENT config"""
    if hasattr(config, 'VGT_MEGACAP_ADJUSTMENT') and config.VGT_MEGACAP_ADJUSTMENT.get('enabled'):
        return set(config.VGT_MEGACAP_ADJUSTMENT.get('tickers', {}).keys())
    return set()


async def fetch_closing_prices(ib, tickers):
    """Fetch today's closing prices for all tickers"""
    closing_prices = {}
    
    logger.info(f"Fetching closing prices for {len(tickers)} tickers...")
    
    for ticker in sorted(tickers):
        try:
            contract = Stock(ticker, 'SMART', 'USD')
            
            try:
                qualified = await asyncio.wait_for(
                    ib.qualifyContractsAsync(contract),
                    timeout=10.0
                )
            except asyncio.TimeoutError:
                logger.warning(f"  {ticker}: Timeout qualifying")
                continue
            
            if not qualified:
                logger.warning(f"Could not qualify {ticker}")
                continue
            
            try:
                bars = await asyncio.wait_for(
                    ib.reqHistoricalDataAsync(
                        qualified[0],
                        endDateTime='',
                        durationStr='1 D',
                        barSizeSetting='1 day',
                        whatToShow='TRADES',
                        useRTH=True
                    ),
                    timeout=15.0
                )
            except asyncio.TimeoutError:
                logger.warning(f"  {ticker}: Timeout fetching data")
                continue
            
            if bars and len(bars) > 0:
                close_price = bars[-1].close
                closing_prices[ticker] = close_price
                logger.debug(f"  {ticker}: ${close_price:.2f}")
            else:
                logger.warning(f"  {ticker}: No data returned")
        
        except Exception as e:
            logger.warning(f"  {ticker}: Error - {e}")
        
        await asyncio.sleep(0.1)
    
    logger.info(f"✓ Successfully fetched {len(closing_prices)}/{len(tickers)} closing prices")
    return closing_prices


def _load_delisted_tickers():
    """Load the set of confirmed delisted tickers from config.delisted_tickers_file()."""
    filepath = config.delisted_tickers_file()
    if not os.path.exists(filepath):
        return set()
    try:
        with open(filepath, 'r') as f:
            data = json.load(f)
        if isinstance(data, list):
            return set(data)
        return set(data.get('tickers', []))
    except (json.JSONDecodeError, Exception):
        return set()


def _save_delisted_tickers(failed_tickers, threshold=3):
    """
    Track failure counts and only confirm a ticker as delisted after
    it has failed on at least `threshold` separate runs.

    JSON format: {"tickers": [...], "failure_counts": {"TICK": n, ...}}
    """
    filepath = config.delisted_tickers_file()
    confirmed = set()
    failure_counts = {}

    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            # Support legacy format (plain list)
            if isinstance(data, list):
                confirmed = set(data)
            else:
                confirmed = set(data.get('tickers', []))
                failure_counts = data.get('failure_counts', {})
        except (json.JSONDecodeError, Exception):
            pass

    # Increment counts for this run's failures
    newly_confirmed = []
    for t in failed_tickers:
        if t in confirmed:
            continue
        failure_counts[t] = failure_counts.get(t, 0) + 1
        if failure_counts[t] >= threshold:
            confirmed.add(t)
            newly_confirmed.append(t)
            failure_counts.pop(t, None)

    with open(filepath, 'w') as f:
        json.dump({
            'tickers': sorted(confirmed),
            'failure_counts': failure_counts,
        }, f, indent=2)

    logger.info(f"Delisted tickers file updated: {len(confirmed)} confirmed "
                f"({len(newly_confirmed)} new), {len(failure_counts)} pending")


async def fetch_market_caps_async(ib, tickers, max_workers=6):
    """
    Fetch market caps for all tickers using yfinance with true parallelism.

    Uses ThreadPoolExecutor to issue individual yf.Ticker().fast_info calls
    concurrently, with a semaphore to limit concurrent requests and a single
    retry for transient failures.

    Parameters are kept for interface compatibility (ib is unused).
    Returns (market_caps dict in millions, failed_tickers list).
    """
    import yfinance as yf
    from concurrent.futures import ThreadPoolExecutor

    sorted_tickers = sorted(tickers)
    logger.info(f"Fetching market caps via yfinance for {len(sorted_tickers)} tickers "
                f"(parallel, {max_workers} workers)...")

    def _get_mcap(ticker):
        try:
            mcap = yf.Ticker(ticker).fast_info.market_cap
            return ticker, mcap
        except Exception:
            return ticker, None

    semaphore = asyncio.Semaphore(max_workers)
    loop = asyncio.get_event_loop()

    async def _fetch_with_semaphore(executor, ticker):
        async with semaphore:
            return await loop.run_in_executor(executor, _get_mcap, ticker)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        tasks = [_fetch_with_semaphore(executor, t) for t in sorted_tickers]
        results = await asyncio.gather(*tasks)

    # First pass — collect successes and failures
    market_caps = {}
    first_pass_failed = []
    for ticker, mcap in results:
        if mcap is not None and mcap > 0:
            market_caps[ticker] = mcap / 1_000_000
        else:
            first_pass_failed.append(ticker)

    # Retry failed tickers once after a short delay
    if first_pass_failed:
        logger.info(f"Retrying {len(first_pass_failed)} failed tickers after 2s delay...")
        await asyncio.sleep(2)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            retry_tasks = [_fetch_with_semaphore(executor, t) for t in first_pass_failed]
            retry_results = await asyncio.gather(*retry_tasks)

        for ticker, mcap in retry_results:
            if mcap is not None and mcap > 0:
                market_caps[ticker] = mcap / 1_000_000

    failed_tickers = [t for t in sorted_tickers if t not in market_caps]

    logger.info(f"✓ Fetched {len(market_caps)}/{len(sorted_tickers)} market caps "
                f"({len(failed_tickers)} failed)")

    # Persist failed tickers to delisted file
    if failed_tickers:
        _save_delisted_tickers(failed_tickers)

    return market_caps, failed_tickers


def save_daily_data(closing_prices, market_caps, output_file, capture_date=None):
    """
    Save closing prices and market caps to Excel file.
    
    ENHANCED: Now stores Capture_Date explicitly for staleness detection.
    
    Parameters:
    -----------
    closing_prices : dict
    market_caps : dict
    output_file : str
    capture_date : date, optional
        The trading date this data represents. If None, uses last trading day.
    """
    from datetime import date
    
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if capture_date is None:
            capture_date = get_last_trading_day()
        elif isinstance(capture_date, datetime):
            capture_date = capture_date.date()
        
        capture_date_str = capture_date.strftime('%Y-%m-%d')
        
        closes_data = []
        for ticker, price in sorted(closing_prices.items()):
            closes_data.append({
                'Ticker': ticker,
                'Close_Price': price,
                'Capture_Date': capture_date_str,  # NEW: explicit date
                'Timestamp': timestamp
            })
        closes_df = pd.DataFrame(closes_data)
        
        mcaps_data = []
        for ticker, mcap in sorted(market_caps.items()):
            mcaps_data.append({
                'Ticker': ticker,
                'Market_Cap_M': mcap,
                'Capture_Date': capture_date_str,
                'Timestamp': timestamp
            })
        mcaps_df = pd.DataFrame(mcaps_data)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            closes_df.to_excel(writer, sheet_name='Closing_Prices', index=False)
            mcaps_df.to_excel(writer, sheet_name='Market_Caps', index=False)
        
        logger.info(f"✓ Saved {len(closes_data)} closing prices and "
                    f"{len(mcaps_data)} market caps to {output_file}")
        logger.info(f"  Capture date: {capture_date_str}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error saving data: {e}")
        return False


def load_closing_prices(closing_prices_file=None, check_staleness=True, warn_if_stale=True):
    """
    Load closing prices from file.
    
    ENHANCED: Now checks for stale data and warns.
    
    BACKWARD COMPATIBLE: This function is imported by workflow_v9c.py
    
    Parameters:
    -----------
    closing_prices_file : str, optional
        Path to file
    check_staleness : bool
        If True, check if data is stale (default: True)
    warn_if_stale : bool
        If True, print warning if stale (default: True)
    
    Returns:
    --------
    dict : {ticker: close_price}
    """
    if closing_prices_file is None:
        closing_prices_file = config.closing_prices_file()
    
    # Check staleness first
    if check_staleness:
        staleness = check_closing_prices_staleness(closing_prices_file, warn=warn_if_stale)
        if staleness['is_stale'] and staleness['trading_days_stale'] > 0:
            logger.warning(f"STALE DATA: {staleness['message']}")
    
    try:
        if not os.path.exists(closing_prices_file):
            logger.warning(f"File not found: {closing_prices_file}")
            return {}
        
        df = pd.read_excel(closing_prices_file, sheet_name='Closing_Prices')
        
        closing_prices = {}
        for _, row in df.iterrows():
            closing_prices[row['Ticker']] = row['Close_Price']
        
        logger.info(f"✓ Loaded {len(closing_prices)} closing prices")
        return closing_prices
        
    except Exception as e:
        logger.error(f"Error loading closing prices: {e}")
        return {}


def load_market_caps(market_caps_file=None):
    """Load market caps from file"""
    if market_caps_file is None:
        market_caps_file = config.closing_prices_file()
    
    try:
        if not os.path.exists(market_caps_file):
            logger.warning(f"File not found: {market_caps_file}")
            return {}
        
        df = pd.read_excel(market_caps_file, sheet_name='Market_Caps')
        
        market_caps = {}
        for _, row in df.iterrows():
            market_caps[row['Ticker']] = row['Market_Cap_M']
        
        logger.info(f"✓ Loaded {len(market_caps)} market caps")
        return market_caps
        
    except Exception as e:
        logger.error(f"Error loading market caps: {e}")
        return {}


def save_portfolio_with_stops(portfolio_df, options_df, portfolio_file):
    """Save portfolio with updated stop order info"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        wb.remove(wb.active)
        
        ws_portfolio = wb.create_sheet('Portfolio')
        for r in dataframe_to_rows(portfolio_df, index=False, header=True):
            ws_portfolio.append(r)
        
        ws_options = wb.create_sheet('Options')
        for r in dataframe_to_rows(options_df, index=False, header=True):
            ws_options.append(r)
        
        wb.save(portfolio_file)
        logger.info(f"✓ Saved portfolio with stop order info")
        return True
        
    except Exception as e:
        logger.error(f"Error saving portfolio: {e}")
        return False


async def capture_closes_and_mcaps(ib=None, update_stops=True, verbose=True):
    """
    Capture closing prices and market caps.
    
    Parameters:
    -----------
    ib : IB, optional
        Existing IB connection. If None, will connect.
    update_stops : bool
        If True, update stop loss orders
    verbose : bool
        Print detailed output
    
    Returns:
    --------
    dict : Result with success status and counts
    """
    result = {
        'success': False,
        'prices_fetched': 0,
        'mcaps_fetched': 0,
        'stops_updated': 0,
        'message': ''
    }
    
    own_connection = False
    
    try:
        if verbose:
            print("\n" + "=" * 70)
            print("CLOSING PRICES & MARKET CAPS")
            print("=" * 70)
        
        # Get tickers
        portfolio_tickers = get_active_tickers()
        universe_tickers = get_universe_tickers()
        index_etfs = {'VO', 'VGT', 'VIS', 'VHT', 'VCR', 'VFH', 'VDE'}
        megacap_tickers = get_megacap_tickers()
        
        tickers_for_closes = portfolio_tickers.union(index_etfs).union(megacap_tickers)
        
        if verbose:
            print(f"Portfolio tickers: {len(portfolio_tickers)}")
            print(f"Index ETFs: {len(index_etfs)}")
            print(f"Universe tickers (for mcaps): {len(universe_tickers)}")
        
        # Connect if needed
        if ib is None or not ib.isConnected():
            if verbose:
                print("\nConnecting to IBKR...")
            ib, connected = await ch.connect_ib_async(
                host=config.data_capture_host(),
                port=config.data_capture_port()
            )
            if not connected:
                result['message'] = "Failed to connect to IBKR"
                return result
            own_connection = True
            if verbose:
                print("✓ Connected")
        
        # Fetch closing prices
        if verbose:
            print("\nFetching closing prices...")
        closing_prices = await fetch_closing_prices(ib, tickers_for_closes)
        result['prices_fetched'] = len(closing_prices)
        
        # Fetch market caps
        if verbose:
            print("\nFetching market caps...")
        market_caps, _ = await fetch_market_caps_async(ib, universe_tickers)
        result['mcaps_fetched'] = len(market_caps)
        
        # Update stop losses
        if update_stops and STOP_LOSS_AVAILABLE and config.enable_short_squeeze_protection():
            if verbose:
                print("\nUpdating stop losses...")
            
            try:
                portfolio_df = pd.read_excel(config.portfolio_file(), sheet_name='Portfolio')
                options_df = pd.read_excel(config.portfolio_file(), sheet_name='Options')
                
                if not portfolio_df.empty:
                    index_prices = {etf: closing_prices.get(etf) for etf in index_etfs 
                                   if etf in closing_prices}
                    
                    stop_result = await update_stop_losses(
                        ib, portfolio_df, index_prices, verbose=verbose
                    )
                    
                    result['stops_updated'] = stop_result.get('stops_updated', 0)
                    
                    if result['stops_updated'] > 0:
                        save_portfolio_with_stops(portfolio_df, options_df, config.portfolio_file())
            except Exception as e:
                logger.warning(f"Stop loss update failed: {e}")
        
        # Save data
        success = save_daily_data(closing_prices, market_caps, config.closing_prices_file())
        
        if success:
            result['success'] = True
            result['message'] = f"Saved {len(closing_prices)} prices, {len(market_caps)} mcaps"
        
        # Print index ETF closes
        if verbose and closing_prices:
            print("\nIndex ETF Closes:")
            for etf in sorted(index_etfs):
                if etf in closing_prices:
                    print(f"  {etf}: ${closing_prices[etf]:.2f}")
        
    except Exception as e:
        result['message'] = f"Error: {e}"
        logger.error(f"capture_closes_and_mcaps error: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if own_connection and ib and ib.isConnected():
            ch.disconnect_ib(ib)
    
    return result


# ============================================================================
# PART 2: ANALYST DATA ARCHIVER (Alpha Vantage)
# ============================================================================

def init_analyst_database():
    """Initialize SQLite database for analyst data archive."""
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    # Table: Universe
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS universe (
            ticker TEXT PRIMARY KEY,
            security_name TEXT,
            sector TEXT,
            industry TEXT,
            source_etfs TEXT,
            first_seen DATE,
            last_seen DATE,
            is_active INTEGER DEFAULT 1
        )
    ''')
    
    # Table: Earnings Estimates
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS earnings_estimates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetch_date DATE,
            fetch_timestamp DATETIME,
            ticker TEXT,
            fiscal_date_ending TEXT,
            consensus_eps_estimate REAL,
            number_of_analysts INTEGER,
            reported_eps REAL,
            surprise REAL,
            surprise_percentage REAL,
            report_date TEXT,
            UNIQUE(fetch_date, ticker, fiscal_date_ending)
        )
    ''')
    
    # Table: Company Overview
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS company_overview (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetch_date DATE,
            fetch_timestamp DATETIME,
            ticker TEXT,
            market_cap REAL,
            pe_ratio REAL,
            peg_ratio REAL,
            book_value REAL,
            eps REAL,
            eps_next_quarter REAL,
            quarterly_earnings_growth REAL,
            quarterly_revenue_growth REAL,
            analyst_target_price REAL,
            analyst_rating_strong_buy INTEGER,
            analyst_rating_buy INTEGER,
            analyst_rating_hold INTEGER,
            analyst_rating_sell INTEGER,
            analyst_rating_strong_sell INTEGER,
            fifty_two_week_high REAL,
            fifty_two_week_low REAL,
            fifty_day_ma REAL,
            two_hundred_day_ma REAL,
            shares_outstanding REAL,
            shares_float REAL,
            shares_short REAL,
            short_ratio REAL,
            short_percent_outstanding REAL,
            short_percent_float REAL,
            percent_insiders REAL,
            percent_institutions REAL,
            forward_pe REAL,
            price_to_sales REAL,
            price_to_book REAL,
            ev_to_revenue REAL,
            ev_to_ebitda REAL,
            beta REAL,
            week_52_change REAL,
            dividend_yield REAL,
            ex_dividend_date TEXT,
            dividend_date TEXT,
            UNIQUE(fetch_date, ticker)
        )
    ''')
    
    # Table: Earnings Calendar
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS earnings_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetch_date DATE,
            fetch_timestamp DATETIME,
            ticker TEXT,
            report_date TEXT,
            fiscal_date_ending TEXT,
            estimate REAL,
            currency TEXT,
            UNIQUE(fetch_date, ticker, report_date)
        )
    ''')
    
    # Table: Archive Run Log
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS archive_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_timestamp DATETIME,
            run_type TEXT,
            tickers_processed INTEGER,
            tickers_success INTEGER,
            tickers_failed INTEGER,
            duration_seconds REAL,
            notes TEXT
        )
    ''')
    
    # Create indexes
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_earnings_ticker_date ON earnings_estimates(ticker, fetch_date)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_overview_ticker_date ON company_overview(ticker, fetch_date)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_calendar_ticker ON earnings_calendar(ticker, report_date)')
    
    conn.commit()
    conn.close()
    
    logger.info("✓ Analyst database initialized")


def _safe_float(val):
    """Safely convert to float"""
    if val is None or val == 'None' or val == '-' or val == '':
        return None
    try:
        return float(val)
    except:
        return None


def _safe_int(val):
    """Safely convert to int"""
    if val is None or val == 'None' or val == '-' or val == '':
        return None
    try:
        return int(val)
    except:
        return None


def get_etf_holdings_simple(ticker_symbol, api_key=ALPHA_VANTAGE_API_KEY):
    """Fetch ETF holdings for universe building."""
    
    url = f'https://www.alphavantage.co/query?function=ETF_PROFILE&symbol={ticker_symbol}&apikey={api_key}&outputsize=full'
    
    try:
        response = requests.get(url, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            
            if "Note" in data or "Information" in data:
                logger.warning(f"API limit reached for {ticker_symbol}")
                return None
            
            holdings = []
            if 'holdings' in data:
                for holding in data['holdings']:
                    ticker = holding.get('symbol', '')
                    if ticker and ticker != 'N/A':
                        holdings.append({
                            'ticker': ticker,
                            'security': holding.get('description', 'N/A'),
                            'weight': float(holding.get('weight', 0))
                        })
            
            return holdings
    except Exception as e:
        logger.error(f"Error fetching {ticker_symbol} holdings: {e}")
    
    return None


def update_analyst_universe():
    """Update the universe of tickers from all ETFs and all version parameter files."""

    logger.info("Updating analyst universe...")

    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()

    today = datetime.now().date()
    all_tickers = {}

    for etf in BASE_ETFS:
        logger.info(f"  Fetching {etf} holdings...")
        holdings = get_etf_holdings_simple(etf)

        if holdings:
            logger.info(f"    ✓ {len(holdings)} holdings")
            for h in holdings:
                ticker = h['ticker']
                if ticker not in all_tickers:
                    all_tickers[ticker] = {
                        'security': h['security'],
                        'etfs': [etf]
                    }
                else:
                    if etf not in all_tickers[ticker]['etfs']:
                        all_tickers[ticker]['etfs'].append(etf)
        else:
            logger.warning(f"    ✗ Failed to fetch {etf}")

        time.sleep(AV_API_DELAY)

    # Update database from ETF holdings
    for ticker, info in all_tickers.items():
        etfs_str = ','.join(info['etfs'])

        cursor.execute('SELECT first_seen FROM universe WHERE ticker = ?', (ticker,))
        result = cursor.fetchone()

        if result:
            cursor.execute('''
                UPDATE universe
                SET security_name = ?, source_etfs = ?, last_seen = ?, is_active = 1
                WHERE ticker = ?
            ''', (info['security'], etfs_str, today, ticker))
        else:
            cursor.execute('''
                INSERT INTO universe (ticker, security_name, source_etfs, first_seen, last_seen, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
            ''', (ticker, info['security'], etfs_str, today, today))

    conn.commit()

    # Merge cumulative version tickers into universe
    cumulative_tickers, versions_found = get_cumulative_version_tickers()
    active_version_tickers = get_universe_tickers()  # Current active version

    if cumulative_tickers:
        added_count = 0
        for ticker in cumulative_tickers:
            cursor.execute('SELECT first_seen FROM universe WHERE ticker = ?', (ticker,))
            result = cursor.fetchone()

            if result:
                # Update last_seen for tickers in the current active version
                if ticker in active_version_tickers:
                    cursor.execute('''
                        UPDATE universe SET last_seen = ?, is_active = 1
                        WHERE ticker = ?
                    ''', (today, ticker))
            else:
                # New ticker from a version parameters file
                cursor.execute('''
                    INSERT INTO universe (ticker, source_etfs, first_seen, last_seen, is_active)
                    VALUES (?, ?, ?, ?, 1)
                ''', (ticker, 'parameters', today, today))
                added_count += 1

        conn.commit()
        logger.info(f"  Version parameters: {added_count} new tickers added from {len(versions_found)} versions")

    cursor.execute('SELECT COUNT(*) FROM universe WHERE is_active = 1')
    total = cursor.fetchone()[0]

    conn.close()

    logger.info(f"✓ Universe updated: {total} active tickers")
    return total


def get_analyst_universe_tickers():
    """Get list of all active tickers in analyst universe."""
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('SELECT ticker FROM universe WHERE is_active = 1 ORDER BY ticker')
    tickers = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    return tickers


def should_refresh_analyst_universe():
    """Check if analyst universe needs refreshing."""
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('SELECT MAX(last_seen) FROM universe')
    result = cursor.fetchone()
    conn.close()
    
    if result[0] is None:
        return True
    
    last_update = datetime.strptime(result[0], '%Y-%m-%d').date()
    days_since = (datetime.now().date() - last_update).days
    
    return days_since >= UNIVERSE_REFRESH_DAYS


def fetch_company_overview_av(ticker, api_key=ALPHA_VANTAGE_API_KEY):
    """Fetch company overview from Alpha Vantage."""
    
    url = f'https://www.alphavantage.co/query?function=OVERVIEW&symbol={ticker}&apikey={api_key}'
    
    try:
        response = requests.get(url, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            
            if "Note" in data or "Information" in data:
                return None, "API_LIMIT"
            
            if 'Symbol' not in data:
                return None, "NO_DATA"
            
            overview = {
                'market_cap': _safe_float(data.get('MarketCapitalization')),
                'pe_ratio': _safe_float(data.get('PERatio')),
                'peg_ratio': _safe_float(data.get('PEGRatio')),
                'book_value': _safe_float(data.get('BookValue')),
                'eps': _safe_float(data.get('EPS')),
                'eps_next_quarter': _safe_float(data.get('EPSNextQuarter')),
                'quarterly_earnings_growth': _safe_float(data.get('QuarterlyEarningsGrowthYOY')),
                'quarterly_revenue_growth': _safe_float(data.get('QuarterlyRevenueGrowthYOY')),
                'analyst_target_price': _safe_float(data.get('AnalystTargetPrice')),
                'analyst_rating_strong_buy': _safe_int(data.get('AnalystRatingStrongBuy')),
                'analyst_rating_buy': _safe_int(data.get('AnalystRatingBuy')),
                'analyst_rating_hold': _safe_int(data.get('AnalystRatingHold')),
                'analyst_rating_sell': _safe_int(data.get('AnalystRatingSell')),
                'analyst_rating_strong_sell': _safe_int(data.get('AnalystRatingStrongSell')),
                'fifty_two_week_high': _safe_float(data.get('52WeekHigh')),
                'fifty_two_week_low': _safe_float(data.get('52WeekLow')),
                'fifty_day_ma': _safe_float(data.get('50DayMovingAverage')),
                'two_hundred_day_ma': _safe_float(data.get('200DayMovingAverage')),
                'shares_outstanding': _safe_float(data.get('SharesOutstanding')),
                'shares_float': _safe_float(data.get('SharesFloat')),
                'shares_short': _safe_float(data.get('SharesShort')),
                'short_ratio': _safe_float(data.get('ShortRatio')),
                'short_percent_outstanding': _safe_float(data.get('ShortPercentOutstanding')),
                'short_percent_float': _safe_float(data.get('ShortPercentFloat')),
                'percent_insiders': _safe_float(data.get('PercentInsiders')),
                'percent_institutions': _safe_float(data.get('PercentInstitutions')),
                'forward_pe': _safe_float(data.get('ForwardPE')),
                'price_to_sales': _safe_float(data.get('PriceToSalesRatioTTM')),
                'price_to_book': _safe_float(data.get('PriceToBookRatio')),
                'ev_to_revenue': _safe_float(data.get('EVToRevenue')),
                'ev_to_ebitda': _safe_float(data.get('EVToEBITDA')),
                'beta': _safe_float(data.get('Beta')),
                'week_52_change': _safe_float(data.get('52WeekChange')),
                'dividend_yield': _safe_float(data.get('DividendYield')),
                'ex_dividend_date': data.get('ExDividendDate'),
                'dividend_date': data.get('DividendDate'),
                'sector': data.get('Sector'),
                'industry': data.get('Industry'),
            }
            
            return overview, "SUCCESS"
            
    except Exception as e:
        return None, f"ERROR: {e}"
    
    return None, "UNKNOWN"


def fetch_earnings_data_av(ticker, api_key=ALPHA_VANTAGE_API_KEY):
    """Fetch earnings estimates from Alpha Vantage."""
    
    url = f'https://www.alphavantage.co/query?function=EARNINGS&symbol={ticker}&apikey={api_key}'
    
    try:
        response = requests.get(url, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            
            if "Note" in data or "Information" in data:
                return None, "API_LIMIT"
            
            if 'quarterlyEarnings' not in data:
                return None, "NO_DATA"
            
            earnings = []
            for q in data.get('quarterlyEarnings', []):
                earnings.append({
                    'fiscal_date_ending': q.get('fiscalDateEnding'),
                    'reported_date': q.get('reportedDate'),
                    'reported_eps': _safe_float(q.get('reportedEPS')),
                    'estimated_eps': _safe_float(q.get('estimatedEPS')),
                    'surprise': _safe_float(q.get('surprise')),
                    'surprise_percentage': _safe_float(q.get('surprisePercentage')),
                })
            
            return earnings, "SUCCESS"
            
    except Exception as e:
        return None, f"ERROR: {e}"
    
    return None, "UNKNOWN"


def fetch_earnings_calendar_av(api_key=ALPHA_VANTAGE_API_KEY):
    """Fetch upcoming earnings calendar from Alpha Vantage."""
    
    url = f'https://www.alphavantage.co/query?function=EARNINGS_CALENDAR&horizon=3month&apikey={api_key}'
    
    try:
        response = requests.get(url, timeout=30)
        
        if response.status_code == 200:
            if 'symbol' in response.text[:100]:
                df = pd.read_csv(StringIO(response.text))
                return df, "SUCCESS"
            else:
                return None, "API_LIMIT"
                
    except Exception as e:
        return None, f"ERROR: {e}"
    
    return None, "UNKNOWN"


def store_overview_data(ticker, overview, fetch_date, fetch_timestamp):
    """Store company overview data in database."""
    
    if not overview:
        return False
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            INSERT OR REPLACE INTO company_overview 
            (fetch_date, fetch_timestamp, ticker, market_cap, pe_ratio, peg_ratio,
             book_value, eps, eps_next_quarter, quarterly_earnings_growth,
             quarterly_revenue_growth, analyst_target_price, analyst_rating_strong_buy,
             analyst_rating_buy, analyst_rating_hold, analyst_rating_sell,
             analyst_rating_strong_sell, fifty_two_week_high, fifty_two_week_low,
             fifty_day_ma, two_hundred_day_ma, shares_outstanding, shares_float,
             shares_short, short_ratio, short_percent_outstanding, short_percent_float,
             percent_insiders, percent_institutions, forward_pe, price_to_sales,
             price_to_book, ev_to_revenue, ev_to_ebitda, beta, week_52_change,
             dividend_yield, ex_dividend_date, dividend_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            fetch_date, fetch_timestamp, ticker,
            overview['market_cap'], overview['pe_ratio'], overview['peg_ratio'],
            overview['book_value'], overview['eps'], overview['eps_next_quarter'],
            overview['quarterly_earnings_growth'], overview['quarterly_revenue_growth'],
            overview['analyst_target_price'], overview['analyst_rating_strong_buy'],
            overview['analyst_rating_buy'], overview['analyst_rating_hold'],
            overview['analyst_rating_sell'], overview['analyst_rating_strong_sell'],
            overview['fifty_two_week_high'], overview['fifty_two_week_low'],
            overview['fifty_day_ma'], overview['two_hundred_day_ma'],
            overview['shares_outstanding'], overview['shares_float'],
            overview['shares_short'], overview['short_ratio'],
            overview['short_percent_outstanding'], overview['short_percent_float'],
            overview['percent_insiders'], overview['percent_institutions'],
            overview['forward_pe'], overview['price_to_sales'], overview['price_to_book'],
            overview['ev_to_revenue'], overview['ev_to_ebitda'], overview['beta'],
            overview['week_52_change'], overview['dividend_yield'],
            overview['ex_dividend_date'], overview['dividend_date']
        ))
        
        if overview.get('sector') and overview.get('industry'):
            cursor.execute('''
                UPDATE universe SET sector = ?, industry = ? WHERE ticker = ?
            ''', (overview['sector'], overview['industry'], ticker))
        
        conn.commit()
        conn.close()
        return True
        
    except Exception as e:
        conn.close()
        return False


def store_earnings_data(ticker, earnings_list, fetch_date, fetch_timestamp):
    """Store earnings data in database."""
    
    if not earnings_list:
        return 0
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    stored = 0
    for e in earnings_list:
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO earnings_estimates 
                (fetch_date, fetch_timestamp, ticker, fiscal_date_ending, 
                 consensus_eps_estimate, reported_eps, surprise, surprise_percentage, report_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                fetch_date, fetch_timestamp, ticker, e['fiscal_date_ending'],
                e['estimated_eps'], e['reported_eps'], e['surprise'], 
                e['surprise_percentage'], e['reported_date']
            ))
            stored += 1
        except:
            pass
    
    conn.commit()
    conn.close()
    
    return stored


def store_earnings_calendar_av(calendar_df, fetch_date, fetch_timestamp):
    """Store earnings calendar from Alpha Vantage."""
    
    if calendar_df is None or len(calendar_df) == 0:
        return 0
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    stored = 0
    for _, row in calendar_df.iterrows():
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO earnings_calendar 
                (fetch_date, fetch_timestamp, ticker, report_date, fiscal_date_ending, estimate, currency)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                fetch_date, fetch_timestamp,
                row.get('symbol'), row.get('reportDate'),
                row.get('fiscalDateEnding'), _safe_float(row.get('estimate')),
                row.get('currency')
            ))
            stored += 1
        except:
            pass
    
    conn.commit()
    conn.close()
    
    return stored


def log_archive_run(run_type, processed, success, failed, duration, notes=''):
    """Log an archive run."""
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO archive_runs (run_timestamp, run_type, tickers_processed, 
                                  tickers_success, tickers_failed, duration_seconds, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (datetime.now(), run_type, processed, success, failed, duration, notes))
    
    conn.commit()
    conn.close()


def run_analyst_archive(max_tickers=None, skip_existing_today=True, verbose=True):
    """
    Run analyst data archive job.
    
    Parameters:
    -----------
    max_tickers : int, optional
        Limit number of tickers (for testing)
    skip_existing_today : bool
        Skip tickers already fetched today
    verbose : bool
        Print progress updates
    
    Returns:
    --------
    tuple : (success_count, failed_count)
    """
    start_time = datetime.now()
    fetch_date = start_time.date().isoformat()
    fetch_timestamp = start_time.isoformat()
    
    if verbose:
        print("\n" + "=" * 70)
        print("ANALYST DATA ARCHIVE")
        print(f"Timestamp: {fetch_timestamp}")
        print("=" * 70)
    
    # Initialize database
    init_analyst_database()
    
    # Check universe
    if should_refresh_analyst_universe():
        if verbose:
            print("Refreshing analyst universe...")
        update_analyst_universe()
    
    # Get tickers and filter out delisted
    tickers = get_analyst_universe_tickers()
    delisted = _load_delisted_tickers()
    if delisted:
        pre_filter = len(tickers)
        tickers = [t for t in tickers if t not in delisted]
        skipped_delisted = pre_filter - len(tickers)
        if skipped_delisted > 0:
            logger.info(f"Skipped {skipped_delisted} delisted tickers")
            if verbose:
                print(f"Skipped {skipped_delisted} delisted tickers")

    if max_tickers:
        tickers = tickers[:max_tickers]

    if verbose:
        print(f"Tickers to process: {len(tickers)}")
    
    # Check already fetched
    if skip_existing_today:
        conn = sqlite3.connect(ANALYST_DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            'SELECT DISTINCT ticker FROM company_overview WHERE fetch_date = ?', 
            (fetch_date,)
        )
        already_fetched = set(row[0] for row in cursor.fetchall())
        conn.close()
        
        tickers = [t for t in tickers if t not in already_fetched]
        if verbose:
            print(f"Skipping {len(already_fetched)} already fetched today")
            print(f"Remaining to fetch: {len(tickers)}")
    
    if len(tickers) == 0:
        if verbose:
            print("Nothing to fetch - all tickers already processed today")
        return 0, 0
    
    # Process tickers
    success_count = 0
    fail_count = 0
    api_limit_hit = False
    
    for i, ticker in enumerate(tickers):
        if api_limit_hit:
            if verbose:
                print(f"\n⚠️ API limit hit - stopping. Processed {i} tickers.")
            break
        
        if verbose and (i + 1) % 50 == 0:
            elapsed = (datetime.now() - start_time).total_seconds()
            rate = i / elapsed * 60
            print(f"  Progress: {i+1}/{len(tickers)} ({rate:.1f}/min)")
        
        # Fetch overview
        overview, status = fetch_company_overview_av(ticker)
        
        if status == "API_LIMIT":
            api_limit_hit = True
            fail_count += 1
            continue
        elif status == "SUCCESS" and overview:
            if store_overview_data(ticker, overview, fetch_date, fetch_timestamp):
                success_count += 1
            else:
                fail_count += 1
        else:
            fail_count += 1
        
        time.sleep(AV_API_DELAY)
        
        # Fetch earnings (every 3rd ticker)
        if i % 3 == 0:
            earnings, e_status = fetch_earnings_data_av(ticker)
            
            if e_status == "API_LIMIT":
                api_limit_hit = True
            elif e_status == "SUCCESS" and earnings:
                store_earnings_data(ticker, earnings, fetch_date, fetch_timestamp)
            
            time.sleep(AV_API_DELAY)
    
    # Fetch earnings calendar (single API call)
    if verbose:
        print("\nFetching earnings calendar...")
    calendar_df, cal_status = fetch_earnings_calendar_av()
    if cal_status == "SUCCESS" and calendar_df is not None:
        universe = set(get_analyst_universe_tickers())
        calendar_df = calendar_df[calendar_df['symbol'].isin(universe)]
        stored = store_earnings_calendar_av(calendar_df, fetch_date, fetch_timestamp)
        if verbose:
            print(f"✓ Stored {stored} upcoming earnings announcements")
    
    duration = (datetime.now() - start_time).total_seconds()
    
    if verbose:
        print("\n" + "-" * 70)
        print("ANALYST ARCHIVE COMPLETE")
        print("-" * 70)
        print(f"Duration: {duration:.1f}s ({duration/60:.1f}min)")
        print(f"Success: {success_count}")
        print(f"Failed: {fail_count}")
        if api_limit_hit:
            print("⚠️ API limit hit - run again later to complete")
    
    log_archive_run('daily', len(tickers), success_count, fail_count, duration,
                    'API_LIMIT_HIT' if api_limit_hit else '')
    
    return success_count, fail_count


def get_analyst_archive_stats():
    """Get statistics about the analyst archive."""
    
    conn = sqlite3.connect(ANALYST_DB_FILE)
    cursor = conn.cursor()
    
    stats = {}
    
    cursor.execute('SELECT COUNT(*) FROM universe WHERE is_active = 1')
    stats['active_tickers'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(DISTINCT ticker) FROM company_overview')
    stats['tickers_with_overview'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(DISTINCT fetch_date) FROM company_overview')
    stats['overview_days'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT MIN(fetch_date), MAX(fetch_date) FROM company_overview')
    dates = cursor.fetchone()
    stats['overview_date_range'] = f"{dates[0]} to {dates[1]}" if dates[0] else "No data"
    
    cursor.execute('SELECT COUNT(DISTINCT ticker) FROM earnings_estimates')
    stats['tickers_with_earnings'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM earnings_estimates')
    stats['total_earnings_records'] = cursor.fetchone()[0]
    
    conn.close()
    
    return stats


def print_analyst_archive_stats():
    """Print formatted analyst archive statistics."""
    
    stats = get_analyst_archive_stats()
    
    print("\n" + "=" * 60)
    print("ANALYST ARCHIVE STATISTICS")
    print("=" * 60)
    print(f"Active tickers: {stats['active_tickers']}")
    print(f"Tickers with overview: {stats['tickers_with_overview']}")
    print(f"Overview data days: {stats['overview_days']}")
    print(f"Date range: {stats['overview_date_range']}")
    print(f"Tickers with earnings: {stats['tickers_with_earnings']}")
    print(f"Total earnings records: {stats['total_earnings_records']}")


# ============================================================================
# PART 3: EARNINGS CALENDAR (IBKR Wall Street Horizons)
# ============================================================================

def parse_wsh_earnings_json(json_data, ticker_map=None):
    """
    Parse JSON response from IBKR Wall Street Horizon event data.

    Parameters:
    -----------
    json_data : str
        JSON string from getWshEventData
    ticker_map : dict, optional
        Unused, kept for backward compatibility.

    Returns:
    --------
    list : Earnings event dicts with ticker, date, time info
    """
    earnings_list = []

    logger.info(f"Parser input type: {type(json_data)}")
    logger.info(f"Parser input sample: {str(json_data)[:500]}")

    try:
        data = json.loads(json_data)

        logger.info(f"Parsed type: {type(data)}, length: {len(data) if isinstance(data, list) else 'not list'}")
        logger.info(f"First event keys: {list(data[0].keys()) if isinstance(data, list) and data else 'empty'}")

        # WSH returns a JSON array of events
        events = data if isinstance(data, list) else [data]

        for event in events:
            try:
                # Filter for earnings events only
                if event.get('event_type') != 'wshe_eps':
                    continue

                # Extract ticker from nested data.company.contract
                event_data = event.get('data', {})
                company = event_data.get('company', {})
                ticker = company.get('contract')

                if not ticker:
                    continue

                # Extract date from announce_datetime first, fall back to index_date
                # announce_datetime format: "20260219T120000+0000"
                announce_dt = event_data.get('announce_datetime', '')
                index_date_str = event.get('index_date', '')

                event_date = None
                announce_hour_utc = None

                if announce_dt:
                    try:
                        parsed_dt = datetime.strptime(str(announce_dt).strip(), '%Y%m%dT%H%M%S%z')
                        event_date = parsed_dt.replace(tzinfo=None)
                        announce_hour_utc = parsed_dt.hour
                    except ValueError:
                        # Try extracting just the date portion (first 8 chars)
                        try:
                            event_date = datetime.strptime(str(announce_dt).strip()[:8], '%Y%m%d')
                        except ValueError:
                            pass

                if event_date is None and index_date_str:
                    try:
                        event_date = datetime.strptime(str(index_date_str).strip(), '%Y%m%d')
                    except ValueError:
                        pass

                if event_date is None:
                    continue

                # Determine pre/post market from announce_datetime hour (UTC)
                # Pre-market announcements: before 14:30 UTC (09:30 ET market open)
                # Post-market announcements: 20:00 UTC or later (16:00 ET market close)
                # Mid-day (e.g. T120000 = noon UTC = 7/8am ET) → pre-market
                if announce_hour_utc is not None:
                    if announce_hour_utc >= 20:
                        report_time = 'post-market'
                        trading_date = event_date + timedelta(days=1)
                        while trading_date.weekday() >= 5:
                            trading_date += timedelta(days=1)
                    else:
                        report_time = 'pre-market'
                        trading_date = event_date
                else:
                    report_time = 'pre-market'
                    trading_date = event_date

                # Extract quarter/period info from data if available
                quarter = (event_data.get('fiscal_quarter') or
                           event_data.get('fiscalQuarter') or
                           event_data.get('quarter') or '')

                # Determine status
                current_date = datetime.now()
                if event_date < current_date - timedelta(days=30):
                    status = 'Past'
                elif event_date > current_date + timedelta(days=60):
                    status = 'Future'
                elif event_date < current_date:
                    status = 'Past'
                else:
                    status = 'Upcoming'

                earnings_list.append({
                    'ticker': ticker,
                    'reportDate': event_date.strftime('%Y-%m-%d'),
                    'reportTime': report_time,
                    'tradingDateAffected': trading_date.strftime('%Y-%m-%d'),
                    'quarter': quarter,
                    'status': status
                })

            except Exception as e:
                logger.debug(f"Error parsing WSH event: {e}")
                continue

    except json.JSONDecodeError as e:
        logger.debug(f"WSH JSON parsing error: {e}")
    except Exception as e:
        logger.debug(f"WSH parsing error: {e}")

    return earnings_list


async def fetch_earnings_calendar_ibkr(ib=None, verbose=True):
    """
    Fetch earnings calendar from IBKR Wall Street Horizons using WSH API.

    Uses reqWshEventData (replaces deprecated reqFundamentalData CalendarReport).

    Parameters:
    -----------
    ib : IB, optional
        Existing IB connection. If None, will connect.
    verbose : bool
        Print progress updates

    Returns:
    --------
    dict : Result with success status and dataframe
    """
    result = {
        'success': False,
        'tickers_processed': 0,
        'earnings_found': 0,
        'output_file': None,
        'message': ''
    }

    own_connection = False

    try:
        if verbose:
            print("\n" + "=" * 70)
            print("EARNINGS CALENDAR (IBKR Wall Street Horizons)")
            print("=" * 70)

        # Get tickers from parameters
        tickers = list(get_universe_tickers())
        if verbose:
            print(f"Loading earnings for {len(tickers)} tickers...")

        # Connect if needed
        if ib is None or not ib.isConnected():
            if verbose:
                print("Connecting to IBKR...")
            ib, connected = await ch.connect_ib_async(
                host=config.data_capture_host(),
                port=config.data_capture_port()
            )
            if not connected:
                result['message'] = "Failed to connect to IBKR"
                return result
            own_connection = True
            if verbose:
                print("✓ Connected")

        # Date range
        start_date = datetime.now() - timedelta(days=60)
        end_date = datetime.now() + timedelta(days=60)

        if verbose:
            print(f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        # Step 1: Qualify all contracts to get conIds
        if verbose:
            print("Qualifying contracts...")

        ticker_map = {}  # conId -> ticker
        conid_list = []

        for idx, ticker in enumerate(tickers):
            try:
                contract = Stock(ticker, 'SMART', 'USD')
                try:
                    contracts = await asyncio.wait_for(
                        ib.qualifyContractsAsync(contract),
                        timeout=10.0
                    )
                except asyncio.TimeoutError:
                    continue

                if contracts and contracts[0].conId:
                    con_id = contracts[0].conId
                    ticker_map[con_id] = ticker
                    conid_list.append(str(con_id))

                if (idx + 1) % 200 == 0 and verbose:
                    print(f"  Qualified: {idx + 1}/{len(tickers)}")

                await asyncio.sleep(0.02)

            except Exception as e:
                continue

        if verbose:
            print(f"  ✓ Qualified {len(conid_list)}/{len(tickers)} contracts")

        # Step 2: Subscribe to WSH meta data first (required before event data)
        try:
            meta = await asyncio.wait_for(
                ib.getWshMetaDataAsync(),
                timeout=10.0
            )
            if verbose:
                print(f"  ✓ WSH meta data subscribed")
        except Exception as e:
            logger.warning(f"WSH meta data subscription failed: {e}")

        # Step 3: Fetch WSH earnings data in batches
        all_earnings_data = []
        batch_size = 50

        if verbose:
            print(f"Fetching WSH earnings data...")

        for batch_start in range(0, len(conid_list), batch_size):
            batch = conid_list[batch_start:batch_start + batch_size]

            try:
                wsh_filter = json.dumps({
                    "country": "All",
                    "watchlist": batch,
                    "limit_region": len(batch) * 4,
                    "limit": len(batch) * 4,
                    "wshe_ed": "true"
                })

                wsh_request = WshEventData(
                    filter=wsh_filter,
                    startDate=start_date.strftime('%Y%m%d'),
                    endDate=end_date.strftime('%Y%m%d'),
                )

                try:
                    event_json = await asyncio.wait_for(
                        ib.getWshEventDataAsync(wsh_request),
                        timeout=30.0
                    )
                except asyncio.TimeoutError:
                    logger.warning(f"WSH batch timeout at offset {batch_start}")
                    continue

                if event_json:
                    logger.info(f"WSH response type: {type(event_json)}")
                    logger.info(f"WSH response full: {str(event_json)[:2000]}")
                    # Log first response for debugging
                    if batch_start == 0 and verbose:
                        preview = str(event_json)[:300]
                        logger.info(f"WSH response sample: {preview}")

                    events = parse_wsh_earnings_json(event_json, ticker_map)
                    all_earnings_data.extend(events)

            except Exception as e:
                logger.warning(f"WSH batch error at offset {batch_start}: {e}")
                continue

            if verbose and (batch_start + batch_size) % 200 == 0:
                print(f"  Progress: {min(batch_start + batch_size, len(conid_list))}/{len(conid_list)} "
                      f"({len(all_earnings_data)} earnings found)")

            await asyncio.sleep(0.5)

        # If batch approach returned nothing, try individual conId queries as fallback
        if not all_earnings_data and conid_list:
            if verbose:
                print("  Batch returned no data, trying individual queries...")

            for idx, con_id_str in enumerate(conid_list[:20]):  # Test first 20
                try:
                    wsh_request = WshEventData(
                        conId=int(con_id_str),
                        startDate=start_date.strftime('%Y%m%d'),
                        endDate=end_date.strftime('%Y%m%d'),
                    )

                    event_json = await asyncio.wait_for(
                        ib.getWshEventDataAsync(wsh_request),
                        timeout=15.0
                    )

                    if event_json:
                        logger.info(f"WSH response type: {type(event_json)}")
                        logger.info(f"WSH response full: {str(event_json)[:2000]}")
                        if idx == 0 and verbose:
                            # Print first response to help debug format
                            preview = str(event_json)[:500]
                            print(f"  WSH response sample: {preview}")

                        events = parse_wsh_earnings_json(event_json, ticker_map)
                        all_earnings_data.extend(events)

                    await asyncio.sleep(0.3)

                except Exception as e:
                    logger.debug(f"WSH individual query error for conId {con_id_str}: {e}")
                    continue

            # If individual queries work, continue with the rest
            if all_earnings_data and len(conid_list) > 20:
                if verbose:
                    print(f"  Individual queries working, processing remaining...")
                for idx, con_id_str in enumerate(conid_list[20:], start=20):
                    try:
                        wsh_request = WshEventData(
                            conId=int(con_id_str),
                            startDate=start_date.strftime('%Y%m%d'),
                            endDate=end_date.strftime('%Y%m%d'),
                        )

                        event_json = await asyncio.wait_for(
                            ib.getWshEventDataAsync(wsh_request),
                            timeout=15.0
                        )

                        if event_json:
                            events = parse_wsh_earnings_json(event_json, ticker_map)
                            all_earnings_data.extend(events)

                        if (idx + 1) % 100 == 0 and verbose:
                            print(f"  Progress: {idx + 1}/{len(conid_list)} "
                                  f"({len(all_earnings_data)} earnings found)")

                        await asyncio.sleep(0.3)

                    except Exception as e:
                        continue

        result['tickers_processed'] = len(tickers)
        result['earnings_found'] = len(all_earnings_data)

        # Save to Excel
        if all_earnings_data:
            df = pd.DataFrame(all_earnings_data)
            df['reportDate'] = pd.to_datetime(df['reportDate'])
            df = df.sort_values('reportDate')
            df['reportDate'] = df['reportDate'].dt.strftime('%Y-%m-%d')

            output_file = config.earnings_calendar_file()

            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Earnings Calendar', index=False)

            result['success'] = True
            result['output_file'] = output_file
            result['message'] = f"Saved {len(df)} earnings to {output_file}"

            if verbose:
                print(f"\n✓ Saved {len(df)} earnings announcements")
                print(f"  Output: {output_file}")

                # Show upcoming
                upcoming = df[pd.to_datetime(df['reportDate']) >= datetime.now()]
                upcoming = upcoming[pd.to_datetime(upcoming['reportDate']) <= datetime.now() + timedelta(days=14)]
                if len(upcoming) > 0:
                    print(f"\nUpcoming earnings (next 14 days): {len(upcoming)}")
                    print(upcoming[['ticker', 'reportDate', 'reportTime']].head(10).to_string(index=False))
        else:
            result['message'] = "No earnings data found"
            if verbose:
                print("\n⚠️ No earnings data found")

    except Exception as e:
        result['message'] = f"Error: {e}"
        logger.error(f"fetch_earnings_calendar_ibkr error: {e}")
        import traceback
        traceback.print_exc()

    finally:
        if own_connection and ib and ib.isConnected():
            ch.disconnect_ib(ib)

    return result


# ============================================================================
# UNIFIED DAILY CAPTURE
# ============================================================================

async def run_daily_capture(
    capture_closes=True,
    capture_mcaps=True,
    update_stops=True,
    run_analyst_archive_flag=True,
    analyst_max_tickers=None,
    fetch_earnings=True,
    verbose=True
):
    """
    Run all daily data capture tasks in a single call.
    
    Parameters:
    -----------
    capture_closes : bool
        Capture closing prices (default: True)
    capture_mcaps : bool
        Capture market caps (default: True)
    update_stops : bool
        Update stop loss orders (default: True)
    run_analyst_archive_flag : bool
        Run analyst data archive (default: True)
    analyst_max_tickers : int, optional
        Limit analyst tickers (for testing)
    fetch_earnings : bool
        Fetch earnings calendar from IBKR (default: True)
    verbose : bool
        Print detailed output (default: True)
    
    Returns:
    --------
    dict : Combined results from all tasks
    
    Usage in Jupyter:
    -----------------
    from daily_data_capture import run_daily_capture
    result = await run_daily_capture()
    """
    
    start_time = datetime.now()
    
    results = {
        'success': True,
        'closes_mcaps': None,
        'analyst_archive': None,
        'earnings_calendar': None,
        'total_duration': 0,
        'message': ''
    }
    
    if verbose:
        print("=" * 80)
        print("DAILY DATA CAPTURE - UNIFIED RUN")
        print(f"Timestamp: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 80)
        print("\nTasks:")
        if capture_closes or capture_mcaps:
            print(f"  • Closing prices & market caps (IBKR)")
        if update_stops:
            print(f"  • Stop loss updates")
        if run_analyst_archive_flag:
            print(f"  • Analyst data archive (Alpha Vantage)")
        if fetch_earnings:
            print(f"  • Earnings calendar (IBKR WSH)")
        print()
    
    ib = None
    
    try:
        # ================================================================
        # TASK 1: Closing Prices & Market Caps (needs IBKR)
        # ================================================================
        if capture_closes or capture_mcaps:
            if verbose:
                print("\n" + "-" * 80)
                print("TASK 1: CLOSING PRICES & MARKET CAPS")
                print("-" * 80)
            
            # Connect to IBKR
            ib, connected = await ch.connect_ib_async(
                host=config.data_capture_host(),
                port=config.data_capture_port()
            )

            if connected:
                closes_result = await capture_closes_and_mcaps(
                    ib=ib,
                    update_stops=update_stops,
                    verbose=verbose
                )
                results['closes_mcaps'] = closes_result
                
                if not closes_result['success']:
                    results['success'] = False
            else:
                if verbose:
                    print("❌ Failed to connect to IBKR")
                results['closes_mcaps'] = {'success': False, 'message': 'IBKR connection failed'}
                results['success'] = False
        
        # ================================================================
        # TASK 2: Earnings Calendar (needs IBKR)
        # ================================================================
        if fetch_earnings:
            if verbose:
                print("\n" + "-" * 80)
                print("TASK 2: EARNINGS CALENDAR")
                print("-" * 80)
            
            # Reuse IBKR connection if available
            if ib is None or not ib.isConnected():
                ib, connected = await ch.connect_ib_async(
                    host=config.data_capture_host(),
                    port=config.data_capture_port()
                )
            else:
                connected = True
            
            if connected:
                earnings_result = await fetch_earnings_calendar_ibkr(
                    ib=ib,
                    verbose=verbose
                )
                results['earnings_calendar'] = earnings_result
                
                if not earnings_result['success']:
                    # Non-critical failure - don't fail overall
                    if verbose:
                        print(f"⚠️ Earnings calendar: {earnings_result['message']}")
            else:
                results['earnings_calendar'] = {'success': False, 'message': 'IBKR connection failed'}
        
        # Disconnect IBKR before long-running analyst archive
        if ib and ib.isConnected():
            ch.disconnect_ib(ib)
            ib = None
        
        # ================================================================
        # TASK 3: Analyst Data Archive (Alpha Vantage - long running)
        # ================================================================
        if run_analyst_archive_flag:
            if verbose:
                print("\n" + "-" * 80)
                print("TASK 3: ANALYST DATA ARCHIVE")
                print("-" * 80)
                print("This may take 30-60+ minutes for full universe...")
            
            success, failed = run_analyst_archive(
                max_tickers=analyst_max_tickers,
                skip_existing_today=True,
                verbose=verbose
            )
            
            results['analyst_archive'] = {
                'success': success > 0 or failed == 0,
                'tickers_success': success,
                'tickers_failed': failed
            }
            
            if verbose:
                print_analyst_archive_stats()
    
    except Exception as e:
        results['success'] = False
        results['message'] = f"Error: {e}"
        logger.error(f"run_daily_capture error: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if ib and ib.isConnected():
            ch.disconnect_ib(ib)
    
    # Calculate total duration
    duration = (datetime.now() - start_time).total_seconds()
    results['total_duration'] = duration
    
    # Summary
    if verbose:
        print("\n" + "=" * 80)
        print("DAILY DATA CAPTURE COMPLETE")
        print("=" * 80)
        print(f"Total duration: {duration:.1f}s ({duration/60:.1f}min)")
        
        if results['closes_mcaps']:
            cm = results['closes_mcaps']
            print(f"\nClosing Prices & MCaps:")
            print(f"  Prices: {cm.get('prices_fetched', 0)}")
            print(f"  MCaps: {cm.get('mcaps_fetched', 0)}")
            print(f"  Stops: {cm.get('stops_updated', 0)}")
        
        if results['earnings_calendar']:
            ec = results['earnings_calendar']
            print(f"\nEarnings Calendar:")
            print(f"  Tickers: {ec.get('tickers_processed', 0)}")
            print(f"  Earnings: {ec.get('earnings_found', 0)}")
        
        if results['analyst_archive']:
            aa = results['analyst_archive']
            print(f"\nAnalyst Archive:")
            print(f"  Success: {aa.get('tickers_success', 0)}")
            print(f"  Failed: {aa.get('tickers_failed', 0)}")
        
        print("=" * 80)
    
    return results


# ============================================================================
# BACKFILL CAPABILITY
# ============================================================================

async def backfill_closing_prices(ib=None, target_date=None, verbose=True):
    """
    Backfill closing prices for a missed day.
    
    Use this if you missed running the daily capture and need to 
    set the baseline for correct daily return calculations.
    
    Parameters:
    -----------
    ib : IB connection, optional
        If None, will connect
    target_date : date, optional
        The date to backfill. If None, uses previous trading day.
    verbose : bool
        Print progress
    
    Returns:
    --------
    dict : Result info
    
    Usage in Jupyter:
    -----------------
    from daily_data_capture import backfill_closing_prices
    
    # Backfill previous trading day
    await backfill_closing_prices()
    
    # Backfill specific date
    from datetime import date
    await backfill_closing_prices(target_date=date(2026, 1, 10))
    """
    from datetime import date
    
    result = {
        'success': False,
        'target_date': None,
        'prices_fetched': 0,
        'message': ''
    }
    
    if target_date is None:
        target_date = get_previous_trading_day()
    elif isinstance(target_date, datetime):
        target_date = target_date.date()
    
    result['target_date'] = target_date
    
    if verbose:
        print(f"\n{'='*70}")
        print(f"BACKFILL CLOSING PRICES")
        print(f"{'='*70}")
        print(f"Target date: {target_date}")
    
    close_connection = False
    
    try:
        if ib is None or not ib.isConnected():
            if verbose:
                print("Connecting to IBKR...")
            ib, connected = await ch.connect_ib_async(
                host=config.data_capture_host(),
                port=config.data_capture_port()
            )
            close_connection = True
            
            if not connected:
                result['message'] = 'Failed to connect to IBKR'
                if verbose:
                    print(f"❌ {result['message']}")
                return result
        
        # Get tickers
        active_tickers = get_active_tickers()
        index_tickers = set(config.index_etfs()) | {'VO'}
        megacap_tickers = get_megacap_tickers()
        all_tickers = active_tickers | index_tickers | megacap_tickers
        
        if verbose:
            print(f"Fetching {len(all_tickers)} tickers for {target_date}...")
        
        # Fetch with specific end date
        end_datetime = f"{target_date.strftime('%Y%m%d')} 23:59:59"
        
        closing_prices = {}
        failed = []
        
        for ticker in sorted(all_tickers):
            try:
                contract = Stock(ticker, 'SMART', 'USD')
                qualified = await asyncio.wait_for(
                    ib.qualifyContractsAsync(contract), timeout=10.0
                )
                
                if not qualified:
                    failed.append(ticker)
                    continue
                
                bars = await asyncio.wait_for(
                    ib.reqHistoricalDataAsync(
                        qualified[0],
                        endDateTime=end_datetime,
                        durationStr='1 D',
                        barSizeSetting='1 day',
                        whatToShow='TRADES',
                        useRTH=True
                    ),
                    timeout=15.0
                )
                
                if bars and len(bars) > 0:
                    closing_prices[ticker] = bars[-1].close
                    if verbose and len(closing_prices) % 25 == 0:
                        print(f"  Progress: {len(closing_prices)} fetched...")
                else:
                    failed.append(ticker)
                
            except Exception as e:
                logger.debug(f"{ticker}: {e}")
                failed.append(ticker)
            
            await asyncio.sleep(0.05)
        
        result['prices_fetched'] = len(closing_prices)
        
        if verbose:
            print(f"\n✓ Fetched {len(closing_prices)}/{len(all_tickers)} prices")
            if failed:
                print(f"  Failed: {len(failed)} tickers")
        
        if closing_prices:
            # Save with explicit capture date
            save_success = save_daily_data(
                closing_prices, {},  # Empty market caps for backfill
                config.closing_prices_file(),
                capture_date=target_date
            )
            
            if save_success:
                result['success'] = True
                result['message'] = f'Backfilled {len(closing_prices)} prices for {target_date}'
                if verbose:
                    print(f"\n✓ {result['message']}")
                    print(f"  Saved to: {config.closing_prices_file()}")
            else:
                result['message'] = 'Failed to save prices'
                if verbose:
                    print(f"❌ {result['message']}")
        else:
            result['message'] = 'No prices fetched'
            if verbose:
                print(f"❌ {result['message']}")
    
    except Exception as e:
        result['message'] = f'Error: {e}'
        if verbose:
            print(f"❌ {result['message']}")
        import traceback
        traceback.print_exc()
    
    finally:
        if close_connection and ib and ib.isConnected():
            ch.disconnect_ib(ib)
    
    return result


# ============================================================================
# BACKWARD COMPATIBILITY ALIASES
# ============================================================================

# These maintain compatibility with existing imports
async def capture_daily_data_async(*args, **kwargs):
    """Backward compatible wrapper for capture_closes_and_mcaps"""
    return await capture_closes_and_mcaps(*args, **kwargs)


def capture_daily_closes_interactive(*args, **kwargs):
    """Backward compatible sync wrapper"""
    return asyncio.run(capture_closes_and_mcaps(*args, **kwargs))


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point for command line usage"""
    parser = argparse.ArgumentParser(
        description='Daily Data Capture - Unified module for closes, mcaps, analyst data, and earnings'
    )
    parser.add_argument('--no-closes', action='store_true', help='Skip closing prices')
    parser.add_argument('--no-mcaps', action='store_true', help='Skip market caps')
    parser.add_argument('--no-stops', action='store_true', help='Skip stop loss updates')
    parser.add_argument('--no-analyst', action='store_true', help='Skip analyst archive')
    parser.add_argument('--no-earnings', action='store_true', help='Skip earnings calendar')
    parser.add_argument('--analyst-max', type=int, default=None, help='Limit analyst tickers')
    
    args = parser.parse_args()
    
    result = asyncio.run(run_daily_capture(
        capture_closes=not args.no_closes,
        capture_mcaps=not args.no_mcaps,
        update_stops=not args.no_stops,
        run_analyst_archive_flag=not args.no_analyst,
        analyst_max_tickers=args.analyst_max,
        fetch_earnings=not args.no_earnings,
        verbose=True
    ))
    
    return result['success']


if __name__ == "__main__":
    main()