"""
Delisting and acquisition handler for portfolio positions.

When a portfolio ticker gets acquired or delisted, this module closes
counterpart legs of affected pairs, liquidates acquirer shares received,
and records trades as completed with a delisting exit reason.

STATUS: live
"""

import sys
import os
import asyncio
import logging
from datetime import datetime
from typing import Optional, List, Dict, Tuple

import pandas as pd
import numpy as np

from src.shared import config

from src.shared import config_helper as ch
from ib_insync import IB, Stock, MarketOrder, LimitOrder

logger = logging.getLogger(__name__)


# ============================================================================
# PORTFOLIO ANALYSIS
# ============================================================================

def find_affected_pairs(portfolio_df: pd.DataFrame, delisted_ticker: str) -> pd.DataFrame:
    """
    Find all pairs in portfolio containing the delisted ticker.
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current portfolio
    delisted_ticker : str
        The ticker that was delisted/acquired
    
    Returns:
    --------
    DataFrame : Subset of portfolio containing the delisted ticker
    """
    if portfolio_df.empty:
        return pd.DataFrame()
    
    delisted_ticker = delisted_ticker.upper().strip()
    
    # Find pairs where delisted ticker is Co1 or Co2
    mask = (portfolio_df['Co1'].str.upper() == delisted_ticker) | \
           (portfolio_df['Co2'].str.upper() == delisted_ticker)
    
    affected = portfolio_df[mask].copy()
    
    if not affected.empty:
        # Add helper columns
        affected['Delisted_Is_Co1'] = affected['Co1'].str.upper() == delisted_ticker
        affected['Counterpart'] = affected.apply(
            lambda r: r['Co2'] if r['Delisted_Is_Co1'] else r['Co1'],
            axis=1
        )
        affected['Counterpart_Qty'] = affected.apply(
            lambda r: r['Quantity2'] if r['Delisted_Is_Co1'] else r['Quantity1'],
            axis=1
        )
        affected['Counterpart_Direction'] = affected.apply(
            lambda r: _get_counterpart_direction(r['Tail'], r['Delisted_Is_Co1']),
            axis=1
        )
    
    return affected


def _get_counterpart_direction(tail: str, delisted_is_co1: bool) -> str:
    """
    Determine if counterpart is long or short based on tail and position.
    
    L-tail: Co1 long, Co2 short
    U-tail: Co1 short, Co2 long
    """
    tail = tail.strip().upper() if isinstance(tail, str) else 'L'
    
    if tail == 'L':
        # L-tail: Co1 is long, Co2 is short
        return 'SHORT' if delisted_is_co1 else 'LONG'
    else:
        # U-tail: Co1 is short, Co2 is long
        return 'LONG' if delisted_is_co1 else 'SHORT'


def check_for_acquirer_shares(ib: IB, acquirer_ticker: str) -> Tuple[bool, int]:
    """
    Check if acquirer shares exist in TWS account.
    
    Returns:
    --------
    Tuple[bool, int] : (has_shares, quantity)
    """
    if not ib or not ib.isConnected():
        return False, 0
    
    acquirer_ticker = acquirer_ticker.upper().strip()
    
    for position in ib.positions():
        if position.contract.symbol.upper() == acquirer_ticker:
            return True, int(position.position)
    
    return False, 0


# ============================================================================
# EXECUTION
# ============================================================================

async def close_counterpart_position(
    ib: IB,
    ticker: str,
    quantity: int,
    direction: str,
    use_limit: bool = True,
    limit_buffer_pct: float = 0.005
) -> Dict:
    """
    Close a counterpart position.
    
    Parameters:
    -----------
    ib : IB
        Connected IB instance
    ticker : str
        Ticker to close
    quantity : int
        Number of shares (always positive)
    direction : str
        'LONG' or 'SHORT' - indicates current position direction
    use_limit : bool
        Use limit order (safer) vs market order
    limit_buffer_pct : float
        Buffer for limit price (default 0.5%)
    
    Returns:
    --------
    dict : Execution result
    """
    result = {
        'ticker': ticker,
        'quantity': quantity,
        'direction': direction,
        'success': False,
        'fill_price': None,
        'message': ''
    }
    
    if quantity <= 0:
        result['message'] = 'Invalid quantity'
        return result
    
    try:
        # Create and qualify contract
        contract = Stock(ticker, 'SMART', 'USD')
        qualified = ib.qualifyContracts(contract)
        
        if not qualified:
            result['message'] = f'Could not qualify contract for {ticker}'
            return result
        
        contract = qualified[0]
        
        # Determine order side (opposite of current position)
        # If we're LONG, we SELL to close
        # If we're SHORT, we BUY to close
        side = 'SELL' if direction == 'LONG' else 'BUY'
        
        if use_limit:
            # Get market data for limit pricing
            ticker_data = ib.reqMktData(contract, '', False, False)
            await asyncio.sleep(0.5)
            
            bid = ticker_data.bid if ticker_data.bid and ticker_data.bid > 0 else None
            ask = ticker_data.ask if ticker_data.ask and ticker_data.ask > 0 else None
            
            ib.cancelMktData(contract)
            
            # Calculate limit price
            if side == 'SELL':
                if bid and bid > 0:
                    limit_price = round(bid * (1 - limit_buffer_pct), 2)
                else:
                    # Fall back to market
                    use_limit = False
            else:  # BUY
                if ask and ask > 0:
                    limit_price = round(ask * (1 + limit_buffer_pct), 2)
                else:
                    use_limit = False
        
        # Place order
        if use_limit:
            order = LimitOrder(side, quantity, limit_price, tif='DAY')
            logger.info(f"  Placing LIMIT order: {side} {quantity} {ticker} @ ${limit_price:.2f}")
        else:
            order = MarketOrder(side, quantity, tif='DAY')
            logger.info(f"  Placing MARKET order: {side} {quantity} {ticker}")
        
        trade = ib.placeOrder(contract, order)
        
        # Wait for fill
        max_wait = 30 if use_limit else 15
        for i in range(max_wait):
            await asyncio.sleep(1)
            
            if trade.orderStatus.status == 'Filled':
                result['success'] = True
                result['fill_price'] = trade.orderStatus.avgFillPrice
                result['message'] = f'Filled at ${result["fill_price"]:.2f}'
                logger.info(f"  ✓ {ticker}: Filled {quantity} @ ${result['fill_price']:.2f}")
                return result
            
            if trade.orderStatus.filled > 0 and i % 5 == 0:
                logger.info(f"  Partial fill: {int(trade.orderStatus.filled)}/{quantity}")
        
        # Check final status
        if trade.orderStatus.filled > 0:
            result['success'] = True
            result['fill_price'] = trade.orderStatus.avgFillPrice
            result['quantity'] = int(trade.orderStatus.filled)
            result['message'] = f'Partial fill: {result["quantity"]} @ ${result["fill_price"]:.2f}'
        else:
            # Cancel unfilled order
            ib.cancelOrder(trade.order)
            result['message'] = 'Order not filled - cancelled'
        
    except Exception as e:
        result['message'] = f'Error: {e}'
        logger.error(f"Error closing {ticker}: {e}")
    
    return result


async def liquidate_acquirer_shares(
    ib: IB,
    acquirer_ticker: str,
    quantity: int = None
) -> Dict:
    """
    Liquidate acquirer shares received from acquisition.
    
    Parameters:
    -----------
    ib : IB
        Connected IB instance
    acquirer_ticker : str
        Acquirer ticker to liquidate
    quantity : int, optional
        Shares to sell. If None, sells all.
    
    Returns:
    --------
    dict : Execution result
    """
    acquirer_ticker = acquirer_ticker.upper().strip()
    
    # Check current position
    has_shares, current_qty = check_for_acquirer_shares(ib, acquirer_ticker)
    
    if not has_shares or current_qty <= 0:
        return {
            'ticker': acquirer_ticker,
            'success': False,
            'message': f'No {acquirer_ticker} shares found in account'
        }
    
    # Determine quantity to sell
    sell_qty = quantity if quantity and quantity <= current_qty else current_qty
    
    logger.info(f"Liquidating {sell_qty} shares of acquirer {acquirer_ticker}")
    
    # Close as a long position
    result = await close_counterpart_position(
        ib=ib,
        ticker=acquirer_ticker,
        quantity=sell_qty,
        direction='LONG',
        use_limit=True
    )
    
    result['acquirer'] = True
    return result


# ============================================================================
# COMPLETED TRADES RECORDING
# ============================================================================

def record_delisting_completion(
    affected_trades: pd.DataFrame,
    counterpart_results: List[Dict],
    delisted_ticker: str,
    acquirer_ticker: str = None,
    deal_notes: str = None,
    completed_trades_path: str = None
) -> bool:
    """
    Record affected trades to completed trades file with delisting exit reason.
    
    Parameters:
    -----------
    affected_trades : DataFrame
        Trades affected by delisting
    counterpart_results : list
        Results from closing counterpart positions
    delisted_ticker : str
        The delisted ticker
    acquirer_ticker : str, optional
        Acquirer ticker if known
    deal_notes : str, optional
        Notes about the deal (e.g., "Cash and stock deal, 0.5 MSFT per share")
    completed_trades_path : str, optional
        Path to completed trades file
    
    Returns:
    --------
    bool : Success status
    """
    if completed_trades_path is None:
        completed_trades_path = config.completed_trades_file()
    
    if affected_trades.empty:
        logger.warning("No trades to record")
        return False
    
    try:
        # Load existing completed trades
        try:
            existing_df = pd.read_excel(completed_trades_path, sheet_name='Completed Trades')
        except FileNotFoundError:
            existing_df = pd.DataFrame()
        
        # Create completion records
        completion_records = []
        
        for idx, trade in affected_trades.iterrows():
            # Find counterpart result
            counterpart = trade.get('Counterpart', '')
            cp_result = next(
                (r for r in counterpart_results if r['ticker'] == counterpart),
                None
            )
            
            # Build exit notes
            exit_notes = f"Delisting: {delisted_ticker}"
            if acquirer_ticker:
                exit_notes += f" (acquired by {acquirer_ticker})"
            if deal_notes:
                exit_notes += f" - {deal_notes}"
            if cp_result:
                if cp_result['success']:
                    exit_notes += f" | Counterpart {counterpart} closed @ ${cp_result.get('fill_price', 0):.2f}"
                else:
                    exit_notes += f" | Counterpart {counterpart} close FAILED: {cp_result.get('message', '')}"
            
            record = {
                'Tag': trade.get('Tag'),
                'Pair': trade.get('Pair'),
                'Co1': trade.get('Co1'),
                'Co2': trade.get('Co2'),
                'Tail': trade.get('Tail'),
                'Index': trade.get('Index'),
                'Trade Initiation Date': trade.get('Trade Initiation Date'),
                'Trade Termination Date': datetime.now().strftime('%Y-%m-%d'),
                'Exit Reason': 'Delisting',
                'Exit Notes': exit_notes,
                'Quantity1': trade.get('Quantity1'),
                'Quantity2': trade.get('Quantity2'),
                'Co1 at Initiation': trade.get('Co1 at Initiation'),
                'Co2 at Initiation': trade.get('Co2 at Initiation'),
                'Index at Initiation': trade.get('Index at Initiation'),
                'Counterpart_Exit_Price': cp_result.get('fill_price') if cp_result else None,
                'Alpha_Return': np.nan,  # Cannot calculate due to delisting
                'Version': trade.get('Version', 'V9'),
            }
            
            completion_records.append(record)
        
        # Create DataFrame
        new_completions = pd.DataFrame(completion_records)
        
        # Append to existing
        if not existing_df.empty:
            # Ensure columns match
            for col in new_completions.columns:
                if col not in existing_df.columns:
                    existing_df[col] = np.nan
            
            combined = pd.concat([existing_df, new_completions], ignore_index=True)
        else:
            combined = new_completions
        
        # Save
        combined.to_excel(completed_trades_path, sheet_name='Completed Trades', index=False)
        
        logger.info(f"✓ Recorded {len(completion_records)} delisting completions to {completed_trades_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error recording delisting completions: {e}")
        import traceback
        traceback.print_exc()
        return False


def remove_from_portfolio(
    portfolio_df: pd.DataFrame,
    affected_trades: pd.DataFrame,
    portfolio_file: str = None
) -> pd.DataFrame:
    """
    Remove affected trades from portfolio.
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Full portfolio
    affected_trades : DataFrame
        Trades to remove
    portfolio_file : str, optional
        Path to save updated portfolio
    
    Returns:
    --------
    DataFrame : Updated portfolio
    """
    if portfolio_file is None:
        portfolio_file = config.portfolio_file()
    
    if affected_trades.empty:
        return portfolio_df
    
    # Get tags to remove
    tags_to_remove = set(affected_trades['Tag'].astype(str))
    
    # Filter out affected trades
    remaining = portfolio_df[~portfolio_df['Tag'].astype(str).isin(tags_to_remove)].copy()
    
    logger.info(f"Removed {len(tags_to_remove)} trades from portfolio")
    logger.info(f"Portfolio: {len(portfolio_df)} → {len(remaining)} positions")
    
    return remaining


# ============================================================================
# MAIN HANDLER
# ============================================================================

async def handle_delisting(
    delisted_ticker: str,
    acquirer_ticker: str = None,
    deal_notes: str = None,
    auto_liquidate_acquirer: bool = False,
    dry_run: bool = False,
    verbose: bool = True
) -> Dict:
    """
    Handle a delisted/acquired ticker in the portfolio.
    
    This is the main entry point for dealing with delistings.
    
    Parameters:
    -----------
    delisted_ticker : str
        The ticker that was delisted/acquired
    acquirer_ticker : str, optional
        The acquirer ticker (if known, for liquidation)
    deal_notes : str, optional
        Notes about the deal terms
    auto_liquidate_acquirer : bool
        Automatically liquidate acquirer shares if found
    dry_run : bool
        If True, show what would happen without executing
    verbose : bool
        Print detailed output
    
    Returns:
    --------
    dict : Results summary
    
    Usage:
    ------
    # Basic usage - interactive
    result = await handle_delisting('ATVI')
    
    # With acquirer liquidation
    result = await handle_delisting(
        'ATVI',
        acquirer_ticker='MSFT',
        deal_notes='Cash and stock deal, 0.5 MSFT per share + $50 cash',
        auto_liquidate_acquirer=True
    )
    
    # Dry run to see what would happen
    result = await handle_delisting('ATVI', dry_run=True)
    """
    
    delisted_ticker = delisted_ticker.upper().strip()
    
    result = {
        'delisted_ticker': delisted_ticker,
        'acquirer_ticker': acquirer_ticker,
        'affected_pairs': 0,
        'counterpart_closes': [],
        'acquirer_liquidation': None,
        'portfolio_updated': False,
        'completed_recorded': False,
        'success': False,
        'message': ''
    }
    
    if verbose:
        print("=" * 80)
        print(f"DELISTING HANDLER: {delisted_ticker}")
        print("=" * 80)
        if dry_run:
            print("*** DRY RUN MODE - No trades will be executed ***")
        print()
    
    ib = None
    
    try:
        # ================================================================
        # STEP 1: Load Portfolio and Find Affected Pairs
        # ================================================================
        if verbose:
            print("-" * 60)
            print("STEP 1: Finding Affected Pairs")
            print("-" * 60)
        
        portfolio_df = pd.read_excel(config.portfolio_file(), sheet_name='Portfolio')
        options_df = pd.read_excel(config.portfolio_file(), sheet_name='Options')
        
        affected = find_affected_pairs(portfolio_df, delisted_ticker)
        result['affected_pairs'] = len(affected)
        
        if affected.empty:
            result['message'] = f'No pairs found containing {delisted_ticker}'
            if verbose:
                print(f"  ✓ No pairs found containing {delisted_ticker}")
                print("  Nothing to do.")
            result['success'] = True
            return result
        
        if verbose:
            print(f"  Found {len(affected)} pairs containing {delisted_ticker}:")
            print()
            for _, trade in affected.iterrows():
                print(f"    • {trade['Pair']} ({trade['Tail']}-tail, {trade['Index']})")
                print(f"      Counterpart: {trade['Counterpart']} ({trade['Counterpart_Direction']})")
                print(f"      Qty to close: {int(trade['Counterpart_Qty'])}")
                print()
        
        # ================================================================
        # STEP 2: Connect to IBKR
        # ================================================================
        if not dry_run:
            if verbose:
                print("-" * 60)
                print("STEP 2: Connecting to IBKR")
                print("-" * 60)
            
            ib, connected = await ch.connect_ib_async()
            
            if not connected:
                result['message'] = 'Failed to connect to IBKR'
                if verbose:
                    print("  ❌ Failed to connect to IBKR")
                return result
            
            if verbose:
                print("  ✓ Connected to IBKR")
        
        # ================================================================
        # STEP 3: Close Counterpart Positions
        # ================================================================
        if verbose:
            print()
            print("-" * 60)
            print("STEP 3: Closing Counterpart Positions")
            print("-" * 60)
        
        counterpart_results = []
        
        for _, trade in affected.iterrows():
            counterpart = trade['Counterpart']
            qty = int(trade['Counterpart_Qty'])
            direction = trade['Counterpart_Direction']
            
            if verbose:
                print(f"\n  Processing: {counterpart}")
                print(f"    Direction: {direction}, Quantity: {qty}")
            
            if dry_run:
                # Simulate success
                cp_result = {
                    'ticker': counterpart,
                    'quantity': qty,
                    'direction': direction,
                    'success': True,
                    'fill_price': None,
                    'message': '[DRY RUN] Would close position'
                }
                if verbose:
                    print(f"    [DRY RUN] Would {'SELL' if direction == 'LONG' else 'BUY'} {qty} shares")
            else:
                cp_result = await close_counterpart_position(
                    ib=ib,
                    ticker=counterpart,
                    quantity=qty,
                    direction=direction
                )
                
                if verbose:
                    if cp_result['success']:
                        print(f"    ✓ Closed @ ${cp_result['fill_price']:.2f}")
                    else:
                        print(f"    ❌ Failed: {cp_result['message']}")
            
            counterpart_results.append(cp_result)
        
        result['counterpart_closes'] = counterpart_results
        
        # ================================================================
        # STEP 4: Check for / Liquidate Acquirer Shares
        # ================================================================
        if acquirer_ticker and not dry_run:
            if verbose:
                print()
                print("-" * 60)
                print(f"STEP 4: Checking for Acquirer Shares ({acquirer_ticker})")
                print("-" * 60)
            
            has_shares, acq_qty = check_for_acquirer_shares(ib, acquirer_ticker)
            
            if has_shares:
                if verbose:
                    print(f"  Found {acq_qty} shares of {acquirer_ticker}")
                
                if auto_liquidate_acquirer:
                    if verbose:
                        print(f"  Liquidating acquirer shares...")
                    
                    acq_result = await liquidate_acquirer_shares(ib, acquirer_ticker)
                    result['acquirer_liquidation'] = acq_result
                    
                    if verbose:
                        if acq_result['success']:
                            print(f"  ✓ Liquidated @ ${acq_result['fill_price']:.2f}")
                        else:
                            print(f"  ❌ Failed: {acq_result['message']}")
                else:
                    if verbose:
                        print(f"  ⚠️  Acquirer shares NOT auto-liquidated")
                        print(f"      Set auto_liquidate_acquirer=True to liquidate")
                    result['acquirer_liquidation'] = {
                        'ticker': acquirer_ticker,
                        'quantity': acq_qty,
                        'success': False,
                        'message': 'Auto-liquidation disabled'
                    }
            else:
                if verbose:
                    print(f"  No {acquirer_ticker} shares found in account")
        elif acquirer_ticker and dry_run:
            if verbose:
                print()
                print("-" * 60)
                print(f"STEP 4: Acquirer Share Check ({acquirer_ticker})")
                print("-" * 60)
                print(f"  [DRY RUN] Would check for and optionally liquidate {acquirer_ticker}")
        
        # ================================================================
        # STEP 5: Record to Completed Trades
        # ================================================================
        if verbose:
            print()
            print("-" * 60)
            print("STEP 5: Recording Completed Trades")
            print("-" * 60)
        
        # Separate trades by close outcome — only successfully closed
        # trades are recorded and removed. Failed closes stay in the
        # portfolio for retry.
        successfully_closed = []
        failed_closes = []
        for idx_trade, (_, trade) in enumerate(affected.iterrows()):
            cp_result = counterpart_results[idx_trade]
            if cp_result['success']:
                successfully_closed.append(trade)
            else:
                failed_closes.append(trade)

        successfully_closed_df = pd.DataFrame(successfully_closed) if successfully_closed else pd.DataFrame()
        failed_closes_df = pd.DataFrame(failed_closes) if failed_closes else pd.DataFrame()

        if dry_run:
            if verbose:
                print(f"  [DRY RUN] Would record {len(affected)} trades with Exit Reason: Delisting")
            result['completed_recorded'] = True
        else:
            if successfully_closed_df.empty:
                if verbose:
                    print("  No trades to record (all closes failed)")
                result['completed_recorded'] = True
            else:
                # Only record trades whose counterpart close succeeded
                successful_results = [r for r in counterpart_results if r['success']]
                recorded = record_delisting_completion(
                    affected_trades=successfully_closed_df,
                    counterpart_results=successful_results,
                    delisted_ticker=delisted_ticker,
                    acquirer_ticker=acquirer_ticker,
                    deal_notes=deal_notes
                )
                result['completed_recorded'] = recorded

                if verbose:
                    if recorded:
                        print(f"  ✓ Recorded {len(successfully_closed_df)} trades to completed trades")
                    else:
                        print("  ❌ Failed to record completed trades")

            if not failed_closes_df.empty:
                if verbose:
                    print(f"  ⚠️ {len(failed_closes_df)} trades NOT recorded (close failed — retained in portfolio for retry)")
                    for _, trade in failed_closes_df.iterrows():
                        print(f"      {trade.get('Pair', '?')} — counterpart {trade.get('Counterpart', '?')}")

        # ================================================================
        # STEP 6: Update Portfolio File
        # ================================================================
        if verbose:
            print()
            print("-" * 60)
            print("STEP 6: Updating Portfolio File")
            print("-" * 60)

        if dry_run:
            if verbose:
                print(f"  [DRY RUN] Would remove {len(affected)} trades from portfolio")
            result['portfolio_updated'] = True
        else:
            # Only remove trades whose counterpart close succeeded.
            # Failed closes stay in the portfolio and retain tracking.
            if successfully_closed_df.empty:
                if verbose:
                    print("  No trades to remove (all closes failed — portfolio unchanged)")
                remaining = portfolio_df
                result['portfolio_updated'] = True
            else:
                remaining = remove_from_portfolio(portfolio_df, successfully_closed_df)

                # Save portfolio
                try:
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows

                    wb = Workbook()
                    wb.remove(wb.active)

                    ws_portfolio = wb.create_sheet('Portfolio')
                    for r in dataframe_to_rows(remaining, index=False, header=True):
                        ws_portfolio.append(r)

                    ws_options = wb.create_sheet('Options')
                    for r in dataframe_to_rows(options_df, index=False, header=True):
                        ws_options.append(r)

                    wb.save(config.portfolio_file())
                    result['portfolio_updated'] = True

                    if verbose:
                        print(f"  ✓ Portfolio updated: {len(portfolio_df)} → {len(remaining)} positions")
                        if not failed_closes_df.empty:
                            print(f"  ⚠️ {len(failed_closes_df)} failed trades retained in portfolio")

                except Exception as e:
                    if verbose:
                        print(f"  ❌ Failed to save portfolio: {e}")

        # ================================================================
        # SUMMARY
        # ================================================================
        successful_closes = sum(1 for r in counterpart_results if r['success'])
        result['success'] = (
            successful_closes == len(affected) and
            result['completed_recorded'] and
            result['portfolio_updated']
        )
        
        if verbose:
            print()
            print("=" * 80)
            print("DELISTING HANDLER SUMMARY")
            print("=" * 80)
            print(f"Delisted Ticker: {delisted_ticker}")
            if acquirer_ticker:
                print(f"Acquirer: {acquirer_ticker}")
            if deal_notes:
                print(f"Deal Notes: {deal_notes}")
            print()
            print(f"Affected Pairs: {len(affected)}")
            print(f"Counterparts Closed: {successful_closes}/{len(affected)}")
            print(f"Completed Trades Recorded: {'Yes' if result['completed_recorded'] else 'No'}")
            print(f"Portfolio Updated: {'Yes' if result['portfolio_updated'] else 'No'}")
            print()
            if result['success']:
                print("✓ DELISTING HANDLED SUCCESSFULLY")
            else:
                print("⚠️ DELISTING HANDLING INCOMPLETE - Check details above")
            print("=" * 80)
        
        result['message'] = 'Delisting handled successfully' if result['success'] else 'Partial success'
        
    except Exception as e:
        result['message'] = f'Error: {e}'
        logger.error(f"handle_delisting error: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if ib and ib.isConnected():
            ch.disconnect_ib(ib)
    
    return result


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def list_potential_delistings(portfolio_df: pd.DataFrame = None, ib: IB = None) -> pd.DataFrame:
    """
    Check portfolio for tickers that may have been delisted.
    
    This attempts to qualify contracts for all portfolio tickers and flags
    any that fail - which could indicate delisting.
    
    Parameters:
    -----------
    portfolio_df : DataFrame, optional
        Portfolio to check. If None, loads from Config.
    ib : IB, optional
        Connected IB instance. If None, connects.
    
    Returns:
    --------
    DataFrame : Tickers that could not be qualified (potential delistings)
    """
    if portfolio_df is None:
        portfolio_df = pd.read_excel(config.portfolio_file(), sheet_name='Portfolio')
    
    if portfolio_df.empty:
        return pd.DataFrame()
    
    # Get all unique tickers
    tickers = set()
    tickers.update(portfolio_df['Co1'].dropna().unique())
    tickers.update(portfolio_df['Co2'].dropna().unique())
    
    failed_tickers = []
    own_connection = False
    
    try:
        if ib is None or not ib.isConnected():
            ib = IB()
            ib.connect('127.0.0.1', config.ibkr_port(), clientId=ch.get_client_id() + 50)
            own_connection = True
        
        for ticker in sorted(tickers):
            try:
                contract = Stock(ticker, 'SMART', 'USD')
                qualified = ib.qualifyContracts(contract)
                
                if not qualified:
                    # Find pairs containing this ticker
                    pairs = portfolio_df[
                        (portfolio_df['Co1'] == ticker) | (portfolio_df['Co2'] == ticker)
                    ]['Pair'].tolist()
                    
                    failed_tickers.append({
                        'Ticker': ticker,
                        'Pairs': ', '.join(pairs),
                        'Pair_Count': len(pairs)
                    })
            except Exception as e:
                failed_tickers.append({
                    'Ticker': ticker,
                    'Pairs': 'Error checking',
                    'Pair_Count': 0,
                    'Error': str(e)
                })
    
    finally:
        if own_connection and ib.isConnected():
            ib.disconnect()
    
    return pd.DataFrame(failed_tickers)


async def check_for_delistings_async(verbose: bool = True) -> pd.DataFrame:
    """
    Async version of delisting check.
    
    Usage in Jupyter:
        from delisting_handler import check_for_delistings_async
        potential = await check_for_delistings_async()
    """
    if verbose:
        print("Checking portfolio for potential delistings...")
    
    portfolio_df = pd.read_excel(config.portfolio_file(), sheet_name='Portfolio')
    
    if portfolio_df.empty:
        if verbose:
            print("  Portfolio is empty")
        return pd.DataFrame()
    
    tickers = set()
    tickers.update(portfolio_df['Co1'].dropna().unique())
    tickers.update(portfolio_df['Co2'].dropna().unique())
    
    if verbose:
        print(f"  Checking {len(tickers)} tickers...")
    
    failed_tickers = []
    
    ib, connected = await ch.connect_ib_async()
    
    if not connected:
        print("  ❌ Could not connect to IBKR")
        return pd.DataFrame()
    
    try:
        for ticker in sorted(tickers):
            try:
                contract = Stock(ticker, 'SMART', 'USD')
                qualified = await asyncio.wait_for(
                    ib.qualifyContractsAsync(contract),
                    timeout=5.0
                )
                
                if not qualified:
                    pairs = portfolio_df[
                        (portfolio_df['Co1'] == ticker) | (portfolio_df['Co2'] == ticker)
                    ]['Pair'].tolist()
                    
                    failed_tickers.append({
                        'Ticker': ticker,
                        'Pairs': ', '.join(pairs),
                        'Pair_Count': len(pairs)
                    })
                    
                    if verbose:
                        print(f"  ⚠️  {ticker}: Could not qualify - possible delisting")
                        
            except asyncio.TimeoutError:
                if verbose:
                    print(f"  ⚠️  {ticker}: Timeout - possible delisting")
                failed_tickers.append({
                    'Ticker': ticker,
                    'Pairs': 'Timeout',
                    'Pair_Count': 0
                })
            except Exception as e:
                pass  # Normal failures are fine
            
            await asyncio.sleep(0.1)
    
    finally:
        ch.disconnect_ib(ib)
    
    result_df = pd.DataFrame(failed_tickers)
    
    if verbose:
        if result_df.empty:
            print("  ✓ All tickers validated successfully")
        else:
            print(f"\n  Found {len(result_df)} potential delistings:")
            print(result_df.to_string(index=False))
    
    return result_df


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Handle delisted/acquired tickers')
    parser.add_argument('ticker', help='Delisted ticker symbol')
    parser.add_argument('--acquirer', help='Acquirer ticker')
    parser.add_argument('--notes', help='Deal notes')
    parser.add_argument('--liquidate', action='store_true', help='Auto-liquidate acquirer shares')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    
    args = parser.parse_args()
    
    result = asyncio.run(handle_delisting(
        delisted_ticker=args.ticker,
        acquirer_ticker=args.acquirer,
        deal_notes=args.notes,
        auto_liquidate_acquirer=args.liquidate,
        dry_run=args.dry_run,
        verbose=True
    ))
