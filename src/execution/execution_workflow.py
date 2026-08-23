"""
Execution workflow for the portfolio management system.

Orchestrates the complete daily workflow: IBKR connection, data loading,
alpha calculations, trade terminations, trade evaluation and execution,
portfolio reconciliation, and post-close updates including daily data capture.

STATUS: live
"""

import sys
import os
import asyncio
import logging
import time
import shutil
from datetime import datetime, time as dtime

import pandas as pd
import numpy as np
import nest_asyncio

# ============================================================================
# MODULE STATE
# ============================================================================
# These are populated by initialize() and used throughout the workflow.
# Keeping them at module level avoids passing them through every function.

_initialized = False
logger = logging.getLogger(__name__)

# Module references (populated after reload)
config = None
pm = None
trade_execution = None
tool_box = None


# ============================================================================
# INITIALIZATION
# ============================================================================

def initialize():
    """
    Initialize the workflow module: reload modules, configure logging,
    set up IBKR error filters. Safe to call multiple times (idempotent).
    """
    global _initialized, config, pm, trade_execution, tool_box

    print("=" * 60)
    print("INITIALIZING WORKFLOW MODULE")
    print("=" * 60)

    # Import/re-import modules to get fresh references
    from src.shared import config as _config_module
    from src.execution import portfolio_management as _pm
    from src.execution import trade_execution as _te
    from src.shared import calculations as _tb

    config = _config_module
    pm = _pm
    trade_execution = _te
    tool_box = _tb

    # Apply nest_asyncio for Jupyter compatibility
    nest_asyncio.apply()

    # Configure logging
    config.setup_logging()

    # Prevent duplicate handlers from Jupyter re-runs
    for handler_logger_name in [None, 'ib_insync', 'ib_insync.wrapper']:
        lg = logging.getLogger(handler_logger_name)
        while len(lg.handlers) > 1:
            lg.handlers.pop()

    # Apply IBKR error filters
    _apply_ibkr_error_filters()

    _initialized = True
    print("=" * 60)
    print("✓ Workflow module initialized")
    print("=" * 60 + "\n")


def _apply_ibkr_error_filters():
    """Apply filters to suppress harmless IBKR error messages"""

    class IBKRErrorFilter(logging.Filter):
        def filter(self, record):
            msg = record.getMessage()
            if 'Error 300' in msg:
                return False
            if 'Error 10147' in msg:
                return False
            if 'Error 321' in msg and 'size value cannot be zero' in msg:
                return False
            return True

    ib_wrapper_logger = logging.getLogger('ib_insync.wrapper')
    # Avoid adding duplicate filters
    if not any(isinstance(f, IBKRErrorFilter) for f in ib_wrapper_logger.filters):
        ib_wrapper_logger.addFilter(IBKRErrorFilter())

    logger.info("✓ IBKR error filters applied (suppressing Error 300, 321, 10147)")


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def print_stage_header(stage_num, stage_name):
    """Print formatted stage header"""
    print("\n" + "=" * 80)
    print(f"STAGE {stage_num}: {stage_name}")
    print("=" * 80)


def save_interim_portfolio(portfolio_df, options_df, label=""):
    """Save interim backup of portfolio (simplified, no analytics)"""
    try:
        if label:
            backup_path = config.PORTFOLIO_FILE.replace('.xlsx', f'_{label}.xlsx')
        else:
            backup_path = config.PORTFOLIO_FILE.replace('.xlsx', '_interim.xlsx')

        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        wb.remove(wb.active)

        ws_portfolio = wb.create_sheet('Portfolio')
        for r in dataframe_to_rows(portfolio_df, index=False, header=True):
            ws_portfolio.append(r)

        ws_options = wb.create_sheet('Options')
        for r in dataframe_to_rows(options_df, index=False, header=True):
            ws_options.append(r)

        wb.save(backup_path)
        logger.info(f"Interim save: {backup_path}")
        return True
    except Exception as e:
        logger.error(f"Interim save failed: {e}")
        return False


def normalize_shortlist_columns(shortlist_df):
    """
    Normalize shortlist column names for compatibility with portfolio management.
    LAM outputs: Ticker1/Ticker2 → Portfolio management expects: Co1/Co2
    """
    df = shortlist_df.copy()
    rename_map = {'Ticker1': 'Co1', 'Ticker2': 'Co2'}
    existing_renames = {k: v for k, v in rename_map.items() if k in df.columns}

    if existing_renames:
        df = df.rename(columns=existing_renames)
        logger.info(f"Normalized column names: {list(existing_renames.keys())} → {list(existing_renames.values())}")

    return df


def check_for_duplicate_tags(portfolio_df):
    """
    Check for duplicate Tags in portfolio (critical safeguard).
    Returns: (has_duplicates: bool, duplicate_details: DataFrame)
    """
    if portfolio_df.empty:
        logger.info("Empty portfolio - no duplicates possible")
        return False, pd.DataFrame()

    tag_counts = portfolio_df['Tag'].value_counts()
    duplicate_tags = tag_counts[tag_counts > 1]

    if duplicate_tags.empty:
        logger.info("✓ No duplicate Tags found in portfolio")
        return False, pd.DataFrame()

    duplicate_details = []
    for tag, count in duplicate_tags.items():
        duplicate_rows = portfolio_df[portfolio_df['Tag'] == tag]
        for idx, row in duplicate_rows.iterrows():
            duplicate_details.append({
                'Tag': tag,
                'Pair': row.get('Pair', 'Unknown'),
                'Co1': row.get('Co1', 'Unknown'),
                'Co2': row.get('Co2', 'Unknown'),
                'Tail': row.get('Tail', 'Unknown'),
                'Index': row.get('Index', 'Unknown'),
                'Trade_Date': row.get('Trade Initiation Date', 'Unknown'),
                'Quantity1': row.get('Quantity1', 0),
                'Quantity2': row.get('Quantity2', 0),
                'Total_Count': count,
                'Row_Index': idx
            })

    return True, pd.DataFrame(duplicate_details)


def remove_duplicate_tags(portfolio_df):
    """Remove duplicate tags, keeping first occurrence. Returns (cleaned_df, removed_df)."""
    if portfolio_df.empty:
        return portfolio_df, pd.DataFrame()

    tag_counts = portfolio_df['Tag'].value_counts()
    duplicate_tags = tag_counts[tag_counts > 1].index.tolist()

    if not duplicate_tags:
        logger.info("No duplicates to remove")
        return portfolio_df, pd.DataFrame()

    removed_rows = []
    for tag in duplicate_tags:
        duplicate_rows = portfolio_df[portfolio_df['Tag'] == tag]
        for idx in duplicate_rows.index[1:]:
            removed_rows.append({
                'Tag': tag,
                'Pair': portfolio_df.loc[idx, 'Pair'],
                'Row_Index': idx
            })

    cleaned_df = portfolio_df.drop_duplicates(subset=['Tag'], keep='first')
    return cleaned_df, pd.DataFrame(removed_rows)


def fix_duplicate_tags_interactive(portfolio_file):
    """Interactive duplicate removal with safety checks"""
    print("\n" + "=" * 80)
    print("DUPLICATE TAG REMOVAL UTILITY")
    print("=" * 80)

    print(f"\nLoading portfolio from: {portfolio_file}")
    portfolio_df = pd.read_excel(portfolio_file, sheet_name='Portfolio')
    options_df = pd.read_excel(portfolio_file, sheet_name='Options')
    print(f"Portfolio loaded: {len(portfolio_df)} positions")

    has_duplicates, duplicate_details = check_for_duplicate_tags(portfolio_df)

    if not has_duplicates:
        print("\n✅ No duplicate tags found - portfolio is clean!")
        return

    print(f"\nFound {len(duplicate_details['Tag'].unique())} duplicate tags:")
    display_cols = ['Tag', 'Pair', 'Tail', 'Index', 'Quantity1', 'Quantity2', 'Row_Index']
    print(duplicate_details[display_cols].to_string(index=False))

    print("\n" + "=" * 80)
    print("REMOVAL STRATEGY:")
    print("  • Keep FIRST occurrence of each duplicate tag (lowest row index)")
    print("  • Remove all subsequent occurrences")
    print("  • Create backup before making changes")
    print("=" * 80)

    user_input = input("\nProceed with duplicate removal? (yes/no): ")
    if user_input.lower() != 'yes':
        print("Cancelled - no changes made")
        return

    backup_file = portfolio_file.replace('.xlsx', '_BACKUP_before_dedup.xlsx')
    print(f"\nCreating backup: {backup_file}")
    shutil.copy2(portfolio_file, backup_file)
    print("✓ Backup created")

    print("\nRemoving duplicates...")
    cleaned_df, removed_df = remove_duplicate_tags(portfolio_df)

    print(f"\n✓ Removed {len(removed_df)} duplicate rows:")
    for _, row in removed_df.iterrows():
        print(f"  - Tag {row['Tag']} ({row['Pair']}) at row {row['Row_Index']}")

    print(f"\nSaving cleaned portfolio...")
    pm.save_portfolio_with_analytics(cleaned_df, options_df, portfolio_file, None)

    print("\n" + "=" * 80)
    print("✅ DUPLICATE REMOVAL COMPLETE")
    print("=" * 80)
    print(f"Original: {len(portfolio_df)} positions")
    print(f"Cleaned:  {len(cleaned_df)} positions")
    print(f"Removed:  {len(removed_df)} duplicates")
    print(f"\nBackup saved to: {backup_file}")
    print("=" * 80)




# ============================================================================
# POST-CLOSE FUNCTIONS
# ============================================================================

async def wait_for_market_close(buffer_seconds=30):
    """
    Wait until US market close (4:00 PM ET + buffer).
    Returns immediately if already past close.
    """
    import pytz

    et = pytz.timezone('US/Eastern')
    close_time = dtime(16, 0, buffer_seconds)

    while True:
        now = datetime.now(et)

        if now.time() >= close_time:
            print(f"✓ Market closed — proceeding at {now.strftime('%H:%M:%S ET')}")
            return

        # Calculate remaining time
        close_dt = datetime.combine(now.date(), close_time)
        now_dt = datetime.combine(now.date(), now.time())
        remaining = (close_dt - now_dt).total_seconds()
        mins = remaining / 60

        print(f"  Waiting for close... {mins:.1f} min remaining    ", end='\r')
        await asyncio.sleep(min(30, remaining))  # Check every 30s or less


async def post_close_update(workflow_result):
    """
    Lightweight post-close refresh:
    1. Re-fetch prices at/near close
    2. Re-save portfolio with closing analytics
    3. Disconnect IB
    4. Run daily data capture (manages its own connection)

    Parameters:
    -----------
    workflow_result : dict
        Return value from run_portfolio_workflow()
    """
    from src.shared.fetch_market_data import fetch_live_prices_batch_v2
    from src.execution.daily_data_capture import load_closing_prices

    print("\n" + "=" * 80)
    print("POST-CLOSE PORTFOLIO UPDATE")
    print("=" * 80)

    ib = workflow_result.get('ib')
    portfolio_df = workflow_result.get('portfolio_df')
    options_portfolio_df = workflow_result.get('options_df')
    connected = ib is not None and ib.isConnected()

    if not connected:
        print("⚠️  No active IB connection — attempting reconnect...")
        try:
            ib, connected = await config.connect_ib_async(timeout=10, max_retries=3)
        except Exception as e:
            print(f"❌ Could not reconnect: {e}")
            connected = False

    if not connected:
        print("❌ Cannot update portfolio without IB connection")
        print("   Skipping to daily data capture (which manages its own connection)...")
    else:
        try:
            # 1. Collect all tickers
            all_tickers = set()
            if portfolio_df is not None and not portfolio_df.empty:
                all_tickers.update(portfolio_df['Co1'].tolist())
                all_tickers.update(portfolio_df['Co2'].tolist())
            all_tickers.update(config.INDEX_ETFS)
            all_tickers.update(['VO', 'IGV'])
            all_tickers = {t for t in all_tickers if t and pd.notna(t)}

            print(f"\nFetching closing prices for {len(all_tickers)} tickers...")

            # 2. Fetch closing prices
            closing_prices = await fetch_live_prices_batch_v2(ib, list(all_tickers), timeout=30)
            simple_prices = {}
            for t, d in closing_prices.items():
                simple_prices[t] = d.get('live_price') if isinstance(d, dict) else d

            valid = sum(1 for p in simple_prices.values() if p and not pd.isna(p) and p > 0)
            print(f"  ✓ {valid}/{len(simple_prices)} valid prices")

            # 3. Load yesterday's closes for return calculation
            yesterday_closes = load_closing_prices(config.CLOSING_PRICES_FILE)
            if not yesterday_closes:
                yesterday_closes = {}

            # 4. Build index prices
            index_prices = {}
            for etf in config.INDEX_ETFS:
                curr = simple_prices.get(etf)
                prev = yesterday_closes.get(etf)
                if curr and prev:
                    index_prices[etf] = {'initial': prev, 'current': curr}

            print(f"  ✓ Index prices for {len(index_prices)} ETFs")

            # 5. Re-save portfolio with closing analytics
            print("\nSaving portfolio with closing prices...")
            success = await pm.save_portfolio_with_analytics(
                portfolio_df, options_portfolio_df,
                config.PORTFOLIO_FILE, ib,
                live_prices=simple_prices,
                index_prices=index_prices,
                opening_prices=yesterday_closes,
                yesterday_closes=yesterday_closes
            )

            if success:
                print("✓ Portfolio re-saved with closing prices")
            else:
                print("❌ Portfolio save failed")

        except Exception as e:
            logger.error(f"Post-close update error: {e}")
            import traceback
            traceback.print_exc()

    # 6. Disconnect IB (daily capture manages its own connection)
    if ib and ib.isConnected():
        print("\nDisconnecting IB (daily capture will reconnect)...")
        config.disconnect_ib(ib)

    # 7. Run daily data capture
    print("\n" + "=" * 80)
    print("DAILY DATA CAPTURE")
    print("=" * 80)

    try:
        from src.execution.daily_data_capture import run_daily_capture
        capture_result = await run_daily_capture(
            capture_closes=True,
            capture_mcaps=True,
            update_stops=True,
            run_analyst_archive_flag=True,
            analyst_max_tickers=None,
            fetch_earnings=True,
            verbose=True
        )
        print("\n✓ Daily data capture complete")
        return capture_result

    except Exception as e:
        logger.error(f"Daily data capture failed: {e}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================================
# MAIN WORKFLOW
# ============================================================================

async def run_portfolio_workflow():
    """
    Complete V9C portfolio management workflow.

    Returns:
    --------
    dict with keys:
        success : bool
        portfolio_df : DataFrame
        options_df : DataFrame
        ib : IB connection (still connected for post-close use)
        simple_prices : dict
        index_prices : dict
        index_prices_current : dict
        yesterday_closes : dict
        evaluated_trades_df : DataFrame
        terminated_df : DataFrame
        workflow_stages : dict (timing breakdown)
    """
    # Ensure module is initialized
    if not _initialized:
        initialize()

    # Lazy imports that depend on initialization
    from src.execution.portfolio_management import (
        check_for_existing_trades,
        save_completed_trades,
    )
    from src.execution.reconciliation import reconcile_with_tws, execute_remedies, format_reconciliation_summary
    from src.shared.fetch_market_data import fetch_live_prices_batch
    from src.execution.daily_data_capture import load_closing_prices
    from src.execution.trade_execution import execute_trades_in_batches, execute_trades_with_aggregation
    from src.execution.stop_loss_protection import (
        detect_stop_loss_orphans, close_orphaned_long, validate_stop_loss_config
    )
    from ib_insync import IB, Stock, MarketOrder, LimitOrder

    start_time = time.time()
    ib = None
    connected = False
    workflow_stages = {}

    # Track state for return value (safe defaults)
    portfolio_df = pd.DataFrame()
    options_portfolio_df = pd.DataFrame()
    simple_prices = {}
    index_prices = {}
    index_prices_current = {}
    yesterday_closes = {}
    evaluated_trades_df = pd.DataFrame()
    terminated_df = pd.DataFrame()
    index_price = None
    completed_trades_path = None

    print("=" * 80)
    print(f"PORTFOLIO MANAGEMENT WORKFLOW - Version {config.VERSION}")
    print("=" * 80)
    print("\nKey Features:")
    print("  ✓ Dynamic beta using CDF weights (W1)")
    print("  ✓ Tool box integration")
    print("  ✓ Sum Deviation CDF-based sizing")
    print("  ✓ SES removed from terminations")
    print("  ✓ Async batch market data fetching")
    print("  ✓ Portfolio reconciliation with TWS")

    if config.DRY_RUN_MODE:
        print("\n⚠️  DRY RUN MODE: Orders will be simulated, not executed")
        print("    Portfolio file will NOT be modified for terminations/new trades")

    print(f"\nAccount: {config.ACCOUNT_TYPE}")
    print(f"Base trade size: ${config.BASE_TRADE_SIZE:,.0f}")
    print(f"Portfolio beta limits: [{config.MIN_PORTFOLIO_BETA:.4f}, {config.MAX_PORTFOLIO_BETA:.4f}]")
    print(f"Target portfolio beta: {config.TARGET_PORTFOLIO_BETA:.4f}")
    print()

    # ========================================================================
    # LOAD TICKER BETAS
    # ========================================================================
    print("Loading individual ticker betas...")
    try:
        ticker_betas = pm.load_ticker_betas()
        print(f"  ✓ Loaded {len(ticker_betas)} ticker betas from SubSector_Beta_Analysis files")
    except Exception as e:
        logger.warning(f"Could not load ticker betas: {e}")
        print(f"  ⚠️  Ticker beta loading failed - will use defaults")

    # ========================================================================
    # STAGE 1: IBKR Connection
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(1, "IBKR Connection")

        ib, connected = await config.connect_ib_async(timeout=10, max_retries=3)

        if connected:
            logger.info(f"✓ Connected to IBKR - {time.time() - stage_time:.2f}s")
        else:
            logger.error("IBKR connection failed")
            ib = None
            if not config.DRY_RUN_MODE:
                proceed = input("\nContinue without IBKR? (y/n): ")
                if proceed.lower() != 'y':
                    return _build_result(False, locals())
            else:
                print("Continuing in DRY RUN mode without IBKR connection")

        workflow_stages['ibkr_connection'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"IBKR connection stage failed: {e}")
        connected = False
        ib = None
        if not config.DRY_RUN_MODE:
            proceed = input("\nContinue without IBKR? (y/n): ")
            if proceed.lower() != 'y':
                return _build_result(False, locals())
        else:
            print("Continuing in DRY RUN mode without IBKR connection")

    # ========================================================================
    # STAGE 2: Load Data
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(2, "Loading Data")

        (shortlist_df, parameters_df, portfolio_df, options_portfolio_df,
         _, completed_trades_path, earnings_dates) = pm.load_data()

        shortlist_df = normalize_shortlist_columns(shortlist_df)

        logger.info(f"✓ Portfolio: {len(portfolio_df)} positions")
        logger.info(f"✓ Options: {len(options_portfolio_df)} positions")
        logger.info(f"✓ Shortlist: {len(shortlist_df)} candidates")
        logger.info(f"✓ Earnings dates: {len(earnings_dates)} tickers")

        portfolio_df = pm.add_trade_dates(portfolio_df)
        workflow_stages['load_data'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Data loading failed: {e}")
        import traceback
        traceback.print_exc()
        return _build_result(False, locals())

    # ========================================================================
    # STAGE 2.5: Duplicate Tag Safeguard
    # ========================================================================
    try:
        print_stage_header("2.5", "Duplicate Tag Safeguard")

        has_duplicates, duplicate_details = check_for_duplicate_tags(portfolio_df)

        if has_duplicates:
            logger.error("❌ DUPLICATE TAGS DETECTED IN PORTFOLIO!")
            print("\n" + "=" * 80)
            print("🚨 CRITICAL: DUPLICATE TAGS FOUND IN PORTFOLIO")
            print("=" * 80)
            print(f"\nFound {len(duplicate_details['Tag'].unique())} duplicate Tags:")

            display_cols = ['Tag', 'Pair', 'Tail', 'Index', 'Trade_Date',
                            'Quantity1', 'Quantity2', 'Total_Count', 'Row_Index']
            print(duplicate_details[display_cols].to_string(index=False))

            print("\n⚠️  DO NOT CONTINUE - Risk of double-executing trades!")
            print("\nWorkflow PAUSED for safety.")
            user_input = input("\nType 'OVERRIDE' to continue anyway (NOT recommended): ")

            if user_input != 'OVERRIDE':
                logger.info("Workflow aborted by user - duplicate tags must be resolved")
                print("\n✓ Workflow safely aborted.")
                if ib and ib.isConnected():
                    config.disconnect_ib(ib)
                return _build_result(False, locals())
            else:
                logger.warning("⚠️  User OVERRIDE: Continuing despite duplicate tags")
        else:
            logger.info("✓ Duplicate tag check passed - all Tags are unique")

    except Exception as e:
        logger.error(f"Duplicate tag check failed: {e}")
        import traceback
        traceback.print_exc()
        user_input = input("Continue anyway? (y/n): ")
        if user_input.lower() != 'y':
            if ib and ib.isConnected():
                config.disconnect_ib(ib)
            return _build_result(False, locals())

    # ========================================================================
    # STAGE 3: Fetch Market Data (Async Batch)
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(3, "Fetching Market Data (Fast Snapshot Mode)")

        if connected and ib and ib.isConnected():
            # Collect all unique tickers
            all_tickers = set()
            if not portfolio_df.empty:
                all_tickers.update(portfolio_df['Co1'].tolist())
                all_tickers.update(portfolio_df['Co2'].tolist())
            if not shortlist_df.empty:
                all_tickers.update(shortlist_df['Co1'].tolist())
                all_tickers.update(shortlist_df['Co2'].tolist())

            all_tickers.update(['VO', 'IGV'])
            all_tickers.update(config.INDEX_ETFS)

            # Add megacap tickers
            from src.shared.calculations import get_megacap_tickers_for_fetch
            megacap_config = getattr(config, 'MEGACAP_ADJUSTMENT_CONFIG', {})
            if megacap_config.get('enabled', False):
                megacap_tickers = get_megacap_tickers_for_fetch()
                all_tickers.update(megacap_tickers)
                logger.info(f"Added {len(megacap_tickers)} megacap tickers for sector ETF adjustments")

            all_tickers = {t for t in all_tickers if t and pd.notna(t)}
            logger.info(f"Fetching prices for {len(all_tickers)} tickers...")

            # Fetch live prices (fast snapshot)
            from src.shared.fetch_market_data import fetch_live_prices_batch_v2
            live_prices = await fetch_live_prices_batch_v2(ib, list(all_tickers), timeout=30)

            yesterday_closes = load_closing_prices(config.CLOSING_PRICES_FILE)
            if not yesterday_closes:
                logger.warning("⚠️  No yesterday's closing prices available")
                yesterday_closes = {}

            # Fetch yesterday's close for megacaps if needed
            if megacap_config.get('enabled', False):
                megacap_tickers_list = list(get_megacap_tickers_for_fetch())
                missing_megacaps = [t for t in megacap_tickers_list if t not in yesterday_closes]

                if missing_megacaps:
                    logger.info(f"Fetching yesterday's close for {len(missing_megacaps)} megacaps")
                    for ticker in missing_megacaps:
                        try:
                            contract = Stock(ticker, 'SMART', 'USD')
                            qualified = ib.qualifyContracts(contract)
                            if qualified:
                                bars = ib.reqHistoricalData(
                                    qualified[0], endDateTime='', durationStr='2 D',
                                    barSizeSetting='1 day', whatToShow='TRADES', useRTH=True
                                )
                                if bars and len(bars) >= 2:
                                    yesterday_closes[ticker] = bars[-2].close
                                elif bars and len(bars) == 1:
                                    yesterday_closes[ticker] = bars[0].close
                        except Exception as e:
                            logger.warning(f"Could not fetch yesterday's close for {ticker}: {e}")

            # Get VO price
            index_price = live_prices.get('VO', {}).get('live_price')

            logger.info(f"✓ Fetched {len(live_prices)} prices")
            logger.info(f"  VO: ${index_price:.2f}" if index_price else "  VO: MISSING")

            # Create simple_prices format
            simple_prices = {}
            for ticker, data in live_prices.items():
                simple_prices[ticker] = data.get('live_price') if isinstance(data, dict) else data

            # Retry missing prices
            missing_price_tickers = [t for t, p in simple_prices.items()
                                     if p is None or pd.isna(p) or p <= 0]

            if missing_price_tickers:
                logger.warning(f"⚠️  {len(missing_price_tickers)} tickers have missing prices, retrying...")
                for ticker in missing_price_tickers:
                    try:
                        contract = Stock(ticker, 'SMART', 'USD')
                        qualified = ib.qualifyContracts(contract)
                        if qualified:
                            md = ib.reqMktData(qualified[0], '', False, False)
                            ib.sleep(3)
                            price = md.last
                            if not price or pd.isna(price) or price <= 0:
                                if md.bid and md.ask and md.bid > 0 and md.ask > 0:
                                    price = (md.bid + md.ask) / 2
                            if not price or pd.isna(price) or price <= 0:
                                if md.close and md.close > 0:
                                    price = md.close
                            ib.cancelMktData(qualified[0])
                            if price and not pd.isna(price) and price > 0:
                                simple_prices[ticker] = price
                                logger.info(f"   ✓ {ticker}: Retry successful, price=${price:.2f}")
                    except Exception as e:
                        logger.warning(f"   ✗ {ticker}: Retry error - {e}")

            valid_prices = sum(1 for p in simple_prices.values()
                               if p is not None and not pd.isna(p) and p > 0)
            logger.info(f"✓ {valid_prices}/{len(simple_prices)} valid prices")

            # Collect index prices for sector ETFs
            index_prices = {}
            for index_ticker in config.INDEX_ETFS:
                current_price = simple_prices.get(index_ticker)
                yesterday_price = yesterday_closes.get(index_ticker)
                if current_price and yesterday_price:
                    index_prices[index_ticker] = {
                        'initial': yesterday_price,
                        'current': current_price
                    }

            index_prices_current = {etf: data['current'] for etf, data in index_prices.items()}
            logger.info(f"✓ Index prices for {len(index_prices)} sector ETFs")

        else:
            logger.warning("No IBKR connection - using placeholder values")
            live_prices = {}
            index_price = config.VO_DEFAULT_PRICE

        workflow_stages['fetch_market_data'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Market data fetch failed: {e}")
        import traceback
        traceback.print_exc()
        live_prices = {}
        index_price = getattr(config, 'VO_DEFAULT_PRICE', 250)

    # ========================================================================
    # STAGE 4: Update Alpha Returns (Dynamic Beta)
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(4, "Updating Alpha Returns (Dynamic Beta with W1)")

        if not portfolio_df.empty:
            portfolio_df = pm.update_live_alpha_returns(
                portfolio_df, parameters_df,
                live_prices=simple_prices,
                fallback_to_previous=True
            )

            alpha_values = portfolio_df['Live Alpha Return (%)'].values
            valid_alphas = alpha_values[~np.isnan(alpha_values) & (alpha_values != 0)]

            if len(valid_alphas) > 0:
                logger.info(f"✓ Alpha statistics:")
                logger.info(f"  Valid: {len(valid_alphas)}/{len(alpha_values)}")
                logger.info(f"  Min: {np.min(valid_alphas):.4f}%  Max: {np.max(valid_alphas):.4f}%  Mean: {np.mean(valid_alphas):.4f}%")
        else:
            logger.info("Empty portfolio - no alphas to update")

        workflow_stages['update_alpha_returns'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Alpha update failed: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # STAGE 5: Evaluate Terminations (Date & Earnings Only)
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(5, "Evaluating Terminations (NO SES)")

        # Step 5.1: Stop Loss Orphan Detection
        if config.ENABLE_SHORT_SQUEEZE_PROTECTION and connected and ib and ib.isConnected():
            print("\n--- Step 5.1: Stop Loss Orphan Detection ---")

            orphaned_df, remaining_after_orphans_df = await detect_stop_loss_orphans(portfolio_df, ib)

            if not orphaned_df.empty:
                print(f"\n⚠️  STOP LOSS ORPHANS DETECTED: {len(orphaned_df)} positions")

                if config.DRY_RUN_MODE:
                    print("\n  [DRY RUN] Orphaned positions detected but will NOT be processed:")
                    for idx, orphan in orphaned_df.iterrows():
                        pair = orphan.get('Pair', 'Unknown')
                        tag = orphan.get('Tag')
                        tail = orphan.get('Tail', 'L').strip().upper()
                        long_ticker = orphan.get('Co1') if tail == 'L' else orphan.get('Co2')
                        long_qty = orphan.get('Quantity1', 0) if tail == 'L' else orphan.get('Quantity2', 0)
                        print(f"    - {pair} (Tag: {tag}): Would close {long_qty} {long_ticker}")
                    print("\n  ⚠️  Portfolio file will NOT be modified")
                else:
                    orphan_completed = []
                    for idx, orphan in orphaned_df.iterrows():
                        pair = orphan.get('Pair', 'Unknown')
                        tag = orphan.get('Tag')
                        print(f"\n  Closing orphaned long for {pair} (Tag: {tag})...")

                        success = await close_orphaned_long(ib, orphan, simple_prices)
                        if success:
                            print(f"    ✓ Closed orphaned long position")
                            orphan_completed.append(orphan.to_dict())
                        else:
                            print(f"    ❌ Failed to close orphaned long")

                    if orphan_completed:
                        portfolio_df = remaining_after_orphans_df

                        logger.info("Saving stop-loss orphans to daily terminated trades file...")
                        orphan_df_to_save = pd.DataFrame(orphan_completed)
                        if 'Exit Reason' not in orphan_df_to_save.columns:
                            orphan_df_to_save['Exit Reason'] = 'Stop Loss Triggered'

                        pm.save_daily_terminated_trades(orphan_df_to_save, simple_prices, yesterday_closes)
                        save_completed_trades(orphan_df_to_save, simple_prices,
                                              completed_trades_path, index_prices=index_prices_current)
            else:
                print("  ✓ No stop loss orphans detected")

        elif config.ENABLE_SHORT_SQUEEZE_PROTECTION:
            print("\n--- Step 5.1: Stop Loss Orphan Detection ---")
            print("  ⚠️  Skipped - no IBKR connection")

        # Step 5.2: Normal Termination Evaluation
        print("\n--- Step 5.2: Scheduled Terminations ---")

        if not portfolio_df.empty:
            portfolio_size_before = len(portfolio_df)

            to_terminate_df, remaining_df = pm.evaluate_trade_terminations(
                portfolio_df, parameters_df,
                earnings_dates=earnings_dates, ib=ib
            )

            terminated_df = to_terminate_df  # Store for return value
            terminated_count = len(to_terminate_df)
            logger.info(f"✓ Evaluated {portfolio_size_before} trades → {terminated_count} terminated")

            if terminated_count > 0:
                display_cols = [c for c in ['Pair', 'Tail', 'Index', 'Trade Date', 'Days Held', 'Exit Reason']
                                if c in to_terminate_df.columns]
                print(to_terminate_df[display_cols].to_string(index=False))

                if config.DRY_RUN_MODE:
                    logger.info("⚠️  DRY RUN MODE - Terminations NOT executed, portfolio unchanged")
                else:
                    if connected and ib and ib.isConnected():
                        from src.execution.trade_execution import execute_terminations

                        success = await execute_terminations(
                            to_terminate_df=to_terminate_df,
                            ib=ib, live_prices=simple_prices
                        )

                        if success:
                            logger.info(f"✓ Successfully executed {terminated_count} terminations")

                            pm.save_daily_terminated_trades(to_terminate_df, simple_prices, yesterday_closes)
                            save_completed_trades(to_terminate_df, simple_prices,
                                                  completed_trades_path, index_prices=index_prices_current)
                            portfolio_df = remaining_df
                        else:
                            logger.error("❌ Some terminations failed - keeping trades in portfolio")
                    else:
                        logger.error("❌ Cannot execute terminations - no IBKR connection")

                if not config.DRY_RUN_MODE:
                    save_interim_portfolio(portfolio_df, options_portfolio_df, "after_terminations")
        else:
            logger.info("Empty portfolio - no terminations to evaluate")

        workflow_stages['evaluate_terminations'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Termination evaluation failed: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # STAGE 6: Evaluate Trades (with Duplicate Check)
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(6, "Evaluating Trades (with Duplicate Check)")

        if not shortlist_df.empty:
            logger.info("Checking for duplicate trades in portfolio...")
            original_count = len(shortlist_df)
            shortlist_df = check_for_existing_trades(shortlist_df, portfolio_df)
            filtered_count = original_count - len(shortlist_df)

            if filtered_count > 0:
                logger.warning(f"⚠️  Filtered out {filtered_count} duplicate trades")

            if shortlist_df.empty:
                logger.info("No new trades to evaluate (all were duplicates)")
                evaluated_trades_df = pd.DataFrame()
            else:
                logger.info(f"Proceeding with {len(shortlist_df)} unique trades")

                # Evaluate remaining trades
                if not shortlist_df.empty:
                    evaluated_trades_df = await pm.evaluate_trades(
                        shortlist_df, parameters_df, portfolio_df,
                        ib=ib, live_prices=simple_prices,
                        index_prices=index_prices_current
                    )

                    logger.info(f"✓ Evaluated trades: {len(evaluated_trades_df)} approved")

                    if not evaluated_trades_df.empty:
                        display_cols = [c for c in ['Pair', 'Tail', 'Index', 'Quantity1', 'Quantity2',
                                                     'Position_Multiplier', 'Beta', 'sum_dev_bucket']
                                        if c in evaluated_trades_df.columns]
                        print(evaluated_trades_df[display_cols].to_string())
                else:
                    evaluated_trades_df = pd.DataFrame()
        else:
            logger.info("Empty shortlist - no trades to evaluate")
            evaluated_trades_df = pd.DataFrame()

        workflow_stages['evaluate_trades'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Trade evaluation failed: {e}")
        import traceback
        traceback.print_exc()
        evaluated_trades_df = pd.DataFrame()

    # ========================================================================
    # STAGE 7: Execute Approved Trades
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(7, "Executing Approved Trades (Batch Mode)")

        if not evaluated_trades_df.empty:
            logger.info(f"Executing {len(evaluated_trades_df)} approved trades...")

            if config.DRY_RUN_MODE:
                logger.info("⚠️  DRY RUN MODE - Trades NOT executed, portfolio unchanged")
                for _, trade in evaluated_trades_df.iterrows():
                    logger.info(f"      - {trade.get('Pair', 'Unknown')} ({trade.get('Tail', '?')}-tail, {trade.get('Index', '?')})")
            else:
                if connected and ib and ib.isConnected():
                    if config.ENABLE_ORDER_AGGREGATION:
                        logger.info("Using ORDER AGGREGATION")
                        execution_summary = await execute_trades_with_aggregation(
                            evaluated_trades_df,
                            ib, simple_prices, index_price
                        )
                    else:
                        logger.info("Using STANDARD EXECUTION")
                        execution_summary = await execute_trades_in_batches(
                            evaluated_trades_df,
                            ib, simple_prices, index_price
                        )

                    # Record successful trades to portfolio
                    if not execution_summary.empty:
                        successful = execution_summary[
                            execution_summary['Status'].isin(['Executed', 'Partial'])
                        ]
                        if not successful.empty:
                            successful_pairs = successful['Pair'].tolist()
                            executed_trades = evaluated_trades_df[
                                evaluated_trades_df['Pair'].isin(successful_pairs)
                            ].copy()
                            executed_trades = pm.add_trade_dates(executed_trades)
                            portfolio_df = pm.append_executed_trades(
                                portfolio_df, executed_trades, parameters_df
                            )
                        logger.info(f"Successfully executed: {len(successful)}/{len(evaluated_trades_df)} trades")
                else:
                    logger.error("❌ Cannot execute - no IBKR connection")

            logger.info(f"✓ Portfolio now has: {len(portfolio_df)} total positions")

            if not config.DRY_RUN_MODE:
                save_interim_portfolio(portfolio_df, options_portfolio_df, "after_new_trades")
        else:
            logger.info("No trades to execute")

        workflow_stages['execute_trades'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Trade execution failed: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # EXECUTION SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("TRADE EXECUTION SUMMARY")
    print("=" * 80)

    if not evaluated_trades_df.empty:
        total_approved = len(evaluated_trades_df)

        if config.DRY_RUN_MODE:
            print(f"\n📊 DRY RUN RESULTS:")
            print(f"  Approved for execution: {total_approved}")
            print(f"  ⚠️  NO TRADES EXECUTED (dry run mode)")
        else:
            new_tags = set(evaluated_trades_df['Tag'].astype(str))
            executed_tags = set(portfolio_df['Tag'].astype(str)) & new_tags
            successful = len(executed_tags)
            failed = total_approved - successful

            print(f"\n📊 EXECUTION RESULTS:")
            print(f"  Approved: {total_approved}")
            print(f"  ✓ Executed: {successful} ({successful / total_approved * 100:.1f}%)")
            print(f"  ✗ Failed: {failed} ({failed / total_approved * 100:.1f}%)")

            if successful > 0:
                executed_trades = evaluated_trades_df[evaluated_trades_df['Tag'].astype(str).isin(executed_tags)]
                print(f"\n✓ EXECUTED TRADES:")
                for _, trade in executed_trades.iterrows():
                    print(f"  {trade['Pair']:15s} ({trade['Tail']}-tail, {trade['Index']}): "
                          f"${trade['Total_Notional']:,.0f}, Beta={trade['Beta']:+.4f}")

                total_notional = executed_trades['Total_Notional'].sum()
                print(f"\n  Total deployed: ${total_notional:,.0f}")
                print(f"  L-tail: {sum(executed_trades['Tail'] == 'L')} | U-tail: {sum(executed_trades['Tail'] == 'U')}")
    else:
        print("\n⚠️  No trades were approved for execution")

    print("=" * 80)

    # Place initial stop orders for newly executed trades
    if config.ENABLE_SHORT_SQUEEZE_PROTECTION and not config.DRY_RUN_MODE:
        if connected and ib and ib.isConnected() and not evaluated_trades_df.empty:
            print("\n--- Placing initial stop orders ---")
            from src.execution.stop_loss_protection import update_stop_losses

            new_tags = set(evaluated_trades_df['Tag'].astype(str))
            new_trades_df = portfolio_df[portfolio_df['Tag'].astype(str).isin(new_tags)].copy()

            if not new_trades_df.empty:
                stop_result = await update_stop_losses(ib, new_trades_df, index_prices, verbose=True)
                logger.info(f"✓ Placed {stop_result['stops_placed']} initial stop orders")

                for idx, row in new_trades_df.iterrows():
                    tag = str(row['Tag'])
                    portfolio_idx = portfolio_df[portfolio_df['Tag'].astype(str) == tag].index
                    if len(portfolio_idx) > 0:
                        portfolio_df.loc[portfolio_idx[0], 'Stop_Order_ID'] = row.get('Stop_Order_ID')
                        portfolio_df.loc[portfolio_idx[0], 'Stop_Price'] = row.get('Stop_Price')

    # ========================================================================
    # STAGE 8: Portfolio Beta Check
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(8, "Portfolio Beta Check")

        if not portfolio_df.empty:
            total_value = 0
            weighted_beta = 0
            for _, trade in portfolio_df.iterrows():
                trade_value = abs(trade.get('Trade Value Co1 ($)', 0)) + abs(trade.get('Trade Value Co2 ($)', 0))
                trade_beta = trade.get('Beta', 0)
                total_value += trade_value
                weighted_beta += trade_beta * trade_value

            current_beta = weighted_beta / total_value if total_value > 0 else 0
            logger.info(f"Current portfolio beta: {current_beta:.4f}")

            if current_beta > config.MAX_PORTFOLIO_BETA:
                logger.warning(f"Portfolio beta {current_beta:.4f} exceeds max {config.MAX_PORTFOLIO_BETA:.4f}")
            elif current_beta < config.MIN_PORTFOLIO_BETA:
                logger.warning(f"Portfolio beta {current_beta:.4f} below min {config.MIN_PORTFOLIO_BETA:.4f}")
            else:
                logger.info(f"✓ Portfolio beta {current_beta:.4f} within limits")
        else:
            logger.info("Empty portfolio - no beta check needed")

        workflow_stages['beta_check'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Portfolio beta check stage failed: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # STAGE 9: Portfolio Reconciliation
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(9, "Portfolio Reconciliation")

        if connected and ib and ib.isConnected() and not config.DRY_RUN_MODE:
            reconciliation = reconcile_with_tws(portfolio_df, options_portfolio_df, ib, live_prices=simple_prices)

            total_discrepancies = sum(len(reconciliation[k]) for k in [
                "missing_in_tws", "missing_in_portfolio", "quantity_mismatch",
                "options_missing_in_tws", "options_missing_in_portfolio", "options_quantity_mismatch"
            ])

            if total_discrepancies == 0:
                print("\n✓ Portfolio and TWS fully reconciled!")
            else:
                print(f"\n⚠️ Found {total_discrepancies} discrepancies")
                format_reconciliation_summary(reconciliation, live_prices=simple_prices)

                if len(reconciliation["missing_in_portfolio"]) > 0 or len(reconciliation["quantity_mismatch"]) > 0:
                    user_input = input("\nExecute remedies? (y/n): ")
                    if user_input.lower() == 'y':
                        execute_remedies(reconciliation, ib)
        else:
            reason = "DRY RUN MODE" if config.DRY_RUN_MODE else "requires connection"
            print(f"Skipping portfolio reconciliation - {reason}")

        workflow_stages['portfolio_reconciliation'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Error in portfolio reconciliation: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # STAGE 10: Final Portfolio Save
    # ========================================================================
    try:
        stage_time = time.time()
        print_stage_header(10, "Saving Portfolio")

        if config.DRY_RUN_MODE:
            logger.info("⚠️  DRY RUN MODE - Skipping portfolio save")
            save_success = True
        else:
            save_success = await pm.save_portfolio_with_analytics(
                portfolio_df, options_portfolio_df,
                config.PORTFOLIO_FILE, ib,
                live_prices=simple_prices,
                index_prices=index_prices,
                opening_prices=yesterday_closes,
                yesterday_closes=yesterday_closes
            )

            if save_success:
                logger.info(f"✓ Portfolio saved to {config.PORTFOLIO_FILE}")
            else:
                logger.error("Portfolio save failed")

        workflow_stages['save_portfolio'] = time.time() - stage_time

    except Exception as e:
        logger.error(f"Portfolio save failed: {e}")
        import traceback
        traceback.print_exc()

    # ========================================================================
    # FINAL: Summary (IB stays connected for post-close use)
    # ========================================================================
    total_time = time.time() - start_time

    print("\n" + "=" * 80)
    print("WORKFLOW SUMMARY")
    print("=" * 80)

    if config.DRY_RUN_MODE:
        print("\n⚠️  DRY RUN MODE - No changes were made to portfolio or TWS")

    print(f"Total execution time: {total_time:.2f}s")
    print("\nTime breakdown by stage:")
    for stage, st in sorted(workflow_stages.items(), key=lambda x: x[1], reverse=True):
        pct = (st / total_time * 100)
        print(f"  {stage:.<30} {st:>6.2f}s ({pct:>5.1f}%)")

    print(f"\nCurrent portfolio:")
    print(f"  Trades: {len(portfolio_df)}")
    print(f"  Options: {len(options_portfolio_df)}")

    summary = pm.calculate_summary(portfolio_df, options_portfolio_df)
    print(f"  Total nominal: ${summary['Total Nominal Size']:,.2f}")
    print(f"  Portfolio beta: {summary['Portfolio Beta']:.4f}")
    print(f"  Net beta: {summary['Net Beta']:.4f}")

    # NOTE: IB stays connected for post-close use. Disconnect happens in
    # post_close_update() or must be done manually if skipping post-close.
    print("\n✓ IB connection kept alive for post-close update")
    print("  Call: await post_close_update(result)")
    print("  Or:   config.disconnect_ib(result['ib'])")

    print("\n" + "=" * 80)
    print("WORKFLOW COMPLETE")
    print("=" * 80)

    return _build_result(True, locals())


def _build_result(success, local_vars):
    """Build standardized result dict from workflow local variables."""
    return {
        'success': success,
        'portfolio_df': local_vars.get('portfolio_df', pd.DataFrame()),
        'options_df': local_vars.get('options_portfolio_df', pd.DataFrame()),
        'ib': local_vars.get('ib'),
        'connected': local_vars.get('connected', False),
        'simple_prices': local_vars.get('simple_prices', {}),
        'index_prices': local_vars.get('index_prices', {}),
        'index_prices_current': local_vars.get('index_prices_current', {}),
        'yesterday_closes': local_vars.get('yesterday_closes', {}),
        'evaluated_trades_df': local_vars.get('evaluated_trades_df', pd.DataFrame()),
        'terminated_df': local_vars.get('terminated_df', pd.DataFrame()),
        'workflow_stages': local_vars.get('workflow_stages', {}),
    }


# ============================================================================
# STANDALONE EXECUTION
# ============================================================================

if __name__ == "__main__":
    initialize()

    print("=" * 80)
    print(f"PORTFOLIO WORKFLOW - Version {config.VERSION}")
    print("=" * 80)
    print("\nCore functionality:")
    print("  ✓ Data loading + Market data fetching (async batch)")
    print("  ✓ Alpha calculations with dynamic beta")
    print("  ✓ Terminations (date & earnings)")
    print("  ✓ Trade evaluation with CDF sizing")
    print("  ✓ Trade execution + Portfolio reconciliation")
    print("  ✓ Post-close update + Daily data capture")
    print()

    if config.DRY_RUN_MODE:
        print("⚠️  DRY RUN MODE ENABLED")
        print()

    response = input("Run workflow? (y/n): ")
    if response.lower() == 'y':
        asyncio.run(run_portfolio_workflow())
    else:
        print("Workflow cancelled")
