"""
Portfolio management module with constraint enforcement and greedy selection.

Handles data loading, alpha return calculations (dynamic beta with W1 weights),
trade evaluation with multi-constraint optimization (leverage, beta limits,
ticker concentration, spread ceilings), trade terminations (date-based,
earnings-based, early exit), portfolio analytics, and position sizing.

STATUS: live (constraint enforcement and greedy selection in full;
        position sizing behind interface)
"""

import os
import re
import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Project imports
from src.shared import config
from src.shared import config_helper as Config_Helper
from src.shared import constraints as Constraints
from src.shared import calculations as Tool_Box
from src.shared.calculations import BetaDataManager, get_subsector_manager, load_subsector_indices

# NOTE: Trade_Execution is NOT imported at module level to avoid a circular
# dependency (TE previously imported PM). Functions that need TE utilities
# use lazy imports inside their function bodies. See
# docs/decisions/circular_dependency_fix.md for rationale.

logger = logging.getLogger(__name__)

# ============================================================================
# BETA MANAGER INITIALIZATION
# ============================================================================

beta_manager = BetaDataManager()
_beta_manager_initialized = False

def initialize_beta_manager(parameters_file=None):
    """
    Load beta coefficients once at startup
    
    Parameters:
    -----------
    parameters_file : str, optional
        Path to parameters file. If None, uses config.parameters_file()
    """
    global _beta_manager_initialized
    
    if _beta_manager_initialized:
        logger.debug("Beta manager already initialized, skipping")
        return
    
    if parameters_file is None:
        parameters_file = config.parameters_file()
    
    logger.info(f"Initializing beta manager from: {parameters_file}")
    
    # Load betas
    beta_manager.load_betas(parameters_file)
    
    # DEBUG: Check if betas actually loaded
    try:
        subsector_count = len(beta_manager._subsector_betas) if hasattr(beta_manager, '_subsector_betas') else 0
        treasury_count = len(beta_manager._treasury_betas) if hasattr(beta_manager, '_treasury_betas') else 0
        
        logger.info(f"Beta manager loaded: {subsector_count} subsector betas, {treasury_count} treasury betas")
        
        # Show sample
        if subsector_count > 0:
            sample = list(beta_manager._subsector_betas.items())[:3]
            logger.info(f"Sample subsector betas: {sample}")
        else:
            logger.warning("⚠️  No subsector betas loaded - check BetaDataManager.load_betas()")
            
    except Exception as e:
        logger.error(f"Error checking beta manager state: {e}")
    
    _beta_manager_initialized = True
    logger.info("Beta manager initialization complete")


# ============================================================================
# INDIVIDUAL TICKER BETA LOADING
# ============================================================================
# Loads actual ticker betas from SubSector_Beta_Analysis.xlsx files
# Used for accurate alpha calculations in portfolio monitoring

_ticker_betas_cache = {}  # {ticker: market_beta}
_ticker_betas_loaded = False

def load_ticker_betas():
    """
    Load individual ticker betas from SubSector_Beta_Analysis.xlsx files.
    
    Reads from: {config.beta_files_dir()}/{INDEX}/{INDEX}_SubSector_Beta_Analysis.xlsx
    Sheet: "Traditional Beta Summary"
    Columns: Ticker, market_beta
    
    Returns:
    --------
    dict : {ticker: market_beta}
    """
    global _ticker_betas_cache, _ticker_betas_loaded
    
    if _ticker_betas_loaded:
        logger.debug(f"Ticker betas already loaded ({len(_ticker_betas_cache)} tickers)")
        return _ticker_betas_cache
    
    logger.info("Loading individual ticker betas from SubSector_Beta_Analysis files...")
    
    # Get base directory from config
    if not hasattr(config, 'BETA_FILES_DIR'):
        logger.error("config.beta_files_dir() not defined - cannot load ticker betas")
        return {}
    
    beta_dir = config.beta_files_dir()
    
    if not os.path.exists(beta_dir):
        logger.error(f"Beta files directory not found: {beta_dir}")
        return {}
    
    # Load betas for each index
    indexes_loaded = 0
    tickers_loaded = 0
    errors = []
    
    for index_ticker in config.index_etfs():
        file_path = os.path.join(beta_dir, index_ticker, f"{index_ticker}_SubSector_Beta_Analysis.xlsx")
        
        if not os.path.exists(file_path):
            logger.warning(f"Beta file not found: {file_path}")
            continue
        
        try:
            # Read the Traditional Beta Summary sheet
            df = pd.read_excel(file_path, sheet_name='Traditional Beta Summary')
            
            # Find Ticker and market_beta columns
            ticker_col = None
            beta_col = None
            
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if col_lower == 'ticker':
                    ticker_col = col
                elif col_lower == 'market_beta':
                    beta_col = col
            
            if ticker_col is None or beta_col is None:
                logger.warning(f"Required columns not found in {file_path}. Found: {list(df.columns)}")
                continue
            
            # Load betas
            count_before = len(_ticker_betas_cache)
            for _, row in df.iterrows():
                ticker = row[ticker_col]
                beta = row[beta_col]
                
                if pd.notna(ticker) and pd.notna(beta):
                    ticker = str(ticker).strip().upper()
                    _ticker_betas_cache[ticker] = float(beta)
            
            count_added = len(_ticker_betas_cache) - count_before
            tickers_loaded += count_added
            indexes_loaded += 1
            logger.info(f"  {index_ticker}: Loaded {count_added} ticker betas")
            
        except Exception as e:
            errors.append(f"{index_ticker}: {e}")
            logger.error(f"Error loading betas from {file_path}: {e}")
    
    _ticker_betas_loaded = True
    
    logger.info(f"✓ Loaded {tickers_loaded} ticker betas from {indexes_loaded} index files")
    if errors:
        logger.warning(f"Errors encountered: {errors}")
    
    return _ticker_betas_cache


def get_ticker_beta(ticker, default=1.0):
    """
    Get the market beta for a specific ticker.
    
    Parameters:
    -----------
    ticker : str
        Stock ticker symbol
    default : float
        Default beta to return if ticker not found (default: 1.0)
    
    Returns:
    --------
    float : market beta
    """
    global _ticker_betas_cache, _ticker_betas_loaded
    
    # Ensure betas are loaded
    if not _ticker_betas_loaded:
        load_ticker_betas()
    
    ticker = str(ticker).strip().upper()
    
    if ticker in _ticker_betas_cache:
        return _ticker_betas_cache[ticker]
    else:
        logger.error(f"⚠️ Ticker beta not found: {ticker} - using default {default}")
        return default


# ============================================================================
# INDIVIDUAL TICKER BETA LOADING (imported from tool_box)
# ============================================================================

from src.shared.calculations import (
    load_ticker_betas_from_files as load_ticker_betas,
    get_single_ticker_beta as get_ticker_beta,
    get_all_ticker_betas as get_cached_ticker_betas,
    clear_ticker_betas_cache,
    calculate_megacap_adjusted_return,
    get_adjusted_index_return,
    calculate_spread_quality_score,
    calculate_sum_dev_extremity_score,
    calculate_composite_priority_score,
    categorize_ticker_beta,
    get_ticker_beta_buckets,
    categorize_pair_beta,
    get_pair_beta_buckets
)

# Backward compatibility aliases for beta categorization
def categorize_beta_detailed(beta):
    """Alias for categorize_pair_beta - kept for backward compatibility"""
    return categorize_pair_beta(beta)

def get_all_beta_buckets():
    """Alias for get_pair_beta_buckets - kept for backward compatibility"""
    return get_pair_beta_buckets()


# ============================================================================
# DATA LOADING
# ============================================================================

def load_data():
    """
    Load all required data files for portfolio management
    
    CRITICAL: Uses config paths ONLY - no hardcoding
    
    Returns:
        tuple: (shortlist_df, parameters_dict, portfolio_df, options_portfolio_df,
                available_liquidity, completed_trades_path, earnings_dates)
    """
    logger.info("Loading data files...")
    
    # Use config paths - NO HARDCODING
    files = {
        'portfolio': config.portfolio_file(),
        'shortlist': config.shortlist_file(),
        'parameters': config.parameters_file(),
        'completed_trades': config.completed_trades_file(),
        'earnings': config.earnings_calendar_file()
    }
    
    # Verify files exist
    for name, path in files.items():
        if not os.path.exists(path):
            raise FileNotFoundError(f"Required file not found: {name} at {path}")
    
    # Load parameters (all sheets)
    parameters_dict = pd.read_excel(files['parameters'], sheet_name=None)
    logger.info(f"Loaded parameters with sheets: {list(parameters_dict.keys())}")
    
    # Initialize beta manager
    initialize_beta_manager(files['parameters'])

    # Load portfolio
    try:
        excel_data = pd.ExcelFile(files['portfolio'])
        if 'Portfolio' in excel_data.sheet_names:
            portfolio_df = pd.read_excel(excel_data, sheet_name='Portfolio')
        else:
            portfolio_df = create_empty_portfolio()
            
        if 'Options' in excel_data.sheet_names:
            options_portfolio_df = pd.read_excel(excel_data, sheet_name='Options')
        else:
            options_portfolio_df = create_empty_options_portfolio()
            
    except Exception as e:
        logger.error(f"Error loading portfolio: {e}")
        portfolio_df = create_empty_portfolio()
        options_portfolio_df = create_empty_options_portfolio()
    
    # Load shortlist
    shortlist_df = pd.read_excel(files['shortlist'], sheet_name='Shortlist')
    logger.info(f"Loaded shortlist: {len(shortlist_df)} entries")
    
    # Load earnings dates
    earnings_dates = load_earnings_calendar(files['earnings'])
    
    # Get available liquidity
    # PLACEHOLDER: This should connect to IBKR
    available_liquidity = None
    logger.info("Available liquidity will be fetched from IBKR in workflow")
    
    return (shortlist_df, parameters_dict, portfolio_df, options_portfolio_df,
            available_liquidity, files['completed_trades'], earnings_dates)

def create_empty_portfolio():
    """Create empty portfolio with portfolio columns - CLEANED UP VERSION"""
    columns = [
        
        # Version tracking
        'Version',
        
        # Core identification
        'Tag', 'Pair', 'Co1', 'Co2', 'Index',
        
        # Position sizing
        'Quantity1', 'Quantity2',
        'Trade Value Co1 ($)', 'Trade Value Co2 ($)',
        
        # Entry data
        'Trade Initiation Date',
        'Co1 at Initiation', 'Co2 at Initiation',
        'Index at Initiation', 'Treasury at Initiation',
        'Entry_Spread_BPS',
        
        # Strategy configuration  
        'Tail', 'W1', 'W2', 'Position_Multiplier',
        'Sum_Dev_Bucket', 'Sum_Dev_CDF', 'Sum_Dev_Value',
        'Beta',
        
        # LAM signals - Raw values (matching LAM output names)
        'Weighted_Score', 'Index_Bias', 'Composite_Score',
        'Volume_Ratio', 'Rolling_Intraday_Vol', 'Volume_Dominance',
        'Last_Hour_Vol', 'IV_Percentile',
        
        # LAM signals - Percentiles (for comparison with rejected trades)
        'Volume_Ratio_Pct', 'Intraday_Vol_Pct', 'Volume_Dom_Pct',
        'Last_Hour_Pct', 'IV_Pct_Pct',

        # Exit data (for active trades)
        'Trade Termination Date',
        
        # Live analytics
        'Live Alpha Return (%)',
        
        # Stop loss tracking (NEW)
        'Stop_Order_ID',      # TWS order ID for stop order
        'Stop_Price',         # Current stop price level
    ]
    return pd.DataFrame(columns=columns)

def create_empty_completed_trades():
    """
    Create empty DataFrame with proper completed trades schema
    
    This defines the canonical columns for completed trades analysis.
    """
    columns = [
        # Core identifiers
        'Version', 'Tag', 'Pair', 'Co1', 'Co2', 'Index', 'Tail',
        
        # Dates
        'Trade Initiation Date', 'Trade Termination Date', 'Holding_Days', 'Exit_Reason',
        
        # Entry prices
        'Co1 at Initiation', 'Co2 at Initiation', 'Index at Initiation', 'Treasury at Initiation',
        
        # Exit prices  
        'Co1 at Exit', 'Co2 at Exit', 'Index at Exit', 'Treasury at Exit',
        
        # Position details
        'Quantity1', 'Quantity2', 'Trade Value Co1 ($)', 'Trade Value Co2 ($)', 'Total_Notional',
        
        # Strategy parameters
        'W1', 'W2', 'Beta', 'Position_Multiplier',
        'Sum_Dev_Value', 'Sum_Dev_CDF', 'Sum_Dev_Bucket',
        
        # Individual returns
        'Co1_Return_Pct', 'Co2_Return_Pct', 'Index_Return_Pct',
        
        # Individual alpha returns
        'Co1_Alpha_Pct', 'Co2_Alpha_Pct',
        
        # Trade alpha return
        'Final_Alpha_Return_Pct',
        
        # Entry quality metrics
        'Entry_Spread_BPS', 'Weighted_Score', 'Index_Bias', 'Composite_Score',
        
        # Volume metrics
        'Volume_Ratio', 'Rolling_Intraday_Vol', 'Volume_Dominance', 'IV_Percentile',
    ]
    return pd.DataFrame(columns=columns)

def check_for_existing_trades(shortlist_df, portfolio_df):
    """
    Check shortlist against portfolio to prevent duplicate trades
    """
    if portfolio_df.empty:
        shortlist_df['Existing'] = 0
        logger.info("Portfolio empty - no duplicates to check")
        return shortlist_df
    
    # Convert Tags to same type for comparison
    portfolio_tags = set(portfolio_df['Tag'].astype(str))
    
    # Mark duplicates
    shortlist_df['Existing'] = shortlist_df['Tag'].astype(str).isin(portfolio_tags).astype(int)
    
    # Report duplicates
    num_duplicates = shortlist_df['Existing'].sum()
    
    if num_duplicates > 0:
        print(f"⚠️  Found {num_duplicates} trades already in portfolio:")
        duplicates = shortlist_df[shortlist_df['Existing'] == 1]
        for _, dup in duplicates.iterrows():
            print(f"    {dup['Pair']} (Tag: {dup['Tag']})")
        print(f"⚠️  Filtered out {num_duplicates} duplicate trades")
    else:
        logger.info("✓ No duplicate trades found")
    
    # CRITICAL FIX: Return only non-duplicates!
    return shortlist_df[shortlist_df['Existing'] == 0]

def save_completed_trades(terminated_trades_df, live_prices, completed_trades_file, index_prices=None):
    """
    Save terminated trades to completed trades file with exit details
    
    Parameters:
    -----------
    terminated_trades_df : DataFrame
        Trades being terminated
    live_prices : dict
        Current market prices for exit prices
    completed_trades_file : str
        Path to completed trades file
    index_prices : dict, optional
        Current index prices {index: price} for calculating index returns
        e.g., {'VGT': 765.50, 'VIS': 310.20, ...}
    """
    try:
        # Load existing completed trades
        if os.path.exists(completed_trades_file):
            try:
                existing_df = pd.read_excel(completed_trades_file)
            except:
                existing_df = create_empty_completed_trades()
        else:
            existing_df = create_empty_completed_trades()
        
        # Define the columns we want to keep (clean schema)
        CLEAN_COLUMNS = [
            # Core identifiers
            'Version', 'Tag', 'Pair', 'Co1', 'Co2', 'Index', 'Tail',
            
            # Dates
            'Trade Initiation Date', 'Trade Termination Date', 'Holding_Days', 'Exit_Reason',
            
            # Entry prices
            'Co1 at Initiation', 'Co2 at Initiation', 'Index at Initiation', 'Treasury at Initiation',
            
            # Exit prices  
            'Co1 at Exit', 'Co2 at Exit', 'Index at Exit', 'Treasury at Exit',
            
            # Position details
            'Quantity1', 'Quantity2', 'Trade Value Co1 ($)', 'Trade Value Co2 ($)', 'Total_Notional',
            
            # Strategy parameters
            'W1', 'W2', 'Beta', 'Position_Multiplier',
            'Sum_Dev_Value', 'Sum_Dev_CDF', 'Sum_Dev_Bucket',
            
            # Individual returns
            'Co1_Return_Pct', 'Co2_Return_Pct', 'Index_Return_Pct',
            
            # Individual alpha returns
            'Co1_Alpha_Pct', 'Co2_Alpha_Pct',
            
            # Trade alpha return
            'Final_Alpha_Return_Pct',
            
            # Entry quality metrics
            'Entry_Spread_BPS', 'Weighted_Score', 'Index_Bias', 'Composite_Score',
            
            # Volume metrics
            'Volume_Ratio', 'Rolling_Intraday_Vol', 'Volume_Dominance', 'IV_Percentile',
        ]
        
        # Enrich terminated trades with exit data
        enriched_trades = []
        
        for _, trade in terminated_trades_df.iterrows():
            trade_dict = {}
            
            # Copy only the columns we want
            for col in CLEAN_COLUMNS:
                if col in trade.index:
                    trade_dict[col] = trade[col]
                else:
                    trade_dict[col] = np.nan
            
            # Handle Exit Reason (may have different column name)
            if pd.isna(trade_dict.get('Exit_Reason')):
                trade_dict['Exit_Reason'] = trade.get('Exit Reason', np.nan)
            
            # Add exit prices
            co1 = trade['Co1']
            co2 = trade['Co2']
            trade_dict['Co1 at Exit'] = live_prices.get(co1, np.nan)
            trade_dict['Co2 at Exit'] = live_prices.get(co2, np.nan)
            
            # Add index exit price
            trade_index = trade.get('Index', '')
            if index_prices and trade_index in index_prices:
                trade_dict['Index at Exit'] = index_prices[trade_index]
            else:
                # Try to get from live_prices
                trade_dict['Index at Exit'] = live_prices.get(trade_index, np.nan)
            
            # Calculate holding period
            if pd.notna(trade.get('Trade Initiation Date')) and pd.notna(trade.get('Trade Termination Date')):
                init_date = pd.to_datetime(trade['Trade Initiation Date'])
                term_date = pd.to_datetime(trade['Trade Termination Date'])
                trade_dict['Holding_Days'] = (term_date - init_date).days
            
            # Calculate Co1 return
            co1_init = trade.get('Co1 at Initiation')
            co1_exit = trade_dict.get('Co1 at Exit')
            if pd.notna(co1_exit) and pd.notna(co1_init) and co1_init != 0:
                trade_dict['Co1_Return_Pct'] = (co1_exit - co1_init) / co1_init * 100
            
            # Calculate Co2 return
            co2_init = trade.get('Co2 at Initiation')
            co2_exit = trade_dict.get('Co2 at Exit')
            if pd.notna(co2_exit) and pd.notna(co2_init) and co2_init != 0:
                trade_dict['Co2_Return_Pct'] = (co2_exit - co2_init) / co2_init * 100
            
            # Calculate Index return
            index_init = trade.get('Index at Initiation')
            index_exit = trade_dict.get('Index at Exit')
            if pd.notna(index_exit) and pd.notna(index_init) and index_init != 0:
                trade_dict['Index_Return_Pct'] = (index_exit - index_init) / index_init * 100
            
            # Get strategy parameters
            w1 = trade.get('W1', 0.5)
            w2 = trade.get('W2', 0.5)
            beta = trade.get('Beta', 0)
            tail = str(trade.get('Tail', 'L')).upper()
            
            co1_ret = trade_dict.get('Co1_Return_Pct', 0) or 0
            co2_ret = trade_dict.get('Co2_Return_Pct', 0) or 0
            index_ret = trade_dict.get('Index_Return_Pct', 0) or 0
            
            # Calculate individual alpha returns
            # Co Alpha = Co Return - Beta * Index Return (simplified - uses pair beta)
            # In reality, each stock has its own beta, but we approximate with pair beta
            if pd.notna(trade_dict.get('Co1_Return_Pct')) and pd.notna(index_ret):
                trade_dict['Co1_Alpha_Pct'] = co1_ret - beta * index_ret
            
            if pd.notna(trade_dict.get('Co2_Return_Pct')) and pd.notna(index_ret):
                trade_dict['Co2_Alpha_Pct'] = co2_ret - beta * index_ret
            
            # Calculate final alpha return
            # For L-tail: Alpha = W1*Co1_Ret - W2*Co2_Ret - Beta*Index_Ret
            # For U-tail: Alpha = W2*Co2_Ret - W1*Co1_Ret - Beta*Index_Ret
            if tail == 'L':
                nominal = w1 * co1_ret - w2 * co2_ret
            else:
                nominal = w2 * co2_ret - w1 * co1_ret
            
            alpha = nominal - beta * index_ret
            
            # Use calculated alpha, or fall back to Live Alpha Return if available
            if pd.notna(alpha) and abs(alpha) < 100:  # Sanity check
                trade_dict['Final_Alpha_Return_Pct'] = alpha
            elif pd.notna(trade.get('Live Alpha Return (%)')):
                trade_dict['Final_Alpha_Return_Pct'] = trade['Live Alpha Return (%)']
            
            enriched_trades.append(trade_dict)
        
        # Convert to DataFrame
        new_completed = pd.DataFrame(enriched_trades)
        
        # Ensure column order
        new_completed = new_completed.reindex(columns=CLEAN_COLUMNS)
        
        # Append to existing (also reindex existing to match)
        for col in CLEAN_COLUMNS:
            if col not in existing_df.columns:
                existing_df[col] = np.nan
        
        existing_df = existing_df.reindex(columns=CLEAN_COLUMNS)
        
        updated_completed = pd.concat([existing_df, new_completed], ignore_index=True)
        
        # Save
        updated_completed.to_excel(completed_trades_file, index=False)
        
        logger.info(f"✓ Saved {len(new_completed)} completed trades to {completed_trades_file}")
        logger.info(f"  Total completed trades: {len(updated_completed)}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error saving completed trades: {e}")
        import traceback
        traceback.print_exc()
        return False

def load_earnings_calendar(earnings_file):
    """
    Load earnings calendar with proper date/time handling
    
    Returns:
        dict: {ticker: effective_date} accounting for post-market timing
    """
    try:
        earnings_df = pd.read_excel(earnings_file)
        earnings_df.columns = earnings_df.columns.str.strip()
        
        earnings_dates = {}
        for _, row in earnings_df.iterrows():
            ticker = row['ticker'].upper().strip()
            report_date = pd.to_datetime(row['reportDate']).date()
            report_time = row.get('reportTime', 'pre-market')
            
            # Adjust for post-market reports
            if pd.notna(report_time) and isinstance(report_time, str):
                if 'post' in report_time.lower():
                    effective_date = report_date + timedelta(days=1)
                    # Skip weekends
                    while effective_date.weekday() >= 5:
                        effective_date += timedelta(days=1)
                    earnings_dates[ticker] = effective_date
                else:
                    earnings_dates[ticker] = report_date
            else:
                earnings_dates[ticker] = report_date
        
        logger.info(f"Loaded {len(earnings_dates)} earnings dates")
        return earnings_dates
        
    except Exception as e:
        logger.error(f"Error loading earnings calendar: {e}")
        return {}

# ============================================================================
# CONSTRAINT CHECKING HELPERS
# ============================================================================

# calculate_ticker_exposures and check_ticker_concentration_constraint
# have moved to Constraints.py.

# Scoring functions imported from tool_box:
# - calculate_spread_quality_score
# - calculate_sum_dev_extremity_score
# - calculate_composite_priority_score

# After calculate_composite_priority_score function (around line 395)
# Add these two functions:

def diagnose_execution_spread(ticker1, ticker2, live_prices, W1, W2, 
                              max_spread, pair_name="Unknown", ticker_bid_ask=None):
    """
    Detailed spread calculation diagnostic for execution stage
    
    Parameters:
    -----------
    ticker1, ticker2 : str
        Ticker symbols
    live_prices : dict
        Either simple prices {ticker: price} or detailed {ticker: {'bid': x, 'ask': y}}
    W1, W2 : float
        Leg weights
    max_spread : float
        Maximum allowed spread (decimal)
    pair_name : str
        Display name for the pair
    ticker_bid_ask : dict, optional
        Explicit bid/ask dict {ticker: {'bid': x, 'ask': y}} - preferred source
    """
    print(f"\n{'='*80}")
    print(f"EXECUTION SPREAD DIAGNOSTIC: {pair_name} ({ticker1}/{ticker2})")
    print(f"{'='*80}")
    
    # Try ticker_bid_ask first (most reliable source during evaluation)
    if ticker_bid_ask:
        price_data1 = ticker_bid_ask.get(ticker1, {})
        price_data2 = ticker_bid_ask.get(ticker2, {})
    else:
        # Fall back to live_prices
        price_data1 = live_prices.get(ticker1, {})
        price_data2 = live_prices.get(ticker2, {})
    
    # Ticker 1
    if isinstance(price_data1, dict):
        bid1 = price_data1.get('bid')
        ask1 = price_data1.get('ask')
        mid1 = price_data1.get('live_price') or price_data1.get('mid')
        if mid1 is None and bid1 is not None and ask1 is not None:
            mid1 = (bid1 + ask1) / 2
    else:
        bid1 = ask1 = mid1 = None
    
    print(f"\n{ticker1}:")
    print(f"  Bid:      ${bid1:.2f}" if bid1 else "  Bid:      MISSING")
    print(f"  Ask:      ${ask1:.2f}" if ask1 else "  Ask:      MISSING")
    print(f"  Mid:      ${mid1:.2f}" if mid1 else "  Mid:      MISSING")
    print(f"  Weight:   {W1:.4f} ({W1*100:.1f}%)")
    
    if bid1 and ask1 and mid1:
        spread1 = (ask1 - bid1) / mid1
        spread1_bps = spread1 * 10000
        weighted_spread1 = W1 * spread1
        weighted_spread1_bps = weighted_spread1 * 10000
        print(f"  Spread:   {spread1:.6f} ({spread1_bps:.1f} bps)")
        print(f"  Weighted: {weighted_spread1:.6f} ({weighted_spread1_bps:.1f} bps)")
    else:
        spread1 = None
        print(f"  Spread:   CANNOT CALCULATE")
    
    # Ticker 2
    if isinstance(price_data2, dict):
        bid2 = price_data2.get('bid')
        ask2 = price_data2.get('ask')
        mid2 = price_data2.get('live_price') or price_data2.get('mid')
        if mid2 is None and bid2 is not None and ask2 is not None:
            mid2 = (bid2 + ask2) / 2
    else:
        bid2 = ask2 = mid2 = None
    
    print(f"\n{ticker2}:")
    print(f"  Bid:      ${bid2:.2f}" if bid2 else "  Bid:      MISSING")
    print(f"  Ask:      ${ask2:.2f}" if ask2 else "  Ask:      MISSING")
    print(f"  Mid:      ${mid2:.2f}" if mid2 else "  Mid:      MISSING")
    print(f"  Weight:   {W2:.4f} ({W2*100:.1f}%)")
    
    if bid2 and ask2 and mid2:
        spread2 = (ask2 - bid2) / mid2
        spread2_bps = spread2 * 10000
        weighted_spread2 = W2 * spread2
        weighted_spread2_bps = weighted_spread2 * 10000
        print(f"  Spread:   {spread2:.6f} ({spread2_bps:.1f} bps)")
        print(f"  Weighted: {weighted_spread2:.6f} ({weighted_spread2_bps:.1f} bps)")
    else:
        spread2 = None
        print(f"  Spread:   CANNOT CALCULATE")
    
    # Pair spread (weighted)
    print(f"\nPAIR SPREAD CALCULATION (Execution):")
    print(f"  Method: Position-weighted by W1/W2")
    
    if spread1 is not None and spread2 is not None:
        weighted_pair_spread = (W1 * spread1) + (W2 * spread2)
        weighted_pair_spread_bps = weighted_pair_spread * 10000
        
        print(f"  Formula: ({W1:.4f} × {spread1:.6f}) + ({W2:.4f} × {spread2:.6f})")
        print(f"  Result:  {weighted_pair_spread:.6f} ({weighted_pair_spread_bps:.1f} bps)")
        
        # Compare to hurdle
        hurdle_bps = max_spread * 10000
        print(f"\n  Hurdle:  {max_spread:.6f} ({hurdle_bps:.1f} bps)")
        
        if weighted_pair_spread <= max_spread:
            margin = (max_spread - weighted_pair_spread) * 10000
            print(f"  Status:  ✓ PASS (spread {weighted_pair_spread_bps:.1f} bps <= hurdle {hurdle_bps:.1f} bps)")
            print(f"           Margin: {margin:.1f} bps below hurdle")
        else:
            excess = (weighted_pair_spread - max_spread) * 10000
            print(f"  Status:  ✗ FAIL (spread {weighted_pair_spread_bps:.1f} bps > hurdle {hurdle_bps:.1f} bps)")
            print(f"           Excess: {excess:.1f} bps above hurdle")
    else:
        print(f"  Result:  CANNOT CALCULATE (missing ticker data)")
        print(f"  Status:  ✗ FAIL (insufficient data)")
    
    print(f"{'='*80}\n")


def generate_spread_report(passed_pairs, rejected_pairs, pairs_missing_data=None):
    """
    Generate summary report of spread filtering results
    
    Parameters:
    -----------
    passed_pairs : list
        List of pairs that passed spread filter
    rejected_pairs : list
        List of rejected trades with spread info
    pairs_missing_data : list, optional
        List of pairs that were skipped due to missing market data
    """
    print(f"\n{'='*80}")
    print(f"SPREAD FILTER SUMMARY - EXECUTION STAGE")
    print(f"{'='*80}")
    
    # Separate spread-related rejections from other rejections
    spread_too_wide = [r for r in rejected_pairs if 'spread' in r.get('reason', '').lower() and 'too wide' in r.get('reason', '').lower()]
    
    # Count pairs missing data
    missing_data_count = len(pairs_missing_data) if pairs_missing_data else 0
    
    # Total includes: passed + spread too wide + missing data
    total_initial_candidates = len(passed_pairs) + len(spread_too_wide) + missing_data_count
    
    if total_initial_candidates == 0:
        print("No pairs evaluated for spread")
        return
    
    print(f"\nTotal pairs requiring spread data: {total_initial_candidates}")
    print(f"  ├─ Successfully fetched spread data: {len(passed_pairs) + len(spread_too_wide)}")
    print(f"  │    ├─ Passed spread filter: {len(passed_pairs)}")
    print(f"  │    └─ Failed (spread too wide): {len(spread_too_wide)}")
    if missing_data_count > 0:
        print(f"  └─ ⚠️  Missing market data: {missing_data_count} ({missing_data_count/total_initial_candidates*100:.1f}%)")
    else:
        print(f"  └─ Missing market data: 0")
    
    if passed_pairs:
        spreads = [p.get('weighted_spread', 0) * 10000 for p in passed_pairs if 'weighted_spread' in p]
        if spreads:
            print(f"\nPassed pairs - Spread distribution (bps):")
            print(f"  Min:    {min(spreads):.1f}")
            print(f"  Mean:   {np.mean(spreads):.1f}")
            print(f"  Median: {np.median(spreads):.1f}")
            print(f"  Max:    {max(spreads):.1f}")
    
    if spread_too_wide:
        print(f"\nSpread too wide - Examples:")
        for i, failure in enumerate(spread_too_wide[:5], 1):
            print(f"  {i}. {failure['pair']}: {failure['reason']}")
        if len(spread_too_wide) > 5:
            print(f"  ... and {len(spread_too_wide) - 5} more")
    
    if pairs_missing_data and missing_data_count > 0:
        print(f"\nMissing market data - Examples:")
        for i, missing in enumerate(pairs_missing_data[:5], 1):
            print(f"  {i}. {missing['pair']}: {missing['reason']}")
        if missing_data_count > 5:
            print(f"  ... and {missing_data_count - 5} more")
    
    print(f"{'='*80}\n")

# ============================================================================
# PORTFOLIO DATE MANAGEMENT
# ============================================================================

def add_trade_dates(portfolio_df):
    """Add/update trade dates and initialize tracking columns"""
    if portfolio_df.empty:
        return portfolio_df
    
    today = datetime.today().date()
    
    # Ensure date columns exist
    if 'Trade Initiation Date' not in portfolio_df.columns:
        portfolio_df['Trade Initiation Date'] = today
    else:
        portfolio_df['Trade Initiation Date'] = pd.to_datetime(
            portfolio_df['Trade Initiation Date'], errors='coerce'
        ).dt.date.fillna(today)
    
    if 'Trade Termination Date' not in portfolio_df.columns:
        portfolio_df['Trade Termination Date'] = today + timedelta(days=config.max_holding_days())
    else:
        portfolio_df['Trade Termination Date'] = pd.to_datetime(
            portfolio_df['Trade Termination Date'], errors='coerce'
        ).dt.date.fillna(today + timedelta(days=config.max_holding_days()))
    
    # Initialize tracking columns if missing
    tracking_columns = [
        'Index Return', 'Co1 Return (%)', 'Co2 Return (%)', 'Exit Reason'
    ]
    
    for col in tracking_columns:
        if col not in portfolio_df.columns:
            portfolio_df[col] = np.nan
    
    logger.info("Updated trade dates and tracking columns")
    return portfolio_df

# ============================================================================
# ALPHA CALCULATIONS - USING TOOL_BOX WITH W1
# ============================================================================

def update_live_alpha_returns(portfolio_df, parameters_df, live_prices=None,
                              current_dgs10=None, fallback_to_previous=True,
                              index_prices=None):
    """
    Update alpha returns using tool_box calculations with dynamic beta
    
    V9.2 UPDATE: Handles both V9 (two-factor) and V9.2 (single-factor) trades.
    - V9 trades: Use VO index + DGS10 treasury with market betas
    - V9.2 trades: Use parent ETF index with sub-sector betas (no treasury)
    
    CRITICAL: Uses W1 from portfolio records for proper weighting
    
    Parameters:
    -----------
    index_prices : dict, optional
        Current ETF prices {etf_ticker: price} for V9.2 calculations
    """
    if portfolio_df.empty:
        logger.info("Empty portfolio - no alpha updates needed")
        return portfolio_df
    
    logger.info("Updating live alpha returns...")
    
    # PLACEHOLDER: Fetch prices if not provided
    if live_prices is None:
        logger.warning("No live prices provided - using placeholder")
        live_prices = {}
    
    if index_prices is None:
        index_prices = {}
    
    # V9 fallbacks
    if current_dgs10 is None:
        current_dgs10 = config.dgs10_default_yield()
    
    current_vo_price = live_prices.get('VO', config.vo_default_price())
    
    # Load V9.2 sub-sector manager if any V9.2/V9.3 trades exist
    subsector_manager = None
    # Check for V9.2/V9.3 trades - handle both string and numeric Version values
    has_v92_trades = False
    if 'Version' in portfolio_df.columns:
        versions_as_str = portfolio_df['Version'].astype(str)
        has_v92_trades = versions_as_str.isin(['V9.2', '9.2', 'V9.3', '9.3']).any()
    
    if has_v92_trades:
        subsector_manager = get_subsector_manager()
        # Load all ETFs if not already loaded
        load_subsector_indices()
        logger.info("Loaded sub-sector manager for V9.2 trades")
    
    success_count = 0
    v9_count = 0
    v92_count = 0
    missing_data_count = 0
    
    for index, row in portfolio_df.iterrows():
        try:
            # Get trade details
            ticker1 = row['Co1']
            ticker2 = row['Co2']
            tail = row.get('Tail', 'L').strip().upper()
            trade_version = str(row.get('Version', 'V9'))  # Default to V9 for legacy trades
            trade_index = row.get('Index', 'VGT')  # Parent ETF for the trade
            
            # Get initiation prices
            initial_price_co1 = row['Co1 at Initiation']
            initial_price_co2 = row['Co2 at Initiation']
            initial_index = row['Index at Initiation']
            
            # Get current prices
            current_price_co1 = live_prices.get(ticker1)
            current_price_co2 = live_prices.get(ticker2)
            
            if pd.isna(current_price_co1) or pd.isna(current_price_co2):
                logger.warning(f"Missing prices for {ticker1}/{ticker2}")
                portfolio_df.at[index, 'Live Alpha Return (%)'] = 0.0
                missing_data_count += 1
                continue
            
            # Get W1 weight from portfolio (CRITICAL for both V9 and V9.2)
            W1 = row.get('W1', 0.5)
            if pd.isna(W1):
                logger.warning(f"Missing W1 for {row.get('Pair')}, using 0.5")
                W1 = 0.5
            W2 = 1 - W1
            
            # Determine direction from tail
            if tail == 'L':
                direction_co1 = 1
                direction_co2 = -1
            else:  # U
                direction_co1 = -1
                direction_co2 = 1
            
            # ================================================================
            # V9.2/V9.3: Single-factor model with market betas (stock vs parent ETF)
            # ================================================================
            if trade_version in ['V9.2', '9.2', 'V9.3', '9.3']:
                # Get current parent ETF price
                # index_prices structure: {etf: {'initial': price, 'current': price}} or {etf: price}
                etf_data = index_prices.get(trade_index)
                if isinstance(etf_data, dict):
                    current_index_price = etf_data.get('current')
                else:
                    current_index_price = etf_data or live_prices.get(trade_index)
                
                if current_index_price is None or pd.isna(current_index_price):
                    logger.warning(f"Missing {trade_index} price for {trade_version} trade {ticker1}/{ticker2}")
                    portfolio_df.at[index, 'Live Alpha Return (%)'] = 0.0
                    missing_data_count += 1
                    continue
                
                # Get market betas (stock vs parent ETF) for interpretable alpha
                beta_co1 = Tool_Box.get_single_ticker_beta(ticker1, fallback=1.0)
                beta_co2 = Tool_Box.get_single_ticker_beta(ticker2, fallback=1.0)
                
                # V9.2/V9.3 alpha calculation: single-factor with parent ETF
                # Alpha = weighted_stock_return - weighted_beta * index_return
                alpha = Tool_Box.calculate_live_alpha_return_v92(
                    ticker1, ticker2,
                    initial_price_co1, initial_price_co2,
                    current_price_co1, current_price_co2,
                    initial_index, current_index_price,  # Use parent ETF for both
                    initial_index, current_index_price,  # Same index for both tickers
                    beta_co1, beta_co2,
                    W1, W2,
                    direction_co1, direction_co2
                )
                
                v92_count += 1
            
            # ================================================================
            # V9: Two-factor model with VO index + treasury
            # ================================================================
            else:
                # Handle FVX -> DGS10 transition
                initial_dgs10 = row.get('Treasury at Initiation')
                if pd.isna(initial_dgs10):
                    initial_dgs10 = row.get('FVX at Initiation', config.dgs10_default_yield())
                
                # Get V9 betas
                beta_vo_co1 = beta_manager.get_vgt_beta(ticker1)
                beta_treasury_co1 = beta_manager.get_treasury_beta(ticker1)
                beta_vo_co2 = beta_manager.get_vgt_beta(ticker2)
                beta_treasury_co2 = beta_manager.get_treasury_beta(ticker2)
                
                # V9 alpha calculation: two-factor model
                alpha = Tool_Box.calculate_live_alpha_return(
                    ticker1, ticker2,
                    initial_price_co1, initial_price_co2,
                    current_price_co1, current_price_co2,
                    initial_index, current_vo_price,
                    initial_dgs10, current_dgs10,
                    beta_vo_co1, beta_treasury_co1,
                    beta_vo_co2, beta_treasury_co2,
                    W1, W2,
                    direction_co1, direction_co2
                )
                
                v9_count += 1
            
            portfolio_df.at[index, 'Live Alpha Return (%)'] = alpha
            success_count += 1
            
        except Exception as e:
            logger.error(f"Error calculating alpha for row {index}: {e}")
            portfolio_df.at[index, 'Live Alpha Return (%)'] = 0.0
    
    logger.info(f"Updated {success_count} alphas (V9: {v9_count}, V9.2: {v92_count}), {missing_data_count} missing data")
    return portfolio_df

async def evaluate_trades(shortlist_df, parameters_df, portfolio_df,
                         ib=None, live_prices=None,
                         index_prices=None, dgs10_price=None):
    # DIAGNOSTIC - remove after debugging
    print(f"\n🔍 SHORTLIST DIAGNOSTIC:")
    print(f"   Columns: {shortlist_df.columns.tolist()}")
    print(f"   First row Index_Bias: {shortlist_df['Index_Bias'].iloc[0] if 'Index_Bias' in shortlist_df.columns else 'MISSING'}")
    print(f"   First row Rolling_Intraday_Vol: {shortlist_df['Rolling_Intraday_Vol'].iloc[0] if 'Rolling_Intraday_Vol' in shortlist_df.columns else 'MISSING'}")
    print(f"   First row IV_Percentile: {shortlist_df['IV_Percentile'].iloc[0] if 'IV_Percentile' in shortlist_df.columns else 'MISSING'}")
    """
    Evaluate trades with multi-constraint optimization
    
    NEW V9C FEATURES:
    - Composite priority scoring (weighted_score + spread + sum_dev)
    - Greedy selection respecting ALL constraints simultaneously:
      * Liquidity limits
      * Portfolio beta limits (min/max)
      * Ticker concentration limits (long/short separate)
      * Spread ceiling
    - Prioritizes best trades first (by composite score)
    
    Parameters:
    -----------
    index_prices : dict, optional
        Dict of sector ETF prices {etf: price}
        e.g., {'VGT': 762.50, 'VIS': 310.20, 'VHT': 294.35, 'VCR': 385.10}
    """
    from src.execution.trade_execution import calculate_weighted_spread

    logger.info("=" * 80)
    logger.info("EVALUATING TRADES (MULTI-CONSTRAINT OPTIMIZATION)")
    logger.info("=" * 80)

    if shortlist_df.empty:
        logger.info("Empty shortlist - no trades to evaluate")
        return pd.DataFrame()
    
    if live_prices is None:
        live_prices = {}
    
    if index_prices is None:
        index_prices = {}
    
    # ========================================================================
    # PHASE 1: Read Sum Deviation from Shortlist
    # ========================================================================
    
    logger.info("\nPhase 1: Reading sum deviation from shortlist...")
    
    required_cols = ['Sum_Dev_Bucket', 'Sum_Dev_CDF', 'Sum_Dev_Value']
    missing_cols = [col for col in required_cols if col not in shortlist_df.columns]
    
    if missing_cols:
        logger.error(f"Missing required columns from LAM: {missing_cols}")
        return pd.DataFrame()
    
    shortlist_df = shortlist_df.copy()
    shortlist_df['is_tradeable'] = False
    shortlist_df['trigger_type'] = shortlist_df['Tail'].map({'L': 'lower', 'U': 'upper'})
    
    for idx, row in shortlist_df.iterrows():
        bucket = row['Sum_Dev_Bucket']
        trigger_type = row['trigger_type']
        
        if pd.notna(bucket):
            is_tradeable = Constraints.is_tradeable_bucket(trigger_type, bucket)
            shortlist_df.at[idx, 'is_tradeable'] = is_tradeable
    
    total_pairs = len(shortlist_df)
    tradeable = shortlist_df['is_tradeable'].sum()
    
    logger.info(f"  Total pairs: {total_pairs}")
    logger.info(f"  Tradeable: {tradeable}")
    
    # ========================================================================
    # PHASE 2: Filter to Tradeable Pairs
    # ========================================================================
    
    logger.info("\nPhase 2: Filtering to tradeable pairs...")
    
    tradeable_shortlist = shortlist_df[shortlist_df['is_tradeable']].copy()
    
    if tradeable_shortlist.empty:
        logger.warning("No tradeable pairs after sum dev filtering")
        return pd.DataFrame()
    
    # ========================================================================
    # PHASE 3: Load Trading Parameters
    # ========================================================================
    
    logger.info("\nPhase 3: Loading trading parameters...")
    
    base_trade_size = config.base_trade_size()
    
    logger.info(f"  Base trade size: ${base_trade_size:,.2f}")
    logger.info(f"  Portfolio beta limits: [{config.min_portfolio_beta():.4f}, {config.max_portfolio_beta():.4f}]")
    logger.info(f"  Max long ticker concentration: {config.max_long_ticker_concentration():.1%}")
    logger.info(f"  Max short ticker concentration: {config.max_short_ticker_concentration():.1%}")
    logger.info(f"  Max spread: {config.max_spread_bps()} bps")
    
    # ========================================================================
    # Load Market Caps for VHT Small Cap Constraint
    # ========================================================================
    market_caps = {}
    if hasattr(config, 'VHT_SMALL_CAP_CONSTRAINT') and config.VHT_SMALL_CAP_CONSTRAINT.get('enabled', False):
        try:
            from src.execution.daily_data_capture import load_market_caps
            market_caps = load_market_caps()
            vht_threshold = config.VHT_SMALL_CAP_CONSTRAINT.get('mcap_threshold_millions', 5000)
            vht_adjustment = config.VHT_SMALL_CAP_CONSTRAINT.get('concentration_reduction', 0.005)
            logger.info(f"  VHT small cap constraint: MCap < ${vht_threshold}M → concentration limit reduced by {vht_adjustment:.1%}")
            logger.info(f"    Loaded {len(market_caps)} market caps for constraint checking")
        except Exception as e:
            logger.warning(f"  Could not load market caps for VHT constraint: {e}")
    
    # Calculate total portfolio value
    if not portfolio_df.empty:
        try:
            total_portfolio_value = portfolio_df[['Trade Value Co1 ($)', 'Trade Value Co2 ($)']].abs().sum().sum()
        except:
            total_portfolio_value = base_trade_size * 50  # Conservative estimate
    else:
        total_portfolio_value = base_trade_size * 50  # Conservative estimate for empty portfolio
    
    logger.info(f"  Current portfolio value: ${total_portfolio_value:,.2f}")
    
    # ========================================================================
    # Calculate Leverage Capacity
    # ========================================================================
    logger.info(f"  Current portfolio value: ${total_portfolio_value:,.2f}")
    
    if ib and ib.isConnected():
        from src.execution.trade_execution import get_account_summary_values
        account_values = get_account_summary_values(ib)
        account_currency = account_values.get('Currency', 'USD')
        
        # Convert to USD if account is in GBP
        if account_currency == 'GBP':
            gbp_usd_rate = Config_Helper.get_gbp_usd_rate(ib)
            net_liq_usd = account_values['NetLiquidation'] * gbp_usd_rate
            current_gross_usd = account_values['GrossPositionValue'] * gbp_usd_rate
        else:
            net_liq_usd = account_values['NetLiquidation']
            current_gross_usd = account_values['GrossPositionValue']
        
        # Calculate maximum allowed gross exposure
        max_gross_usd = net_liq_usd * config.max_account_leverage()
        
        # Available capacity (with 95% safety buffer)
        leverage_capacity_usd = (max_gross_usd - current_gross_usd) * 0.95
        leverage_capacity_usd = max(0, leverage_capacity_usd)
        
        # Calculate leverage from portfolio data (matches evaluate_trades logic)
        if not portfolio_df.empty:
            total_notional = 0
            for _, row in portfolio_df.iterrows():
                co1_value = abs(row.get('Trade Value Co1 ($)', 0) or 0)
                co2_value = abs(row.get('Trade Value Co2 ($)', 0) or 0)
                total_notional += co1_value + co2_value

            # Use USD-converted net liquidation for leverage calc
            # Get net liquidation for leverage calculation
            net_liquidation = account_values.get('NetLiquidation', 0)
            currency = account_values.get('Currency', 'USD')
            if currency == 'GBP':
                gbp_usd_rate = Config_Helper.get_gbp_usd_rate(ib)
                net_liq_for_leverage = net_liquidation * gbp_usd_rate
            else:
                net_liq_for_leverage = net_liquidation

            account_values['Leverage'] = total_notional / net_liq_for_leverage if net_liq_for_leverage > 0 else 0
            current_leverage = account_values['Leverage']
        else:
            account_values['Leverage'] = 0
            current_leverage = 0

        logger.info(f"\n  Leverage status ({account_currency} account):")
        logger.info(f"    Net liquidation: ${net_liq_usd:,.0f} USD")
        logger.info(f"    Current gross: ${current_gross_usd:,.0f} USD")
        logger.info(f"    Current leverage: {current_leverage:.2f}x")
        logger.info(f"    Max leverage: {config.max_account_leverage():.2f}x")
        logger.info(f"    Available capacity: ${leverage_capacity_usd:,.0f} USD")
    else:
        # Fallback if no connection (dry run mode)
        # Use a generous default - IBKR will reject if insufficient in live mode
        leverage_capacity_usd = base_trade_size * 100
        current_gross_usd = total_portfolio_value
        net_liq_usd = total_portfolio_value  # Needed for projected leverage calc
        logger.warning("  No IBKR connection - using generous capacity default (dry run)")
    
    # ========================================================================
    # PHASE 4: Fetch Live Prices and Calculate Spreads
    # ========================================================================
    
    logger.info("\nPhase 4: Fetching live prices and calculating spreads...")
    
    # Get unique tickers
    all_tickers = set(tradeable_shortlist['Co1'].tolist() + 
                     tradeable_shortlist['Co2'].tolist())
    
    logger.info(f"  Fetching prices for {len(all_tickers)} tickers...")
    
    # Fetch bid/ask for all tickers with retry logic
    ticker_bid_ask = {}
    failed_tickers = []
    
    close_ib = False
    if ib is None:
        from src.execution.trade_execution import connect_ibkr
        ib = connect_ibkr()
        close_ib = True

    from src.execution.trade_execution import fetch_market_data
    try:
        # First pass - standard fetch
        for ticker in all_tickers:
            bid, ask = fetch_market_data(ib, ticker)
            if bid is not None and ask is not None and bid > 0 and ask > 0:
                ticker_bid_ask[ticker] = {'bid': bid, 'ask': ask}
                mid = (bid + ask) / 2
                live_prices[ticker] = mid
            else:
                failed_tickers.append(ticker)
        
        # Retry pass for failed tickers with longer timeout
        if failed_tickers:
            logger.info(f"  Retrying {len(failed_tickers)} tickers with extended timeout...")
            from ib_insync import Stock
            
            retry_success = 0
            for ticker in failed_tickers:
                try:
                    contract = Stock(ticker, 'SMART', 'USD')
                    qualified = ib.qualifyContracts(contract)
                    
                    if qualified:
                        # Request market data with longer wait
                        md = ib.reqMktData(qualified[0], '', False, False)
                        ib.sleep(2.0)  # Extended wait for data
                        
                        bid = md.bid if md.bid and md.bid > 0 else None
                        ask = md.ask if md.ask and md.ask > 0 else None
                        
                        # If still no bid/ask, try one more sleep
                        if not bid or not ask:
                            ib.sleep(1.5)  # Additional wait
                            bid = md.bid if md.bid and md.bid > 0 else None
                            ask = md.ask if md.ask and md.ask > 0 else None
                        
                        ib.cancelMktData(qualified[0])
                        
                        if bid and ask and bid > 0 and ask > 0:
                            ticker_bid_ask[ticker] = {'bid': bid, 'ask': ask}
                            mid = (bid + ask) / 2
                            live_prices[ticker] = mid
                            retry_success += 1
                            logger.debug(f"    ✓ Retry success: {ticker} bid=${bid:.2f} ask=${ask:.2f}")
                        else:
                            logger.warning(f"    ✗ Retry failed: {ticker} (bid={bid}, ask={ask})")
                except Exception as e:
                    logger.warning(f"    ✗ Retry error for {ticker}: {e}")
            
            if retry_success > 0:
                logger.info(f"    Recovered {retry_success}/{len(failed_tickers)} tickers on retry")
            
            # Update failed_tickers list to only those that truly failed
            failed_tickers = [t for t in failed_tickers if t not in ticker_bid_ask]
            
            if failed_tickers:
                logger.warning(f"  ⚠️  Still missing data for {len(failed_tickers)} tickers: {failed_tickers}")
    finally:
        if close_ib:
            ib.disconnect()
    
    logger.info(f"  Successfully fetched {len(ticker_bid_ask)}/{len(all_tickers)} prices")
    
    # Calculate spreads for all candidates
    candidates_with_spreads = []
    pairs_missing_data = []  # Track pairs skipped due to missing market data
    
    for idx, row in tradeable_shortlist.iterrows():
        ticker1 = row['Co1']
        ticker2 = row['Co2']
        trigger_type = row['trigger_type']
        bucket = row['Sum_Dev_Bucket']
        
        # Get strategy weights
        W1, W2 = Constraints.get_leg_weights(trigger_type, bucket)
        
        # Get bid/ask
        if ticker1 not in ticker_bid_ask or ticker2 not in ticker_bid_ask:
            missing_tickers = []
            if ticker1 not in ticker_bid_ask:
                missing_tickers.append(ticker1)
            if ticker2 not in ticker_bid_ask:
                missing_tickers.append(ticker2)
            pairs_missing_data.append({
                'pair': row['Pair'],
                'missing_tickers': missing_tickers,
                'reason': f"Missing market data for: {', '.join(missing_tickers)}"
            })
            logger.warning(f"  Skipping {row['Pair']}: missing market data for {', '.join(missing_tickers)}")
            continue
        
        bid1 = ticker_bid_ask[ticker1]['bid']
        ask1 = ticker_bid_ask[ticker1]['ask']
        bid2 = ticker_bid_ask[ticker2]['bid']
        ask2 = ticker_bid_ask[ticker2]['ask']
        
        # Calculate weighted spread
        weighted_spread = calculate_weighted_spread(bid1, ask1, bid2, ask2, W1, W2)
        
        # Add to candidate
        candidate = row.to_dict()
        
        # CRITICAL: Ensure Tag is preserved (it's the unique pair ID from Parameters)
        if 'Tag' not in candidate or pd.isna(candidate['Tag']):
            logger.error(f"  ⚠️  Missing Tag for {row['Pair']}!")
            candidate['Tag'] = f"MISSING_TAG_{row['Pair']}"
        
        candidate['weighted_spread'] = weighted_spread
        candidate['weighted_spread_bps'] = weighted_spread * 10000
        candidate['W1'] = W1
        candidate['W2'] = W2
        candidates_with_spreads.append(candidate)
    
    logger.info(f"  Calculated spreads for {len(candidates_with_spreads)} candidates")
    if pairs_missing_data:
        logger.warning(f"  ⚠️  {len(pairs_missing_data)} pairs skipped due to missing market data")
    
    if not candidates_with_spreads:
        logger.warning("No candidates with valid spreads")
        return pd.DataFrame()
    
    # ========================================================================
    # PHASE 5: Calculate Composite Priority Scores
    # ========================================================================
    
    logger.info("\nPhase 5: Calculating composite priority scores...")
    
    for candidate in candidates_with_spreads:
        # Get quality metrics
        weighted_score = candidate.get('Weighted_Score', 0.5)
        sum_dev_cdf = candidate.get('Sum_Dev_CDF', 0.5)
        weighted_spread = candidate.get('weighted_spread', 0.01)
        
        # Calculate component scores
        spread_quality = calculate_spread_quality_score(weighted_spread)
        sum_dev_extremity = calculate_sum_dev_extremity_score(sum_dev_cdf)
        
        # Calculate composite score
        composite_score = calculate_composite_priority_score(
            weighted_score, spread_quality, sum_dev_extremity
        )
        
        # Store scores
        candidate['spread_quality_score'] = spread_quality
        candidate['sum_dev_extremity_score'] = sum_dev_extremity
        candidate['composite_priority_score'] = composite_score
    
    logger.info(f"  Priority scores calculated")
    logger.info(f"  Priority weights: {config.priority_weights()}")
    
    # ========================================================================
    # PHASE 6: Sort by Composite Priority (Best First)
    # ========================================================================
    
    logger.info("\nPhase 6: Sorting by priority...")
    
    candidates_with_spreads.sort(
        key=lambda x: x['composite_priority_score'],
        reverse=True
    )
    
    logger.info(f"  Top 10 candidates by priority:")
    for i, candidate in enumerate(candidates_with_spreads[:10], 1):
        logger.info(f"    {i:2d}. {candidate['Pair']:15s} "
                   f"Score: {candidate['composite_priority_score']:.4f} "
                   f"(Spread: {candidate['weighted_spread_bps']:5.1f} bps)")
    
    # ========================================================================
    # PHASE 7: Greedy Selection with Multi-Constraint Checking
    # ========================================================================

    logger.info("\nPhase 7: Greedy selection with constraint checking...")

    approved_trades = []
    rejected_trades = []

    # Initialize tracking
    cumulative_new_exposure = 0
    long_exposures, short_exposures = Constraints.calculate_ticker_exposures(portfolio_df, [])

    # Calculate current portfolio beta
    # Initialize all variables FIRST (fixes empty portfolio crash)
    current_portfolio_value = 0
    current_weighted_beta = 0
    current_portfolio_beta = 0

    if not portfolio_df.empty:
        for _, existing_trade in portfolio_df.iterrows():
            trade_value = abs(existing_trade.get('Trade Value Co1 ($)', 0)) + \
                         abs(existing_trade.get('Trade Value Co2 ($)', 0))
            trade_beta = existing_trade.get('Beta', 0)

            current_portfolio_value += trade_value
            current_weighted_beta += trade_beta * trade_value

        if current_portfolio_value > 0:
            current_portfolio_beta = current_weighted_beta / current_portfolio_value

    logger.info(f"\nInitial state:")
    logger.info(f"  Portfolio beta: {current_portfolio_beta:.4f}")
    logger.info(f"  Portfolio value: ${current_portfolio_value:,.0f}")
    logger.info(f"  Leverage capacity: ${leverage_capacity_usd:,.0f}")

    # Pre-calculate IGV exposure for Constraint 3.9
    igv_exposure_limit_pct = getattr(config, 'IGV_EXPOSURE_LIMIT_PCT', None)
    igv_allow_corrective = getattr(config, 'IGV_ALLOW_CORRECTIVE_TRADES', True)

    if igv_exposure_limit_pct is not None:
        current_igv_exposure = Constraints.calculate_igv_exposure(portfolio_df)
        current_igv_net = current_igv_exposure['net_igv_exposure']
        current_igv_gross = current_igv_exposure['gross_exposure']
        logger.info(f"  IGV beta exposure: ${current_igv_net:+,.0f} ({current_igv_exposure['exposure_pct']:+.1%} of gross)")
        logger.info(f"  IGV exposure limit: ±{igv_exposure_limit_pct:.0%} of gross")

    # Process candidates in priority order
    for candidate_idx, candidate in enumerate(candidates_with_spreads, 1):
        pair = candidate['Pair']
        ticker1 = candidate['Co1']
        ticker2 = candidate['Co2']
        trigger_type = candidate['trigger_type']
        bucket = candidate['Sum_Dev_Bucket']
        weighted_spread = candidate['weighted_spread']
        
        # ====================================================================
        # CONSTRAINT 1: Spread Ceiling
        # ====================================================================
        
        if weighted_spread > config.max_spread_decimal():
            # Show diagnostic - use W1/W2 from candidate, pass ticker_bid_ask for accurate display
            diagnose_execution_spread(ticker1, ticker2, live_prices, 
                                      candidate.get('W1', 0.5), candidate.get('W2', 0.5), 
                                      config.max_spread_decimal(), pair,
                                      ticker_bid_ask=ticker_bid_ask)
            
            rejected_trades.append({
                'pair': pair,
                'reason': f'Spread too wide ({weighted_spread*10000:.1f} bps > {config.max_spread_decimal()*10000:.1f} bps)'
            })
            continue
        
        # ====================================================================
        # Get Trade Details
        # ====================================================================
        
        # Get prices
        price1 = live_prices.get(ticker1)
        price2 = live_prices.get(ticker2)
        
        if pd.isna(price1) or pd.isna(price2):
            rejected_trades.append({'pair': pair, 'reason': 'Missing prices'})
            continue
        
        # Price validation
        if price1 < config.min_stock_price() or price1 > config.max_stock_price():
            print(f"  ⚠️  {pair}: {ticker1} price ${price1:.2f} outside valid range [${config.min_stock_price():.2f}, ${config.max_stock_price():.2f}]")
            rejected_trades.append({'pair': pair, 'reason': f'{ticker1} price ${price1:.2f} out of range'})
            continue
        
        if price2 < config.min_stock_price() or price2 > config.max_stock_price():
            print(f"  ⚠️  {pair}: {ticker2} price ${price2:.2f} outside valid range [${config.min_stock_price():.2f}, ${config.max_stock_price():.2f}]")
            rejected_trades.append({'pair': pair, 'reason': f'{ticker2} price ${price2:.2f} out of range'})
            continue
        
        # Get strategy configuration
        W1 = candidate['W1']
        W2 = candidate['W2']
        position_mult = Constraints.get_position_multiplier(trigger_type, bucket)
        
        # Calculate quantities
        adjusted_notional = base_trade_size * position_mult
        notional1 = adjusted_notional * W1
        notional2 = adjusted_notional * W2
        
        qty1 = int(notional1 / price1)
        qty2 = int(notional2 / price2)
        
        # CRITICAL: Ensure quantities are non-zero (prevents Error 321)
        # Note: We only check for zero - share count doesn't matter as long as
        # we can achieve the approximate target dollar value
        if qty1 == 0 or qty2 == 0:
            rejected_trades.append({
                'pair': pair, 
                'reason': f'Zero quantity calculated (qty1={qty1}, qty2={qty2}) - stock price too high for trade size'
            })
            continue
        
        actual_notional1 = qty1 * price1
        actual_notional2 = qty2 * price2
        total_notional = actual_notional1 + actual_notional2
        
        # Position size validation - REMOVED
        # Let CDF-based sizing determine all position sizes naturally
        # No artificial floor or ceiling constraints
        
        # ====================================================================
        # CONSTRAINT 2: Leverage Capacity
        # ====================================================================
        
        projected_new_exposure = cumulative_new_exposure + total_notional
        
        if projected_new_exposure > leverage_capacity_usd:
            remaining_capacity = leverage_capacity_usd - cumulative_new_exposure
            projected_leverage = (current_gross_usd + projected_new_exposure) / net_liq_usd
            rejected_trades.append({
                'pair': pair,
                'reason': f'Would exceed leverage limit (need ${total_notional:,.0f}, capacity ${remaining_capacity:,.0f}, projected {projected_leverage:.2f}x)'
            })
            continue
        
        # ====================================================================
        # CONSTRAINT 3: Portfolio Beta
        # ====================================================================
        
        # Get VO betas for portfolio-level market exposure constraint
        beta_vo_co1 = beta_manager.get_vo_beta(ticker1)
        beta_vo_co2 = beta_manager.get_vo_beta(ticker2)
        
        # Determine direction
        if trigger_type == 'lower':
            direction_co1 = 1   # Long
            direction_co2 = -1  # Short
        else:  # upper
            direction_co1 = -1  # Short
            direction_co2 = 1   # Long
        
        # Calculate pair beta using VO betas (overall market exposure)
        pair_beta = (W1 * direction_co1 * beta_vo_co1) + (W2 * direction_co2 * beta_vo_co2)
        
        # Calculate projected portfolio beta
        new_portfolio_value = current_portfolio_value + total_notional
        new_weighted_beta = (current_weighted_beta + pair_beta * total_notional)
        new_portfolio_beta = new_weighted_beta / new_portfolio_value if new_portfolio_value > 0 else 0
        
        # Determine which beta limits to use
        if current_portfolio_value < config.min_portfolio_value_for_strict_constraints():
            min_beta = config.portfolio_building_min_beta()
            max_beta = config.portfolio_building_max_beta()
            in_building_phase = True
        else:
            min_beta = config.min_portfolio_beta()
            max_beta = config.max_portfolio_beta()
            in_building_phase = False
        
        # Calculate distances for corrective trade logic
        distance_before = abs(current_portfolio_beta - config.target_portfolio_beta())
        distance_after = abs(new_portfolio_beta - config.target_portfolio_beta())
        is_getting_closer = distance_after < distance_before
        
        # Log for first 3 candidates to see what's happening
        if candidate_idx <= 3:
            logger.info(f"\n  Beta check for {pair} (candidate #{candidate_idx}):")
            logger.info(f"    Pair beta: {pair_beta:+.4f}")
            logger.info(f"    Current portfolio beta: {current_portfolio_beta:.4f}")
            logger.info(f"    New portfolio beta: {new_portfolio_beta:.4f}")
            logger.info(f"    Limits: [{min_beta:.4f}, {max_beta:.4f}]")
            logger.info(f"    Target: {config.target_portfolio_beta():.4f}")
            logger.info(f"    Distance before: {distance_before:.4f}")
            logger.info(f"    Distance after: {distance_after:.4f}")
            logger.info(f"    Getting closer to target? {is_getting_closer}")
            logger.info(f"    ALLOW_CORRECTIVE_TRADES: {config.allow_corrective_trades()}")
            if in_building_phase:
                logger.info(f"    (Building phase - using relaxed limits)")
        
        # Beta Constraint Check
        if min_beta <= new_portfolio_beta <= max_beta:
            # New beta is within limits - ALWAYS ALLOW
            if candidate_idx <= 3:
                logger.info(f"    ✓ ALLOW: New beta within limits")
            # Continue to next constraint
        
        elif new_portfolio_beta > max_beta:
            # NEW BETA EXCEEDS MAX
            currently_above_max = current_portfolio_beta > max_beta
            
            if currently_above_max and config.allow_corrective_trades():
                # Portfolio already above max, corrective trades enabled
                if is_getting_closer:
                    logger.info(f"  ✓ ALLOWING corrective trade: {pair}")
                    logger.info(f"    Portfolio beta: {current_portfolio_beta:.4f} → {new_portfolio_beta:.4f}")
                    logger.info(f"    Still above max ({max_beta:.4f}), but moving toward target ({config.target_portfolio_beta():.4f})")
                    logger.info(f"    Distance improvement: {distance_before:.4f} → {distance_after:.4f}")
                    # Continue to next constraint
                else:
                    # Trade moves us AWAY from target - REJECT
                    rejected_trades.append({
                        'pair': pair,
                        'reason': f'Beta {new_portfolio_beta:.4f} > max {max_beta:.4f}, not corrective (distance: {distance_before:.4f} → {distance_after:.4f})'
                    })
                    if candidate_idx <= 3:
                        logger.info(f"    ✗ REJECT: Would move farther from target")
                    continue
            else:
                # Portfolio within limits OR corrective trades disabled
                if not config.allow_corrective_trades():
                    reason = f'Beta {new_portfolio_beta:.4f} > max {max_beta:.4f} (corrective trades disabled)'
                else:
                    reason = f'Beta {new_portfolio_beta:.4f} > max {max_beta:.4f} (would exceed limit)'
                
                rejected_trades.append({'pair': pair, 'reason': reason})
                if candidate_idx <= 3:
                    logger.info(f"    ✗ REJECT: {reason}")
                continue
        
        elif new_portfolio_beta < min_beta:
            # NEW BETA BELOW MIN
            currently_below_min = current_portfolio_beta < min_beta
            
            if currently_below_min and config.allow_corrective_trades():
                # Portfolio already below min, corrective trades enabled
                if is_getting_closer:
                    logger.info(f"  ✓ ALLOWING corrective trade: {pair}")
                    logger.info(f"    Portfolio beta: {current_portfolio_beta:.4f} → {new_portfolio_beta:.4f}")
                    logger.info(f"    Still below min ({min_beta:.4f}), but moving toward target ({config.target_portfolio_beta():.4f})")
                    logger.info(f"    Distance improvement: {distance_before:.4f} → {distance_after:.4f}")
                    # Continue to next constraint
                else:
                    # Trade moves us AWAY from target - REJECT
                    rejected_trades.append({
                        'pair': pair,
                        'reason': f'Beta {new_portfolio_beta:.4f} < min {min_beta:.4f}, not corrective'
                    })
                    if candidate_idx <= 3:
                        logger.info(f"    ✗ REJECT: Would move farther from target")
                    continue
            else:
                # Portfolio within limits OR corrective trades disabled
                if not config.allow_corrective_trades():
                    reason = f'Beta {new_portfolio_beta:.4f} < min {min_beta:.4f} (corrective trades disabled)'
                else:
                    reason = f'Beta {new_portfolio_beta:.4f} < min {min_beta:.4f} (would go below limit)'
                
                rejected_trades.append({'pair': pair, 'reason': reason})
                if candidate_idx <= 3:
                    logger.info(f"    ✗ REJECT: {reason}")
                continue

        # ====================================================================
        # CONSTRAINT 3.5: INDEX-SPECIFIC BETA LIMITS (if index concentrated)
        # ====================================================================

        # Calculate current index exposures
        if config.enable_index_concentration_limits():
            # Build index breakdown from current portfolio + approved trades
            current_index_exposures = {}

            # Process existing portfolio
            if not portfolio_df.empty:
                for _, existing_trade in portfolio_df.iterrows():
                    idx = existing_trade.get('Index', 'Unknown')
                    if pd.isna(idx):
                        idx = 'Unknown'

                    trade_value = abs(existing_trade.get('Trade Value Co1 ($)', 0)) + \
                                 abs(existing_trade.get('Trade Value Co2 ($)', 0))
                    trade_beta = existing_trade.get('Beta', 0)

                    if idx not in current_index_exposures:
                        current_index_exposures[idx] = {
                            'net_exposure': 0,
                            'weighted_beta': 0,
                            'gross_exposure': 0
                        }

                    tail = existing_trade.get('Tail', 'L').strip().upper()
                    value1 = abs(existing_trade.get('Trade Value Co1 ($)', 0))
                    value2 = abs(existing_trade.get('Trade Value Co2 ($)', 0))

                    if tail == 'L':
                        net_exp = value1 - value2
                    else:
                        net_exp = value2 - value1

                    current_index_exposures[idx]['net_exposure'] += net_exp
                    current_index_exposures[idx]['weighted_beta'] += trade_beta * trade_value
                    current_index_exposures[idx]['gross_exposure'] += trade_value

            # Process approved trades from this run
            for approved in approved_trades:
                idx = approved.get('Index', 'Unknown')
                if pd.isna(idx):
                    idx = 'Unknown'

                trade_value = approved['Total_Notional']
                trade_beta = approved['Beta']

                if idx not in current_index_exposures:
                    current_index_exposures[idx] = {
                        'net_exposure': 0,
                        'weighted_beta': 0,
                        'gross_exposure': 0
                    }

                trigger_type = approved['trigger_type']
                W1 = approved['W1']
                W2 = approved['W2']
                value1 = approved['Trade Value Co1 ($)']
                value2 = approved['Trade Value Co2 ($)']

                if trigger_type == 'lower':
                    net_exp = value1 - value2
                else:
                    net_exp = value2 - value1

                current_index_exposures[idx]['net_exposure'] += net_exp
                current_index_exposures[idx]['weighted_beta'] += trade_beta * trade_value
                current_index_exposures[idx]['gross_exposure'] += trade_value

            # Calculate dollar beta for each index
            # weighted_beta is already Σ(pair_beta × trade_value) = dollar beta!
            total_dollar_beta = sum(
                abs(data['weighted_beta'])
                for data in current_index_exposures.values()
            )

            # Get current candidate's index
            candidate_index = candidate.get('Index')

            if candidate_index:
                # Get current index data (may not exist yet if this is first trade for this index)
                idx_data = current_index_exposures.get(candidate_index, {
                    'net_exposure': 0,
                    'weighted_beta': 0,
                    'gross_exposure': 0
                })

                # Current state (BEFORE this trade)
                idx_dollar_beta = idx_data['weighted_beta']
                idx_pct_of_total = abs(idx_dollar_beta) / total_dollar_beta if total_dollar_beta > 0 else 0

                # Calculate POST-TRADE state
                # This is the key fix: check what concentration WOULD BE after adding this trade
                new_idx_dollar_beta = idx_dollar_beta + pair_beta * total_notional
                # Note: total_dollar_beta uses abs() of each index's weighted beta
                # We need to recalculate what total would be after this trade
                new_total_dollar_beta = total_dollar_beta - abs(idx_dollar_beta) + abs(new_idx_dollar_beta)
                idx_pct_of_total_after = abs(new_idx_dollar_beta) / new_total_dollar_beta if new_total_dollar_beta > 0 else 0

                # Check if this index WOULD exceed concentration threshold AFTER this trade
                # This prevents trades that push an index from 24% to 35% from slipping through
                if idx_pct_of_total_after > config.max_index_dollar_beta_concentration():
                    # This index is concentrated - enforce index-specific beta limits

                    # For index-specific limit checking, we need AVERAGE beta per dollar
                    idx_net_beta = idx_data['weighted_beta'] / idx_data['gross_exposure'] if idx_data['gross_exposure'] > 0 else 0

                    # Calculate what the NEW index beta would be with this trade
                    new_idx_net_exp = idx_data['net_exposure'] + (
                        actual_notional1 - actual_notional2 if trigger_type == 'lower' 
                        else actual_notional2 - actual_notional1
                    )
                    new_idx_gross_exp = idx_data['gross_exposure'] + total_notional
                    new_idx_weighted_beta = idx_data['weighted_beta'] + pair_beta * total_notional

                    new_idx_beta = new_idx_weighted_beta / new_idx_gross_exp if new_idx_gross_exp > 0 else 0

                    # Target for index is 0.0 (market-neutral within the index)
                    idx_target_beta = 0.0

                    # Calculate distances for corrective trade logic
                    idx_distance_before = abs(idx_net_beta - idx_target_beta)
                    idx_distance_after = abs(new_idx_beta - idx_target_beta)
                    idx_is_getting_closer = idx_distance_after < idx_distance_before

                    # Check against index-specific limits (with corrective trade logic)
                    if config.index_specific_min_beta() <= new_idx_beta <= config.index_specific_max_beta():
                        # Within limits - ALWAYS ALLOW
                        if candidate_idx <= 3:
                            logger.info(f"  Index-specific check for {candidate_index}:")
                            logger.info(f"    Index concentration: {idx_pct_of_total:.1%} → {idx_pct_of_total_after:.1%} (threshold: {config.max_index_dollar_beta_concentration():.0%})")
                            logger.info(f"    Current index beta: {idx_net_beta:.4f}")
                            logger.info(f"    New index beta: {new_idx_beta:.4f}")
                            logger.info(f"    Index limits: [{config.index_specific_min_beta():.4f}, {config.index_specific_max_beta():.4f}]")
                            logger.info(f"    ✓ Within index limits")
                        # Continue to next constraint

                    elif new_idx_beta > config.index_specific_max_beta():
                        # NEW INDEX BETA EXCEEDS MAX
                        currently_above_max = idx_net_beta > config.index_specific_max_beta()

                        if currently_above_max and config.allow_corrective_trades():
                            # Index already above max, check if corrective
                            if idx_is_getting_closer:
                                logger.info(f"  ✓ ALLOWING corrective trade for {candidate_index}: {pair}")
                                logger.info(f"    Index concentration: {idx_pct_of_total:.1%} → {idx_pct_of_total_after:.1%}")
                                logger.info(f"    Index beta: {idx_net_beta:.4f} → {new_idx_beta:.4f}")
                                logger.info(f"    Still above max ({config.index_specific_max_beta():.4f}), but moving toward target (0.0)")
                                logger.info(f"    Distance improvement: {idx_distance_before:.4f} → {idx_distance_after:.4f}")
                                # Continue to next constraint
                            else:
                                # Trade moves us AWAY from target - REJECT
                                rejected_trades.append({
                                    'pair': pair,
                                    'reason': (f'{candidate_index} index would be concentrated ({idx_pct_of_total_after:.1%}), '
                                              f'beta {new_idx_beta:.4f} > max {config.index_specific_max_beta():.4f}, not corrective')
                                })
                                continue
                        else:
                            # Index within limits OR corrective trades disabled
                            rejected_trades.append({
                                'pair': pair,
                                'reason': (f'{candidate_index} index would be concentrated ({idx_pct_of_total_after:.1%}), '
                                          f'beta {new_idx_beta:.4f} > max {config.index_specific_max_beta():.4f}')
                            })
                            continue

                    elif new_idx_beta < config.index_specific_min_beta():
                        # NEW INDEX BETA BELOW MIN
                        currently_below_min = idx_net_beta < config.index_specific_min_beta()

                        if currently_below_min and config.allow_corrective_trades():
                            # Index already below min, check if corrective
                            if idx_is_getting_closer:
                                logger.info(f"  ✓ ALLOWING corrective trade for {candidate_index}: {pair}")
                                logger.info(f"    Index concentration: {idx_pct_of_total:.1%} → {idx_pct_of_total_after:.1%}")
                                logger.info(f"    Index beta: {idx_net_beta:.4f} → {new_idx_beta:.4f}")
                                logger.info(f"    Still below min ({config.index_specific_min_beta():.4f}), but moving toward target (0.0)")
                                logger.info(f"    Distance improvement: {idx_distance_before:.4f} → {idx_distance_after:.4f}")
                                # Continue to next constraint
                            else:
                                # Trade moves us AWAY from target - REJECT
                                rejected_trades.append({
                                    'pair': pair,
                                    'reason': (f'{candidate_index} index would be concentrated ({idx_pct_of_total_after:.1%}), '
                                              f'beta {new_idx_beta:.4f} < min {config.index_specific_min_beta():.4f}, not corrective')
                                })
                                continue
                        else:
                            # Index within limits OR corrective trades disabled
                            rejected_trades.append({
                                'pair': pair,
                                'reason': (f'{candidate_index} index would be concentrated ({idx_pct_of_total_after:.1%}), '
                                          f'beta {new_idx_beta:.4f} < min {config.index_specific_min_beta():.4f}')
                            })
                            continue
                    
        # ====================================================================
        # CONSTRAINT 3.7: Max Gross Exposure Concentration by Index
        # ====================================================================
        # Prevents any single index from dominating portfolio gross exposure.
        # Unlike the dollar-beta concentration check in 3.5, this uses raw
        # gross exposure (notional), which is more stable and not affected
        # by beta sign cancellations.
        
        max_index_gross_pct = getattr(config, 'MAX_INDEX_GROSS_EXPOSURE_PCT', None)
        
        if max_index_gross_pct is not None and candidate_index:
            # Calculate total gross exposure (existing + approved)
            gross_by_index = {}
            
            # From existing portfolio
            if not portfolio_df.empty:
                for _, existing_trade in portfolio_df.iterrows():
                    idx = existing_trade.get('Index', 'Unknown')
                    if pd.isna(idx):
                        idx = 'Unknown'
                    tv = abs(existing_trade.get('Trade Value Co1 ($)', 0)) + \
                         abs(existing_trade.get('Trade Value Co2 ($)', 0))
                    gross_by_index[idx] = gross_by_index.get(idx, 0) + tv
            
            # From previously approved trades this run
            for approved in approved_trades:
                idx = approved.get('Index', 'Unknown')
                if pd.isna(idx):
                    idx = 'Unknown'
                gross_by_index[idx] = gross_by_index.get(idx, 0) + approved['Total_Notional']
            
            total_gross_exposure = sum(gross_by_index.values())
            candidate_idx_gross = gross_by_index.get(candidate_index, 0)
            
            # Calculate what concentration WOULD BE after this trade
            new_candidate_gross = candidate_idx_gross + total_notional
            new_total_gross = total_gross_exposure + total_notional
            new_gross_pct = new_candidate_gross / new_total_gross if new_total_gross > 0 else 0
            
            if new_gross_pct > max_index_gross_pct:
                current_gross_pct = candidate_idx_gross / total_gross_exposure if total_gross_exposure > 0 else 0
                
                rejected_trades.append({
                    'pair': pair,
                    'reason': (f'{candidate_index} gross exposure would be {new_gross_pct:.1%} '
                              f'(limit {max_index_gross_pct:.0%}), '
                              f'currently {current_gross_pct:.1%}')
                })
                if candidate_idx <= 3:
                    logger.info(f"  ✗ REJECT: {candidate_index} gross exposure {new_gross_pct:.1%} > {max_index_gross_pct:.0%}")
                continue

        # ====================================================================
        # CONSTRAINT 3.9: IGV Beta Exposure Limit
        # ====================================================================
        # Limits net IGV beta-weighted dollar exposure as a percentage of gross.
        # Uses betas from igv_beta_analysis.xlsx covering the full investable
        # universe (not just IGV constituents). This is the primary risk
        # management for software/IGV exposure; the options hedge (Stage 8)
        # serves as backup for exposure drift from terminations.
        
        if igv_exposure_limit_pct is not None:
            # Build pending trades list (already-approved this run + this candidate)
            pending_for_igv = []
            for prev_trade in approved_trades:
                pending_for_igv.append({
                    'co1': prev_trade.get('ticker1', ''),
                    'co2': prev_trade.get('ticker2', ''),
                    'tail': prev_trade.get('tail', 'L'),
                    'notional1': prev_trade.get('actual_notional1', 0),
                    'notional2': prev_trade.get('actual_notional2', 0),
                })
            # Add current candidate
            pending_for_igv.append({
                'co1': ticker1,
                'co2': ticker2,
                'tail': 'L' if trigger_type == 'lower' else 'U',
                'notional1': abs(candidate.get('actual_notional1', base_trade_size * candidate.get('W1', 0.5))),
                'notional2': abs(candidate.get('actual_notional2', base_trade_size * (1 - candidate.get('W1', 0.5)))),
            })
            
            new_igv_exposure = Constraints.calculate_igv_exposure(
                portfolio_df, pending_trades=pending_for_igv
            )
            new_igv_net = new_igv_exposure['net_igv_exposure']
            new_igv_gross = new_igv_exposure['gross_exposure']
            new_igv_pct = new_igv_net / new_igv_gross if new_igv_gross > 0 else 0

            if abs(new_igv_pct) > igv_exposure_limit_pct:
                # Check if this trade is corrective (reduces exposure toward limit)
                is_corrective = False
                if igv_allow_corrective:
                    # Corrective = the new exposure is closer to zero than the current
                    current_igv_with_approved = Constraints.calculate_igv_exposure(
                        portfolio_df, pending_trades=pending_for_igv[:-1]
                    )
                    current_net_before = current_igv_with_approved['net_igv_exposure']
                    is_corrective = abs(new_igv_net) < abs(current_net_before)
                
                if is_corrective:
                    if candidate_idx <= 3:
                        logger.info(f"  ✓ IGV exposure {new_igv_pct:+.1%} > ±{igv_exposure_limit_pct:.0%} "
                                   f"but trade is corrective (${current_net_before:+,.0f} → ${new_igv_net:+,.0f})")
                else:
                    rejected_trades.append({
                        'pair': pair,
                        'reason': f'IGV beta exposure {new_igv_pct:+.1%} would exceed ±{igv_exposure_limit_pct:.0%} limit '
                                  f'(${new_igv_net:+,.0f})'
                    })
                    if candidate_idx <= 3:
                        logger.info(f"  ✗ REJECT: IGV exposure would be {new_igv_pct:+.1%} "
                                   f"(${new_igv_net:+,.0f}), limit ±{igv_exposure_limit_pct:.0%}")
                    continue

        # ====================================================================
        # CONSTRAINT 4 & 5: Ticker Concentration (Long and Short)
        # ====================================================================
        # Added corrective trade logic: Allow trades that REDUCE max concentration
        # even when portfolio is already over-concentrated
        
        # Count total trades (existing + approved in this run)
        num_existing_trades = len(portfolio_df) if not portfolio_df.empty else 0
        num_total_trades = num_existing_trades + len(approved_trades)
        
        # Skip concentration checks for first N trades
        # With <10 tickers, concentration metrics are meaningless
        if num_total_trades < config.min_trades_for_concentration_check():
            logger.debug(f"  Skipping concentration check (trade #{num_total_trades + 1}/{config.min_trades_for_concentration_check()})")
        else:
            # Enough trades now - enforce concentration limits
            
            # Determine which tickers are long/short for this trade
            if trigger_type == 'lower':
                long_ticker = ticker1
                short_ticker = ticker2
                long_notional = actual_notional1
                short_notional = actual_notional2
            else:  # upper
                long_ticker = ticker2
                short_ticker = ticker1
                long_notional = actual_notional2
                short_notional = actual_notional1
            
            # Use relaxed or strict limits based on portfolio size
            if current_portfolio_value < config.min_portfolio_value_for_strict_constraints():
                max_long_concentration = config.portfolio_building_max_long_concentration()
                max_short_concentration = config.portfolio_building_max_short_concentration()
                logger.debug(f"  Using relaxed concentration limits: {max_long_concentration:.0%}/{max_short_concentration:.0%}")
            else:
                max_long_concentration = config.max_long_ticker_concentration()
                max_short_concentration = config.max_short_ticker_concentration()
                logger.debug(f"  Using strict concentration limits: {max_long_concentration:.0%}/{max_short_concentration:.0%}")
            
            # ================================================================
            # Calculate CURRENT max concentration (before this trade)
            # ================================================================
            current_max_long_conc = 0
            current_max_short_conc = 0
            
            if current_portfolio_value > 0:
                for ticker, exposure in long_exposures.items():
                    conc = exposure / current_portfolio_value
                    current_max_long_conc = max(current_max_long_conc, conc)
                
                for ticker, exposure in short_exposures.items():
                    conc = exposure / current_portfolio_value
                    current_max_short_conc = max(current_max_short_conc, conc)
            
            # ================================================================
            # Calculate NEW max concentration (after this trade)
            # ================================================================
            # Copy exposures and add this trade's exposure
            projected_long_exposures = long_exposures.copy()
            projected_short_exposures = short_exposures.copy()
            
            projected_long_exposures[long_ticker] = projected_long_exposures.get(long_ticker, 0) + long_notional
            projected_short_exposures[short_ticker] = projected_short_exposures.get(short_ticker, 0) + short_notional
            
            # Calculate max concentration after trade
            new_max_long_conc = 0
            new_max_short_conc = 0
            
            if new_portfolio_value > 0:
                for ticker, exposure in projected_long_exposures.items():
                    conc = exposure / new_portfolio_value
                    new_max_long_conc = max(new_max_long_conc, conc)
                
                for ticker, exposure in projected_short_exposures.items():
                    conc = exposure / new_portfolio_value
                    new_max_short_conc = max(new_max_short_conc, conc)
            
            # ================================================================
            # Check long ticker concentration with corrective logic
            # ================================================================
            current_long_exposure = long_exposures.get(long_ticker, 0)
            new_long_exposure = current_long_exposure + long_notional
            long_concentration = new_long_exposure / new_portfolio_value if new_portfolio_value > 0 else 0
            
            # Check if this specific ticker would exceed limit
            if long_concentration > max_long_concentration:
                # This ticker would be over-concentrated
                # BUT: Allow if it's a corrective trade (reduces max concentration)
                
                is_corrective_long = (current_max_long_conc > max_long_concentration and 
                                      new_max_long_conc < current_max_long_conc)
                
                allow_corrective = getattr(config, 'ALLOW_CORRECTIVE_TRADES', True)
                
                if is_corrective_long and allow_corrective:
                    logger.info(f"  ✓ ALLOWING corrective concentration trade: {pair}")
                    logger.info(f"    Long {long_ticker} concentration: {long_concentration:.2%} (exceeds {max_long_concentration:.2%})")
                    logger.info(f"    But MAX long concentration reduces: {current_max_long_conc:.2%} → {new_max_long_conc:.2%}")
                    # Allow this trade - it helps even though this ticker is high
                else:
                    rejected_trades.append({
                        'pair': pair,
                        'reason': f'Long {long_ticker} concentration {long_concentration:.2%} exceeds max {max_long_concentration:.2%}'
                    })
                    continue
            
            # ================================================================
            # Check short ticker concentration with corrective logic
            # ================================================================
            current_short_exposure = short_exposures.get(short_ticker, 0)
            new_short_exposure = current_short_exposure + short_notional
            short_concentration = new_short_exposure / new_portfolio_value if new_portfolio_value > 0 else 0
            
            # ================================================================
            # VHT Small Cap Constraint: Tighter limit for small healthcare shorts
            # ================================================================
            effective_max_short = max_short_concentration
            
            if (hasattr(config, 'VHT_SMALL_CAP_CONSTRAINT') and 
                config.VHT_SMALL_CAP_CONSTRAINT.get('enabled', False)):
                
                candidate_index = candidate.get('Index', '')
                
                if candidate_index == 'VHT' and market_caps:
                    mcap_threshold = config.VHT_SMALL_CAP_CONSTRAINT.get('mcap_threshold_millions', 5000)
                    conc_reduction = config.VHT_SMALL_CAP_CONSTRAINT.get('concentration_reduction', 0.005)
                    
                    short_mcap = market_caps.get(short_ticker)
                    
                    if short_mcap is not None and short_mcap < mcap_threshold:
                        effective_max_short = max_short_concentration - conc_reduction
                        logger.debug(f"    VHT small cap constraint: {short_ticker} MCap=${short_mcap:.0f}M < ${mcap_threshold}M")
                        logger.debug(f"      Reducing max short concentration: {max_short_concentration:.2%} → {effective_max_short:.2%}")
            
            # Check if this specific ticker would exceed limit
            if short_concentration > effective_max_short:
                # This ticker would be over-concentrated
                # BUT: Allow if it's a corrective trade (reduces max concentration)
                
                is_corrective_short = (current_max_short_conc > effective_max_short and 
                                       new_max_short_conc < current_max_short_conc)
                
                allow_corrective = getattr(config, 'ALLOW_CORRECTIVE_TRADES', True)
                
                if is_corrective_short and allow_corrective:
                    logger.info(f"  ✓ ALLOWING corrective concentration trade: {pair}")
                    logger.info(f"    Short {short_ticker} concentration: {short_concentration:.2%} (exceeds {effective_max_short:.2%})")
                    logger.info(f"    But MAX short concentration reduces: {current_max_short_conc:.2%} → {new_max_short_conc:.2%}")
                    # Allow this trade - it helps even though this ticker is high
                else:
                    rejected_trades.append({
                        'pair': pair,
                        'reason': f'Short {short_ticker} concentration {short_concentration:.2%} exceeds max {effective_max_short:.2%}'
                    })
                    continue
        
        # ====================================================================
        # CONSTRAINT 6: Per-Run Ticker Limit (NEW)
        # ====================================================================
        
        if config.enable_per_run_ticker_limit():
            # Count how many times each ticker appears in approved trades THIS RUN
            ticker_counts_this_run = {}
            
            for approved_trade in approved_trades:
                t1 = approved_trade['Co1']
                t2 = approved_trade['Co2']
                ticker_counts_this_run[t1] = ticker_counts_this_run.get(t1, 0) + 1
                ticker_counts_this_run[t2] = ticker_counts_this_run.get(t2, 0) + 1
            
            # Check if either ticker in current candidate would exceed limit
            ticker1_count = ticker_counts_this_run.get(ticker1, 0)
            ticker2_count = ticker_counts_this_run.get(ticker2, 0)
            
            max_allowed = config.max_new_positions_per_ticker()
            
            if ticker1_count >= max_allowed:
                rejected_trades.append({
                    'pair': pair,
                    'reason': f'{ticker1} already in {ticker1_count} new positions this run (max: {max_allowed})'
                })
                continue
            
            if ticker2_count >= max_allowed:
                rejected_trades.append({
                    'pair': pair,
                    'reason': f'{ticker2} already in {ticker2_count} new positions this run (max: {max_allowed})'
                })
                continue
            
            # Both tickers OK - log if getting close to limit
            if ticker1_count == max_allowed - 1 or ticker2_count == max_allowed - 1:
                logger.info(f"  {pair}: Ticker approaching limit "
                           f"({ticker1}={ticker1_count}/{max_allowed}, "
                           f"{ticker2}={ticker2_count}/{max_allowed})")
        
        # ====================================================================
        # ALL CONSTRAINTS PASSED - APPROVE TRADE
        # ====================================================================
        
        # Validate Tag
        original_tag = candidate.get('Tag')
        if pd.isna(original_tag):
            logger.error(f"  ⚠️  {pair} has no Tag! Skipping.")
            rejected_trades.append({'pair': pair, 'reason': 'Missing Tag'})
            continue
        
        # Ensure Tag is integer (from Parameters file)
        try:
            tag_int = int(original_tag)
        except:
            logger.error(f"  ⚠️  {pair} has invalid Tag: {original_tag}")
            rejected_trades.append({'pair': pair, 'reason': f'Invalid Tag: {original_tag}'})
            continue
        
        # Calculate concentration
        concentration = total_notional / new_portfolio_value if new_portfolio_value > 0 else 0.0
        
        trade_spec = {
            # Identifiers
            'Tag': int(candidate['Tag']) if pd.notna(candidate.get('Tag')) else None,
            'Pair': pair,
            'Co1': ticker1,
            'Co2': ticker2,
            'Index': candidate.get('Index'),

            # Strategy configuration
            'trigger_type': trigger_type,
            'Tail': candidate['Tail'],
            'Sum_Dev_Bucket': bucket,  # Canonical CamelCase
            'Sum_Dev_Value': candidate.get('Sum_Dev_Value', np.nan),
            'Sum_Dev_CDF': candidate.get('Sum_Dev_CDF', np.nan),
            'W1': W1,
            'W2': W2,
            'Position_Multiplier': position_mult,

            # Position sizing
            'Quantity1': qty1,
            'Quantity2': qty2,
            'Co1 at Initiation': price1,
            'Co2 at Initiation': price2,
            'Trade Value Co1 ($)': actual_notional1,
            'Trade Value Co2 ($)': actual_notional2,
            'Total_Notional': total_notional,

            # Risk metrics
            'Beta': pair_beta,
            'Concentration': concentration,

            # Quality metrics
            'Entry_Spread_BPS': weighted_spread * 10000,
            'Spread_Quality_Score': candidate['spread_quality_score'],
            'Sum_Dev_Extremity_Score': candidate['sum_dev_extremity_score'],
            'Composite_Priority_Score': candidate['composite_priority_score'],

            # Market data at entry
            # Use index_prices dict for sector ETF price, fall back to live_prices
            'Index at Initiation': (
                index_prices.get(candidate.get('Index')) if index_prices 
                else live_prices.get(candidate.get('Index')) if live_prices 
                else np.nan
            ),
            'Treasury at Initiation': dgs10_price if dgs10_price else np.nan,

            # LAM signals - Raw values (CamelCase canonical)
            'Weighted_Score': candidate.get('Weighted_Score', np.nan),
            'Index_Bias': candidate.get('Index_Bias', Config_Helper.get_index_bias(candidate.get('Index'), candidate['Tail'])),
            'Volume_Ratio': candidate.get('Volume_Ratio', np.nan),
            'Rolling_Intraday_Vol': candidate.get('Rolling_Intraday_Vol', np.nan),
            'Volume_Dominance': candidate.get('Volume_Dominance', np.nan),
            'Last_Hour_Vol': candidate.get('Last_Hour_Vol', np.nan),
            'IV_Percentile': candidate.get('IV_Percentile', np.nan),
            
            # LAM signals - Percentiles (for like-for-like comparison with rejected trades)
            'Volume_Ratio_Pct': candidate.get('Volume_Ratio_Pct', np.nan),
            'Intraday_Vol_Pct': candidate.get('Intraday_Vol_Pct', np.nan),
            'Volume_Dom_Pct': candidate.get('Volume_Dom_Pct', np.nan),
            'Last_Hour_Pct': candidate.get('Last_Hour_Pct', np.nan),
            'IV_Pct_Pct': candidate.get('IV_Pct_Pct', np.nan),
            
            # Composite score (matches rejected trades archive)
            'Composite_Score': candidate.get('Composite_Score', np.nan),

            # Status
            'Existing': 0,
            'Evaluated_At': datetime.now()
        }
        
        approved_trades.append(trade_spec)
        
        # Update tracking
        cumulative_new_exposure += total_notional 
        current_weighted_beta += pair_beta * total_notional
        current_portfolio_value += total_notional
        current_portfolio_beta = current_weighted_beta / current_portfolio_value
        
        # Update ticker exposures
        if trigger_type == 'lower':
            long_exposures[ticker1] = long_exposures.get(ticker1, 0) + actual_notional1
            short_exposures[ticker2] = short_exposures.get(ticker2, 0) + actual_notional2
        else:
            short_exposures[ticker1] = short_exposures.get(ticker1, 0) + actual_notional1
            long_exposures[ticker2] = long_exposures.get(ticker2, 0) + actual_notional2
    
    # ========================================================================
    # PHASE 8: Create Summary (Enhanced)
    # ========================================================================

    print("\n" + "="*80)
    print("EVALUATION COMPLETE")
    print("="*80)
    print(f"Candidates evaluated: {len(candidates_with_spreads)}")
    print(f"✓ Approved: {len(approved_trades)}")
    print(f"✗ Rejected: {len(rejected_trades)}")
    print(f"Approval rate: {len(approved_trades)/len(candidates_with_spreads)*100:.1f}%")

    # Show rejection breakdown
    if rejected_trades:
        print("\n🚫 Rejection breakdown:")
        rejection_summary = {}
        beta_rejections = []  # Track beta rejections for detailed breakdown
        
        for reject in rejected_trades:
            reason = reject['reason']
            pair = reject.get('pair', 'Unknown')
            
            # Simplify reason for grouping
            if 'leverage' in reason.lower():
                simple_reason = 'Leverage limit'
            elif 'beta' in reason.lower():
                simple_reason = 'Beta constraint'
                beta_rejections.append({'pair': pair, 'reason': reason})
            elif 'spread' in reason.lower():
                simple_reason = 'Spread above ceiling'
            elif 'concentration' in reason.lower():
                simple_reason = 'Concentration limit'
            elif 'gross exposure' in reason.lower():
                simple_reason = 'Index exposure limit'
            elif 'exposure' in reason.lower() and any(f in reason.upper() for f in ['USMV', 'VLUE', 'MTUM', 'SOXX']):
                simple_reason = 'Factor exposure limit'
            elif 'igv' in reason.lower() and 'exposure' in reason.lower():
                simple_reason = 'IGV exposure limit'
            elif 'already in' in reason.lower() and 'positions this run' in reason.lower():
                simple_reason = 'Per-run ticker limit'
            elif 'price' in reason.lower():
                simple_reason = 'Price out of range'
            elif 'quantity' in reason.lower() or 'zero' in reason.lower():
                simple_reason = 'Quantity too small'
            elif 'leverage' in reason.lower():
                simple_reason = 'Leverage limit'
            else:
                simple_reason = 'Other'

            rejection_summary[simple_reason] = rejection_summary.get(simple_reason, 0) + 1

        for reason, count in sorted(rejection_summary.items(), key=lambda x: x[1], reverse=True):
            pct = count / len(rejected_trades) * 100
            print(f"  {reason:25s}: {count:3d} ({pct:5.1f}%)")
        
        # Detailed beta constraint breakdown
        if beta_rejections:
            print(f"\n  📊 Beta constraint details ({len(beta_rejections)} rejections):")
            
            # Categorize by type
            above_max = [r for r in beta_rejections if '> max' in r['reason']]
            below_min = [r for r in beta_rejections if '< min' in r['reason']]
            
            if above_max:
                print(f"     Above max ({len(above_max)}):")
                for r in above_max[:5]:  # Show first 5
                    # Extract beta value from reason string (case-insensitive)
                    beta_match = re.search(r'[Bb]eta ([\d.-]+)', r['reason'])
                    beta_val = beta_match.group(1) if beta_match else '?'
                    print(f"       {r['pair']}: β={beta_val}")
                if len(above_max) > 5:
                    print(f"       ... and {len(above_max) - 5} more")
            
            if below_min:
                print(f"     Below min ({len(below_min)}):")
                for r in below_min[:5]:  # Show first 5
                    # Extract beta value from reason string (case-insensitive)
                    beta_match = re.search(r'[Bb]eta ([\d.-]+)', r['reason'])
                    beta_val = beta_match.group(1) if beta_match else '?'
                    print(f"       {r['pair']}: β={beta_val}")
                if len(below_min) > 5:
                    print(f"       ... and {len(below_min) - 5} more")

    # Show approved trades summary
    if approved_trades:
        approved_df = pd.DataFrame(approved_trades)

        print(f"\n✓ Approved trades:")
        print(f"  Total notional: ${approved_df['Total_Notional'].sum():,.0f}")
        print(f"  New exposure: ${cumulative_new_exposure:,.0f} of ${leverage_capacity_usd:,.0f} capacity")
        print(f"  Portfolio beta: {current_portfolio_beta:.4f}")
        print(f"  Average beta: {approved_df['Beta'].mean():.4f}")
        print(f"  Beta range: [{approved_df['Beta'].min():.4f}, {approved_df['Beta'].max():.4f}]")

        # Tail breakdown
        l_tail = sum(1 for t in approved_trades if t.get('trigger_type') == 'lower')
        u_tail = sum(1 for t in approved_trades if t.get('trigger_type') == 'upper')
        print(f"  By tail: L={l_tail}, U={u_tail}")

        # Index breakdown
        index_counts = approved_df['Index'].value_counts()
        print(f"  By index:")
        for idx, cnt in index_counts.items():
            print(f"    {idx}: {cnt}")

        print(f"  Average spread: {approved_df['Entry_Spread_BPS'].mean():.1f} bps")
        print(f"  Spread range: [{approved_df['Entry_Spread_BPS'].min():.1f}, {approved_df['Entry_Spread_BPS'].max():.1f}] bps")

        # Show final leverage
        if ib and ib.isConnected():
            final_leverage = (current_gross_usd + cumulative_new_exposure) / net_liq_usd
            print(f"  Final projected leverage: {final_leverage:.2f}x / {config.max_account_leverage():.2f}x")

        # Generate spread report
        generate_spread_report(approved_trades, rejected_trades, pairs_missing_data)
        return approved_df
    else:
        print("\n⚠️  No trades approved")
        # Generate spread report even if no approvals
        generate_spread_report([], rejected_trades, pairs_missing_data)

        return pd.DataFrame()

# ============================================================================
# TRADE TERMINATIONS - NO SES
# ============================================================================

def get_us_market_holidays(year):
    """
    Get US stock market holidays for a given year
    
    Returns a set of dates when US markets are closed.
    Note: This is a simplified list - some holidays have early closes not tracked here.
    """
    from datetime import date
    
    holidays = set()
    
    # New Year's Day (Jan 1, or observed on nearest weekday)
    nyd = date(year, 1, 1)
    if nyd.weekday() == 5:  # Saturday -> Friday
        holidays.add(date(year - 1, 12, 31))
    elif nyd.weekday() == 6:  # Sunday -> Monday
        holidays.add(date(year, 1, 2))
    else:
        holidays.add(nyd)
    
    # MLK Day (3rd Monday of January)
    jan1 = date(year, 1, 1)
    first_monday = jan1 + timedelta(days=(7 - jan1.weekday()) % 7)
    mlk = first_monday + timedelta(weeks=2)
    holidays.add(mlk)
    
    # Presidents Day (3rd Monday of February)
    feb1 = date(year, 2, 1)
    first_monday = feb1 + timedelta(days=(7 - feb1.weekday()) % 7)
    presidents = first_monday + timedelta(weeks=2)
    holidays.add(presidents)
    
    # Good Friday (varies - need Easter calculation)
    # Simplified: skip for now, add manually if needed
    
    # Memorial Day (last Monday of May)
    may31 = date(year, 5, 31)
    memorial = may31 - timedelta(days=(may31.weekday() + 7) % 7)
    if memorial.weekday() != 0:
        memorial = may31 - timedelta(days=may31.weekday())
    holidays.add(memorial)
    
    # Juneteenth (June 19, observed)
    june19 = date(year, 6, 19)
    if june19.weekday() == 5:
        holidays.add(date(year, 6, 18))
    elif june19.weekday() == 6:
        holidays.add(date(year, 6, 20))
    else:
        holidays.add(june19)
    
    # Independence Day (July 4, observed)
    july4 = date(year, 7, 4)
    if july4.weekday() == 5:
        holidays.add(date(year, 7, 3))
    elif july4.weekday() == 6:
        holidays.add(date(year, 7, 5))
    else:
        holidays.add(july4)
    
    # Labor Day (1st Monday of September)
    sep1 = date(year, 9, 1)
    labor = sep1 + timedelta(days=(7 - sep1.weekday()) % 7)
    holidays.add(labor)
    
    # Thanksgiving (4th Thursday of November)
    nov1 = date(year, 11, 1)
    first_thursday = nov1 + timedelta(days=(3 - nov1.weekday() + 7) % 7)
    thanksgiving = first_thursday + timedelta(weeks=3)
    holidays.add(thanksgiving)
    
    # Christmas (Dec 25, observed)
    xmas = date(year, 12, 25)
    if xmas.weekday() == 5:
        holidays.add(date(year, 12, 24))
    elif xmas.weekday() == 6:
        holidays.add(date(year, 12, 26))
    else:
        holidays.add(xmas)
    
    return holidays


def is_market_open(check_date):
    """
    Check if US markets are open on a given date
    
    Returns False for weekends and US market holidays
    """
    # Weekend check
    if check_date.weekday() >= 5:
        return False
    
    # Holiday check
    holidays = get_us_market_holidays(check_date.year)
    if check_date in holidays:
        return False
    
    return True


def get_previous_trading_day(from_date):
    """
    Get the previous trading day before a given date
    
    Skips weekends and US market holidays
    """
    prev_day = from_date - timedelta(days=1)
    while not is_market_open(prev_day):
        prev_day -= timedelta(days=1)
    return prev_day


def get_effective_termination_date(term_date, today):
    """
    Get the effective termination date, accounting for holidays and weekends
    
    If termination date falls on a non-trading day, trades should terminate
    on the LAST trading day BEFORE the termination date.
    
    Also catches any past-due trades (termination date already passed).
    
    Returns:
        tuple: (should_terminate: bool, effective_date: date, reason: str)
    """
    # Past-due: termination date already passed
    if term_date < today:
        return True, term_date, f'Past Due (was {term_date})'
    
    # Termination date is today
    if term_date == today:
        if is_market_open(today):
            return True, today, 'Date Reached'
        else:
            # Today is a holiday/weekend but term_date is today
            # This shouldn't normally happen if we're running on a trading day
            return True, today, 'Date Reached (Holiday)'
    
    # Termination date is in the future
    # Check if it falls on a non-trading day
    if not is_market_open(term_date):
        # Find the last trading day before term_date
        effective_date = get_previous_trading_day(term_date)
        
        # If that effective date is today or earlier, terminate now
        if effective_date <= today:
            return True, effective_date, f'Pre-Holiday Exit (term date {term_date} is non-trading day)'
    
    # Not ready for termination yet
    return False, term_date, None


# ============================================================================
# EARLY EXIT FUNCTIONS
# ============================================================================
# Based on V9.2 early exit analysis (107,615 trades, 2016-2024)
# Take-profit exits that capture profits before reversals
# ============================================================================

def calculate_days_held(trade, reference_date=None):
    """
    Calculate the number of calendar days a trade has been held.
    
    Parameters:
    -----------
    trade : Series or dict
        Trade record with 'Trade Initiation Date'
    reference_date : date, optional
        Date to calculate against (default: today)
    
    Returns:
    --------
    int : Number of days held (0 if calculation fails)
    """
    if reference_date is None:
        reference_date = datetime.today().date()
    
    init_date = trade.get('Trade Initiation Date')
    
    if pd.isna(init_date):
        logger.warning(f"Missing initiation date for trade {trade.get('Pair', 'Unknown')}")
        return 0
    
    # Handle different date formats
    if hasattr(init_date, 'date'):
        init_date = init_date.date()
    elif isinstance(init_date, str):
        init_date = pd.to_datetime(init_date).date()
    
    days_held = (reference_date - init_date).days
    return max(0, days_held)  # Ensure non-negative


def check_early_exit(trade, days_held, cumulative_alpha_pct, config_override=None):
    """
    Check if trade should be exited early based on take-profit rules.
    
    Based on V9.2 early exit analysis (107,615 trades, 2016-2024):
    - LOWER (L-tail): Day 7+ at 8% alpha → +0.96% avg improvement
    - UPPER (U-tail): Day 5+ at 8% alpha → +0.78% avg improvement
    
    Parameters:
    -----------
    trade : Series or dict
        Trade record with 'Tail' column
    days_held : int
        Days since entry (1-15+)
    cumulative_alpha_pct : float
        Current cumulative alpha return as PERCENTAGE (e.g., 8.0 = 8%)
    config_override : dict, optional
        Override default EARLY_EXIT_CONFIG
    
    Returns:
    --------
    tuple: (should_exit: bool, exit_reason: str)
    
    Examples:
    ---------
    >>> check_early_exit(trade, days_held=7, cumulative_alpha_pct=9.5)
    (True, 'Day7_TakeProfit_8pct')
    
    >>> check_early_exit(trade, days_held=5, cumulative_alpha_pct=4.0)
    (False, 'Hold - alpha 4.0% below threshold 8.0%')
    """
    # Get configuration
    cfg = config_override or getattr(config, 'EARLY_EXIT_CONFIG', {})
    
    # Master switch check
    if not cfg.get('enabled', False):
        return False, "Early exit disabled"
    
    # Version filter check
    version_filter = cfg.get('version_filter')
    if version_filter is not None:
        trade_version = str(trade.get('Version', '')).strip()
        if trade_version not in version_filter:
            return False, f"Version {trade_version} not in filter {version_filter}"
    
    # Determine strategy type from Tail
    tail = str(trade.get('Tail', 'L')).strip().upper()
    strategy = 'lower' if tail == 'L' else 'upper'
    
    # Get strategy-specific config
    strategy_config = cfg.get(strategy, {})
    
    if not strategy_config.get('enabled', True):
        return False, f"{strategy.upper()} early exit disabled"
    
    # Get thresholds
    exit_day = strategy_config.get('exit_day', 7 if strategy == 'lower' else 5)
    threshold_pct = strategy_config.get('threshold_pct', 8.0)
    
    # Check day requirement
    if days_held < exit_day:
        return False, f"Hold - day {days_held} < exit day {exit_day}"
    
    # Check alpha threshold
    if cumulative_alpha_pct >= threshold_pct:
        reason = f"Day{days_held}_TakeProfit_{int(threshold_pct)}pct"
        return True, reason
    
    return False, f"Hold - alpha {cumulative_alpha_pct:.1f}% below threshold {threshold_pct}%"


# ============================================================================
# TRADE TERMINATION EVALUATION
# ============================================================================

def evaluate_trade_terminations(portfolio_df, parameters_df, earnings_dates=None, ib=None):
    """
    Evaluate which trades should be terminated
    
    Termination priority order:
    1. Force termination (manual override via config.force_terminate_tags())
    2. EARLY EXIT - Take profit based on alpha thresholds
    3. Date-based termination (15-day maturity, including holiday/weekend awareness)
    4. Past-due trades (safety net for any previously missed terminations)
    5. Earnings-based termination
    
    Includes force termination override via config.force_terminate_tags()
    """
    
    if portfolio_df.empty:
        logger.info("No trades in portfolio to evaluate for termination")
        return pd.DataFrame(), portfolio_df
    
    to_terminate = []
    remaining = []
    
    today = datetime.today().date()
    logger.info(f"\nEvaluating terminations for {len(portfolio_df)} trades (today: {today})...")
    
    # Track early exit statistics
    early_exit_checked = 0
    early_exit_triggered = 0
    early_exit_cfg = getattr(config, 'EARLY_EXIT_CONFIG', {})
    
    for idx, trade in portfolio_df.iterrows():
        tag = trade['Tag']
        pair = trade['Pair']
        
        # ====================================================================
        # FORCE TERMINATION CHECK (bypass all other rules)
        # ====================================================================
        
        if config.enable_force_termination():
            # Convert both to strings for comparison (handles int and str Tags)
            force_terminate = str(tag) in [str(t) for t in config.force_terminate_tags()]
            
            if force_terminate:
                logger.warning(f"🚨 FORCE TERMINATING {pair} (Tag: {tag})")
                trade_copy = trade.copy()
                trade_copy['Exit Reason'] = 'Force Terminated (Manual Override)'
                trade_copy['Trade Termination Date'] = today
                to_terminate.append(trade_copy)
                continue
        
        # ====================================================================
        # PRIORITY 2: EARLY EXIT CHECK (Take Profit)
        # ====================================================================
        # Based on V9.2 analysis: Exit early when cumulative alpha exceeds
        # threshold at specified day to lock in profits before reversals
        # - LOWER (L-tail): Day 7+ at 8% alpha → +0.96% avg improvement
        # - UPPER (U-tail): Day 5+ at 8% alpha → +0.78% avg improvement
        # ====================================================================
        
        if early_exit_cfg.get('enabled', False):
            # Calculate days held
            days_held = calculate_days_held(trade, today)
            
            # Get cumulative alpha (already calculated in Stage 4: update_live_alpha_returns)
            cumulative_alpha_pct = trade.get('Live Alpha Return (%)', 0)
            
            if pd.notna(cumulative_alpha_pct) and days_held > 0:
                early_exit_checked += 1
                
                should_exit, exit_reason = check_early_exit(
                    trade, 
                    days_held, 
                    cumulative_alpha_pct
                )
                
                if should_exit:
                    early_exit_triggered += 1
                    
                    # Determine strategy type for logging
                    tail = str(trade.get('Tail', 'L')).strip().upper()
                    strategy = 'LOWER' if tail == 'L' else 'UPPER'
                    
                    logger.info(f"  🎯 EARLY EXIT: {pair} (Tag: {tag})")
                    logger.info(f"     Strategy: {strategy} | Days held: {days_held} | Alpha: {cumulative_alpha_pct:.2f}%")
                    logger.info(f"     Reason: {exit_reason}")
                    
                    trade_copy = trade.copy()
                    trade_copy['Exit Reason'] = f'Early Exit - {exit_reason}'
                    trade_copy['Trade Termination Date'] = today
                    trade_copy['Days Held'] = days_held
                    trade_copy['Exit_Alpha_Pct'] = cumulative_alpha_pct
                    to_terminate.append(trade_copy)
                    continue
                
                elif early_exit_cfg.get('verbose', False):
                    # Log why trade didn't qualify (only in verbose mode)
                    logger.debug(f"  {pair}: {exit_reason}")
        
        # ====================================================================
        # NORMAL TERMINATION LOGIC (Date-based with holiday awareness)
        # ====================================================================
        
        term_date = trade.get('Trade Termination Date')
        
        if pd.notna(term_date):
            # Convert to date if needed
            if isinstance(term_date, str):
                term_date = pd.to_datetime(term_date).date()
            elif isinstance(term_date, pd.Timestamp):
                term_date = term_date.date()
            
            # Check if trade should terminate (handles holidays, weekends, past-due)
            should_terminate, effective_date, reason = get_effective_termination_date(term_date, today)
            
            if should_terminate:
                logger.info(f"  Terminating {pair} (Tag: {tag}): {reason}")
                trade_copy = trade.copy()
                trade_copy['Exit Reason'] = reason
                trade_copy['Days Held'] = calculate_days_held(trade, today)
                to_terminate.append(trade_copy)
                continue
        
        # ====================================================================
        # EARNINGS TERMINATION
        # ====================================================================
        # Exit positions if earnings will be announced:
        # - After hours today (effective date = tomorrow)
        # - Pre-market tomorrow (effective date = tomorrow)
        # ====================================================================
        
        if earnings_dates:
            ticker1 = trade.get('Co1')
            ticker2 = trade.get('Co2')

            # Calculate next trading day
            next_trading_day = today + timedelta(days=1)
            # Skip weekends
            while next_trading_day.weekday() >= 5:
                next_trading_day += timedelta(days=1)

            earnings_triggered = False
            earnings_ticker = None
            earnings_date = None

            # Check ticker1
            if ticker1 in earnings_dates:
                earnings_effective_date = earnings_dates[ticker1]

                # ✅ FIXED LINE:
                if earnings_effective_date in [today, next_trading_day]:  # ← CHANGE TO THIS
                    earnings_triggered = True
                    earnings_ticker = ticker1
                    earnings_date = earnings_effective_date

            # Check ticker2 (if ticker1 doesn't have earnings)
            if not earnings_triggered and ticker2 in earnings_dates:
                earnings_effective_date = earnings_dates[ticker2]

                # ✅ FIXED LINE:
                if earnings_effective_date in [today, next_trading_day]:  # ← CHANGE TO THIS
                    earnings_triggered = True
                    earnings_ticker = ticker2
                    earnings_date = earnings_effective_date
                    logger.info(f"    {ticker2} has upcoming earnings: effective {earnings_effective_date}")

            if earnings_triggered:
                logger.info(f"  Terminating {pair} (Tag: {tag}): {earnings_ticker} earnings on {earnings_date}")
                trade_copy = trade.copy()
                trade_copy['Exit Reason'] = f'Earnings - {earnings_ticker} reports {earnings_date}'
                trade_copy['Trade Termination Date'] = today
                trade_copy['Days Held'] = calculate_days_held(trade, today)
                to_terminate.append(trade_copy)
                continue
        
        # ====================================================================
        # Keep trade in portfolio
        # ====================================================================
        
        remaining.append(trade)
    
    to_terminate_df = pd.DataFrame(to_terminate) if to_terminate else pd.DataFrame()
    remaining_df = pd.DataFrame(remaining) if remaining else pd.DataFrame()
    
    logger.info(f"\nTermination summary:")
    logger.info(f"  To terminate: {len(to_terminate_df)}")
    logger.info(f"  Remaining: {len(remaining_df)}")
    
    # Early exit statistics
    if early_exit_cfg.get('enabled', False):
        logger.info(f"\n  Early Exit Stats:")
        logger.info(f"    Checked: {early_exit_checked} trades")
        logger.info(f"    Triggered: {early_exit_triggered} trades")
        
        if early_exit_triggered > 0 and not to_terminate_df.empty:
            # Break down by exit reason
            early_exits = to_terminate_df[to_terminate_df['Exit Reason'].str.contains('Early Exit', na=False)]
            if not early_exits.empty:
                logger.info(f"    Details:")
                for _, exit_trade in early_exits.iterrows():
                    logger.info(f"      - {exit_trade['Pair']}: {exit_trade.get('Exit_Alpha_Pct', 0):.1f}% alpha, "
                               f"{exit_trade.get('Days Held', 0)} days")
    
    return to_terminate_df, remaining_df

# ============================================================================
# TRADE EXECUTION
# ============================================================================

def append_executed_trades(portfolio_df, executed_trades_df, parameters_df):
    """
    Append executed trades to portfolio, preserving W1, signals, and configuration
    """
    if executed_trades_df is None or executed_trades_df.empty:
        logger.info("No trades to append")
        return portfolio_df
    
    logger.info(f"Appending {len(executed_trades_df)} executed trades")
    
    # Ensure required V9 columns exist
    required_columns = ['W1', 'W2', 'Position_Multiplier', 'trigger_type', 'Sum_Dev_Bucket']
    for col in required_columns:
        if col not in executed_trades_df.columns:
            logger.warning(f"Missing column in executed trades: {col}")
    
    # Set Existing flag
    executed_trades_df['Existing'] = 1
    
    # Add Version column from config (V9.2 trades will be marked as such)
    executed_trades_df['Version'] = config.version()
    
    # Calculate dynamic beta for each trade
    for idx, row in executed_trades_df.iterrows():
        co1 = row['Co1']
        co2 = row['Co2']
        tail = row.get('Tail', 'L')
        W1 = row.get('W1', 0.5)
        
        # Calculate dynamic pair beta
        pair_beta = Tool_Box.calculate_dynamic_pair_beta(
            co1, co2, W1, tail, parameters_file=None  # Already loaded
        )
        executed_trades_df.at[idx, 'Beta'] = pair_beta
        
        logger.info(f"Trade {row['Pair']}: Beta={pair_beta:.4f}, W1={W1:.2%}")
    
    # Concatenate with portfolio - fix FutureWarning
    if portfolio_df.empty:
        updated_portfolio = executed_trades_df.copy()
    else:
        if portfolio_df.empty:
            updated_portfolio = executed_trades_df.copy()
        else:
            updated_portfolio = pd.concat([portfolio_df, executed_trades_df], ignore_index=True)
    
    logger.info(f"Portfolio updated: {len(portfolio_df)} → {len(updated_portfolio)} trades")
    return updated_portfolio

# ============================================================================
# PORTFOLIO SUMMARY
# ============================================================================

def calculate_summary(portfolio_df, options_portfolio_df, live_prices=None, 
                     index_prices=None, opening_prices=None):
    """
    Calculate comprehensive portfolio summary with index breakdowns and returns
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current portfolio
    options_portfolio_df : DataFrame
        Options positions
    live_prices : dict, optional
        Current market prices {ticker: price}
    index_prices : dict, optional
        Current index ETF prices {index_ticker: {'initial': price, 'current': price}}
    
    Returns:
    --------
    dict : Comprehensive summary metrics
    """
    if portfolio_df.empty:
        return {
            'Total Nominal Size': 0,
            'Largest % of Total': 0,
            'Portfolio Beta': 0,
            'Net Beta': 0,
            'Index Breakdown': pd.DataFrame(),
            'Index Returns': pd.DataFrame(),
            'Portfolio Returns': {},
            'Highest Index Allocation': 'N/A'
        }
    
    # ========================================================================
    # Basic Portfolio Metrics
    # ========================================================================
    
    # Calculate total nominal size
    total_nominal = portfolio_df[['Trade Value Co1 ($)', 'Trade Value Co2 ($)']].abs().sum().sum()
    
    # Calculate ticker exposures
    ticker_exposures = Tool_Box.calculate_ticker_exposure(portfolio_df)
    largest_pct = (ticker_exposures.abs().max() / total_nominal * 100) if total_nominal > 0 else 0
    
    # Calculate portfolio beta using VO betas (overall market exposure)
    total_weighted_beta_vo = 0
    for idx, row in portfolio_df.iterrows():
        co1 = row['Co1']
        co2 = row['Co2']
        tail = row.get('Tail', 'L')
        W1 = row.get('W1', 0.5)
        
        # Use VO betas for portfolio-level market exposure
        pair_beta = Tool_Box.calculate_dynamic_pair_beta(co1, co2, W1, tail, beta_type='vo')
        
        trade_value = abs(row.get('Trade Value Co1 ($)', 0)) + abs(row.get('Trade Value Co2 ($)', 0))
        total_weighted_beta_vo += pair_beta * trade_value
    
    portfolio_beta = total_weighted_beta_vo / total_nominal if total_nominal > 0 else 0
    
    # Net Beta (VO) = Portfolio Beta (VO) minus any VO-specific hedges
    # Currently no VO options are held. IGV put hedges affect IGV exposure,
    # not VO beta — their effect is shown in the IGV Exposure section.
    net_beta = portfolio_beta
    
    index_breakdown = calculate_index_breakdown(portfolio_df)
    
    index_returns = calculate_index_returns(index_prices) if index_prices else pd.DataFrame()
    
    # Portfolio Returns (Intraday)
    portfolio_returns = calculate_portfolio_returns(portfolio_df, live_prices, opening_prices) if live_prices else {}
    
    # Highest index allocation by gross exposure
    highest_index_alloc = 'N/A'
    if isinstance(index_breakdown, pd.DataFrame) and not index_breakdown.empty and 'Gross_Exposure_Pct' in index_breakdown.columns:
        max_row = index_breakdown.loc[index_breakdown['Gross_Exposure_Pct'].idxmax()]
        highest_index_alloc = f"{max_row['Index']} ({max_row['Gross_Exposure_Pct']:.1f}%)"
    
    logger.info(f"Summary: Nominal=${total_nominal:,.2f}, Beta={portfolio_beta:.4f}, Net Beta={net_beta:.4f}")
    
    return {
        'Total Nominal Size': total_nominal,
        'Largest % of Total': largest_pct,
        'Portfolio Beta': portfolio_beta,
        'Net Beta': net_beta,
        'Index Breakdown': index_breakdown,
        'Index Returns': index_returns,
        'Portfolio Returns': portfolio_returns,
        'Highest Index Allocation': highest_index_alloc
    }

def calculate_index_breakdown(portfolio_df):
    """
    Calculate breakdown by index: ticker counts, exposures, and beta
    
    **UPDATED**: Now includes Dollar Beta Exposure, % of Total Portfolio Beta,
    L/U trade counts, and ALL indexes from config.index_etfs() (even with 0 trades)
    
    Returns:
    --------
    DataFrame with columns: 
        Index, Trade_Count, L_Trades, U_Trades, Ticker_Count, Gross_Exposure, 
        Net_Exposure, Net_Beta, Dollar_Beta_Exposure, Pct_Of_Total_Dollar_Beta
    """
    # Initialize all indexes from config (ensures VFH etc. appear even with 0 trades)
    index_data = {}
    for idx in config.index_etfs():
        index_data[idx] = {
            'tickers': set(),
            'gross_exposure': 0,
            'long_exposure': 0,
            'short_exposure': 0,
            'weighted_beta': 0,
            'trade_count': 0,
            'l_trade_count': 0,
            'u_trade_count': 0
        }
    
    if not portfolio_df.empty:
        for _, trade in portfolio_df.iterrows():
            index = trade.get('Index', 'Unknown')
            if pd.isna(index):
                index = 'Unknown'
            
            ticker1 = trade['Co1']
            ticker2 = trade['Co2']
            value1 = abs(trade.get('Trade Value Co1 ($)', 0))
            value2 = abs(trade.get('Trade Value Co2 ($)', 0))
            tail = trade.get('Tail', 'L').strip().upper()
            W1 = trade.get('W1', 0.5)
            
            # Initialize index entry if not in config.index_etfs()
            if index not in index_data:
                index_data[index] = {
                    'tickers': set(),
                    'gross_exposure': 0,
                    'long_exposure': 0,
                    'short_exposure': 0,
                    'weighted_beta': 0,
                    'trade_count': 0,
                    'l_trade_count': 0,
                    'u_trade_count': 0
                }
            
            # Count trades by tail type
            index_data[index]['trade_count'] += 1
            if tail == 'L':
                index_data[index]['l_trade_count'] += 1
            else:
                index_data[index]['u_trade_count'] += 1
            
            # Add tickers
            index_data[index]['tickers'].add(ticker1)
            index_data[index]['tickers'].add(ticker2)
            
            # Add exposures
            index_data[index]['gross_exposure'] += value1 + value2
            
            if tail == 'L':
                # Long ticker1, short ticker2
                index_data[index]['long_exposure'] += value1
                index_data[index]['short_exposure'] += value2
            else:
                # Short ticker1, long ticker2
                index_data[index]['short_exposure'] += value1
                index_data[index]['long_exposure'] += value2
            
            # Calculate pair beta using market betas (vs parent ETF) for index breakdown
            pair_beta = Tool_Box.calculate_dynamic_pair_beta(ticker1, ticker2, W1, tail, beta_type='market')
            trade_value = value1 + value2
            index_data[index]['weighted_beta'] += pair_beta * trade_value
    
    # Convert to DataFrame
    breakdown_list = []
    total_dollar_beta = 0  # Track total for percentage calculation
    
    for index, data in index_data.items():
        net_exposure = data['long_exposure'] - data['short_exposure']
        net_beta = data['weighted_beta'] / data['gross_exposure'] if data['gross_exposure'] > 0 else 0
        
        # Calculate dollar beta exposure
        dollar_beta_exposure = data['weighted_beta']
        total_dollar_beta += abs(dollar_beta_exposure)
        
        breakdown_list.append({
            'Index': index,
            'Trade_Count': data['trade_count'],
            'L_Trades': data['l_trade_count'],
            'U_Trades': data['u_trade_count'],
            'Ticker_Count': len(data['tickers']),
            'Gross_Exposure': data['gross_exposure'],
            'Net_Exposure': net_exposure,
            'Net_Beta': net_beta,
            'Dollar_Beta_Exposure': dollar_beta_exposure
        })
    
    breakdown_df = pd.DataFrame(breakdown_list)
    
    # Add percentage of total dollar beta
    if total_dollar_beta > 0:
        breakdown_df['Pct_Of_Total_Dollar_Beta'] = (
            breakdown_df['Dollar_Beta_Exposure'].abs() / total_dollar_beta * 100
        )
    else:
        breakdown_df['Pct_Of_Total_Dollar_Beta'] = 0
    
    # Add percentage of total gross exposure
    total_gross_exposure = breakdown_df['Gross_Exposure'].sum()
    if total_gross_exposure > 0:
        breakdown_df['Gross_Exposure_Pct'] = (
            breakdown_df['Gross_Exposure'] / total_gross_exposure * 100
        )
    else:
        breakdown_df['Gross_Exposure_Pct'] = 0
    
    # Sort by Index name for consistency
    if not breakdown_df.empty:
        breakdown_df = breakdown_df.sort_values('Index')
    
    return breakdown_df

def calculate_index_returns(index_prices):
    """
    Calculate intraday returns for index ETFs
    
    Parameters:
    -----------
    index_prices : dict
        Format: {ticker: {'initial': price, 'current': price}}
    
    Returns:
    --------
    DataFrame with columns: Index, Initial_Price, Current_Price, Return_Pct
    """
    if not index_prices:
        return pd.DataFrame()
    
    returns_list = []
    
    for ticker, prices in index_prices.items():
        initial = prices.get('initial')
        current = prices.get('current')
        
        if initial and current and initial > 0:
            return_pct = ((current - initial) / initial) * 100
            
            returns_list.append({
                'Index': ticker,
                'Initial_Price': initial,
                'Current_Price': current,
                'Return_Pct': return_pct
            })
    
    returns_df = pd.DataFrame(returns_list)
    
    # Sort by ticker
    if not returns_df.empty:
        returns_df = returns_df.sort_values('Index')
    
    return returns_df

def calculate_portfolio_returns(portfolio_df, live_prices, opening_prices=None):
    """
    Calculate portfolio returns (cap-weighted and equal-weighted)
    
    **FIXED**: Properly populates Index_Return_Pct column
    **REMOVED**: VGT_Ex_Megacaps column (no longer needed)
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current portfolio
    live_prices : dict
        Current prices {ticker: price}
    opening_prices : dict, optional
        TODAY's opening prices {ticker: price}
    
    Returns:
    --------
    dict : {'cap_weighted': float, 'equal_weighted': float, 'by_index': DataFrame}
    """
    if portfolio_df.empty or not live_prices:
        return {
            'cap_weighted': 0.0, 
            'equal_weighted': 0.0,
            'by_index': pd.DataFrame()
        }
    
    trade_returns = []
    trade_values = []
    trade_indexes = []
    
    for _, trade in portfolio_df.iterrows():
        ticker1 = trade['Co1']
        ticker2 = trade['Co2']
        tail = trade.get('Tail', 'L').strip().upper()
        index = trade.get('Index', 'Unknown')
        
        # Get current prices
        current_price1 = live_prices.get(ticker1)
        current_price2 = live_prices.get(ticker2)
        
        # Get TODAY's opening prices (or use current as fallback)
        if opening_prices:
            open_price1 = opening_prices.get(ticker1, current_price1)
            open_price2 = opening_prices.get(ticker2, current_price2)
        else:
            logger.warning("No opening prices provided - cannot calculate accurate intraday returns")
            open_price1 = current_price1
            open_price2 = current_price2
        
        # Skip if missing data
        if pd.isna(open_price1) or pd.isna(open_price2):
            continue
        if pd.isna(current_price1) or pd.isna(current_price2):
            continue
        
        # Calculate individual ticker returns (INTRADAY)
        return1 = ((current_price1 - open_price1) / open_price1) if open_price1 > 0 else 0
        return2 = ((current_price2 - open_price2) / open_price2) if open_price2 > 0 else 0
        
        # Get weights
        W1 = trade.get('W1', 0.5)
        W2 = 1 - W1
        
        # Calculate pair return based on tail (weighted by position size)
        if tail == 'L':
            # Long ticker1, short ticker2
            pair_return = (W1 * return1) - (W2 * return2)
        else:
            # Short ticker1, long ticker2
            pair_return = -(W1 * return1) + (W2 * return2)
        
        # Store for weighted averaging
        trade_value = abs(trade.get('Trade Value Co1 ($)', 0)) + abs(trade.get('Trade Value Co2 ($)', 0))
        trade_returns.append(pair_return)
        trade_values.append(trade_value)
        trade_indexes.append(index)
    
    if not trade_returns:
        return {
            'cap_weighted': 0.0, 
            'equal_weighted': 0.0,
            'by_index': pd.DataFrame()
        }
    
    # Calculate cap-weighted return
    total_value = sum(trade_values)
    cap_weighted = sum(r * v for r, v in zip(trade_returns, trade_values)) / total_value if total_value > 0 else 0
    
    # Calculate equal-weighted return
    equal_weighted = sum(trade_returns) / len(trade_returns)
    
    # Calculate returns by index
    index_returns_data = {}
    for ret, val, idx in zip(trade_returns, trade_values, trade_indexes):
        if idx not in index_returns_data:
            index_returns_data[idx] = {
                'weighted_return_sum': 0,
                'raw_return_sum': 0,
                'total_value': 0, 
                'count': 0
            }
        
        index_returns_data[idx]['weighted_return_sum'] += ret * val
        index_returns_data[idx]['raw_return_sum'] += ret
        index_returns_data[idx]['total_value'] += val
        index_returns_data[idx]['count'] += 1
    
    # Create by_index DataFrame
    by_index_list = []
    for idx, data in index_returns_data.items():
        # Cap-weighted return
        cap_weighted_idx = (data['weighted_return_sum'] / data['total_value'] * 100) if data['total_value'] > 0 else 0
        
        # Equal-weighted return
        equal_weighted_idx = (data['raw_return_sum'] / data['count'] * 100) if data['count'] > 0 else 0
        
        # ====================================================================
        # Calculate index ETF return (megacap-adjusted for VGT)
        # ====================================================================
        index_return_pct = None
        
        # idx here is the sector index (VGT, VIS, VCR, VHT, VFH)
        if opening_prices and live_prices:
            # Use megacap-adjusted return for all indices
            adjusted_return = get_adjusted_index_return(idx, live_prices, opening_prices)
            index_return_pct = adjusted_return * 100  # Convert to percentage
        
        row_data = {
            'Index': idx,
            'Trade_Count': data['count'],
            'Cap_Weighted_Return_Pct': cap_weighted_idx,
            'Equal_Weighted_Return_Pct': equal_weighted_idx,
            'Index_Return_Pct': index_return_pct if index_return_pct is not None else 0.0
        }
        
        # **REMOVED**: No more VGT_Ex_Megacaps column
        # All indexes treated uniformly
        
        by_index_list.append(row_data)
    
    by_index_df = pd.DataFrame(by_index_list)
    if not by_index_df.empty:
        by_index_df = by_index_df.sort_values('Index')
    
    # Convert to percentage
    return {
        'cap_weighted': cap_weighted * 100,
        'equal_weighted': equal_weighted * 100,
        'by_index': by_index_df
    }

# ============================================================================
# PORTFOLIO ANALYTICS
# ============================================================================
# Beta categorization functions imported from tool_box:
# - categorize_ticker_beta (for individual ticker betas, always positive)
# - categorize_pair_beta (for pair betas, can be negative)
# - get_ticker_beta_buckets, get_pair_beta_buckets

def calculate_ticker_exposure_detailed(portfolio_df):
    """Calculate detailed exposure by ticker"""
    ticker_data = {}
    
    for _, row in portfolio_df.iterrows():
        co1 = row['Co1']
        co2 = row['Co2']
        value1 = row.get('Trade Value Co1 ($)', 0)
        value2 = row.get('Trade Value Co2 ($)', 0)
        tail = row.get('Tail', 'L').strip().upper()
        
        # Process Co1
        if co1 not in ticker_data:
            ticker_data[co1] = {'Long': 0, 'Short': 0, 'Net': 0, 'Gross': 0}
        
        if tail == 'L':
            ticker_data[co1]['Long'] += abs(value1)
        else:
            ticker_data[co1]['Short'] += abs(value1)
        
        # Process Co2
        if co2 not in ticker_data:
            ticker_data[co2] = {'Long': 0, 'Short': 0, 'Net': 0, 'Gross': 0}
        
        if tail == 'L':
            ticker_data[co2]['Short'] += abs(value2)
        else:
            ticker_data[co2]['Long'] += abs(value2)
    
    # Calculate net and gross for each ticker
    exposure_list = []
    for ticker, data in ticker_data.items():
        data['Net'] = data['Long'] - data['Short']
        data['Gross'] = data['Long'] + data['Short']
        
        exposure_list.append({
            'Ticker': ticker,
            'Long_Exposure': data['Long'],
            'Short_Exposure': data['Short'],
            'Net_Exposure': data['Net'],
            'Gross_Exposure': data['Gross'],
            'Direction': 'Long' if data['Net'] > 0 else 'Short' if data['Net'] < 0 else 'Neutral'
        })
    
    return pd.DataFrame(exposure_list).sort_values('Gross_Exposure', ascending=False)

def fetch_today_open_prices(ib, tickers):
    """
    DEPRECATED: Use yesterday's closes instead of today's opens.
    
    Today's return should be calculated as: yesterday_close → current_price
    This includes overnight gaps which are real returns.
    
    Kept for backward compatibility but no longer called by workflow.
    
    Fetch today's opening prices for a list of tickers
    Uses historical bars with 1-day duration to get today's open
    """
    import warnings
    warnings.warn(
        "fetch_today_open_prices is deprecated. Use yesterday's closes instead to include overnight gaps.",
        DeprecationWarning
    )
    
    from ib_insync import Stock, util
    import pandas as pd
    from datetime import datetime
    
    opening_prices = {}
    
    for ticker in tickers:
        try:
            contract = Stock(ticker, 'SMART', 'USD')
            qualified = ib.qualifyContracts(contract)
            
            if not qualified:
                continue
            
            # Get today's bar
            bars = ib.reqHistoricalData(
                qualified[0],
                endDateTime='',
                durationStr='1 D',
                barSizeSetting='1 day',
                whatToShow='TRADES',
                useRTH=True
            )
            
            if bars:
                opening_prices[ticker] = bars[0].open
            
        except Exception as e:
            logger.warning(f"Could not fetch open price for {ticker}: {e}")
    
    return opening_prices


# ============================================================================
# FUNCTION 1: TICKER-LEVEL PERFORMANCE CALCULATION
# ============================================================================

def calculate_ticker_performance_detailed(portfolio_df, live_prices, opening_prices, 
                                         yesterday_closes):
    """
    Calculate performance metrics for individual tickers
    
    Returns are DIRECTION-ADJUSTED (P&L perspective):
    - For long positions: positive return = profit
    - For short positions: returns are flipped (negative ticker return = profit)
    
    This means the Today_Alpha_% and Today_Nominal_% columns show actual
    portfolio contribution, and should aggregate to match portfolio-level returns.
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current active portfolio
    live_prices : dict
        Current market prices {ticker: price}
    opening_prices : dict
        Today's opening prices {ticker: price}
    yesterday_closes : dict
        Yesterday's closing prices {ticker: price}
    
    Returns:
    --------
    DataFrame with columns:
        - Ticker
        - Direction (Long/Short net direction)
        - Net_Exposure ($)
        - Active_Positions (count)
        - Beta (actual ticker beta)
        - Cumulative_Alpha_% (direction-adjusted)
        - Cumulative_Nominal_% (direction-adjusted)
        - Today_Alpha_% (direction-adjusted)
        - Today_Nominal_% (direction-adjusted)
    """
    if portfolio_df.empty:
        return pd.DataFrame()
    
    # Load actual ticker betas from SubSector_Beta_Analysis files
    ticker_betas = get_cached_ticker_betas()
    logger.info(f"Using {len(ticker_betas)} actual ticker betas for performance calculation")
    
    ticker_data = {}
    
    # Initialize for each unique ticker
    all_tickers = set(portfolio_df['Co1'].dropna().tolist() + 
                     portfolio_df['Co2'].dropna().tolist())
    
    for ticker in all_tickers:
        ticker_data[ticker] = {
            'positions': [],
            'long_exposure': 0,
            'short_exposure': 0,
            'cumulative_alpha': [],
            'cumulative_nominal': [],
            'today_alpha': [],
            'today_nominal': [],
            'beta': ticker_betas.get(ticker.upper(), 1.0)  # Store actual beta
        }
    
    # Collect data for each ticker across all positions
    for _, row in portfolio_df.iterrows():
        co1 = row['Co1']
        co2 = row['Co2']
        tail = row.get('Tail', 'L').strip().upper()
        
        value1 = abs(row.get('Trade Value Co1 ($)', 0))
        value2 = abs(row.get('Trade Value Co2 ($)', 0))
        
        pair_id = row.get('Pair', 'Unknown')
        
        # Get prices
        current_co1 = live_prices.get(co1)
        current_co2 = live_prices.get(co2)
        init_co1 = row.get('Co1 at Initiation')
        init_co2 = row.get('Co2 at Initiation')
        open_co1 = opening_prices.get(co1)
        open_co2 = opening_prices.get(co2)
        yesterday_co1 = yesterday_closes.get(co1)
        yesterday_co2 = yesterday_closes.get(co2)
        
        # Get ACTUAL ticker betas (not estimated from pair beta)
        co1_beta = ticker_betas.get(co1.upper(), 1.0)
        co2_beta = ticker_betas.get(co2.upper(), 1.0)
        
        # Get index for beta adjustment - VERSION AWARE
        # V9 trades use VO index, V9.2 trades use sector ETF
        trade_version = str(row.get('Version', 'V9'))
        trade_index = row.get('Index', 'VO')
        init_index = row.get('Index at Initiation')
        
        # Use correct current index based on version
        if trade_version in ['9.0C', 'V9', '9.0']:
            # V9 uses VO as the market factor
            current_index_price = live_prices.get('VO')
        else:
            # V9.2 uses sector ETF
            current_index_price = live_prices.get(trade_index)
        
        # Also get today's index open for intraday alpha - VERSION AWARE
        if trade_version in ['9.0C', 'V9', '9.0']:
            index_open = opening_prices.get('VO')
            index_yesterday = yesterday_closes.get('VO')
        else:
            index_open = opening_prices.get(trade_index)
            index_yesterday = yesterday_closes.get(trade_index)
        
        # Calculate megacap-adjusted index return (for V9.2+ trades)
        if trade_version not in ['9.0C', 'V9', '9.0'] and index_open and index_open > 0:
            # Use megacap-adjusted return for all indices
            adjusted_index_intraday = get_adjusted_index_return(trade_index, live_prices, opening_prices) * 100
        elif index_open and current_index_price and index_open > 0:
            adjusted_index_intraday = ((current_index_price - index_open) / index_open * 100)
        else:
            adjusted_index_intraday = 0
        
        # ====================================================================
        # CO1 ANALYSIS
        # ====================================================================
        
        if pd.notna(current_co1) and pd.notna(init_co1) and init_co1 > 0:
            ticker_data[co1]['positions'].append(pair_id)
            
            # Cumulative nominal return (since initiation)
            cumul_nominal_co1 = ((current_co1 - init_co1) / init_co1 * 100)
            
            # Cumulative alpha return using ACTUAL ticker beta
            if pd.notna(current_index_price) and pd.notna(init_index) and init_index > 0:
                index_return = (current_index_price - init_index) / init_index
                beta_contribution = co1_beta * index_return * 100
                cumul_alpha_co1 = cumul_nominal_co1 - beta_contribution
            else:
                cumul_alpha_co1 = cumul_nominal_co1
            
            # Today's returns
            if pd.notna(open_co1) and pd.notna(yesterday_co1) and yesterday_co1 > 0:
                today_nominal_co1 = ((current_co1 - yesterday_co1) / yesterday_co1 * 100)
                today_intraday_return = ((current_co1 - open_co1) / open_co1 * 100) if open_co1 > 0 else 0
                
                # Calculate today's alpha using ACTUAL ticker beta and megacap-adjusted index
                today_alpha_co1 = today_intraday_return - (co1_beta * adjusted_index_intraday)
            else:
                today_nominal_co1 = 0
                today_alpha_co1 = 0
            
            # Track direction based on tail
            if tail == 'L':
                ticker_data[co1]['long_exposure'] += value1
                direction_multiplier = 1  # Long: positive alpha is good
            else:
                ticker_data[co1]['short_exposure'] += value1
                direction_multiplier = -1  # Short: negative alpha is good, flip sign
            
            # Apply direction to returns - shows P&L contribution perspective
            ticker_data[co1]['cumulative_alpha'].append(cumul_alpha_co1 * direction_multiplier)
            ticker_data[co1]['cumulative_nominal'].append(cumul_nominal_co1 * direction_multiplier)
            ticker_data[co1]['today_alpha'].append(today_alpha_co1 * direction_multiplier)
            ticker_data[co1]['today_nominal'].append(today_nominal_co1 * direction_multiplier)
        
        # ====================================================================
        # CO2 ANALYSIS
        # ====================================================================
        
        if pd.notna(current_co2) and pd.notna(init_co2) and init_co2 > 0:
            ticker_data[co2]['positions'].append(pair_id)
            
            # Cumulative nominal return
            cumul_nominal_co2 = ((current_co2 - init_co2) / init_co2 * 100)
            
            # Cumulative alpha return using ACTUAL ticker beta
            if pd.notna(current_index_price) and pd.notna(init_index) and init_index > 0:
                index_return = (current_index_price - init_index) / init_index
                beta_contribution = co2_beta * index_return * 100
                cumul_alpha_co2 = cumul_nominal_co2 - beta_contribution
            else:
                cumul_alpha_co2 = cumul_nominal_co2
            
            # Today's returns
            if pd.notna(open_co2) and pd.notna(yesterday_co2) and yesterday_co2 > 0:
                today_nominal_co2 = ((current_co2 - yesterday_co2) / yesterday_co2 * 100)
                today_intraday_return = ((current_co2 - open_co2) / open_co2 * 100) if open_co2 > 0 else 0
                
                # Calculate today's alpha using ACTUAL ticker beta and megacap-adjusted index
                today_alpha_co2 = today_intraday_return - (co2_beta * adjusted_index_intraday)
            else:
                today_nominal_co2 = 0
                today_alpha_co2 = 0
            
            # Track direction - Co2 is opposite of Co1
            if tail == 'L':
                ticker_data[co2]['short_exposure'] += value2
                direction_multiplier = -1  # Short: flip sign
            else:
                ticker_data[co2]['long_exposure'] += value2
                direction_multiplier = 1  # Long: keep sign
            
            # Apply direction to returns - shows P&L contribution perspective
            ticker_data[co2]['cumulative_alpha'].append(cumul_alpha_co2 * direction_multiplier)
            ticker_data[co2]['cumulative_nominal'].append(cumul_nominal_co2 * direction_multiplier)
            ticker_data[co2]['today_alpha'].append(today_alpha_co2 * direction_multiplier)
            ticker_data[co2]['today_nominal'].append(today_nominal_co2 * direction_multiplier)
    
    # Create output DataFrame
    ticker_performance = []
    
    for ticker, data in ticker_data.items():
        if not data['positions']:
            continue
        
        net_exposure = data['long_exposure'] - data['short_exposure']
        direction = 'Long' if net_exposure > 0 else 'Short' if net_exposure < 0 else 'Neutral'
        
        # Calculate averages
        cumul_alpha_avg = np.mean(data['cumulative_alpha']) if data['cumulative_alpha'] else 0
        cumul_nominal_avg = np.mean(data['cumulative_nominal']) if data['cumulative_nominal'] else 0
        today_alpha_avg = np.mean(data['today_alpha']) if data['today_alpha'] else 0
        today_nominal_avg = np.mean(data['today_nominal']) if data['today_nominal'] else 0
        
        ticker_performance.append({
            'Ticker': ticker,
            'Direction': direction,
            'Net_Exposure_$': net_exposure,
            'Active_Positions': len(data['positions']),
            'Beta': data.get('beta', 1.0),  # Include actual beta
            'Cumulative_Alpha_%': cumul_alpha_avg,
            'Cumulative_Nominal_%': cumul_nominal_avg,
            'Today_Alpha_%': today_alpha_avg,
            'Today_Nominal_%': today_nominal_avg
        })
    
    result_df = pd.DataFrame(ticker_performance)
    if not result_df.empty:
        result_df = result_df.sort_values('Net_Exposure_$', ascending=False, key=abs)
    
    return result_df


# ============================================================================
# OPTIONS P&L CALCULATION
# ============================================================================

def load_options_portfolio(file_path=None):
    """
    Load options from Portfolio.xlsx Options sheet.
    
    Returns:
    --------
    DataFrame : Options positions
    """
    if file_path is None:
        file_path = config.portfolio_file()
    
    try:
        if os.path.exists(file_path):
            options_df = pd.read_excel(file_path, sheet_name='Options')
            if not options_df.empty:
                logger.debug(f"Loaded {len(options_df)} options from portfolio")
            return options_df
    except Exception as e:
        logger.debug(f"Could not load options: {e}")
    
    return pd.DataFrame()


def calculate_options_pnl(options_df, live_prices, yesterday_closes, igv_price=None):
    """
    Calculate options P&L for the day using delta-based estimation.
    
    Since we don't have reliable yesterday option prices, we use:
    P&L ≈ Delta × Contracts × 100 × IGV_Move
    
    This is accurate for small moves and avoids stale price issues.
    
    Parameters:
    -----------
    options_df : DataFrame
        Options from Options sheet
    live_prices : dict
        Current market prices (must include IGV)
    yesterday_closes : dict
        Yesterday's closing prices (must include IGV)
    igv_price : float, optional
        Current IGV price (overrides live_prices if provided)
    
    Returns:
    --------
    dict : {
        'total_pnl': float,           # Total options P&L ($)
        'starting_value': float,       # Yesterday's total value
        'current_value': float,        # Today's total value
        'delta_pnl': float,           # P&L from delta exposure (primary)
        'positions': list,            # Per-position breakdown
    }
    """
    result = {
        'total_pnl': 0.0,
        'starting_value': 0.0,
        'current_value': 0.0,
        'delta_pnl': 0.0,
        'positions': [],
    }
    
    if options_df is None or options_df.empty:
        return result
    
    # Get IGV prices - try multiple sources
    if igv_price is None:
        igv_price = live_prices.get('IGV') if live_prices else None
    
    igv_yesterday = yesterday_closes.get('IGV') if yesterday_closes else None
    
    # If IGV not in provided dicts, try to fetch/estimate
    if igv_price is None or igv_yesterday is None:
        try:
            # Try to fetch IGV price
            from src.execution.trade_execution import fetch_market_data
            igv_data = fetch_market_data(['IGV'])
            if 'IGV' in igv_data and igv_data['IGV'].get('last'):
                if igv_price is None:
                    igv_price = igv_data['IGV']['last']
                if igv_yesterday is None:
                    # Estimate yesterday from close if available
                    igv_yesterday = igv_data['IGV'].get('close', igv_price)
        except:
            pass
        
        # Last resort: use VGT as proxy (IGV ≈ VGT × 0.11)
        if igv_price is None and live_prices:
            vgt_price = live_prices.get('VGT')
            if vgt_price:
                igv_price = vgt_price * 0.11  # Approximate ratio
                logger.debug(f"Using VGT proxy for IGV: ${igv_price:.2f}")
        
        if igv_yesterday is None and yesterday_closes:
            vgt_yesterday = yesterday_closes.get('VGT')
            if vgt_yesterday:
                igv_yesterday = vgt_yesterday * 0.11
    
    if not igv_price or not igv_yesterday:
        logger.warning("Cannot calculate options P&L: missing IGV prices")
        return result
    
    igv_move = igv_price - igv_yesterday
    
    logger.info(f"Options P&L calc: IGV {igv_yesterday:.2f} → {igv_price:.2f} (move: {igv_move:+.2f})")
    
    for _, opt in options_df.iterrows():
        symbol = opt.get('Symbol', '')
        contracts = opt.get('Contracts', 0)
        
        if contracts == 0:
            continue
        
        # Get delta (negative for puts)
        delta = opt.get('Current Delta', -0.5)
        if pd.isna(delta):
            delta = -0.5 if opt.get('Symbol', '').upper() == 'IGV' else 0
        
        # Delta-based P&L: delta × contracts × 100 × IGV_move
        # For puts with negative delta, when IGV falls, P&L is positive
        position_pnl = delta * abs(contracts) * 100 * igv_move
        
        # Estimate values for tracking
        current_price = opt.get('Current Price', 0)
        if pd.isna(current_price):
            current_price = opt.get('Purchase Price', 0)
        
        current_val = current_price * abs(contracts) * 100 if current_price else 0
        
        result['delta_pnl'] += position_pnl
        result['current_value'] += current_val
        
        result['positions'].append({
            'symbol': symbol,
            'strike': opt.get('Strike'),
            'contracts': contracts,
            'delta': delta,
            'igv_move': igv_move,
            'pnl': position_pnl,
        })
    
    # Use delta P&L as total (most reliable)
    result['total_pnl'] = result['delta_pnl']
    
    if result['total_pnl'] != 0:
        logger.info(f"Options P&L (delta-based): ${result['total_pnl']:+,.2f}")
    
    return result


# ============================================================================
# FUNCTION 2: DAILY PORTFOLIO PERFORMANCE CALCULATION
# ============================================================================

def calculate_daily_portfolio_performance(portfolio_df, live_prices, yesterday_closes,
                                          net_liquidation_usd=None, total_nominal=None,
                                          index_prices=None):
    """
    Calculate daily portfolio performance metrics INCLUDING terminated trades.
    
    CRITICAL FIX: This now loads terminated trades from the daily file and includes
    their intraday P&L in the total. Without this, days with terminations would show
    incorrect daily returns (missing the terminated trades' contribution).
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current active portfolio (terminated trades excluded)
    live_prices : dict
        Current market prices
    yesterday_closes : dict
        Yesterday's closing prices
    net_liquidation_usd : float, optional
        Starting net liquidation value in USD (for % of NAV calculation)
    total_nominal : float, optional
        Total nominal portfolio size (for cap-weighted returns)
    index_prices : dict, optional
        Index ETF prices {ticker: {'initial': float, 'current': float}}
    
    Returns:
    --------
    dict with keys:
        - daily_alpha_return: Portfolio alpha return (%)
        - daily_nominal_return: Portfolio nominal return (%)
        - nominal_return_pct_nav: Nominal return as % of starting NAV
        - vo_return: VO index return (%)
        - alpha_win_rate: % of pairs with positive alpha
        - nominal_win_rate: % of pairs with positive nominal return
        - terminated_trades_included: Number of terminated trades in calculation
        - terminated_pnl: P&L from terminated trades ($)
    """
    # ========== DEBUG: Print input data status ==========
    print("\n" + "="*60)
    print("DEBUG: calculate_daily_portfolio_performance INPUTS")
    print("="*60)
    print(f"  live_prices: {len(live_prices) if live_prices else 0} entries")
    print(f"  yesterday_closes: {len(yesterday_closes) if yesterday_closes else 0} entries")
    print(f"  index_prices: {len(index_prices) if index_prices else 0} entries")
    
    # Check sector ETFs
    sector_etfs = ['VGT', 'VIS', 'VCR', 'VHT', 'VFH']
    print("\n  Sector ETF availability:")
    for etf in sector_etfs:
        lp = live_prices.get(etf) if live_prices else None
        yc = yesterday_closes.get(etf) if yesterday_closes else None
        ip = index_prices.get(etf) if index_prices else None
        print(f"    {etf}: live={lp}, yesterday={yc}, index_prices={ip}")
    print("="*60 + "\n")
    # ========== END DEBUG ==========
    
    # ========================================================================
    # CRITICAL: Load terminated trades from daily file
    # ========================================================================
    terminated_df = load_daily_terminated_trades()
    terminated_metrics = calculate_terminated_trades_intraday_metrics(terminated_df, index_prices)
    
    if terminated_metrics['trade_count'] > 0:
        print("="*60)
        print(f"INCLUDING {terminated_metrics['trade_count']} TERMINATED TRADES IN DAILY PERFORMANCE")
        print(f"  Terminated trades intraday P&L: ${terminated_metrics['total_intraday_pnl']:,.2f}")
        print(f"  Terminated trades value: ${terminated_metrics['total_intraday_value']:,.2f}")
        print("="*60 + "\n")
    
    # ========================================================================
    # Calculate metrics for ACTIVE portfolio (existing logic)
    # ========================================================================
    
    if portfolio_df.empty and terminated_metrics['trade_count'] == 0:
        return {
            'daily_alpha_return': 0.0,
            'daily_nominal_return': 0.0,
            'nominal_return_pct_nav': 0.0,
            'vo_return': 0.0,
            'alpha_win_rate': 0.0,
            'nominal_win_rate': 0.0,
            'terminated_trades_included': 0,
            'terminated_pnl': 0.0
        }
    
    # Load actual ticker betas
    ticker_betas = get_cached_ticker_betas()
    
    pair_returns = []
    pair_pnls = []
    index_fallback_count = 0
    alpha_equals_nominal_count = 0
    
    opening_prices = yesterday_closes  # Use yesterday's close as proxy
    
    for _, row in portfolio_df.iterrows():
        co1 = row['Co1']
        co2 = row['Co2']
        tail = row.get('Tail', 'L').strip().upper()
        W1 = row.get('W1', 0.5)
        W2 = 1 - W1
        
        value1 = abs(row.get('Trade Value Co1 ($)', 0))
        value2 = abs(row.get('Trade Value Co2 ($)', 0))
        qty1 = abs(row.get('Quantity1', 0))
        qty2 = abs(row.get('Quantity2', 0))
        
        co1_beta = ticker_betas.get(co1.upper(), 1.0) if ticker_betas else 1.0
        co2_beta = ticker_betas.get(co2.upper(), 1.0) if ticker_betas else 1.0
        
        current_co1 = live_prices.get(co1)
        current_co2 = live_prices.get(co2)
        yesterday_co1 = yesterday_closes.get(co1)
        yesterday_co2 = yesterday_closes.get(co2)
        
        if pd.isna(current_co1) or pd.isna(current_co2) or \
           pd.isna(yesterday_co1) or pd.isna(yesterday_co2):
            continue
        
        if yesterday_co1 <= 0 or yesterday_co2 <= 0:
            continue
        
        return_co1 = (current_co1 - yesterday_co1) / yesterday_co1
        return_co2 = (current_co2 - yesterday_co2) / yesterday_co2
        
        trade_index = row.get('Index', 'VO')
        
        index_current = live_prices.get(trade_index)
        index_yesterday = yesterday_closes.get(trade_index)
        
        used_fallback = False
        if index_prices and trade_index in index_prices:
            idx_data = index_prices[trade_index]
            if index_current is None or pd.isna(index_current):
                index_current = idx_data.get('current')
                used_fallback = True
            if index_yesterday is None or pd.isna(index_yesterday):
                index_yesterday = idx_data.get('initial')
                used_fallback = True
        
        if used_fallback:
            index_fallback_count += 1
        
        if tail == 'L':
            pair_nominal_return = (W1 * return_co1) - (W2 * return_co2)
            pnl_co1 = qty1 * (current_co1 - yesterday_co1)
            pnl_co2 = qty2 * (yesterday_co2 - current_co2)
        else:
            pair_nominal_return = -(W1 * return_co1) + (W2 * return_co2)
            pnl_co1 = qty1 * (yesterday_co1 - current_co1)
            pnl_co2 = qty2 * (current_co2 - yesterday_co2)
        
        pair_pnl = pnl_co1 + pnl_co2
        
        if pd.notna(index_current) and pd.notna(index_yesterday) and index_yesterday > 0:
            # Use megacap-adjusted return for all indices
            index_return = calculate_megacap_adjusted_return(
                trade_index, 
                (index_current - index_yesterday) / index_yesterday,
                live_prices, 
                opening_prices
            )
            
            co1_alpha = return_co1 - (co1_beta * index_return)
            co2_alpha = return_co2 - (co2_beta * index_return)
            
            if tail == 'L':
                pair_alpha_return = (W1 * co1_alpha) - (W2 * co2_alpha)
            else:
                pair_alpha_return = -(W1 * co1_alpha) + (W2 * co2_alpha)
        else:
            pair_alpha_return = pair_nominal_return
            alpha_equals_nominal_count += 1
        
        pair_returns.append({
            'nominal': pair_nominal_return,
            'alpha': pair_alpha_return,
            'value': value1 + value2
        })
        pair_pnls.append(pair_pnl)
    
    # Log diagnostics
    total_pairs_processed = len(pair_returns)
    if index_fallback_count > 0:
        logger.info(f"Daily Performance: Used index_prices fallback for {index_fallback_count}/{total_pairs_processed} pairs")
    
    if alpha_equals_nominal_count > 0:
        logger.warning(f"⚠️  {alpha_equals_nominal_count}/{total_pairs_processed} pairs have alpha=nominal (missing index)")
    
    # ========================================================================
    # COMBINE ACTIVE AND TERMINATED TRADE METRICS
    # ========================================================================
    
    # Calculate VO return
    vo_current = live_prices.get('VO')
    vo_yesterday = yesterday_closes.get('VO')
    
    if pd.notna(vo_current) and pd.notna(vo_yesterday) and vo_yesterday > 0:
        vo_return = ((vo_current - vo_yesterday) / vo_yesterday) * 100
    else:
        vo_return = 0.0
    
    # Total value includes both active and terminated trades
    active_value = sum(p['value'] for p in pair_returns)
    terminated_value = terminated_metrics['total_intraday_value']
    total_value = active_value + terminated_value
    
    # Calculate combined returns (value-weighted)
    if total_value > 0:
        # Active trades contribution
        active_nominal_weighted = sum(p['nominal'] * p['value'] for p in pair_returns)
        active_alpha_weighted = sum(p['alpha'] * p['value'] for p in pair_returns)
        
        # Terminated trades contribution (already value-weighted in metrics)
        terminated_nominal_weighted = terminated_metrics['weighted_nominal_return']
        terminated_alpha_weighted = terminated_metrics['weighted_alpha_return']
        
        # Combined value-weighted returns
        daily_nominal_return = (active_nominal_weighted + terminated_nominal_weighted) / total_value * 100
        daily_alpha_return = (active_alpha_weighted + terminated_alpha_weighted) / total_value * 100
    else:
        daily_nominal_return = 0.0
        daily_alpha_return = 0.0
    
    # Calculate nominal return as % of NAV
    # CRITICAL: Include terminated trades P&L
    active_pnl = sum(pair_pnls)
    terminated_pnl = terminated_metrics['total_intraday_pnl']
    total_pnl = active_pnl + terminated_pnl
    
    # ========================================================================
    # OPTIONS P&L CALCULATION
    # ========================================================================
    options_df = load_options_portfolio()
    options_pnl_data = calculate_options_pnl(
        options_df, 
        live_prices, 
        yesterday_closes,
        igv_price=live_prices.get('IGV')
    )
    
    options_pnl = options_pnl_data['total_pnl']
    hedged_pnl = total_pnl + options_pnl  # Gross P&L + Options P&L
    
    if net_liquidation_usd and net_liquidation_usd > 0:
        nominal_return_pct_nav = (total_pnl / net_liquidation_usd) * 100  # Gross return
        options_pnl_pct_nav = (options_pnl / net_liquidation_usd) * 100   # Options impact
        hedged_return_pct_nav = (hedged_pnl / net_liquidation_usd) * 100  # Net return
    else:
        nominal_return_pct_nav = 0.0
        options_pnl_pct_nav = 0.0
        hedged_return_pct_nav = 0.0
    
    # Win rates (active trades only - terminated trades don't have "today's" signal)
    alpha_returns = [p['alpha'] for p in pair_returns]
    nominal_returns = [p['nominal'] for p in pair_returns]
    alpha_wins = sum(1 for r in alpha_returns if r > 0)
    nominal_wins = sum(1 for r in nominal_returns if r > 0)
    total_pairs = len(pair_returns)
    
    alpha_win_rate = (alpha_wins / total_pairs * 100) if total_pairs > 0 else 0
    nominal_win_rate = (nominal_wins / total_pairs * 100) if total_pairs > 0 else 0
    
    # ========== DEBUG ==========
    print("\n" + "="*60)
    print("DAILY PERFORMANCE CALCULATION SUMMARY")
    print("="*60)
    print(f"  Active trades: {len(pair_returns)}")
    print(f"  Terminated trades: {terminated_metrics['trade_count']}")
    print(f"  Active P&L: ${active_pnl:,.2f}")
    print(f"  Terminated P&L: ${terminated_pnl:,.2f}")
    print(f"  GROSS P&L: ${total_pnl:,.2f}")
    if options_pnl != 0:
        print(f"  Options P&L: ${options_pnl:+,.2f}")
        print(f"  HEDGED P&L: ${hedged_pnl:,.2f}")
    print(f"  NAV: ${net_liquidation_usd:,.2f}" if net_liquidation_usd else "  NAV: Not provided")
    print(f"  Gross % of NAV: {nominal_return_pct_nav:.4f}%")
    if options_pnl != 0:
        print(f"  Options Impact: {options_pnl_pct_nav:+.4f}%")
        print(f"  Hedged % of NAV: {hedged_return_pct_nav:.4f}%")
    print(f"  Daily Alpha %: {daily_alpha_return:.4f}%")
    print("="*60 + "\n")
    # ========== END DEBUG ==========
    
    return {
        'daily_alpha_return': daily_alpha_return,
        'daily_nominal_return': daily_nominal_return,
        'nominal_return_pct_nav': nominal_return_pct_nav,
        'options_pnl': options_pnl,
        'options_pnl_pct_nav': options_pnl_pct_nav,
        'hedged_return_pct_nav': hedged_return_pct_nav,
        'vo_return': vo_return,
        'alpha_win_rate': alpha_win_rate,
        'nominal_win_rate': nominal_win_rate,
        'terminated_trades_included': terminated_metrics['trade_count'],
        'terminated_pnl': terminated_pnl
    }


# format_reconciliation_summary and _get_price_from_dict moved to Reconciliation.py.

async def create_portfolio_analytics(portfolio_df, ib=None):
    """
    Generate analytics dictionary for the Analytics tab in Portfolio.xlsx.
    
    Returns dict with keys:
        Generated, Position_Summary, Beta_Buckets, Ticker_Exposure,
        Beta_Distribution, Median_MCAP_Long_M, Median_MCAP_Short_M
    """
    from datetime import datetime
    
    analytics = {
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    if portfolio_df.empty:
        empty_df = pd.DataFrame()
        analytics.update({
            'Position_Summary': pd.DataFrame({'Metric': [], 'Value': []}),
            'Beta_Buckets': empty_df,
            'Ticker_Exposure': empty_df,
            'Beta_Distribution': empty_df,
            'Median_MCAP_Long_M': 0,
            'Median_MCAP_Short_M': 0,
        })
        return analytics
    
    # --- Position Summary ---
    total_trades = len(portfolio_df)
    
    val_co1 = portfolio_df['Trade Value Co1 ($)'].astype(float) if 'Trade Value Co1 ($)' in portfolio_df.columns else pd.Series(dtype=float)
    val_co2 = portfolio_df['Trade Value Co2 ($)'].astype(float) if 'Trade Value Co2 ($)' in portfolio_df.columns else pd.Series(dtype=float)
    tails = portfolio_df.get('Tail', pd.Series(['L'] * len(portfolio_df))).str.strip().str.upper()
    
    # Trade values are always stored positive — use Tail to determine direction
    # L-tail: Co1 is long, Co2 is short
    # U-tail: Co1 is short, Co2 is long
    is_l_tail = (tails == 'L')
    total_long = (val_co1[is_l_tail].sum() + val_co2[~is_l_tail].sum())
    total_short = (val_co2[is_l_tail].sum() + val_co1[~is_l_tail].sum())
    total_nominal = abs(val_co1).sum() + abs(val_co2).sum()
    
    betas = pd.to_numeric(portfolio_df.get('Beta', pd.Series(dtype=float)), errors='coerce')
    avg_beta = betas.mean() if not betas.empty else 0
    
    # Beta-weighted portfolio beta
    trade_notionals = abs(val_co1) + abs(val_co2)
    total_notional_sum = trade_notionals.sum()
    portfolio_beta = (betas * trade_notionals).sum() / total_notional_sum if total_notional_sum > 0 else 0
    
    l_tail = is_l_tail.sum()
    u_tail = (~is_l_tail).sum()
    
    summary_data = {
        'Metric': [
            'Total Trades', 'L-Tail', 'U-Tail',
            'Total Long ($)', 'Total Short ($)', 'Total Nominal ($)',
            'Average Trade Beta', 'Portfolio Beta (notional-weighted)',
        ],
        'Value': [
            total_trades, l_tail, u_tail,
            round(total_long, 2), round(total_short, 2), round(total_nominal, 2),
            round(avg_beta, 4), round(portfolio_beta, 4),
        ]
    }
    
    # Add index breakdown
    if 'Index' in portfolio_df.columns:
        for idx_name, count in portfolio_df['Index'].value_counts().items():
            summary_data['Metric'].append(f'  {idx_name} trades')
            summary_data['Value'].append(count)
    
    analytics['Position_Summary'] = pd.DataFrame(summary_data)
    
    # --- Beta Buckets (by individual ticker, using VO betas) ---
    # Each ticker gets its VO beta, signed by position direction:
    #   Long positions → positive beta (exposed to market upside)
    #   Short positions → negative beta (exposed to market downside)
    # If a ticker appears on both sides, exposures accumulate with sign.
    ticker_beta_entries = {}  # {ticker: {'beta': float, 'long_exposure': float, 'short_exposure': float}}
    
    for _, row in portfolio_df.iterrows():
        co1 = str(row.get('Co1', '')).strip()
        co2 = str(row.get('Co2', '')).strip()
        tail = str(row.get('Tail', 'L')).strip().upper()
        v1 = abs(float(row.get('Trade Value Co1 ($)', 0) or 0))
        v2 = abs(float(row.get('Trade Value Co2 ($)', 0) or 0))
        
        # Get individual VO betas (always positive)
        beta1 = get_ticker_beta(co1)
        beta2 = get_ticker_beta(co2)
        
        # Determine long/short assignment
        if tail == 'L':
            long_ticker, short_ticker = co1, co2
            long_beta, short_beta = beta1, beta2
            long_val, short_val = v1, v2
        else:
            long_ticker, short_ticker = co2, co1
            long_beta, short_beta = beta2, beta1
            long_val, short_val = v2, v1
        
        # Accumulate per-ticker (ticker may appear in multiple pairs, even on different sides)
        for ticker, beta, exposure, is_long in [
            (long_ticker, long_beta, long_val, True),
            (short_ticker, short_beta, short_val, False),
        ]:
            if ticker not in ticker_beta_entries:
                ticker_beta_entries[ticker] = {'beta': beta, 'long_exposure': 0, 'short_exposure': 0}
            if is_long:
                ticker_beta_entries[ticker]['long_exposure'] += exposure
            else:
                ticker_beta_entries[ticker]['short_exposure'] += exposure
    
    # Calculate signed beta: if predominantly short, negate the beta
    ticker_betas_list = []
    ticker_exposures_list = []
    for ticker, data in ticker_beta_entries.items():
        net_direction = data['long_exposure'] - data['short_exposure']
        total_exposure = data['long_exposure'] + data['short_exposure']
        signed_beta = data['beta'] if net_direction >= 0 else -data['beta']
        ticker_betas_list.append(signed_beta)
        ticker_exposures_list.append(total_exposure)
    
    ticker_betas_array = np.array(ticker_betas_list)
    ticker_exposures_array = np.array(ticker_exposures_list)
    
    beta_buckets_list = []
    bucket_edges = [
        (-999, -1.25), (-1.25, -1.0), (-1.0, -0.75), (-0.75, -0.5), (-0.5, -0.25), (-0.25, 0),
        (0, 0.25), (0.25, 0.5), (0.5, 0.75), (0.75, 1.0), (1.0, 1.25), (1.25, 999)
    ]
    bucket_labels = [
        '< -1.25', '-1.25 to -1.00', '-1.00 to -0.75', '-0.75 to -0.50', '-0.50 to -0.25', '-0.25 to 0.00',
        '0.00 to 0.25', '0.25 to 0.50', '0.50 to 0.75', '0.75 to 1.00', '1.00 to 1.25', '> 1.25'
    ]
    
    for (lo, hi), label in zip(bucket_edges, bucket_labels):
        mask = (ticker_betas_array >= lo) & (ticker_betas_array < hi)
        bucket_notional = ticker_exposures_array[mask].sum()
        bucket_count = mask.sum()
        beta_buckets_list.append({
            'Beta_Bucket': label,
            'Total_Value': round(float(bucket_notional), 2),
            'Ticker_Count': int(bucket_count),
        })
    
    analytics['Beta_Buckets'] = pd.DataFrame(beta_buckets_list)
    
    # --- Ticker Exposure (top tickers by total absolute exposure) ---
    ticker_values = {}
    for _, row in portfolio_df.iterrows():
        co1 = row.get('Co1', '')
        co2 = row.get('Co2', '')
        v1 = abs(float(row.get('Trade Value Co1 ($)', 0) or 0))
        v2 = abs(float(row.get('Trade Value Co2 ($)', 0) or 0))
        ticker_values[co1] = ticker_values.get(co1, 0) + v1
        ticker_values[co2] = ticker_values.get(co2, 0) + v2
    
    exposure_df = pd.DataFrame([
        {'Ticker': t, 'Total_Exposure': round(v, 2)}
        for t, v in sorted(ticker_values.items(), key=lambda x: -x[1])
    ])
    analytics['Ticker_Exposure'] = exposure_df
    
    # --- Beta Distribution (per-trade detail) ---
    beta_detail_rows = []
    for _, row in portfolio_df.iterrows():
        pair = row.get('Pair', f"{row.get('Co1','')}-{row.get('Co2','')}")
        beta_detail_rows.append({
            'Pair': pair,
            'Tail': row.get('Tail', ''),
            'Index': row.get('Index', ''),
            'Raw_Beta': round(float(row.get('Beta', 0) or 0), 4),
            'Notional': round(abs(float(row.get('Trade Value Co1 ($)', 0) or 0)) +
                              abs(float(row.get('Trade Value Co2 ($)', 0) or 0)), 2),
        })
    analytics['Beta_Distribution'] = pd.DataFrame(beta_detail_rows)
    
    # --- Median MCAP ---
    median_long = 0
    median_short = 0
    try:
        from src.execution.daily_data_capture import load_market_caps
        # Try with file path argument first, then without
        mcap_cache = None
        try:
            mcap_cache = load_market_caps(config.closing_prices_file())
        except TypeError:
            # Function might not take a file path argument
            try:
                mcap_cache = load_market_caps()
            except Exception as e2:
                print(f"  [MCAP DEBUG] load_market_caps() also failed: {e2}")
        
        if mcap_cache:
            print(f"  [MCAP DEBUG] Loaded {len(mcap_cache)} market caps, sample: {list(mcap_cache.items())[:3]}")
            long_mcaps = []
            short_mcaps = []
            for _, row in portfolio_df.iterrows():
                co1, co2 = row.get('Co1', ''), row.get('Co2', '')
                tail = str(row.get('Tail', 'L')).strip().upper()
                
                if tail == 'L':
                    long_ticker, short_ticker = co1, co2
                else:
                    long_ticker, short_ticker = co2, co1
                
                if long_ticker in mcap_cache:
                    long_mcaps.append(mcap_cache[long_ticker])
                if short_ticker in mcap_cache:
                    short_mcaps.append(mcap_cache[short_ticker])
            
            if long_mcaps:
                median_long = float(np.median(long_mcaps))
            if short_mcaps:
                median_short = float(np.median(short_mcaps))
            
            print(f"  [MCAP DEBUG] Long: {len(long_mcaps)} tickers, median=${median_long:,.0f}M")
            print(f"  [MCAP DEBUG] Short: {len(short_mcaps)} tickers, median=${median_short:,.0f}M")
        else:
            print(f"  [MCAP DEBUG] Market cap cache is empty or None")
    except ImportError as e:
        print(f"  [MCAP DEBUG] Import error: {e}")
    except Exception as e:
        print(f"  [MCAP DEBUG] Error loading market caps: {e}")
        import traceback
        traceback.print_exc()
    
    analytics['Median_MCAP_Long_M'] = median_long
    analytics['Median_MCAP_Short_M'] = median_short
    
    return analytics


async def save_portfolio_with_analytics(portfolio_df, options_portfolio_df, file_path, ib=None,  # ← Add async
                                       live_prices=None, index_prices=None, opening_prices=None,
                                       yesterday_closes=None):
    """
    Enhanced save function that includes the Analytics tab
    
    FIX: Added NaN/inf sanitization before wb.save() to prevent Excel
    "We found a problem with some content" repair warnings. openpyxl writes
    Python float('nan') directly into cells, which Excel considers invalid XML.
    """
    
    if yesterday_closes is None:
        yesterday_closes = {}
    
    try:
        import shutil
        import math
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, Reference
        
        # Create backup
        backup_path = file_path.replace('.xlsx', '_backup.xlsx')
        if os.path.exists(file_path):
            shutil.copy2(file_path, backup_path)
            logger.info(f"Created backup at {backup_path}")

        # ====================================================================
        # Calculate return columns for active trades
        # ====================================================================
        if not portfolio_df.empty and live_prices:
            logger.info("Calculating return columns for active trades...")
            
            for idx, row in portfolio_df.iterrows():
                # Co1 Return (%)
                co1 = row.get('Co1')
                init_co1 = row.get('Co1 at Initiation')
                current_co1 = live_prices.get(co1) if co1 else None
                
                if pd.notna(init_co1) and current_co1 and init_co1 > 0:
                    portfolio_df.at[idx, 'Co1 Return (%)'] = ((current_co1 - init_co1) / init_co1) * 100
                
                # Co2 Return (%)
                co2 = row.get('Co2')
                init_co2 = row.get('Co2 at Initiation')
                current_co2 = live_prices.get(co2) if co2 else None
                
                if pd.notna(init_co2) and current_co2 and init_co2 > 0:
                    portfolio_df.at[idx, 'Co2 Return (%)'] = ((current_co2 - init_co2) / init_co2) * 100
                
                # Index Return (%) - version-aware calculation
                # V9 trades use VO index, V9.2 trades use sector ETF
                trade_version = str(row.get('Version', 'V9'))
                trade_index = row.get('Index')
                init_index = row.get('Index at Initiation')
                
                # Determine correct current index price based on version
                if trade_version in ['9.0C', 'V9', '9.0']:
                    # V9 uses VO as the market factor
                    current_index = live_prices.get('VO')
                else:
                    # V9.2 uses sector ETF (VGT, VIS, etc.)
                    current_index = live_prices.get(trade_index) if trade_index else None
                
                if pd.notna(init_index) and current_index and init_index > 0:
                    portfolio_df.at[idx, 'Index Return (%)'] = ((current_index - init_index) / init_index) * 100
            
            # Log summary
            co1_count = portfolio_df['Co1 Return (%)'].notna().sum() if 'Co1 Return (%)' in portfolio_df.columns else 0
            co2_count = portfolio_df['Co2 Return (%)'].notna().sum() if 'Co2 Return (%)' in portfolio_df.columns else 0
            idx_count = portfolio_df['Index Return (%)'].notna().sum() if 'Index Return (%)' in portfolio_df.columns else 0
            logger.info(f"✓ Calculated returns: Co1={co1_count}, Co2={co2_count}, Index={idx_count}")
        # ====================================================================

        # Generate analytics (don't pass prices here - used in summary calculation below)
        analytics = await create_portfolio_analytics(portfolio_df, ib)  # ← Add await
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Portfolio Sheet - clean up columns before saving
        portfolio_output = portfolio_df.copy()
        
        # Remove internal tracking columns and duplicates
        # Include variations with spaces/underscores and different cases
        columns_to_remove = [
            # Internal tracking
            'Existing', 'Exit Reason', 'Exit_Reason', 'Evaluated_At', 'trigger_type',
            # Spread duplicates (all captured in Entry_Spread_BPS)
            'Weighted_Spread', 'Weighted Spread', 'Weighted_Spread_BPS', 'Weighted Spread BPS',
            'Initiation Spread', 'Initiation_Spread',
            'Achieved Spread', 'Achieved_Spread', 'Achieved Spreads', 'Achieved_Spreads',
            # Termination spread only relevant for completed trades
            'Termination Spread', 'Termination_Spread',
            # LAM signal duplicates (lowercase versions - keep the canonical CamelCase)
            'rolling_intraday_vol_pct', 'rolling_intraday_vol', 
            'iv_percentile_pct', 'iv_percentile',
            'index_bias',  # Keep 'Index_Bias' (CamelCase)
            'Index Bias',  # Space version - keep Index_Bias
            # Sum deviation duplicates (keep Sum_Dev_Bucket CamelCase)
            'sum_dev_bucket', 'sum_deviation_15d',
        ]
        
        # First pass: remove explicitly named columns
        removed_cols = []
        for col in columns_to_remove:
            if col in portfolio_output.columns:
                portfolio_output = portfolio_output.drop(columns=[col])
                removed_cols.append(col)
        
        if removed_cols:
            logger.info(f"Removed {len(removed_cols)} legacy/duplicate columns: {removed_cols}")
        
        # Second pass: remove any columns that are entirely NaN/blank
        # (these are likely duplicate columns that didn't get populated)
        empty_removed = []
        for col in portfolio_output.columns.tolist():
            if portfolio_output[col].isna().all() or (portfolio_output[col] == '').all():
                # Check if a non-empty version of this column exists
                # Only drop if it looks like a duplicate (similar name exists)
                col_normalized = col.lower().replace('_', '').replace(' ', '')
                similar_cols = [c for c in portfolio_output.columns 
                               if c != col and c.lower().replace('_', '').replace(' ', '') == col_normalized]
                if similar_cols:
                    portfolio_output = portfolio_output.drop(columns=[col])
                    empty_removed.append(f"{col} (dup of {similar_cols[0]})")
        
        if empty_removed:
            logger.info(f"Removed {len(empty_removed)} empty duplicate columns: {empty_removed}")
        
        # Third pass: remove duplicate columns (keep first occurrence)
        dup_cols = portfolio_output.columns[portfolio_output.columns.duplicated()].tolist()
        if dup_cols:
            logger.info(f"Removing {len(dup_cols)} exactly-duplicated columns: {dup_cols}")
        portfolio_output = portfolio_output.loc[:, ~portfolio_output.columns.duplicated()]
        
        # Define canonical column order (add any missing columns at end)
        canonical_columns = [
            # Version tracking
            'Version',
            # Core identifiers
            'Tag', 'Pair', 'Co1', 'Co2', 'Tail', 'Index',
            # Position details
            'Quantity1', 'Quantity2', 'W1', 'W2', 'Position_Multiplier',
            # Prices at initiation
            'Co1 at Initiation', 'Co2 at Initiation', 'Index at Initiation', 'Treasury at Initiation',
            # Trade values
            'Trade Value Co1 ($)', 'Trade Value Co2 ($)', 'Total_Notional',
            # Sum deviation
            'Sum_Dev_Value', 'Sum_Dev_CDF', 'Sum_Dev_Bucket',
            # Risk metrics
            'Beta', 'Concentration',
            # Multi-factor exposure tags (USMV, VLUE, MTUM, SOXX)
            'USMV_Long_Beta', 'USMV_Short_Beta', 'USMV_Net_Contrib',
            'VLUE_Long_Beta', 'VLUE_Short_Beta', 'VLUE_Net_Contrib',
            'MTUM_Long_Beta', 'MTUM_Short_Beta', 'MTUM_Net_Contrib',
            'SOXX_Long_Beta', 'SOXX_Short_Beta', 'SOXX_Net_Contrib',
            # LAM signals - Raw values
            'Weighted_Score', 'Index_Bias', 'Composite_Score',
            'Volume_Ratio', 'Rolling_Intraday_Vol', 'Volume_Dominance',
            'Last_Hour_Vol', 'IV_Percentile',
            # LAM signals - Percentiles (for comparison with rejected trades)
            'Volume_Ratio_Pct', 'Intraday_Vol_Pct', 'Volume_Dom_Pct',
            'Last_Hour_Pct', 'IV_Pct_Pct',
            # Quality scores
            'Entry_Spread_BPS', 'Spread_Quality_Score', 'Sum_Dev_Extremity_Score', 'Composite_Priority_Score',
            # Alpha tracking
            'Alpha_Return', 'Cumulative_Alpha',
            # Dates
            'Trade Initiation Date', 'Trade Termination Date',
            # Stop orders
            'Stop_Order_ID', 'Stop_Price'
        ]
        
        # Reorder: canonical columns first (if they exist), then any remaining
        existing_canonical = [col for col in canonical_columns if col in portfolio_output.columns]
        remaining_columns = [col for col in portfolio_output.columns if col not in canonical_columns]
        final_column_order = existing_canonical + remaining_columns
        
        portfolio_output = portfolio_output[final_column_order]
        
        ws_portfolio = wb.create_sheet('Portfolio')
        for r in dataframe_to_rows(portfolio_output, index=False, header=True):
            ws_portfolio.append(r)
        
        # 2. Options Sheet
        ws_options = wb.create_sheet('Options')
        for r in dataframe_to_rows(options_portfolio_df, index=False, header=True):
            ws_options.append(r)
        
        # 3. Summary Sheet
        logger.info("Creating Summary sheet...")
        ws_summary = wb.create_sheet('Summary')
        
        # Get enhanced summary
        summary_data = calculate_summary(portfolio_df, options_portfolio_df, 
                                        live_prices=live_prices, 
                                        index_prices=index_prices,
                                        opening_prices=opening_prices)
        
        # ========================================================================
        # Section 1: Overall Portfolio Metrics
        # ========================================================================
        
        ws_summary['A1'] = 'Portfolio Summary'
        ws_summary['A1'].font = Font(size=14, bold=True)
        
        current_row = 2
        
        # Get account values if ib connection available
        if ib is not None:
            from src.execution.trade_execution import get_account_summary_values
            account_values = get_account_summary_values(ib)

            # Calculate leverage correctly with currency conversion (matching evaluate_trades)
            currency = account_values.get('Currency', 'USD')
            if currency == 'GBP':
                gbp_usd_rate = Config_Helper.get_gbp_usd_rate(ib)
                net_liq_for_leverage = account_values['NetLiquidation'] * gbp_usd_rate
                gross_pos_for_leverage = account_values['GrossPositionValue'] * gbp_usd_rate
            else:
                net_liq_for_leverage = account_values['NetLiquidation']
                gross_pos_for_leverage = account_values['GrossPositionValue']

            # Override the leverage value with correctly calculated one
            account_values['Leverage'] = gross_pos_for_leverage / net_liq_for_leverage if net_liq_for_leverage > 0 else 0
        else:
            account_values = {
                'NetLiquidation': 0,
                'ExcessLiquidity': 0,
                'Leverage': 0,
                'Currency': 'N/A'
            }
        
        currency = account_values.get('Currency', 'USD')
        currency_symbol = '£' if currency == 'GBP' else '$'
        
        basic_metrics = pd.DataFrame({
            'Metric': [
                f'Net Liquidation ({currency})',
                f'Excess Liquidity ({currency})',
                'Current Leverage',
                'Max Allowed Leverage',
                '',  # Spacer row
                'Total Nominal Size',
                'Largest % of Total',
                'Portfolio Beta (VO)',
                'Net Beta (VO)',
                '',  # Spacer row
                'Highest Index Allocation'
            ],
            'Value': [
                f"{currency_symbol}{account_values['NetLiquidation']:,.2f}",
                f"{currency_symbol}{account_values['ExcessLiquidity']:,.2f}",
                f"{account_values['Leverage']:.2f}x",
                f"{config.max_account_leverage():.2f}x",
                '',
                summary_data['Total Nominal Size'],
                summary_data['Largest % of Total'],
                summary_data['Portfolio Beta'],
                summary_data['Net Beta'],
                '',
                summary_data.get('Highest Index Allocation', 'N/A')
            ]
        })
        
        for r in dataframe_to_rows(basic_metrics, index=False, header=True):
            ws_summary.append(r)
            current_row += 1
        
        current_row += 2
        
        # ========================================================================
        # Section 2: Index Breakdown
        # ========================================================================
        
        ws_summary[f'A{current_row}'] = 'Breakdown by Index (Market Betas)'
        ws_summary[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        index_breakdown = summary_data.get('Index Breakdown')
        if isinstance(index_breakdown, pd.DataFrame) and not index_breakdown.empty:
            for r in dataframe_to_rows(index_breakdown, index=False, header=True):
                ws_summary.append(r)
                current_row += 1
        
        current_row += 2
        
        # ========================================================================
        # Section 3: Portfolio Returns (Intraday) - MOVED UP, replaces Index ETF Returns
        # ========================================================================
        
        ws_summary[f'A{current_row}'] = 'Portfolio Returns (Intraday)'
        ws_summary[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        portfolio_returns = summary_data.get('Portfolio Returns', {})
        
        # Overall returns
        portfolio_returns_df = pd.DataFrame({
            'Metric': [
                'Cap-Weighted Return (%)',
                'Equal-Weighted Return (%)'
            ],
            'Value': [
                portfolio_returns.get('cap_weighted', 0.0),
                portfolio_returns.get('equal_weighted', 0.0)
            ]
        })
        
        for r in dataframe_to_rows(portfolio_returns_df, index=False, header=True):
            ws_summary.append(r)
            current_row += 1
        
        current_row += 2
        
        # By-index breakdown
        ws_summary[f'A{current_row}'] = 'Portfolio Returns by Index (Intraday)'
        ws_summary[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        by_index_returns = portfolio_returns.get('by_index')
        if isinstance(by_index_returns, pd.DataFrame) and not by_index_returns.empty:
            for r in dataframe_to_rows(by_index_returns, index=False, header=True):
                ws_summary.append(r)
                current_row += 1
        else:
            ws_summary[f'A{current_row}'] = 'No index-level return data available'
            current_row += 1
        
        current_row += 2
        
        # ========================================================================
        # Section 4: Index ETF Returns (unadjusted vs adjusted)
        # ========================================================================
        
        ws_summary[f'A{current_row}'] = 'Index ETF Returns (Unadjusted vs Adjusted)'
        ws_summary[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        ws_summary[f'A{current_row}'] = '(Adjusted = megacap-adjusted where applicable)'
        ws_summary[f'A{current_row}'].font = Font(italic=True)
        current_row += 1
        
        # Build returns for ALL sector ETFs
        etf_data = []
        sector_etfs_ordered = sorted(config.index_etfs())  # VCR, VFH, VGT, VHT, VIS
        
        if index_prices:
            for etf in sector_etfs_ordered:
                if etf in index_prices:
                    etf_initial = index_prices[etf].get('initial')
                    etf_current = index_prices[etf].get('current')
                    
                    if etf_initial and etf_current and etf_initial > 0:
                        etf_return = ((etf_current - etf_initial) / etf_initial) * 100
                        
                        # Get megacap-adjusted return
                        adjusted_return_pct = etf_return  # Default to unadjusted
                        if live_prices and opening_prices:
                            try:
                                adj_ret = get_adjusted_index_return(etf, live_prices, opening_prices)
                                adjusted_return_pct = adj_ret * 100
                            except Exception:
                                pass  # Fall back to unadjusted
                        
                        etf_data.append({
                            'Index': etf,
                            'Yesterday Close': f"${etf_initial:.2f}",
                            'Current Price': f"${etf_current:.2f}",
                            'Unadjusted (%)': f"{etf_return:.4f}%",
                            'Adjusted (%)': f"{adjusted_return_pct:.4f}%"
                        })
                    else:
                        etf_data.append({
                            'Index': etf,
                            'Yesterday Close': 'N/A',
                            'Current Price': 'N/A',
                            'Unadjusted (%)': 'Incomplete data',
                            'Adjusted (%)': 'Incomplete data'
                        })
                else:
                    etf_data.append({
                        'Index': etf,
                        'Yesterday Close': 'N/A',
                        'Current Price': 'N/A',
                        'Unadjusted (%)': 'Not available',
                        'Adjusted (%)': 'Not available'
                    })
        else:
            etf_data = [{'Index': 'N/A', 'Yesterday Close': 'N/A', 
                        'Current Price': 'N/A', 'Unadjusted (%)': 'No data',
                        'Adjusted (%)': 'No data'}]
        
        etf_df = pd.DataFrame(etf_data)
        for r in dataframe_to_rows(etf_df, index=False, header=True):
            ws_summary.append(r)
            current_row += 1
            
        current_row += 2
        
        # ========================================================================
        # Section 5: IGV Beta Exposure (Pre-Trade Constraint)
        # ========================================================================

        igv_exposure_limit = getattr(config, 'IGV_EXPOSURE_LIMIT_PCT', None)

        if igv_exposure_limit is not None:
            ws_summary[f'A{current_row}'] = 'IGV Beta Exposure'
            ws_summary[f'A{current_row}'].font = Font(size=14, bold=True)
            current_row += 1

            # Calculate IGV exposure via Constraints
            igv_exp = Constraints.calculate_igv_exposure(portfolio_df)
            igv_net = igv_exp['net_igv_exposure']
            igv_gross = igv_exp['gross_exposure']
            igv_pct = igv_exp['exposure_pct']

            ws_summary.cell(row=current_row, column=1, value='Net IGV $ Exposure')
            ws_summary.cell(row=current_row, column=2, value=f'${igv_net:+,.0f}')
            current_row += 1
            ws_summary.cell(row=current_row, column=1, value='Long IGV Contribution')
            ws_summary.cell(row=current_row, column=2, value=f'${igv_exp.get("long_igv_contrib", 0):+,.0f}')
            current_row += 1
            ws_summary.cell(row=current_row, column=1, value='Short IGV Contribution')
            ws_summary.cell(row=current_row, column=2, value=f'${igv_exp.get("short_igv_contrib", 0):+,.0f}')
            current_row += 1
            ws_summary.cell(row=current_row, column=1, value='IGV Exposure %')
            ws_summary.cell(row=current_row, column=2, value=f'{igv_pct:+.1%}')
            current_row += 2

            ws_summary.cell(row=current_row, column=1, value='Exposure Limit')
            ws_summary.cell(row=current_row, column=2, value=f'±{igv_exposure_limit:.0%} of gross')
            current_row += 1
            within_limit = abs(igv_pct) <= igv_exposure_limit
            ws_summary.cell(row=current_row, column=1, value='Constraint Status')
            ws_summary.cell(row=current_row, column=2,
                            value='✓ Within limit' if within_limit else f'⚠ BREACH ({igv_pct:+.1%})')
            current_row += 2
            
        # 4. Analytics Sheet - NEW
        logger.info("Creating Analytics sheet...")
        ws_analytics = wb.create_sheet('Analytics')
        
        # Add title
        ws_analytics['A1'] = 'Portfolio Analytics Dashboard'
        ws_analytics['A1'].font = Font(size=16, bold=True)
        ws_analytics['A2'] = f"Generated: {analytics['Generated']}"
        ws_analytics['A2'].font = Font(italic=True)
        
        current_row = 4
        
        # Section 1: Position Summary
        ws_analytics[f'A{current_row}'] = 'Position Summary'
        ws_analytics[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        position_summary = analytics['Position_Summary']
        for r in dataframe_to_rows(position_summary, index=False, header=True):
            ws_analytics.append(r)
            current_row += 1
        
        # Format numbers in position summary
        for row in range(current_row - len(position_summary) - 1, current_row):
            if row > current_row - len(position_summary):  # Skip header
                ws_analytics[f'B{row}'].number_format = '#,##0.00'
        
        current_row += 2
        
        # Section 2: Beta Distribution
        ws_analytics[f'A{current_row}'] = 'Beta Distribution Analysis'
        ws_analytics[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        if 'Beta_Buckets' in analytics:
            beta_buckets = analytics['Beta_Buckets']
            
            # Write beta buckets data
            start_row = current_row
            for r in dataframe_to_rows(beta_buckets, index=False, header=True):
                ws_analytics.append(r)
                current_row += 1
            
            # Create a bar chart for beta distribution
            if len(beta_buckets) > 0:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Ticker Beta Distribution by Exposure Value"
                chart.y_axis.title = 'Total Exposure Value ($)'
                chart.x_axis.title = 'Ticker VO Beta Range'
                chart.height = 12  # Make chart taller
                chart.width = 22   # Wide enough for 12 buckets

                # Set data references
                data = Reference(ws_analytics, 
                               min_col=2,  # Total_Value column
                               min_row=start_row, 
                               max_row=current_row - 1,
                               max_col=2)

                categories = Reference(ws_analytics,
                                     min_col=1,  # Beta_Bucket column
                                     min_row=start_row + 1,
                                     max_row=current_row - 1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.shape = 4
                ws_analytics.add_chart(chart, f"F{start_row}")
            
            current_row += 2
        
        # Section 3: Top Exposures by Ticker
        ws_analytics[f'A{current_row}'] = 'Top 10 Ticker Exposures'
        ws_analytics[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        ticker_exposure = analytics['Ticker_Exposure']
        if not ticker_exposure.empty:
            top_10_exposure = ticker_exposure.head(10)
            for r in dataframe_to_rows(top_10_exposure, index=False, header=True):
                ws_analytics.append(r)
                current_row += 1
        
        # Section 4: Detailed Beta Distribution (all positions)
        current_row += 2
        ws_analytics[f'A{current_row}'] = 'Detailed Beta Distribution'
        ws_analytics[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        beta_dist = analytics['Beta_Distribution']
        if not beta_dist.empty:
            # Sort by effective beta
            beta_dist_sorted = beta_dist.sort_values('Raw_Beta')
            for r in dataframe_to_rows(beta_dist_sorted, index=False, header=True):
                ws_analytics.append(r)
                current_row += 1
                
        # Section 5: MCAP Summary
        current_row += 2
        ws_analytics[f'A{current_row}'] = 'Market Cap Summary (Live from IBKR)'
        ws_analytics[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        median_long_b = analytics.get('Median_MCAP_Long_M', 0) / 1000  # Convert millions to billions
        median_short_b = analytics.get('Median_MCAP_Short_M', 0) / 1000
        
        mcap_summary = pd.DataFrame({
            'Metric': ['Median MCAP - Long Positions', 'Median MCAP - Short Positions'],
            'Value ($ Billions)': [f"{median_long_b:.2f}", f"{median_short_b:.2f}"]
        })
        
        for r in dataframe_to_rows(mcap_summary, index=False, header=True):
                    ws_analytics.append(r)
                    current_row += 1
        
        # ========================================================================
        # TICKER PERFORMANCE SHEET - ADD THIS BLOCK HERE
        # ========================================================================
        
        logger.info("Creating Ticker Performance sheet...")
        ws_ticker_perf = wb.create_sheet('Ticker Performance')
        ws_ticker_perf['A1'] = 'Individual Ticker Performance'
        ws_ticker_perf['A1'].font = Font(size=14, bold=True)
        
        ticker_perf_df = calculate_ticker_performance_detailed(
            portfolio_df, live_prices or {}, opening_prices or {}, yesterday_closes
        )
        
        if not ticker_perf_df.empty:
            for r in dataframe_to_rows(ticker_perf_df, index=False, header=True):
                ws_ticker_perf.append(r)
        
        # ========================================================================
        # DAILY PERFORMANCE SHEET - APPEND TODAY'S DATA (not overwrite)
        # ========================================================================
        
        # Get net liquidation for NAV-based return calculation
        net_liquidation_usd = None
        if ib and ib.isConnected():
            try:
                account_values = ib.accountValues()
                for item in account_values:
                    if item.tag == 'NetLiquidation':
                        net_liq = float(item.value)
                        currency = item.currency
                        # Convert to USD if needed
                        if currency == 'GBP':
                            gbp_usd_rate = Config_Helper.get_gbp_usd_rate(ib)
                            net_liquidation_usd = net_liq * gbp_usd_rate
                        else:
                            net_liquidation_usd = net_liq
                        break
            except Exception as e:
                logger.warning(f"Could not get net liquidation: {e}")
        
        # Calculate total nominal for reference
        total_nominal = portfolio_df[['Trade Value Co1 ($)', 'Trade Value Co2 ($)']].abs().sum().sum() if not portfolio_df.empty else 0
        
        today_metrics = calculate_daily_portfolio_performance(
            portfolio_df, live_prices or {}, yesterday_closes,
            net_liquidation_usd=net_liquidation_usd,
            total_nominal=total_nominal,
            index_prices=index_prices
        )
        
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        # Load existing daily performance data from the file being overwritten
        existing_daily_data = []
        try:
            if os.path.exists(file_path):
                existing_wb = load_workbook(file_path)
                if 'Daily Performance' in existing_wb.sheetnames:
                    existing_sheet = existing_wb['Daily Performance']
                    
                    # Get headers to understand column structure
                    headers = [cell.value for cell in existing_sheet[3]]
                    
                    # Read existing data starting from row 4 (after headers)
                    for row in existing_sheet.iter_rows(min_row=4, values_only=True):
                        if row[0] is not None:  # Has a date
                            # Normalize date for comparison
                            row_date = row[0]
                            if hasattr(row_date, 'strftime'):
                                row_date_str = row_date.strftime('%Y-%m-%d')
                            else:
                                row_date_str = str(row_date)[:10]
                            
                            # Don't duplicate today's date
                            if row_date_str != today_date:
                                # Build data dict based on available columns
                                data_row = {'Date': row_date_str}
                                for i, header in enumerate(headers):
                                    if header and i < len(row) and header != 'Date':
                                        data_row[header] = row[i]
                                existing_daily_data.append(data_row)
                    
                    logger.info(f"Loaded {len(existing_daily_data)} historical daily performance rows")
                existing_wb.close()
        except Exception as e:
            logger.warning(f"Could not load existing daily performance data: {e}")
            import traceback
            traceback.print_exc()
        
        logger.info("Creating Daily Performance sheet...")
        ws_daily_perf = wb.create_sheet('Daily Performance')
        ws_daily_perf['A1'] = 'Daily Portfolio Performance'
        ws_daily_perf['A1'].font = Font(size=14, bold=True)
        
        # Headers - expanded columns with options tracking
        headers = ['Date', 'Daily_Alpha_%', 'Gross_Return_%', 'Options_Impact_%', 
                   'Hedged_Return_%', 'VO_Return_%', 'Alpha_Win_Rate_%', 'Nominal_Win_Rate_%']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_daily_perf.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # Write existing historical data first (with backward compatibility)
        current_row = 4
        for data_row in existing_daily_data:
            ws_daily_perf.cell(row=current_row, column=1, value=data_row.get('Date'))
            ws_daily_perf.cell(row=current_row, column=2, value=data_row.get('Daily_Alpha_%'))
            
            # Backward compatibility: old data has Nominal_%_of_NAV, new has Gross_Return_%
            gross_return = data_row.get('Gross_Return_%', data_row.get('Nominal_%_of_NAV', data_row.get('Daily_Nominal_%')))
            ws_daily_perf.cell(row=current_row, column=3, value=gross_return)
            
            # Options impact - default to 0 for historical data without it
            options_impact = data_row.get('Options_Impact_%', 0)
            ws_daily_perf.cell(row=current_row, column=4, value=options_impact)
            
            # Hedged return - calculate if not present
            hedged_return = data_row.get('Hedged_Return_%')
            if hedged_return is None and gross_return is not None:
                hedged_return = (gross_return or 0) + (options_impact or 0)
            ws_daily_perf.cell(row=current_row, column=5, value=hedged_return)
            
            ws_daily_perf.cell(row=current_row, column=6, value=data_row.get('VO_Return_%'))
            ws_daily_perf.cell(row=current_row, column=7, value=data_row.get('Alpha_Win_Rate_%'))
            ws_daily_perf.cell(row=current_row, column=8, value=data_row.get('Nominal_Win_Rate_%'))
            current_row += 1
        
        # Append today's data at the end
        ws_daily_perf.cell(row=current_row, column=1, value=today_date)
        ws_daily_perf.cell(row=current_row, column=2, value=today_metrics['daily_alpha_return'])
        ws_daily_perf.cell(row=current_row, column=3, value=today_metrics['nominal_return_pct_nav'])  # Gross return
        ws_daily_perf.cell(row=current_row, column=4, value=today_metrics.get('options_pnl_pct_nav', 0))  # Options impact
        ws_daily_perf.cell(row=current_row, column=5, value=today_metrics.get('hedged_return_pct_nav', today_metrics['nominal_return_pct_nav']))  # Hedged return
        ws_daily_perf.cell(row=current_row, column=6, value=today_metrics['vo_return'])
        ws_daily_perf.cell(row=current_row, column=7, value=today_metrics['alpha_win_rate'])
        ws_daily_perf.cell(row=current_row, column=8, value=today_metrics['nominal_win_rate'])
        
        logger.info(f"Daily Performance: {len(existing_daily_data)} historical rows + today's data")

        # ====================================================================
        # AUTO-FIT COLUMN WIDTHS & PERCENTAGE FORMATTING FOR ALL SHEETS
        # ====================================================================
        
        logger.info("Applying column auto-fit and percentage formatting...")
        
        # Helper function to auto-fit column widths
        def auto_fit_columns(ws, min_width=8, max_width=50):
            """Auto-fit column widths based on content"""
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, min_width), max_width)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Helper function to format percentage columns
        def format_percentage_columns(ws, pct_keywords=None):
            """Format columns containing percentages consistently as 0.00%"""
            if pct_keywords is None:
                pct_keywords = ['%', 'Pct', 'Return', 'Rate', 'Alpha', 'Nominal', 'Win_Rate']
            
            # For sheets with multiple tables, we need to find all headers
            # Look at each row - if it looks like a header row (has text with keywords), 
            # treat the following rows as data until we hit another header or empty section
            
            current_pct_columns = {}  # {col_idx: True/False}
            
            for row_idx in range(1, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
                
                # Check if this looks like a header row (has string values with keywords)
                is_header_row = False
                has_string_values = False
                for col_idx, val in enumerate(row_values, 1):
                    if val and isinstance(val, str):
                        has_string_values = True
                        if any(kw in val for kw in pct_keywords):
                            is_header_row = True
                            break
                
                if is_header_row and has_string_values:
                    # This is a header row - update our column tracking
                    current_pct_columns = {}
                    for col_idx, val in enumerate(row_values, 1):
                        if val and isinstance(val, str):
                            if any(kw in val for kw in pct_keywords):
                                current_pct_columns[col_idx] = True
                elif current_pct_columns:
                    # This is a data row - apply formatting to percentage columns
                    for col_idx in current_pct_columns:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            # Values are already in percentage form (e.g., 0.568 means 0.568%)
                            cell.number_format = '0.00"%"'
        
        # Apply to all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            auto_fit_columns(ws)
            
            # Apply percentage formatting based on sheet type
            if sheet_name in ['Summary', 'Analytics', 'Daily Performance', 'Ticker Performance']:
                format_percentage_columns(ws)

        # ====================================================================
        # SANITIZE ALL CELLS: Remove NaN/inf values before saving
        # ====================================================================
        # openpyxl writes Python float('nan') and float('inf') directly into
        # cells, which Excel considers invalid XML. This triggers the
        # "We found a problem with some content" repair warning on open.
        # Converting these to None produces valid empty cells instead.
        # ====================================================================
        
        logger.info("Sanitizing workbook cells (removing NaN/inf)...")
        sanitized_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, float) and (math.isnan(cell.value) or math.isinf(cell.value)):
                        cell.value = None
                        sanitized_count += 1
        
        if sanitized_count > 0:
            logger.info(f"  Sanitized {sanitized_count} NaN/inf cells → None")
        else:
            logger.info("  No NaN/inf cells found")

        # ====================================================================
        # SAVE WORKBOOK
        # ====================================================================
        
        logger.info("Saving workbook...")
        wb.save(file_path)
        logger.info(f"✓ Portfolio saved to {file_path}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error saving portfolio: {e}")
        import traceback
        traceback.print_exc()
        return False
    
# ============================================================================
# DAILY TERMINATED TRADES FILE MANAGEMENT
# ============================================================================
# These functions manage a daily file that stores terminated trades.
# This is critical for accurate daily performance calculation - without it,
# terminated trades' intraday P&L would be missing from daily metrics.

def get_daily_terminated_trades_path():
    """Get path to daily terminated trades file from config"""
    if hasattr(config, 'DAILY_TERMINATED_TRADES_FILE'):
        return config.daily_terminated_trades_file()
    else:
        # Fallback: same directory as portfolio file
        portfolio_dir = os.path.dirname(config.portfolio_file())
        return os.path.join(portfolio_dir, 'daily_terminated_trades.xlsx')


def should_overwrite_daily_terminated_file(file_path):
    """
    Check if daily terminated trades file should be overwritten.
    
    Logic:
    - If file doesn't exist: return True (create new)
    - If file exists and was modified TODAY: return False (preserve first run)
    - If file exists and was modified BEFORE today: return True (overwrite)
    
    This ensures we keep the terminated trades from the first workflow run of the day,
    even if the workflow is run multiple times (e.g., for end-of-day updates).
    """
    if not os.path.exists(file_path):
        return True
    
    # Get file modification time
    file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
    file_date = file_mtime.date()
    today = datetime.now().date()
    
    # Only overwrite if file is from a previous day
    return file_date < today


def save_daily_terminated_trades(terminated_trades_df, live_prices, yesterday_closes):
    """
    Save terminated trades to daily file for performance calculation.
    
    Only overwrites if the file is from a previous day (preserves first run of day).
    
    Parameters:
    -----------
    terminated_trades_df : DataFrame
        Trades being terminated this run
    live_prices : dict
        Current execution prices
    yesterday_closes : dict
        Yesterday's closing prices (baseline for intraday P&L)
    
    Returns:
    --------
    bool : True if saved/preserved, False on error
    """
    file_path = get_daily_terminated_trades_path()
    
    # Check if we should overwrite
    if not should_overwrite_daily_terminated_file(file_path):
        logger.info(f"Daily terminated trades file already exists for today - preserving")
        logger.info(f"  File: {file_path}")
        return True
    
    if terminated_trades_df.empty:
        logger.info("No terminated trades to save")
        # Create empty file to mark today's date
        try:
            empty_df = pd.DataFrame(columns=[
                'Tag', 'Pair', 'Co1', 'Co2', 'Index', 'Tail',
                'Quantity1', 'Quantity2', 'W1', 'W2', 'Beta',
                'Trade Value Co1 ($)', 'Trade Value Co2 ($)',
                'Yesterday_Close_Co1', 'Yesterday_Close_Co2',
                'Exit_Price_Co1', 'Exit_Price_Co2',
                'Intraday_PnL', 'Exit_Reason', 'Exit_Timestamp'
            ])
            empty_df.to_excel(file_path, index=False, sheet_name='Terminated Trades')
            logger.info(f"Created empty daily terminated trades file: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error creating empty daily terminated file: {e}")
            return False
    
    try:
        # Enrich terminated trades with prices and P&L
        enriched_trades = []
        total_intraday_pnl = 0
        
        for _, trade in terminated_trades_df.iterrows():
            co1 = trade['Co1']
            co2 = trade['Co2']
            qty1 = abs(trade.get('Quantity1', 0))
            qty2 = abs(trade.get('Quantity2', 0))
            tail = str(trade.get('Tail', 'L')).strip().upper()
            w1 = trade.get('W1', 0.5)
            w2 = trade.get('W2', 1 - w1)
            beta = trade.get('Beta', 0)
            
            # Get prices
            yest_co1 = yesterday_closes.get(co1, 0)
            yest_co2 = yesterday_closes.get(co2, 0)
            exit_co1 = live_prices.get(co1, 0)
            exit_co2 = live_prices.get(co2, 0)
            
            # Calculate intraday P&L (yesterday close → execution)
            if tail == 'L':  # Long co1, short co2
                pnl_co1 = qty1 * (exit_co1 - yest_co1) if yest_co1 > 0 else 0
                pnl_co2 = qty2 * (yest_co2 - exit_co2) if yest_co2 > 0 else 0
            else:  # U tail: Short co1, long co2
                pnl_co1 = qty1 * (yest_co1 - exit_co1) if yest_co1 > 0 else 0
                pnl_co2 = qty2 * (exit_co2 - yest_co2) if yest_co2 > 0 else 0
            
            intraday_pnl = pnl_co1 + pnl_co2
            total_intraday_pnl += intraday_pnl
            
            enriched_trades.append({
                'Tag': trade.get('Tag'),
                'Pair': trade.get('Pair'),
                'Co1': co1,
                'Co2': co2,
                'Index': trade.get('Index', 'VO'),
                'Tail': tail,
                'Quantity1': qty1,
                'Quantity2': qty2,
                'W1': w1,
                'W2': w2,
                'Beta': beta,
                'Trade Value Co1 ($)': abs(trade.get('Trade Value Co1 ($)', 0)),
                'Trade Value Co2 ($)': abs(trade.get('Trade Value Co2 ($)', 0)),
                'Yesterday_Close_Co1': yest_co1,
                'Yesterday_Close_Co2': yest_co2,
                'Exit_Price_Co1': exit_co1,
                'Exit_Price_Co2': exit_co2,
                'Intraday_PnL': intraday_pnl,
                'Exit_Reason': trade.get('Exit Reason', trade.get('Exit_Reason', '')),
                'Exit_Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create DataFrame and save
        daily_df = pd.DataFrame(enriched_trades)
        daily_df.to_excel(file_path, index=False, sheet_name='Terminated Trades')
        
        logger.info(f"✓ Saved {len(enriched_trades)} terminated trades to: {file_path}")
        logger.info(f"  Total intraday P&L from terminated trades: ${total_intraday_pnl:,.2f}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error saving daily terminated trades: {e}")
        import traceback
        traceback.print_exc()
        return False


def load_daily_terminated_trades():
    """
    Load today's terminated trades for daily performance calculation.
    
    Returns:
    --------
    DataFrame : Terminated trades with intraday P&L, or empty DataFrame if none
    """
    file_path = get_daily_terminated_trades_path()
    
    if not os.path.exists(file_path):
        logger.debug("No daily terminated trades file found")
        return pd.DataFrame()
    
    try:
        # Check if file is from today
        file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
        file_date = file_mtime.date()
        today = datetime.now().date()
        
        if file_date != today:
            logger.debug(f"Daily terminated trades file is from {file_date}, not today - ignoring")
            return pd.DataFrame()
        
        df = pd.read_excel(file_path, sheet_name='Terminated Trades')
        
        if df.empty:
            logger.debug("Daily terminated trades file is empty")
            return pd.DataFrame()
        
        logger.info(f"Loaded {len(df)} terminated trades from today's file")
        
        # Calculate total intraday P&L for logging
        if 'Intraday_PnL' in df.columns:
            total_pnl = df['Intraday_PnL'].sum()
            logger.info(f"  Total intraday P&L from terminated trades: ${total_pnl:,.2f}")
        
        return df
        
    except Exception as e:
        logger.warning(f"Error loading daily terminated trades: {e}")
        return pd.DataFrame()


def calculate_terminated_trades_intraday_metrics(terminated_df, index_prices=None):
    """
    Calculate intraday alpha and nominal metrics for terminated trades.
    
    Parameters:
    -----------
    terminated_df : DataFrame
        Terminated trades with Yesterday_Close and Exit_Price columns
    index_prices : dict, optional
        Index prices {ticker: {'initial': float, 'current': float}}
    
    Returns:
    --------
    dict with:
        - total_intraday_pnl: Total P&L in dollars
        - total_intraday_value: Total position value (for weighting)
        - weighted_nominal_return: Value-weighted nominal return
        - weighted_alpha_return: Value-weighted alpha return
        - trade_count: Number of terminated trades
    """
    if terminated_df.empty:
        return {
            'total_intraday_pnl': 0.0,
            'total_intraday_value': 0.0,
            'weighted_nominal_return': 0.0,
            'weighted_alpha_return': 0.0,
            'trade_count': 0
        }
    
    # Load ticker betas for alpha calculation
    ticker_betas = get_cached_ticker_betas()
    
    total_pnl = 0.0
    total_value = 0.0
    weighted_nominal = 0.0
    weighted_alpha = 0.0
    
    for _, trade in terminated_df.iterrows():
        co1 = trade['Co1']
        co2 = trade['Co2']
        tail = str(trade.get('Tail', 'L')).strip().upper()
        w1 = trade.get('W1', 0.5)
        w2 = 1 - w1
        
        # Get prices
        yest_co1 = trade.get('Yesterday_Close_Co1', 0)
        yest_co2 = trade.get('Yesterday_Close_Co2', 0)
        exit_co1 = trade.get('Exit_Price_Co1', 0)
        exit_co2 = trade.get('Exit_Price_Co2', 0)
        
        # Get trade value
        value1 = abs(trade.get('Trade Value Co1 ($)', 0))
        value2 = abs(trade.get('Trade Value Co2 ($)', 0))
        trade_value = value1 + value2
        
        # Skip if missing prices
        if yest_co1 <= 0 or yest_co2 <= 0 or exit_co1 <= 0 or exit_co2 <= 0:
            continue
        
        # Calculate returns
        return_co1 = (exit_co1 - yest_co1) / yest_co1
        return_co2 = (exit_co2 - yest_co2) / yest_co2
        
        # Calculate nominal return based on tail
        if tail == 'L':
            pair_nominal = (w1 * return_co1) - (w2 * return_co2)
        else:
            pair_nominal = -(w1 * return_co1) + (w2 * return_co2)
        
        # Calculate alpha
        trade_index = trade.get('Index', 'VO')
        index_return = 0.0
        
        if index_prices and trade_index in index_prices:
            idx_data = index_prices[trade_index]
            idx_initial = idx_data.get('initial', 0)
            idx_current = idx_data.get('current', 0)
            if idx_initial > 0:
                index_return = (idx_current - idx_initial) / idx_initial
        
        # Get ticker betas
        co1_beta = ticker_betas.get(co1.upper(), 1.0) if ticker_betas else 1.0
        co2_beta = ticker_betas.get(co2.upper(), 1.0) if ticker_betas else 1.0
        
        # Calculate individual alphas
        co1_alpha = return_co1 - (co1_beta * index_return)
        co2_alpha = return_co2 - (co2_beta * index_return)
        
        # Combine alphas based on tail
        if tail == 'L':
            pair_alpha = (w1 * co1_alpha) - (w2 * co2_alpha)
        else:
            pair_alpha = -(w1 * co1_alpha) + (w2 * co2_alpha)
        
        # Accumulate
        intraday_pnl = trade.get('Intraday_PnL', 0)
        total_pnl += intraday_pnl
        total_value += trade_value
        weighted_nominal += pair_nominal * trade_value
        weighted_alpha += pair_alpha * trade_value
    
    return {
        'total_intraday_pnl': total_pnl,
        'total_intraday_value': total_value,
        'weighted_nominal_return': weighted_nominal,
        'weighted_alpha_return': weighted_alpha,
        'trade_count': len(terminated_df)
    }

# IGV direct hedge (put options) retired in V9.4.
# IGV exposure is now managed via the factor shock framework
# (IGV added to VGT and VFH shortlists in Factor_Shock_Detection.py).
# Pre-trade IGV beta exposure constraint remains in Constraints.calculate_igv_exposure.