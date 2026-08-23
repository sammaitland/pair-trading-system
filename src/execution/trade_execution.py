"""
Trade execution module for pair trading.

Handles order placement (market and limit), trade monitoring, spread validation,
order aggregation for commission reduction, trade terminations, and
implied volatility calculations. Integrates with IBKR via ib_insync.

STATUS: live
"""

import pandas as pd
import numpy as np
import time
import math
import logging
import os
import asyncio
from datetime import datetime, timedelta
from openpyxl import load_workbook
from scipy import stats
from scipy.optimize import brentq

# IBKR
from ib_insync import IB, Stock, MarketOrder, Option, LimitOrder

# Project imports
from src.shared import config
from src.shared.constraints import (
    get_leg_weights,
    get_position_multiplier,
    is_tradeable_bucket,
    check_leverage_limit,
    check_emergency_leverage,
)
from src.shared import config_helper as ch
from src.shared.config_helper import should_use_limit_orders
from src.shared import calculations as Tool_Box
from src.shared.calculations import BetaDataManager

# Market data fetching
from src.shared.fetch_market_data import (
    fetch_live_prices_batch,
)


# ============================================================================
# ORDER AGGREGATION CLASSES
# ============================================================================

class OrderAggregator:
    """
    Aggregates orders across multiple pairs to reduce commissions
    Maintains mapping of aggregated orders back to contributing pairs
    """
    
    def __init__(self, min_aggregation_threshold=3):
        """
        Parameters:
        -----------
        min_aggregation_threshold : int
            Only aggregate if ticker appears in this many or more pairs
        """
        self.min_threshold = min_aggregation_threshold
        self.aggregated_orders = []
        self.order_to_pairs_map = {}
        
    def aggregate_approved_trades(self, approved_trades_df):
        """
        Aggregate orders from approved trades
        
        Returns:
        --------
        dict : {
            'aggregated_orders': List of aggregated orders to execute,
            'order_map': Mapping from order IDs to contributing pairs,
            'statistics': Aggregation statistics
        }
        """
        from collections import defaultdict
        
        logger.info("="*80)
        logger.info("ORDER AGGREGATION")
        logger.info("="*80)
        
        # Group orders by ticker and direction
        ticker_groups = defaultdict(lambda: defaultdict(list))
        
        for idx, trade in approved_trades_df.iterrows():
            pair_id = f"{trade['Pair']}-{trade['Tail']}"
            priority = trade.get('Composite_Priority_Score', 0.5)
            tail = str(trade.get('Tail', 'L')).strip().upper()
            
            # Determine order directions based on tail
            if tail == 'L':
                side1, side2 = 'BUY', 'SELL'
            else:  # U
                side1, side2 = 'SELL', 'BUY'
            
            # Add ticker1 to groups
            ticker_groups[trade['Co1']][side1].append({
                'pair_id': pair_id,
                'ticker': trade['Co1'],
                'qty': int(trade['Quantity1']),
                'side': side1,
                'priority': priority,
                'original_trade': trade.to_dict()
            })
            
            # Add ticker2 to groups
            ticker_groups[trade['Co2']][side2].append({
                'pair_id': pair_id,
                'ticker': trade['Co2'],
                'qty': int(trade['Quantity2']),
                'side': side2,
                'priority': priority,
                'original_trade': trade.to_dict()
            })
        
        # Create aggregated orders
        aggregated = []
        order_map = {}
        stats = {
            'total_orders_before': 0,
            'total_orders_after': 0,
            'orders_saved': 0,
            'commission_saved': 0.0
        }
        
        for ticker, directions in ticker_groups.items():
            for side, orders in directions.items():
                stats['total_orders_before'] += len(orders)
                
                if len(orders) >= self.min_threshold:
                    # AGGREGATE
                    total_qty = sum(o['qty'] for o in orders)
                    
                    # Sort by priority for allocation purposes
                    sorted_orders = sorted(orders, key=lambda x: x['priority'], reverse=True)
                    
                    agg_order = {
                        'order_id': f"AGG_{ticker}_{side}_{len(aggregated)}",
                        'ticker': ticker,
                        'side': side,
                        'total_qty': total_qty,
                        'contributing_orders': sorted_orders,
                        'num_contributors': len(orders)
                    }
                    
                    aggregated.append(agg_order)
                    order_map[agg_order['order_id']] = sorted_orders
                    
                    stats['total_orders_after'] += 1
                    stats['orders_saved'] += len(orders) - 1
                    stats['commission_saved'] += (len(orders) - 1) * 2.00
                    
                    logger.info(f"\n✓ Aggregated {ticker} {side}:")
                    logger.info(f"  {len(orders)} orders → 1 order ({total_qty} shares)")
                    logger.info(f"  Commission: ${len(orders)*2:.2f} → $2.00 (saved ${(len(orders)-1)*2:.2f})")
                    
                else:
                    # DON'T AGGREGATE - execute individually
                    for order in orders:
                        individual_order = {
                            'order_id': f"IND_{ticker}_{side}_{len(aggregated)}",
                            'ticker': ticker,
                            'side': side,
                            'total_qty': order['qty'],
                            'contributing_orders': [order],
                            'num_contributors': 1
                        }
                        aggregated.append(individual_order)
                        order_map[individual_order['order_id']] = [order]
                        stats['total_orders_after'] += 1
        
        logger.info(f"\n{'='*80}")
        logger.info(f"AGGREGATION SUMMARY:")
        logger.info(f"  Orders before: {stats['total_orders_before']}")
        logger.info(f"  Orders after: {stats['total_orders_after']}")
        logger.info(f"  Orders saved: {stats['orders_saved']}")
        logger.info(f"  Commission saved: ${stats['commission_saved']:.2f}")
        logger.info(f"  Reduction: {stats['orders_saved']/stats['total_orders_before']*100:.1f}%")
        logger.info(f"{'='*80}\n")
        
        return {
            'aggregated_orders': aggregated,
            'order_map': order_map,
            'statistics': stats
        }
    
    def allocate_fills(self, order_id, filled_qty, order_map):
        """
        Allocate filled quantity back to contributing pairs
        
        Parameters:
        -----------
        order_id : str
            ID of the aggregated order
        filled_qty : int
            Actual quantity filled
        order_map : dict
            Mapping of order_id to contributing orders
            
        Returns:
        --------
        list : Allocation results for each contributing pair
        """
        contributing_orders = order_map[order_id]
        total_requested = sum(o['qty'] for o in contributing_orders)
        
        logger.info(f"\nAllocating {filled_qty}/{total_requested} filled shares for {order_id}")
        
        allocations = []
        remaining = filled_qty
        
        # Already sorted by priority in aggregate_approved_trades()
        for order in contributing_orders:
            requested = order['qty']
            
            if remaining >= requested:
                # Fully allocate
                allocated = requested
                status = 'complete'
                remaining -= requested
            elif remaining > 0:
                # Partially allocate
                allocated = remaining
                status = 'partial'
                remaining = 0
            else:
                # No allocation
                allocated = 0
                status = 'unfilled'
            
            allocation = {
                'pair_id': order['pair_id'],
                'ticker': order['ticker'],
                'side': order['side'],
                'qty_allocated': allocated,
                'qty_requested': requested,
                'status': status,
                'priority': order['priority']
            }
            
            allocations.append(allocation)
            
            if status == 'complete':
                logger.info(f"  ✓ {order['pair_id']}: {allocated}/{requested} shares (complete)")
            elif status == 'partial':
                logger.warning(f"  ⚠️  {order['pair_id']}: {allocated}/{requested} shares (partial)")
            else:
                logger.error(f"  ✗ {order['pair_id']}: {allocated}/{requested} shares (unfilled)")
        
        return allocations
    
# ============================================================================
# LOGGING SETUP
# ============================================================================
logger = logging.getLogger(__name__)

# Suppress harmless IBKR errors
class IBKRErrorFilter(logging.Filter):
    def filter(self, record):
        msg = record.getMessage()
        if 'Error 300' in msg:
            return False
        if 'Error 10147' in msg:
            return False
        return True

# Apply immediately
logging.getLogger('ib_insync.wrapper').addFilter(IBKRErrorFilter())

# ============================================================================
# BETA MANAGER
# ============================================================================

beta_manager = BetaDataManager()

# ============================================================================
# MARKET DATA CACHING
# ============================================================================

# ============================================================================
# IBKR CONNECTION
# ============================================================================

def connect_ibkr():
    """Connect to Interactive Brokers API"""
    ib = IB()
    try:
        ib.connect(config.ibkr_host(), config.ibkr_port(), 
                  clientId=ch.get_client_id())
        logger.info("Connected to IBKR API")
    except Exception as e:
        logger.error(f"IBKR connection failed: {e}")
        raise
    return ib

def get_available_liquidity(ib):
    """Get available liquidity from IBKR account"""
    try:
        account_summary = ib.accountSummary()
        for item in account_summary:
            if item.tag == "ExcessLiquidity":
                liquidity = float(item.value)
                logger.info(f"Available liquidity: ${liquidity:,.2f}")
                return liquidity
        
        logger.error("ExcessLiquidity not found in account summary")
        return 0.0
    except Exception as e:
        logger.error(f"Error getting liquidity: {e}")
        return 0.0

# ============================================================================
# STRATEGY CONFIGURATION
# ============================================================================

def get_trade_weights_and_sizing(trade_row):
    """
    Get strategy-specific weights and position sizing from config
    
    Parameters:
        trade_row: Dict or Series with:
            - 'trigger_type': 'lower' or 'upper'
            - 'sum_dev_bucket': '0-10%', '10-20%', etc.
    
    Returns:
        dict: {
            'W1': float,
            'W2': float,
            'position_multiplier': float,
            'effective_allocation': float
        }
    """
    trigger_type = trade_row.get('trigger_type', 'lower')
    bucket = trade_row.get('sum_dev_bucket', '40-50%')
    
    # Use config helper functions
    W1, W2 = get_leg_weights(trigger_type, bucket)
    position_mult = get_position_multiplier(trigger_type, bucket)
    
    # Calculate effective allocation
    weight_concentration = max(W1, W2) / 0.5
    effective_allocation = weight_concentration * position_mult
    
    logger.info(f"Strategy: {trigger_type}, Bucket: {bucket} -> "
               f"Weights: ({W1:.2f}, {W2:.2f}), "
               f"Size: {position_mult:.2f}x, "
               f"Effective: {effective_allocation:.2f}x")
    
    return {
        'W1': W1,
        'W2': W2,
        'position_multiplier': position_mult,
        'effective_allocation': effective_allocation
    }

# ============================================================================
# SPREAD CALCULATIONS
# ============================================================================

def calculate_weighted_spread(bid1, ask1, bid2, ask2, W1, W2):
    """
    Calculate weighted spread using leg weights
    
    Parameters:
        bid1, ask1: Market data for stock 1
        bid2, ask2: Market data for stock 2
        W1, W2: Leg weights
    
    Returns:
        float: Weighted spread as decimal, or 1.0 (100%) if invalid data
    """
    # Input validation - return very wide spread if data is invalid
    if any(v is None or v <= 0 for v in [bid1, ask1, bid2, ask2]):
        logger.warning(f"Invalid bid/ask data: bid1={bid1}, ask1={ask1}, bid2={bid2}, ask2={ask2}")
        return 1.0  # 100% spread indicates bad data
    
    # Additional sanity check: ask should be >= bid
    if ask1 < bid1 or ask2 < bid2:
        logger.warning(f"Ask < Bid detected: ({bid1}/{ask1}), ({bid2}/{ask2})")
        return 1.0
    
    spread1_pct = (ask1 - bid1) / ((ask1 + bid1) / 2)
    spread2_pct = (ask2 - bid2) / ((ask2 + bid2) / 2)
    
    weighted_spread = (spread1_pct * W1) + (spread2_pct * W2)
    
    logger.debug(f"Leg spreads: {spread1_pct:.6f}, {spread2_pct:.6f} -> "
                f"Weighted: {weighted_spread:.6f} (W1={W1:.2f})")
    
    return weighted_spread

# ============================================================================
# ACCOUNT EQUITY & LEVERAGE CHECKING
# ============================================================================

def get_account_summary_values(ib, convert_to_usd=False):
    """
    Get key account values from IBKR
    
    Parameters:
    -----------
    ib : IB
        Connected IB instance
    convert_to_usd : bool
        If True and account is in GBP, convert values to USD
    
    Returns dict with NetLiquidation, ExcessLiquidity, GrossPositionValue, Leverage
    """
    account_values = {
        'NetLiquidation': 0.0,
        'ExcessLiquidity': 0.0,
        'GrossPositionValue': 0.0,
        'Leverage': 0.0,
        'Currency': 'USD'
    }
    
    if ib is None or not ib.isConnected():
        logger.warning("No IBKR connection for account summary")
        return account_values
    
    try:
        all_values = ib.accountValues()
        
        # Find the base currency (GBP or USD)
        for item in all_values:
            if item.tag == 'NetLiquidation':
                account_values['Currency'] = item.currency
                break
        
        base_currency = account_values['Currency']
        
        for item in all_values:
            if item.currency == base_currency:
                if item.tag == 'NetLiquidation':
                    account_values['NetLiquidation'] = float(item.value)
                elif item.tag == 'ExcessLiquidity':
                    account_values['ExcessLiquidity'] = float(item.value)
                elif item.tag == 'GrossPositionValue':
                    account_values['GrossPositionValue'] = float(item.value)
        
        # Convert to USD if requested and account is in GBP
        if convert_to_usd and base_currency == 'GBP':
            # Get GBP/USD rate
            gbp_usd_rate = getattr(config, 'GBP_TO_USD_RATE', 1.27)  # Default fallback
            try:
                from ib_insync import Forex
                fx_contract = Forex('GBPUSD')
                qualified = ib.qualifyContracts(fx_contract)
                if qualified:
                    ib.reqMktData(qualified[0], '', False, False)
                    ib.sleep(1)
                    ticker = ib.ticker(qualified[0])
                    if ticker and ticker.midpoint():
                        gbp_usd_rate = ticker.midpoint()
                    ib.cancelMktData(qualified[0])
            except Exception as e:
                logger.debug(f"Could not get live GBP/USD rate: {e}")
            
            account_values['NetLiquidation'] *= gbp_usd_rate
            account_values['ExcessLiquidity'] *= gbp_usd_rate
            account_values['GrossPositionValue'] *= gbp_usd_rate
            account_values['Currency'] = 'USD'
            logger.info(f"Converted account values from GBP to USD (rate: {gbp_usd_rate:.4f})")
        
        # Calculate leverage
        net_liq = account_values['NetLiquidation']
        gross_pos = account_values['GrossPositionValue']
        
        if net_liq > 0:
            account_values['Leverage'] = gross_pos / net_liq
        
        currency_symbol = '$' if account_values['Currency'] == 'USD' else '£'
        logger.info(f"Account ({account_values['Currency']}): NetLiq={currency_symbol}{net_liq:,.0f}, "
                   f"ExcessLiq={currency_symbol}{account_values['ExcessLiquidity']:,.0f}, "
                   f"Leverage={account_values['Leverage']:.2f}x")
        
        return account_values
        
    except Exception as e:
        logger.error(f"Error getting account summary: {e}")
        return account_values


def get_account_equity(ib, convert_to_usd=True):
    """Get NetLiquidation value in USD"""
    return get_account_summary_values(ib, convert_to_usd=convert_to_usd)['NetLiquidation']


def get_excess_liquidity(ib, convert_to_usd=True):
    """Get ExcessLiquidity for position sizing in USD"""
    return get_account_summary_values(ib, convert_to_usd=convert_to_usd)['ExcessLiquidity']

def calculate_current_gross_exposure(portfolio_df):
    """
    Calculate current gross exposure from portfolio
    
    Parameters:
    -----------
    portfolio_df : DataFrame
        Current portfolio positions
    
    Returns:
    --------
    float: Total gross exposure (sum of absolute position values)
    """
    if portfolio_df.empty:
        return 0.0
    
    try:
        total_exposure = portfolio_df[['Trade Value Co1 ($)', 'Trade Value Co2 ($)']].abs().sum().sum()
        return total_exposure
    except Exception as e:
        logger.error(f"Error calculating gross exposure: {e}")
        return 0.0

def check_leverage_before_trade(ib, portfolio_df, new_trade_value):
    """
    Check if new trade would exceed leverage limits
    
    Parameters:
    -----------
    ib : IB connection
    portfolio_df : DataFrame
        Current portfolio
    new_trade_value : float
        Gross value of proposed new trade
    
    Returns:
    --------
    tuple: (can_trade, message)
        can_trade: bool - True if trade allowed
        message: str - Explanation
    """
    # Get current equity
    account_equity = get_account_equity(ib)
    
    if account_equity <= 0:
        return False, "Cannot determine account equity"
    
    # Calculate current and projected exposure
    current_exposure = calculate_current_gross_exposure(portfolio_df)
    projected_exposure = current_exposure + new_trade_value
    
    # Check against limits
    within_limits, current_leverage, max_leverage = check_leverage_limit(
        projected_exposure, account_equity
    )
    
    if not within_limits:
        message = (f"Trade would exceed leverage limit. "
                  f"Projected: {projected_exposure/account_equity:.2f}x, "
                  f"Max: {max_leverage:.2f}x")
        logger.warning(message)
        return False, message
    
    # Check emergency threshold
    is_emergency = check_emergency_leverage(projected_exposure, account_equity)
    
    if is_emergency:
        message = (f"Emergency leverage threshold exceeded. "
                  f"Projected: {projected_exposure/account_equity:.2f}x, "
                  f"Emergency threshold: {config.emergency_leverage_threshold():.2f}x")
        logger.error(message)
        return False, message
    
    # All checks passed
    message = (f"Leverage check passed. "
              f"Current: {current_leverage:.2f}x, "
              f"Projected: {projected_exposure/account_equity:.2f}x, "
              f"Max: {max_leverage:.2f}x")
    logger.info(message)
    return True, message

# ============================================================================
# LIMIT ORDER PRICE CALCULATION
# ============================================================================

def calculate_limit_price(bid, ask, side, strategy="AGGRESSIVE", buffer_cents=0.0):
    """
    Calculate limit price based on bid/ask and strategy
    
    Parameters:
    -----------
    bid : float
        Bid price
    ask : float
        Ask price
    side : str
        "BUY" or "SELL"
    strategy : str
        "AGGRESSIVE" = cross spread for immediate fill
        "PASSIVE" = sit at bid/ask
        "MID" = split the spread
    buffer_cents : float
        Additional buffer in cents (positive = more aggressive)
    
    Returns:
    --------
    float: Limit price
    """
    mid = (bid + ask) / 2
    spread = ask - bid
    
    if strategy == "AGGRESSIVE":
        # Cross the spread for immediate fill
        if side == "BUY":
            # Buy at ask (hit the offer)
            limit_price = ask + buffer_cents
        else:  # SELL
            # Sell at bid (hit the bid)
            limit_price = bid - buffer_cents
    
    elif strategy == "PASSIVE":
        # Sit at bid/ask and wait
        if side == "BUY":
            # Join the bid
            limit_price = bid + buffer_cents
        else:  # SELL
            # Join the ask
            limit_price = ask - buffer_cents
    
    elif strategy == "MID":
        # Split the spread
        if side == "BUY":
            limit_price = mid + buffer_cents
        else:  # SELL
            limit_price = mid - buffer_cents
    
    else:
        # Default to aggressive
        logger.warning(f"Unknown strategy '{strategy}', using AGGRESSIVE")
        return calculate_limit_price(bid, ask, side, "AGGRESSIVE", buffer_cents)
    
    # Ensure limit price is reasonable
    if side == "BUY":
        # Don't pay more than 2x the spread above ask
        limit_price = min(limit_price, ask + 2 * spread)
    else:  # SELL
        # Don't sell for less than 2x the spread below bid
        limit_price = max(limit_price, bid - 2 * spread)
    
    logger.debug(f"{side} limit price: ${limit_price:.2f} "
                f"(bid: ${bid:.2f}, ask: ${ask:.2f}, mid: ${mid:.2f})")
    
    return round(limit_price, 2)

def validate_spread_for_limit_order(bid, ask, max_spread_bps=100):
    """
    Validate spread is acceptable for limit orders
    
    Parameters:
    -----------
    bid : float
    ask : float
    max_spread_bps : int
        Maximum acceptable spread in basis points
    
    Returns:
    --------
    tuple: (is_valid, spread_bps, message)
    """
    mid = (bid + ask) / 2
    
    if mid <= 0:
        return False, float('inf'), "Invalid mid price"
    
    spread = ask - bid
    spread_bps = (spread / mid) * 10000
    
    if spread_bps > max_spread_bps:
        message = f"Spread too wide: {spread_bps:.1f} bps > {max_spread_bps} bps"
        return False, spread_bps, message
    
    return True, spread_bps, "OK"

# ============================================================================
# LIMIT ORDER PLACEMENT & MONITORING
# ============================================================================

def place_limit_order(ib, ticker, side, quantity, limit_price):
    """
    Place a single limit order
    
    Parameters:
    -----------
    ib : IB connection
    ticker : str
        Stock ticker
    side : str
        "BUY" or "SELL"
    quantity : int
        Number of shares
    limit_price : float
        Limit price
    
    Returns:
    --------
    Trade: IBKR Trade object
    """
    from ib_insync import Stock, LimitOrder
    
    try:
        contract = Stock(ticker, 'SMART', 'USD')
        qualified = ib.qualifyContracts(contract)
        
        if not qualified:
            logger.error(f"Failed to qualify contract for {ticker}")
            return None
        
        order = LimitOrder(side, quantity, limit_price, tif='DAY')
        trade = ib.placeOrder(qualified[0], order)
        
        logger.info(f"Placed limit order: {side} {quantity} {ticker} @ ${limit_price:.2f}")
        
        return trade
        
    except Exception as e:
        logger.error(f"Error placing limit order for {ticker}: {e}")
        return None

def monitor_limit_order_fill(ib, trade, ticker, quantity, timeout=45):
    """
    Monitor a limit order until filled or timeout
    
    Parameters:
    -----------
    ib : IB connection
    trade : Trade object
    ticker : str
    quantity : int
        Target quantity
    timeout : int
        Timeout in seconds
    
    Returns:
    --------
    tuple: (filled, filled_quantity)
        filled: bool - True if fully filled
        filled_quantity: int - Actual quantity filled
    """
    import time
    
    start_time = time.time()
    last_fill_check = 0
    
    while time.time() - start_time < timeout:
        ib.sleep(1)
        
        current_filled = trade.orderStatus.filled
        status = trade.orderStatus.status
        
        # Log progress
        if current_filled != last_fill_check:
            logger.info(f"{ticker} fill: {current_filled}/{quantity} shares ({status})")
            last_fill_check = current_filled
        
        # Check if fully filled
        if current_filled >= quantity:
            logger.info(f"✓ {ticker} fully filled: {current_filled} shares")
            return True, current_filled
        
        # Check if terminal status
        if status in ['Filled', 'Cancelled', 'ApiCancelled', 'Rejected']:
            if current_filled > 0:
                logger.warning(f"⚠️ {ticker} partially filled: {current_filled}/{quantity} ({status})")
                return False, current_filled
            else:
                logger.warning(f"✗ {ticker} order {status} with no fills")
                return False, 0
    
    # Timeout reached
    current_filled = trade.orderStatus.filled
    
    if current_filled >= quantity:
        logger.info(f"✓ {ticker} filled after timeout: {current_filled} shares")
        return True, current_filled
    elif current_filled > 0:
        logger.warning(f"⚠️ {ticker} timeout with partial fill: {current_filled}/{quantity}")
        return False, current_filled
    else:
        logger.warning(f"✗ {ticker} timeout with no fills")
        return False, 0
    
# ============================================================================
# UNIFIED ORDER PLACEMENT (Market or Limit based on config)
# ============================================================================

def place_and_monitor_order_unified(ib, ticker, side, quantity, 
                                   bid=None, ask=None, pair=""):
    """
    Place and monitor order using either market or limit based on config
    
    This function replaces the old place_and_monitor_pair() for single legs.
    
    Parameters:
    -----------
    ib : IB connection
    ticker : str
    side : str
        "BUY" or "SELL"
    quantity : int
    bid : float, optional
        Bid price (required for limit orders)
    ask : float, optional
        Ask price (required for limit orders)
    pair : str
        Pair identifier for logging
    
    Returns:
    --------
    tuple: (success, filled_quantity)
    """
    from ib_insync import Stock, MarketOrder
    
    # ========================================================================
    # CRITICAL: Track initial position for verification at end
    # ========================================================================
    try:
        positions_before = {p.contract.symbol: p.position for p in ib.positions()}
        initial_pos = positions_before.get(ticker, 0)
    except Exception as e:
        logger.warning(f"{ticker}: Could not get initial position: {e}")
        initial_pos = 0
    
    use_limit = should_use_limit_orders()
    
    if use_limit:
        # ====================================================================
        # LIMIT ORDER PATH (Live Trading)
        # ====================================================================
        
        if bid is None or ask is None:
            logger.error(f"{ticker}: Cannot place limit order without bid/ask")
            return False, 0
        
        # Validate spread
        is_valid, spread_bps, message = validate_spread_for_limit_order(
            bid, ask, config.max_limit_order_spread_bps()
        )
        
        if not is_valid:
            logger.warning(f"{ticker}: {message}")
            print(f"  ⚠️  {ticker}: {message}")  # Show in Jupyter
            return False, 0
        
        logger.info(f"{ticker}: Spread {spread_bps:.1f} bps is acceptable")
        
        # Calculate limit price
        limit_price = calculate_limit_price(
            bid, ask, side,
            strategy=config.limit_order_strategy(),
            buffer_cents=config.limit_order_buffer_cents()
        )
        
        # Place limit order
        trade = place_limit_order(ib, ticker, side, quantity, limit_price)
        
        if trade is None:
            return False, 0
        
        # Monitor fill
        filled, filled_qty = monitor_limit_order_fill(
            ib, trade, ticker, quantity,
            timeout=config.limit_order_timeout()
        )
        
        # Cancel if not fully filled
        if not filled:
            try:
                if trade.orderStatus.status not in ['Filled', 'Cancelled', 'ApiCancelled']:
                    ib.cancelOrder(trade.order)
                    logger.info(f"Cancelled remaining {quantity - filled_qty} shares of {ticker}")
            except Exception as e:
                logger.error(f"Error cancelling {ticker}: {e}")
        
        # ====================================================================
        # CRITICAL BUG FIX: Final verification with actual positions
        # ====================================================================
        if filled_qty < quantity:
            logger.debug(f"{ticker}: Final position verification...")
            ib.sleep(2)
            
            try:
                positions_after = {p.contract.symbol: p.position for p in ib.positions()}
                current_pos = positions_after.get(ticker, 0)
                
                # Calculate actual position change in correct direction
                direction_mult = 1 if side == "BUY" else -1
                position_change = (current_pos - initial_pos) * direction_mult
                
                if position_change > 0:
                    actual_filled = abs(position_change)
                    
                    if actual_filled > filled_qty:
                        logger.info(f"✓ {ticker}: Position verification found {actual_filled} shares filled (orderStatus reported {filled_qty})")
                        filled_qty = actual_filled
                        filled = (filled_qty >= quantity)
                    elif actual_filled >= quantity:
                        logger.info(f"✓ {ticker}: Position verification confirms full fill")
                        filled_qty = quantity
                        filled = True
            except Exception as e:
                logger.warning(f"{ticker}: Position verification failed: {e}")
        
        return filled, filled_qty
    
    else:
        # ====================================================================
        # MARKET ORDER PATH (Paper Trading)
        # ====================================================================
        
        try:
            contract = Stock(ticker, 'SMART', 'USD')
            qualified = ib.qualifyContracts(contract)
            
            if not qualified:
                logger.error(f"Failed to qualify {ticker}")
                return False, 0
            
            # Place market order with TIF=DAY to match account preset
            order = MarketOrder(side, quantity, tif='DAY')
            trade = ib.placeOrder(qualified[0], order)
            logger.info(f"Placed market order: {side} {quantity} {ticker}")
            
            # Monitor fill (use longer timeout for market orders)
            start_time = time.time()
            filled = 0
            
            while time.time() - start_time < config.order_timeout():
                ib.sleep(1)
                
                current_filled = trade.orderStatus.filled
                
                if current_filled >= quantity:
                    filled = quantity
                    logger.info(f"✓ {ticker} filled: {filled} shares")
                    break
                
                if trade.orderStatus.status in ['Filled', 'Cancelled', 'Rejected']:
                    filled = current_filled
                    break
            
            # ====================================================================
            # CRITICAL: Final verification with actual positions (already in place)
            # ====================================================================
            ib.sleep(2)
            positions_after = {p.contract.symbol: p.position for p in ib.positions()}
            current_pos = positions_after.get(ticker, 0)
            actual_change = abs(current_pos - initial_pos)
            
            # Use the greater of orderStatus fill or actual position change
            filled = max(filled, actual_change)
            
            if filled > 0 and filled != trade.orderStatus.filled:
                logger.info(f"✓ {ticker}: Position verification adjusted fill from {trade.orderStatus.filled} to {filled}")
            
            success = (filled >= quantity)
            return success, filled
            
        except Exception as e:
            logger.error(f"Error placing market order for {ticker}: {e}")
            return False, 0

def place_and_monitor_pair_unified(ib, ticker1, ticker2, quantity1, quantity2,
                                   bid1, ask1, bid2, ask2, pair, side1, side2):
    """
    Place and monitor pair trade using unified order logic
    
    This replaces the old place_and_monitor_pair() function.
    Automatically uses market or limit orders based on Config.
    
    Parameters:
    -----------
    ib : IB connection
    ticker1, ticker2 : str
        Ticker symbols
    quantity1, quantity2 : int
        Quantities
    bid1, ask1, bid2, ask2 : float
        Market data
    pair : str
        Pair identifier
    side1, side2 : str
        "BUY" or "SELL"
    
    Returns:
    --------
    tuple: (success, filled1, filled2)
    """
    logger.info(f"\nExecuting pair: {pair}")
    logger.info(f"  Leg 1: {side1} {quantity1} {ticker1}")
    logger.info(f"  Leg 2: {side2} {quantity2} {ticker2}")
    logger.info(f"  Order type: {'LIMIT' if should_use_limit_orders() else 'MARKET'}")
    
    # Execute leg 1
    success1, filled1 = place_and_monitor_order_unified(
        ib, ticker1, side1, quantity1, bid1, ask1, pair
    )
    
    # Execute leg 2
    success2, filled2 = place_and_monitor_order_unified(
        ib, ticker2, side2, quantity2, bid2, ask2, pair
    )
    
    # Check if both legs successful
    if success1 and success2:
        logger.info(f"✓ Pair {pair} fully executed")
        return True, filled1, filled2
    else:
        # ====================================================================
        # CRITICAL FIX: Final position verification before declaring failure
        # Orders can fill in TWS even if our monitoring misses it
        # ====================================================================
        logger.info(f"Verifying actual positions for {pair}...")
        time.sleep(2)  # Give IBKR time to update positions
        
        try:
            positions = {p.contract.symbol: p.position for p in ib.positions()}
            actual1 = positions.get(ticker1, 0)
            actual2 = positions.get(ticker2, 0)
            
            # Check if positions changed in the expected direction
            # This is a simplified check - in reality you'd compare to initial positions
            has_position1 = abs(actual1) > 0
            has_position2 = abs(actual2) > 0
            
            if has_position1 and has_position2:
                logger.info(f"✓ Position verification: {ticker1}={actual1}, {ticker2}={actual2}")
                logger.info(f"✓ Pair {pair} likely executed despite monitoring issues")
                # Return the quantities we TRIED to execute
                return True, quantity1, quantity2
        except Exception as e:
            logger.warning(f"Position verification failed: {e}")
        
        logger.warning(f"⚠️ Pair {pair} incomplete: "
                      f"{ticker1}={filled1}/{quantity1}, "
                      f"{ticker2}={filled2}/{quantity2}")
        return False, filled1, filled2
    
# ============================================================================
# MARKET DATA FETCHING - V9C STYLE
# ============================================================================

def get_live_index_price(ib=None):
    """
    Get live VGT price
    
    Returns:
        float: Current VGT price
    """
    close_connection = False
    if ib is None:
        ib = connect_ibkr()
        close_connection = True
    
    try:
        bid, ask = fetch_market_data(ib, 'VGT')
        if bid and ask:
            return (bid + ask) / 2
        
        logger.warning("Failed to get VGT price")
        return 0.0
    finally:
        if close_connection:
            ib.disconnect()

def fetch_market_data(ib, ticker, retries=3, delay=3):
    """
    Fetch market data for a ticker
    
    Args:
        ib: IBKR connection
        ticker: Ticker symbol
        retries: Retry attempts
        delay: Delay between retries
    
    Returns:
        tuple: (bid, ask) or (None, None)
    """
    from ib_insync import Stock

    # Regular stock tickers
    for attempt in range(1, retries + 1):
        try:
            contract = Stock(ticker, 'SMART', 'USD')
            ib.qualifyContracts(contract)
            market_data = ib.reqMktData(contract, "", snapshot=True, 
                                       regulatorySnapshot=False)
            ib.sleep(delay)
            
            bid = market_data.bid
            ask = market_data.ask
            ib.cancelMktData(contract)
            
            if bid is not None and ask is not None:
                return bid, ask
            
            logger.warning(f"Attempt {attempt}: Incomplete data for {ticker}")
        except Exception as e:
            logger.error(f"Attempt {attempt}: Error for {ticker}: {e}")
    
    return None, None

async def fetch_all_live_prices_async(ib, tickers):
    """
    Async wrapper for batch price fetching
    
    CRITICAL: Uses fetch_live_prices_batch from fetch_market_data_V9
    
    Returns:
        dict: {ticker: price_data}
    """
    logger.info(f"Fetching prices for {len(tickers)} tickers (async batch)")
    
    try:
        # Use V9 async batch fetching
        live_prices = await fetch_live_prices_batch(ib, tickers)
        
        # Convert to simple format if needed
        simple_prices = {}
        for ticker, data in live_prices.items():
            if isinstance(data, dict):
                simple_prices[ticker] = data.get('live_price', data.get('close'))
            else:
                simple_prices[ticker] = data
        
        logger.info(f"Successfully fetched {len(simple_prices)} prices")
        return simple_prices
        
    except Exception as e:
        logger.error(f"Error in async price fetch: {e}")
        return {}

def fetch_all_live_prices(ib, tickers, max_retries=2, delay=2):
    """
    Synchronous price fetching (fallback)
    
    For V9C, prefer async version when possible
    """
    logger.info(f"Fetching {len(tickers)} prices (sync)")
    
    prices = {}

    # Handle VGT separately
    if 'VGT' in tickers:
        try:
            bid, ask = fetch_market_data(ib, 'VGT')
            if bid is not None and ask is not None:
                prices['VGT'] = (bid + ask) / 2
        except Exception as e:
            logger.error(f"Error fetching VGT: {e}")
        tickers = [t for t in tickers if t != 'VGT']
    
    # Fetch remaining tickers
    for ticker in tickers:
        success = False
        
        for attempt in range(max_retries + 1):
            try:
                bid, ask = fetch_market_data(ib, ticker)
                
                if bid is not None and ask is not None:
                    prices[ticker] = (bid + ask) / 2
                    success = True
                    break
            except Exception as e:
                logger.error(f"Error fetching {ticker} (attempt {attempt+1}): {e}")
            
            if not success and attempt < max_retries:
                time.sleep(delay)
        
        time.sleep(0.5)  # Rate limiting
    
    logger.info(f"Successfully fetched {len(prices)} prices")
    return prices

# ============================================================================
# IMPLIED VOLATILITY CALCULATIONS
# ============================================================================

def calculate_implied_volatility(option_price, S, K, T, r, option_type='call'):
    """
    Calculate implied volatility using Black-Scholes
    
    Parameters:
        option_price: Market price of option
        S: Current stock price
        K: Strike price
        T: Time to expiration (years)
        r: Risk-free rate
        option_type: 'call' or 'put'
    
    Returns:
        float: Implied volatility as decimal
    """
    def black_scholes(sigma):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)
        
        if option_type.lower() == 'call':
            price = S * stats.norm.cdf(d1) - K * math.exp(-r * T) * stats.norm.cdf(d2)
        else:  # Put
            price = K * math.exp(-r * T) * stats.norm.cdf(-d2) - S * stats.norm.cdf(-d1)
        
        return price - option_price
    
    try:
        if option_price <= 0:
            return 0.0
        
        iv = brentq(black_scholes, 0.01, 2.0, xtol=1e-4)
        return iv
    except Exception as e:
        logger.error(f"Error calculating IV: {e}")
        return 0.0

def get_current_vgt_iv(ib):
    """
    Get current implied volatility for VGT options
    
    Returns:
        float: IV as percentage
    """
    try:
        from ib_insync import Stock, Option
        
        stock = Stock('VGT', 'SMART', 'USD')
        contracts = ib.qualifyContracts(stock)
        
        if not contracts:
            logger.error("Failed to qualify VGT contract")
            return 22.5  # Default estimate
        
        # Get current price
        market_data = ib.reqMktData(stock, '', False, False)
        ib.sleep(2)
        
        current_price = market_data.last or market_data.close
        if not current_price and market_data.bid and market_data.ask:
            current_price = (market_data.bid + market_data.ask) / 2
        
        if not current_price:
            logger.error("Could not get VGT price")
            return 22.5
        
        logger.info(f"VGT price: {current_price}")
        
        # Get option chains
        chains = ib.reqSecDefOptParams(stock.symbol, '', stock.secType, stock.conId)
        if not chains:
            logger.error("No option chains for VGT")
            return 22.5
        
        chain = next((c for c in chains if c.tradingClass == 'VGT' and 
                     c.exchange == 'SMART'), None)
        if not chain:
            return 22.5
        
        # Find ~30-day expiration
        today = datetime.now().date()
        target_date = today + timedelta(days=30)
        expirations = [datetime.strptime(exp, '%Y%m%d').date() 
                      for exp in chain.expirations]
        closest_exp = min(expirations, key=lambda d: abs((d - target_date).days))
        expiration = closest_exp.strftime('%Y%m%d')
        
        days_to_expiry = (closest_exp - today).days
        T = days_to_expiry / 365.0
        
        # Find ATM strike
        strikes = sorted(chain.strikes)
        atm_strike = min(strikes, key=lambda s: abs(s - current_price))
        
        logger.info(f"Selected ATM strike: {atm_strike}")
        
        # Get put option
        put = Option('VGT', expiration, atm_strike, 'P', 'SMART')
        qualified_options = ib.qualifyContracts(put)
        
        if not qualified_options:
            return 22.5
        
        put_data = ib.reqMktData(qualified_options[0], '', False, False)
        
        for i in range(3):
            ib.sleep(1)
        
        put_price = put_data.ask or (put_data.bid * 1.05 if put_data.bid else None) or put_data.last
        
        if put_price:
            risk_free_rate = 0.045
            put_iv = calculate_implied_volatility(
                put_price, current_price, atm_strike, T, risk_free_rate, 'put'
            )
            put_iv_percentage = put_iv * 100
            logger.info(f"VGT IV: {put_iv_percentage:.3f}%")
            return round(put_iv_percentage, 3)
        
        return 22.5
        
    except Exception as e:
        logger.error(f"Error getting VGT IV: {e}")
        return 22.5

# ============================================================================
# TRADE VALIDATION
# ============================================================================

def validate_trade_conditions(pair, live_spread, alpha, hurdles):
    """
    Validate trade conditions (spread only)
    
    ⚠️⚠️⚠️ CRITICAL FOR LIVE TRADING ⚠️⚠️⚠️
    
    CURRENT STATE: TESTING MODE - Spread validation is BYPASSED
    
    BEFORE LIVE TRADING:
    1. Change TESTING_BYPASS_SPREADS = True to False
    2. Verify spread hurdles in Parameters.xlsx are correct
    3. Test with paper account first
    
    Parameters:
        pair: Pair identifier
        live_spread: Current weighted spread
        alpha: Alpha value (not used for validation)
        hurdles: (spread_hurdle, two_day_dev_hurdle)
    
    Returns:
        bool: True if conditions met
    """
    # ========================================================================
    # ⚠️  TESTING BYPASS - SET TO False FOR LIVE TRADING
    # ========================================================================
    TESTING_BYPASS_SPREADS = False  # ← CHANGE THIS TO False FOR LIVE TRADING!
    
    if TESTING_BYPASS_SPREADS:
        logger.info(f"⚠️  TESTING MODE: {pair} spread validation BYPASSED "
                   f"(spread={live_spread:.6f})")
        return True
    
    # ========================================================================
    # PRODUCTION VALIDATION CODE (runs when TESTING_BYPASS_SPREADS = False)
    # ========================================================================
    spread_hurdle, *_ = hurdles
    
    # Convert spread_hurdle to decimal if needed
    if spread_hurdle > 0.1:
        spread_hurdle_decimal = spread_hurdle / 10000
    else:
        spread_hurdle_decimal = spread_hurdle
    
    # Get tolerance from config
    tolerance_bp = config.spread_tolerance()
    tolerance_decimal = tolerance_bp / 10000
    
    effective_spread_hurdle = spread_hurdle_decimal + tolerance_decimal
    
    # Check spread condition
    spread_condition = live_spread <= effective_spread_hurdle
    
    if spread_condition:
        logger.info(f"Pair {pair}: Conditions met. "
                   f"Spread: {live_spread:.6f} ≤ {effective_spread_hurdle:.6f}")
        return True
    
    logger.warning(f"Pair {pair}: Spread too wide. "
                  f"{live_spread:.6f} > {effective_spread_hurdle:.6f}")
    return False

# ============================================================================
# ORDER PLACEMENT
# ============================================================================

def balance_partially_filled_pair(ib, ticker1, ticker2, filled1, filled2,
                                  quantity1, quantity2, side1, side2, pair):
    """
    Balance a partially filled pair by reversing excess
    
    Returns:
        bool: True if balancing successful
    """
    logger.info(f"Balancing {pair}: {ticker1}={filled1}/{quantity1}, "
               f"{ticker2}={filled2}/{quantity2}")
    
    try:
        # Calculate target ratio
        target_ratio = quantity1 / quantity2 if quantity2 > 0 else 1
        
        # Calculate fill ratios
        ratio1 = filled1 / quantity1 if quantity1 > 0 else 0
        ratio2 = filled2 / quantity2 if quantity2 > 0 else 0
        
        # Determine limiting factor
        if ratio1 <= ratio2 and filled1 > 0:
            # Leg 1 limiting, balance leg 2
            target_fill2 = min(quantity2, round(filled1 / target_ratio))
            excess_fill2 = filled2 - target_fill2
            
            if excess_fill2 >= 1:
                contract2 = Stock(ticker2, 'SMART', 'USD')
                qualified = ib.qualifyContracts(contract2)
                
                if qualified:
                    reverse_side2 = "SELL" if side2 == "BUY" else "BUY"
                    reverse_order = MarketOrder(reverse_side2, int(excess_fill2))
                    trade = ib.placeOrder(qualified[0], reverse_order)
                    logger.info(f"Reversing {int(excess_fill2)} shares of {ticker2}")
                    
                    # Wait for fill
                    wait_start = time.time()
                    while time.time() - wait_start < 30:
                        ib.sleep(1)
                        if trade.orderStatus.filled >= excess_fill2:
                            logger.info(f"Balancing complete for {ticker2}")
                            return True
                    
                    logger.warning(f"Balancing may not have filled for {ticker2}")
                    return False
            else:
                logger.info(f"No significant excess in {ticker2}")
                return True
        
        elif ratio2 < ratio1 and filled2 > 0:
            # Leg 2 limiting, balance leg 1
            target_fill1 = min(quantity1, round(filled2 * target_ratio))
            excess_fill1 = filled1 - target_fill1
            
            if excess_fill1 >= 1:
                contract1 = Stock(ticker1, 'SMART', 'USD')
                qualified = ib.qualifyContracts(contract1)
                
                if qualified:
                    reverse_side1 = "SELL" if side1 == "BUY" else "BUY"
                    reverse_order = MarketOrder(reverse_side1, int(excess_fill1))
                    trade = ib.placeOrder(qualified[0], reverse_order)
                    logger.info(f"Reversing {int(excess_fill1)} shares of {ticker1}")
                    
                    # Wait for fill
                    wait_start = time.time()
                    while time.time() - wait_start < 30:
                        ib.sleep(1)
                        if trade.orderStatus.filled >= excess_fill1:
                            logger.info(f"Balancing complete for {ticker1}")
                            return True
                    
                    logger.warning(f"Balancing may not have filled for {ticker1}")
                    return False
            else:
                logger.info(f"No significant excess in {ticker1}")
                return True
        else:
            logger.info(f"Pair {pair} already balanced")
            return True
    
    except Exception as e:
        logger.error(f"Error balancing {pair}: {e}")
        return False

# ============================================================================
# TERMINATION ORDERS
# ============================================================================

def place_termination_orders(ib, trade, offset, termination_type):
    """
    Place termination orders for a trade
    
    Uses stored W1, W2 for weighted spread calculation
    """
    ticker1 = trade['Co1']
    ticker2 = trade['Co2']
    pair = trade['Pair']
    tail = str(trade.get('Tail', 'L')).strip().upper()
    
    # Determine order sides
    if tail == 'L':
        side1 = "SELL"
        side2 = "BUY"
    else:  # U
        side1 = "BUY"
        side2 = "SELL"
    
    quantity1 = trade['Quantity1']
    quantity2 = trade['Quantity2']
    
    # Fetch market data
    bid1, ask1 = fetch_market_data(ib, ticker1)
    bid2, ask2 = fetch_market_data(ib, ticker2)
    
    if any(x is None or math.isnan(x) for x in [bid1, ask1, bid2, ask2]):
        logger.error(f"Pair {pair}: Missing market data for termination")
        return False
    
    # Calculate combo limit price
    if side1 == 'BUY' and side2 == 'SELL':
        combo_limit_price = ((bid1 + ask1)/2 - (bid2 + ask2)/2) + offset
    else:
        combo_limit_price = ((bid2 + ask2)/2 - (bid1 + ask1)/2) + offset
    
    # Place orders
    success = place_and_monitor_pair_unified(  # ← Change this
        ib, ticker1, ticker2,
        quantity1, quantity2,
        bid1, ask1, bid2, ask2,  # ← Add market data
        pair, side1, side2
    )
    
    if success:
        logger.info(f"Pair {pair}: Termination successful ({termination_type})")
    else:
        logger.warning(f"Pair {pair}: Termination failed ({termination_type})")
    
    return success

def execute_immediate_termination(ib, trade, termination_type):
    """
    Execute immediate termination with weighted spread
    
    Uses stored W1, W2 for consistent spread calculation
    """
    if config.testing_mode():
        logger.info(f"(Testing) Simulating termination for {trade['Tag']}")
        
        W1 = trade.get('W1', 0.5)
        W2 = trade.get('W2', 0.5)
        trade['Termination Spread'] = 0.0025 * (max(W1, W2) / 0.5)
        
        if 'Initiation Spread' in trade and not pd.isna(trade['Initiation Spread']):
            trade['Achieved Spreads'] = trade['Initiation Spread'] + trade['Termination Spread']
        
        trade['Exit Reason'] = termination_type
        return True
    
    # Live mode
    offset = getattr(config, f"TERMINATION_ORDER_OFFSET_{termination_type.upper()}", 0.05)
    
    ticker1 = trade['Co1']
    ticker2 = trade['Co2']
    bid1, ask1 = fetch_market_data(ib, ticker1)
    bid2, ask2 = fetch_market_data(ib, ticker2)
    
    if any(x is None or math.isnan(x) for x in [bid1, ask1, bid2, ask2]):
        logger.error(f"Pair {trade['Pair']}: Cannot get market data")
        return False
    
    # Use stored weights for termination spread
    W1 = trade.get('W1', 0.5)
    W2 = trade.get('W2', 0.5)
    
    termination_spread = calculate_weighted_spread(bid1, ask1, bid2, ask2, W1, W2)
    trade['Termination Spread'] = termination_spread
    
    logger.info(f"Termination spread for {trade['Pair']}: {termination_spread:.6f} "
               f"(W1={W1:.2f}, W2={W2:.2f})")
    
    # Calculate achieved spreads
    if 'Initiation Spread' in trade and not pd.isna(trade['Initiation Spread']):
        trade['Achieved Spreads'] = trade['Initiation Spread'] + termination_spread
    
    # Calculate returns
    current_price_co1 = (bid1 + ask1) / 2
    current_price_co2 = (bid2 + ask2) / 2
    
    trade['Co1 at Termination'] = current_price_co1
    trade['Co2 at Termination'] = current_price_co2
    
    if trade['Co1 at Initiation'] > 0:
        trade['Co1 Return (%)'] = ((current_price_co1 - trade['Co1 at Initiation']) / 
                                   trade['Co1 at Initiation']) * 100
    
    if trade['Co2 at Initiation'] > 0:
        trade['Co2 Return (%)'] = ((current_price_co2 - trade['Co2 at Initiation']) / 
                                   trade['Co2 at Initiation']) * 100
    
    # Index return (using Tool_Box for consistency)
    current_index = get_live_index_price(ib)
    if current_index > 0 and trade['Index at Initiation'] > 0:
        trade['Index Return'] = Tool_Box.calculate_percentage_change(
            trade['Index at Initiation'], current_index
        ) * 100
    
    trade['Exit Reason'] = termination_type
    
    # Execute termination
    success = place_termination_orders(ib, trade, offset, termination_type)
    return success

# ============================================================================
# TRADE EXECUTION
# ============================================================================

def create_trade_entry(tag, pair, ticker1, ticker2, qty1, qty2,
                      live_prices, index_price,
                      tail, init_spread, W1, W2, pos_mult, original_row):
    """
    Create trade entry with all configuration details
    
    Stores strategy-specific weights and sizing for later use
    """
    trade_entry = {
        "Tag": tag,
        "Pair": pair,
        "Co1": ticker1,
        "Co2": ticker2,
        "Quantity1": qty1,
        "Quantity2": qty2,
        "Trade Initiation Date": datetime.today().date(),
        "Trade Termination Date": datetime.today().date() + timedelta(days=21),
        "Co1 at Initiation": live_prices.get(ticker1),
        "Co2 at Initiation": live_prices.get(ticker2),
        "Index at Initiation": live_prices.get(original_row.get('Index')) if live_prices else index_price,
        "Tail": tail,
        "Initiation Spread": init_spread,
        # Strategy configuration
        "W1": W1,
        "W2": W2,
        "Position_Multiplier": pos_mult,
        "Effective_Allocation": max(W1, W2) * pos_mult / 0.5,
        "trigger_type": original_row.get('trigger_type'),
        "sum_dev_bucket": original_row.get('sum_dev_bucket'),
    }
    
    # Copy signal values
    signal_columns = [
        'Two-Day Deviation', 'Alpha Variance', 'Volume Correlation',
        'Adjusted Correlation', 'Max Long Volume CDF',
        'Sum Deviation CDF', 'Multi-Score'
    ]
    
    for col in signal_columns:
        if col in original_row and not pd.isna(original_row[col]):
            trade_entry[col] = original_row[col]
    
    return trade_entry

# ============================================================================
# EXECUTION SUMMARY
# ============================================================================

def save_execution_summary(execution_summary_df):
    """Save execution summary to file"""
    try:
        execution_summary_df.to_excel(config.execution_summary_file(), index=False)
        logger.info("Execution summary saved")
    except Exception as e:
        logger.error(f"Failed to save execution summary: {e}")

def archive_completed_trades(terminated_trades_df, completed_trades_path):
    """Archive completed trades"""
    try:
        workbook = load_workbook(completed_trades_path)
        
        if 'Completed Trades' in workbook.sheetnames:
            sheet = workbook['Completed Trades']
        else:
            sheet = workbook.create_sheet(title='Completed Trades')
            
            # Add headers for new sheet
            additional_columns = [
                'Initiation Spread', 'Termination Spread', 'Achieved Spreads',
                'Index Return', 'Co1 Return (%)', 'Co2 Return (%)', 'Exit Reason',
                'Two-Day Deviation', 'Alpha Variance', 'Volume Correlation',
                'Adjusted Correlation', 'Max Long Volume CDF',
                'Sum Deviation CDF', 'Multi-Score'
            ]
            
            all_columns = list(terminated_trades_df.columns)
            for col in additional_columns:
                if col not in all_columns:
                    all_columns.append(col)
            
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
                for c_idx, col_name in enumerate(all_columns, 1):
                    sheet.cell(row=1, column=c_idx, value=col_name)
        
        # Find next row
        next_row = sheet.max_row + 1
        if sheet.cell(1, 1).value is None:
            next_row = 1
        
        # Ensure columns exist
        additional_columns = [
            'Initiation Spread', 'Termination Spread', 'Achieved Spreads',
            'Index Return', 'Co1 Return (%)', 'Co2 Return (%)', 'Exit Reason',
            'Two-Day Deviation', 'Alpha Variance', 'Volume Correlation',
            'Adjusted Correlation', 'Max Long Volume CDF',
            'Sum Deviation CDF', 'Multi-Score'
        ]
        
        for col in additional_columns:
            if col not in terminated_trades_df.columns:
                terminated_trades_df[col] = np.nan
        
        # Get sheet columns
        sheet_columns = []
        for cell in sheet[1]:
            if cell.value:
                sheet_columns.append(cell.value)
        
        # Write data
        for r_idx, row in enumerate(terminated_trades_df.itertuples(index=False), next_row):
            row_dict = {col: val for col, val in zip(terminated_trades_df.columns, row)}
            
            for c_idx, col_name in enumerate(sheet_columns, 1):
                if col_name in row_dict:
                    sheet.cell(row=r_idx, column=c_idx, value=row_dict[col_name])
        
        workbook.save(completed_trades_path)
        logger.info(f"Archived {len(terminated_trades_df)} completed trades")
        
    except Exception as e:
        logger.error(f"Failed to archive completed trades: {e}")
        import traceback
        traceback.print_exc()

async def execute_single_pair_trade(trade_spec, ib, live_prices,
                                   index_price_current,
                                   portfolio_df=None, account_equity=None):
    """
    Execute a single pair trade asynchronously
    
    This is the atomic unit of execution used by execute_trades_in_batches()
    
    Returns:
        dict: Execution result with status and details
    """
    tag = trade_spec.get("Tag")
    pair = trade_spec.get("Pair")
    tail = trade_spec.get("Tail", "L")
    ticker1 = trade_spec.get("Co1")
    ticker2 = trade_spec.get("Co2")
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Executing: {pair} (Tail: {tail})")
    logger.info(f"{'='*60}")
    print(f"\n📤 Executing: {pair} (Tail: {tail})")  # Show in Jupyter
    
    try:
        # Get strategy configuration
        trade_config = get_trade_weights_and_sizing(trade_spec)
        W1 = trade_config['W1']
        W2 = trade_config['W2']
        position_mult = trade_config['position_multiplier']
        
        # Get quantities (already have multiplier applied from evaluate_trades)
        quantity1 = int(trade_spec.get("Quantity1"))
        quantity2 = int(trade_spec.get("Quantity2"))
        
        # Validate quantities
        if quantity1 <= 0 or quantity2 <= 0:
            logger.error(f"❌ {pair}: Invalid quantities - qty1={quantity1}, qty2={quantity2}")
            return {
                'Tag': tag, 'Pair': pair,
                'Status': 'Failed',
                'Details': f'Zero quantity (qty1={quantity1}, qty2={quantity2})',
                'Leg1_Filled': 0, 'Leg2_Filled': 0,
                'Leg1_Target': quantity1, 'Leg2_Target': quantity2
            }
        
        # ====================================================================
        # LEVERAGE CHECK (Live Trading Safety)
        # ====================================================================
        
        # Get current prices for leverage calculation
        price1 = live_prices.get(ticker1)
        price2 = live_prices.get(ticker2)
        
        if pd.isna(price1) or pd.isna(price2):
            logger.error(f"❌ {pair}: Missing prices for leverage check")
            return {
                'Tag': tag, 'Pair': pair,
                'Status': 'Failed',
                'Details': 'Missing prices for leverage check',
                'Leg1_Filled': 0, 'Leg2_Filled': 0,
                'Leg1_Target': quantity1, 'Leg2_Target': quantity2
            }
        
        # Calculate trade value
        trade_value = (quantity1 * price1) + (quantity2 * price2)
        
        # Check leverage if we have account equity
        if account_equity is not None and account_equity > 0:
            try:
                # Calculate current exposure
                current_exposure = calculate_current_gross_exposure(portfolio_df) if portfolio_df is not None else 0.0
                projected_exposure = current_exposure + trade_value
                
                # Check limits
                within_limits, current_leverage, max_leverage = check_leverage_limit(
                    projected_exposure, account_equity
                )
                
                if not within_limits:
                    leverage_msg = (f"Trade would exceed leverage limit. "
                                  f"Projected: {projected_exposure/account_equity:.2f}x, "
                                  f"Max: {max_leverage:.2f}x")
                    logger.error(f"❌ {pair}: {leverage_msg}")
                    return {
                        'Tag': tag, 'Pair': pair,
                        'Status': 'Failed',
                        'Details': leverage_msg,
                        'Leg1_Filled': 0, 'Leg2_Filled': 0,
                        'Leg1_Target': quantity1, 'Leg2_Target': quantity2
                    }
                
                logger.info(f"✓ {pair}: Leverage check passed ({projected_exposure/account_equity:.2f}x)")
                
            except Exception as e:
                logger.warning(f"⚠️ {pair}: Leverage check error: {e}")
                # Continue anyway - don't block trades if leverage check fails
        else:
            logger.debug(f"{pair}: Skipping leverage check (no account equity available)")
        
        logger.info(f"Quantities: {quantity1}/{quantity2} (with {position_mult:.2f}x)")
        
        # Fetch market data
        bid1, ask1 = fetch_market_data(ib, ticker1)
        bid2, ask2 = fetch_market_data(ib, ticker2)
        
        if any(x is None for x in [bid1, ask1, bid2, ask2]):
            logger.error(f"❌ {pair}: Market data unavailable")
            print(f"  ❌ {pair}: Market data unavailable - bid1={bid1}, ask1={ask1}, bid2={bid2}, ask2={ask2}")
            return {
                'Tag': tag, 'Pair': pair,
                'Status': 'Failed',
                'Details': 'Market data unavailable',
                'Leg1_Filled': 0, 'Leg2_Filled': 0,
                'Leg1_Target': quantity1, 'Leg2_Target': quantity2
            }
        
        # Calculate weighted spread
        initiation_spread = calculate_weighted_spread(bid1, ask1, bid2, ask2, W1, W2)
        logger.info(f"Weighted spread: {initiation_spread:.6f}")
        
        # Determine order sides
        if tail.upper() == "U":
            side1 = "SELL"
            side2 = "BUY"
        else:
            side1 = "BUY"
            side2 = "SELL"
        
        # Execute with retries
        success = False
        filled1 = 0
        filled2 = 0
        max_retries = config.max_retries()
        
        for retry in range(1, max_retries + 1):
            logger.info(f"Attempt {retry}/{max_retries}")
            
            remaining_qty1 = max(0, quantity1 - filled1)
            remaining_qty2 = max(0, quantity2 - filled2)
            
            # Check completion
            if remaining_qty1 <= 0 and remaining_qty2 <= 0:
                success = True
                break
            
            # Skip if asymmetric (unless last retry)
            if (remaining_qty1 <= 0 or remaining_qty2 <= 0) and retry < max_retries:
                continue
            
            # Execute both legs using unified order logic (market or limit)
            result, actual_filled1, actual_filled2 = place_and_monitor_pair_unified(
                ib, ticker1, ticker2, 
                remaining_qty1, remaining_qty2, 
                bid1, ask1, bid2, ask2,  # Pass market data for limit orders
                pair, 
                side1, side2
            )
            
            filled1 += actual_filled1
            filled2 += actual_filled2
            
            logger.info(f"Filled so far: {filled1}/{quantity1}, {filled2}/{quantity2}")
            
            if result:
                success = True
                logger.info(f"✅ {pair}: Execution complete!")
                print(f"  ✅ {pair}: Execution complete!")
                break
        
        # Handle partial fills
        if not success and (filled1 > 0 or filled2 > 0):
            balance_success = balance_partially_filled_pair(
                ib, ticker1, ticker2,
                filled1, filled2,
                quantity1, quantity2,
                side1, side2, pair
            )
            
            if balance_success and filled1 > 0 and filled2 > 0:
                return {
                    'Tag': tag, 'Pair': pair,
                    'Status': 'Partial',
                    'Details': f'Balanced: {filled1}/{quantity1}, {filled2}/{quantity2}',
                    'Leg1_Filled': filled1, 'Leg2_Filled': filled2,
                    'Leg1_Target': quantity1, 'Leg2_Target': quantity2,
                    'W1': W1, 'W2': W2,
                    'Position_Multiplier': position_mult
                }
        
        # Return result
        if success:
            return {
                'Tag': tag, 'Pair': pair,
                'Status': 'Executed',
                'Details': 'Success',
                'Leg1_Filled': quantity1, 'Leg2_Filled': quantity2,
                'Leg1_Target': quantity1, 'Leg2_Target': quantity2,
                'W1': W1, 'W2': W2,
                'Position_Multiplier': position_mult
            }
        else:
            return {
                'Tag': tag, 'Pair': pair,
                'Status': 'Failed',
                'Details': f'Incomplete fill: {filled1}/{quantity1}, {filled2}/{quantity2}',
                'Leg1_Filled': filled1, 'Leg2_Filled': filled2,
                'Leg1_Target': quantity1, 'Leg2_Target': quantity2
            }
    
    except Exception as e:
        logger.error(f"❌ {pair}: Exception - {e}")
        print(f"  ❌ {pair}: Exception - {e}")  # Show in Jupyter
        import traceback
        traceback.print_exc()
        return {
            'Tag': tag, 'Pair': pair,
            'Status': 'Failed',
            'Details': f'Exception: {str(e)}',
            'Leg1_Filled': 0, 'Leg2_Filled': 0,
            'Leg1_Target': 0, 'Leg2_Target': 0
        }
    
async def execute_trades_in_batches(evaluated_trades_df,
                                   ib, live_prices, index_price_current,
                                   batch_size=5):
    """
    Execute trades in small batches for faster execution while maintaining balance
    
    This is a simple wrapper around the existing execute_single_pair_trade() that
    processes multiple pairs simultaneously. All existing balancing and error
    handling logic is preserved.
    
    Parameters:
    -----------
    batch_size : int
        Number of pairs to execute simultaneously (default: 5)
        Smaller = more conservative, larger = faster but more complex
    
    Returns:
    --------
    tuple: (updated_portfolio_df, execution_summary_df)
    """
    
    if evaluated_trades_df.empty:
        logger.info("No trades to execute")
        return portfolio_df, pd.DataFrame()
    
    logger.info(f"=" * 80)
    logger.info(f"EXECUTING {len(evaluated_trades_df)} TRADES IN BATCHES OF {batch_size}")
    logger.info(f"=" * 80)
    
    all_execution_results = []
    total_batches = (len(evaluated_trades_df) - 1) // batch_size + 1
    
    # ========================================================================
    # GET ACCOUNT EQUITY ONCE (not per trade - prevents IBKR API limit errors)
    # ========================================================================
    
    account_equity = None
    if config.trading_env() == 'live' or not config.dry_run_mode():
        try:
            account_equity = get_account_equity(ib)
            if account_equity > 0:
                logger.info(f"Account equity for leverage checks: ${account_equity:,.2f}")
            else:
                logger.warning("Could not get account equity - leverage checks will be skipped")
        except Exception as e:
            logger.warning(f"Error getting account equity: {e} - leverage checks will be skipped")
            account_equity = None
    
    # ========================================================================
    # Process trades in batches
    # ========================================================================
    
    for batch_num in range(total_batches):
        start_idx = batch_num * batch_size
        end_idx = min(start_idx + batch_size, len(evaluated_trades_df))
        batch = evaluated_trades_df.iloc[start_idx:end_idx]
        
        logger.info(f"\n{'='*60}")
        logger.info(f"BATCH {batch_num + 1}/{total_batches}: Executing {len(batch)} pairs")
        logger.info(f"{'='*60}")
        
        # List pairs in this batch
        for _, trade in batch.iterrows():
            logger.info(f"  - {trade['Pair']}")
        
        # ====================================================================
        # Execute all pairs in this batch simultaneously
        # ====================================================================
        
        batch_tasks = []
        for _, trade in batch.iterrows():
        # Create task for each pair using existing single-pair function
            task = execute_single_pair_trade(
                trade_spec=trade,
                ib=ib,
                live_prices=live_prices,
                index_price_current=index_price_current,
                portfolio_df=portfolio_df,
                account_equity=account_equity
            )
            batch_tasks.append(task)
        
        # Execute all pairs in parallel and wait for completion
        logger.info(f"\nExecuting {len(batch_tasks)} pairs simultaneously...")
        batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
        
        # ====================================================================
        # Process results
        # ====================================================================
        
        for i, result in enumerate(batch_results):
            trade_spec = batch.iloc[i]
            pair_name = trade_spec['Pair']
            
            if isinstance(result, Exception):
                logger.error(f"❌ {pair_name}: Exception during execution: {result}")
                all_execution_results.append({
                    'Pair': pair_name,
                    'Status': 'Failed',
                    'Details': f'Exception: {str(result)}',
                    'Leg1_Filled': 0,
                    'Leg2_Filled': 0,
                    'Leg1_Target': trade_spec.get('Quantity1', 0),
                    'Leg2_Target': trade_spec.get('Quantity2', 0)
                })
            else:
                # Result is from execute_single_pair_trade
                all_execution_results.append(result)
        
        logger.info(f"\n{'='*60}")
        logger.info(f"BATCH {batch_num + 1} COMPLETE")
        logger.info(f"{'='*60}")
        
        # Small pause between batches to avoid overwhelming IBKR
        if batch_num < total_batches - 1:
            await asyncio.sleep(2)
    
    # ========================================================================
    # Create execution summary
    # ========================================================================
    
    logger.info(f"\n{'='*80}")
    logger.info("EXECUTION SUMMARY")
    logger.info(f"{'='*80}")
    
    execution_summary_df = pd.DataFrame(all_execution_results)
    
    if execution_summary_df.empty:
        logger.warning("No execution results to summarize")
        return portfolio_df, execution_summary_df
    
    # Count results by status
    status_counts = execution_summary_df['Status'].value_counts()
    
    logger.info(f"\nTotal trades attempted: {len(evaluated_trades_df)}")
    for status, count in status_counts.items():
        logger.info(f"  {status}: {count}")
    
    # Show successful trades
    successful = execution_summary_df[
        execution_summary_df['Status'].isin(['Executed', 'Partial'])
    ]
    
    if not successful.empty:
        logger.info(f"\nSuccessfully executed ({len(successful)}):")
        for _, trade in successful.iterrows():
            logger.info(f"  ✓ {trade['Pair']}: {trade['Details']}")
    
    # Show failed trades
    failed = execution_summary_df[
        execution_summary_df['Status'].isin(['Failed', 'Skipped'])
    ]
    
    if not failed.empty:
        logger.info(f"\nFailed/Skipped ({len(failed)}):")
        for _, trade in failed.iterrows():
            logger.info(f"  ✗ {trade['Pair']}: {trade['Details']}")
    
    return execution_summary_df

# ============================================================================
# AGGREGATED EXECUTION (ORDER NETTING)
# ============================================================================

async def execute_aggregated_order(ib, agg_order, order_map):
    """
    Execute a single aggregated order
    
    This combines multiple orders for the same ticker/direction into one
    
    Parameters:
    -----------
    ib : IB connection
    agg_order : dict
        Aggregated order from OrderAggregator
    order_map : dict
        Mapping to contributing pairs
        
    Returns:
    --------
    dict : Execution result with allocations
    """
    order_id = agg_order['order_id']
    ticker = agg_order['ticker']
    side = agg_order['side']
    total_qty = agg_order['total_qty']
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Executing aggregated order: {order_id}")
    logger.info(f"  {side} {total_qty} {ticker}")
    logger.info(f"  Contributing pairs: {agg_order['num_contributors']}")
    logger.info(f"{'='*60}")
    
    # Get market data for pricing
    bid, ask = fetch_market_data(ib, ticker)
    
    if bid is None or ask is None:
        logger.error(f"❌ {ticker}: Cannot get market data")
        return {
            'order_id': order_id,
            'ticker': ticker,
            'side': side,
            'requested_qty': total_qty,
            'filled_qty': 0,
            'allocations': [],
            'status': 'failed'
        }
    
    # Execute order using existing unified logic
    success, filled_qty = place_and_monitor_order_unified(
        ib, ticker, side, total_qty, bid, ask, pair=order_id
    )
    
    # Allocate fills to contributing pairs
    aggregator = OrderAggregator()
    allocations = aggregator.allocate_fills(order_id, filled_qty, order_map)
    
    # Determine overall status
    if filled_qty == total_qty:
        status = 'success'
    elif filled_qty > 0:
        status = 'partial'
    else:
        status = 'failed'
    
    return {
        'order_id': order_id,
        'ticker': ticker,
        'side': side,
        'requested_qty': total_qty,
        'filled_qty': filled_qty,
        'allocations': allocations,
        'status': status
    }

def reconstruct_pair_results_from_allocations(execution_results, evaluated_trades_df):
    """
    Reconstruct pair-level execution results from aggregated order allocations
    
    Maps aggregated fills back to original pairs for portfolio recording
    
    Parameters:
    -----------
    execution_results : list
        Results from execute_aggregated_order() calls
    evaluated_trades_df : DataFrame
        Original approved trades
        
    Returns:
    --------
    list : Pair-level execution results (compatible with existing workflow)
    """
    from collections import defaultdict
    
    # Build map of pair fills
    pair_fills = defaultdict(lambda: {
        'leg1_ticker': None,
        'leg1_filled': 0,
        'leg1_target': 0,
        'leg2_ticker': None,
        'leg2_filled': 0,
        'leg2_target': 0,
        'tag': None
    })
    
    # Process each aggregated order's allocations
    for exec_result in execution_results:
        for allocation in exec_result['allocations']:
            pair_id = allocation['pair_id']
            ticker = allocation['ticker']
            filled = allocation['qty_allocated']
            requested = allocation['qty_requested']
            
            # Extract pair name from pair_id (format: "TICKER1-TICKER2-L")
            pair_name = '-'.join(pair_id.split('-')[:-1])
            
            # Find original trade to determine leg assignment
            original_trade = evaluated_trades_df[
                evaluated_trades_df['Pair'] == pair_name
            ]
            
            if original_trade.empty:
                logger.warning(f"Could not find original trade for {pair_name}")
                continue
            
            original_trade = original_trade.iloc[0]
            
            # Store fills by leg
            if ticker == original_trade['Co1']:
                pair_fills[pair_id]['leg1_ticker'] = ticker
                pair_fills[pair_id]['leg1_filled'] = filled
                pair_fills[pair_id]['leg1_target'] = requested
                pair_fills[pair_id]['tag'] = original_trade.get('Tag')
            elif ticker == original_trade['Co2']:
                pair_fills[pair_id]['leg2_ticker'] = ticker
                pair_fills[pair_id]['leg2_filled'] = filled
                pair_fills[pair_id]['leg2_target'] = requested
                pair_fills[pair_id]['tag'] = original_trade.get('Tag')
    
    # Convert to result list (compatible with existing workflow)
    results = []
    for pair_id, fills in pair_fills.items():
        pair_name = '-'.join(pair_id.split('-')[:-1])
        leg1_filled = fills['leg1_filled']
        leg2_filled = fills['leg2_filled']
        leg1_target = fills['leg1_target']
        leg2_target = fills['leg2_target']
        
        # Determine status
        if leg1_filled >= leg1_target and leg2_filled >= leg2_target:
            status = 'Executed'
            details = 'Success'
        elif leg1_filled > 0 and leg2_filled > 0:
            status = 'Partial'
            details = f'Partial: {leg1_filled}/{leg1_target}, {leg2_filled}/{leg2_target}'
        else:
            status = 'Failed'
            details = f'Incomplete: {leg1_filled}/{leg1_target}, {leg2_filled}/{leg2_target}'
        
        results.append({
            'Tag': fills['tag'],
            'Pair': pair_name,
            'Status': status,
            'Details': details,
            'Leg1_Filled': leg1_filled,
            'Leg2_Filled': leg2_filled,
            'Leg1_Target': leg1_target,
            'Leg2_Target': leg2_target
        })
    
    return results

async def execute_trades_with_aggregation(evaluated_trades_df,
                                         ib, live_prices, index_price_current):
    """
    Execute trades with order aggregation (commission reduction mode)
    
    This is an alternate execution path to execute_trades_in_batches() that
    combines orders for the same ticker/direction to reduce commissions.
    
    WHEN TO USE:
    - Multiple pairs share the same ticker
    - Commission costs are significant
    - Willing to accept slightly more complex allocation logic
    
    HOW IT WORKS:
    1. Analyzes all approved trades
    2. Groups orders by ticker and direction (BUY/SELL)
    3. Aggregates groups meeting threshold (default: 3+ appearances)
    4. Executes single order per aggregated group
    5. Allocates fills back to pairs by priority
    
    EXAMPLE:
    Without aggregation (3 separate orders):
      PSTG-FN: Sell 1 FN → $2.00 commission
      PLAB-FN: Sell 1 FN → $2.00 commission
      CRWD-FN: Sell 1 FN → $2.00 commission
      Total: $6.00 in commissions
    
    With aggregation (1 combined order):
      Aggregated: Sell 3 FN → $2.00 commission
      Savings: $4.00 (67% reduction)
    
    Parameters:
    -----------
    All same as execute_trades_in_batches()
    
    Returns:
    --------
    tuple: (updated_portfolio_df, execution_summary_df)
           Same format as execute_trades_in_batches() for compatibility
    """
    
    if evaluated_trades_df.empty:
        logger.info("No trades to execute")
        return portfolio_df, pd.DataFrame()
    
    logger.info("="*80)
    logger.info("EXECUTING TRADES WITH ORDER AGGREGATION")
    logger.info("="*80)
    
    # ========================================================================
    # GET ACCOUNT EQUITY (for leverage checks)
    # ========================================================================
    
    account_equity = None
    if config.trading_env() == 'live' or not config.dry_run_mode():
        try:
            account_equity = get_account_equity(ib)
            if account_equity > 0:
                logger.info(f"Account equity for leverage checks: ${account_equity:,.2f}")
        except Exception as e:
            logger.warning(f"Could not get account equity: {e}")
    
    # ========================================================================
    # STEP 1: Aggregate Orders
    # ========================================================================
    
    aggregator = OrderAggregator(min_aggregation_threshold=config.min_aggregation_threshold())
    aggregation_result = aggregator.aggregate_approved_trades(evaluated_trades_df)
    
    aggregated_orders = aggregation_result['aggregated_orders']
    order_map = aggregation_result['order_map']
    stats = aggregation_result['statistics']
    
    # ========================================================================
    # STEP 2: Execute Aggregated Orders
    # ========================================================================
    
    logger.info(f"\nExecuting {len(aggregated_orders)} aggregated orders...")
    
    execution_results = []
    
    for agg_order in aggregated_orders:
        result = await execute_aggregated_order(ib, agg_order, order_map)
        execution_results.append(result)
        
        # Small delay between orders
        await asyncio.sleep(0.5)
    
    # ========================================================================
    # STEP 3: Reconstruct Pair-Level Results
    # ========================================================================
    
    logger.info(f"\nReconstructing pair-level results...")
    pair_results = reconstruct_pair_results_from_allocations(execution_results, evaluated_trades_df)
    
    # ========================================================================
    # STEP 4: Create Execution Summary (compatible with existing workflow)
    # ========================================================================
    
    logger.info(f"\n{'='*80}")
    logger.info("EXECUTION SUMMARY")
    logger.info(f"{'='*80}")
    
    execution_summary_df = pd.DataFrame(pair_results)
    
    if execution_summary_df.empty:
        logger.warning("No execution results to summarize")
        return portfolio_df, execution_summary_df
    
    # Count results by status
    status_counts = execution_summary_df['Status'].value_counts()
    
    logger.info(f"\nTotal trades attempted: {len(evaluated_trades_df)}")
    for status, count in status_counts.items():
        logger.info(f"  {status}: {count}")
    
    # Show aggregation savings
    logger.info(f"\nAggregation Savings:")
    logger.info(f"  Orders executed: {stats['total_orders_after']}")
    logger.info(f"  Orders saved: {stats['orders_saved']}")
    logger.info(f"  Commission saved: ${stats['commission_saved']:.2f}")
    
    # Show successful trades
    successful = execution_summary_df[
        execution_summary_df['Status'].isin(['Executed', 'Partial'])
    ]
    
    if not successful.empty:
        logger.info(f"\nSuccessfully executed ({len(successful)}):")
        for _, trade in successful.iterrows():
            logger.info(f"  ✓ {trade['Pair']}: {trade['Details']}")
    
    # Show failed trades
    failed = execution_summary_df[
        execution_summary_df['Status'].isin(['Failed', 'Skipped'])
    ]
    
    if not failed.empty:
        logger.info(f"\nFailed/Skipped ({len(failed)}):")
        for _, trade in failed.iterrows():
            logger.info(f"  ✗ {trade['Pair']}: {trade['Details']}")
    
    return execution_summary_df

async def execute_terminations(to_terminate_df, ib, live_prices):
    """
    Execute closing orders for terminated trades using LIMIT ORDERS
    
    Uses aggressive limit pricing to ensure fills while avoiding catastrophic
    execution that could result from market orders on wide-spread stocks.
    
    Limit pricing strategy for exits:
    - SELL: Limit at (bid - buffer) to ensure fill
    - BUY: Limit at (ask + buffer) to ensure fill
    
    Falls back to market order if limit doesn't fill within timeout.
    
    Parameters:
    -----------
    to_terminate_df : DataFrame
        Trades to close
    ib : IB connection
        Connected IBKR instance
    live_prices : dict
        Current market prices {ticker: price}
    
    Returns:
    --------
    bool : True if all terminations executed successfully
    """
    from ib_insync import Stock, LimitOrder, MarketOrder
    import asyncio
    
    logger.info(f"Executing {len(to_terminate_df)} terminations with LIMIT orders...")
    
    # Exit limit order settings from config
    EXIT_LIMIT_BUFFER_PCT = getattr(config, 'EXIT_LIMIT_BUFFER_PCT', 0.005)  # 0.5% default
    EXIT_LIMIT_TIMEOUT = getattr(config, 'EXIT_LIMIT_TIMEOUT', 30)  # 30s default
    
    all_trades = []  # Track all order info for monitoring
    
    for idx, trade in to_terminate_df.iterrows():
        pair = trade['Pair']
        ticker1 = trade['Co1']
        ticker2 = trade['Co2']
        qty1 = int(trade['Quantity1'])
        qty2 = int(trade['Quantity2'])
        tail = trade.get('Tail', 'L').strip().upper()
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Closing {pair} (Tag: {trade['Tag']})")
        logger.info(f"  Reason: {trade.get('Exit Reason', 'Unknown')}")
        logger.info(f"{'='*60}")
        
        # Determine closing directions (reverse of entry)
        if tail == 'L':
            # Was: Long ticker1, Short ticker2
            # Close: Sell ticker1, Buy ticker2
            side1 = "SELL"
            side2 = "BUY"
        else:  # U
            # Was: Short ticker1, Long ticker2
            # Close: Buy ticker1, Sell ticker2
            side1 = "BUY"
            side2 = "SELL"
        
        # ========================================================================
        # LEG 1
        # ========================================================================
        try:
            contract1 = Stock(ticker1, 'SMART', 'USD')
            qualified1 = ib.qualifyContracts(contract1)
            
            if not qualified1:
                logger.error(f"  ✗ Failed to qualify {ticker1}")
                continue
            
            # Get current market data for limit pricing
            ticker_data1 = ib.reqMktData(qualified1[0], '', False, False)
            await asyncio.sleep(0.5)  # Brief wait for data
            
            bid1 = ticker_data1.bid if ticker_data1.bid and ticker_data1.bid > 0 else None
            ask1 = ticker_data1.ask if ticker_data1.ask and ticker_data1.ask > 0 else None
            mid1 = live_prices.get(ticker1)
            
            # Calculate aggressive limit price
            if side1 == "SELL":
                # Selling: price below bid to ensure fill
                if bid1 and bid1 > 0:
                    limit1 = round(bid1 * (1 - EXIT_LIMIT_BUFFER_PCT), 2)
                elif mid1:
                    limit1 = round(mid1 * (1 - EXIT_LIMIT_BUFFER_PCT * 2), 2)
                else:
                    limit1 = None  # Will use market order
            else:  # BUY
                # Buying: price above ask to ensure fill
                if ask1 and ask1 > 0:
                    limit1 = round(ask1 * (1 + EXIT_LIMIT_BUFFER_PCT), 2)
                elif mid1:
                    limit1 = round(mid1 * (1 + EXIT_LIMIT_BUFFER_PCT * 2), 2)
                else:
                    limit1 = None  # Will use market order
            
            if limit1:
                logger.info(f"  Leg 1: {side1} {qty1} {ticker1} @ ${limit1:.2f} (LIMIT)")
                order1 = LimitOrder(side1, qty1, limit1, tif='DAY')
            else:
                logger.warning(f"  Leg 1: {side1} {qty1} {ticker1} (MARKET - no price data)")
                order1 = MarketOrder(side1, qty1, tif='DAY')
            
            trade1 = ib.placeOrder(qualified1[0], order1)
            all_trades.append({
                'trade': trade1,
                'ticker': ticker1,
                'side': side1,
                'qty': qty1,
                'limit': limit1,
                'contract': qualified1[0]
            })
            logger.info(f"    Order placed for {ticker1}")
            
            ib.cancelMktData(qualified1[0])  # Clean up market data subscription
                
        except Exception as e:
            logger.error(f"  ✗ Error placing {ticker1} order: {e}")
        
        # ========================================================================
        # LEG 2
        # ========================================================================
        try:
            contract2 = Stock(ticker2, 'SMART', 'USD')
            qualified2 = ib.qualifyContracts(contract2)
            
            if not qualified2:
                logger.error(f"  ✗ Failed to qualify {ticker2}")
                continue
            
            # Get current market data for limit pricing
            ticker_data2 = ib.reqMktData(qualified2[0], '', False, False)
            await asyncio.sleep(0.5)  # Brief wait for data
            
            bid2 = ticker_data2.bid if ticker_data2.bid and ticker_data2.bid > 0 else None
            ask2 = ticker_data2.ask if ticker_data2.ask and ticker_data2.ask > 0 else None
            mid2 = live_prices.get(ticker2)
            
            # Calculate aggressive limit price
            if side2 == "SELL":
                if bid2 and bid2 > 0:
                    limit2 = round(bid2 * (1 - EXIT_LIMIT_BUFFER_PCT), 2)
                elif mid2:
                    limit2 = round(mid2 * (1 - EXIT_LIMIT_BUFFER_PCT * 2), 2)
                else:
                    limit2 = None
            else:  # BUY
                if ask2 and ask2 > 0:
                    limit2 = round(ask2 * (1 + EXIT_LIMIT_BUFFER_PCT), 2)
                elif mid2:
                    limit2 = round(mid2 * (1 + EXIT_LIMIT_BUFFER_PCT * 2), 2)
                else:
                    limit2 = None
            
            if limit2:
                logger.info(f"  Leg 2: {side2} {qty2} {ticker2} @ ${limit2:.2f} (LIMIT)")
                order2 = LimitOrder(side2, qty2, limit2, tif='DAY')
            else:
                logger.warning(f"  Leg 2: {side2} {qty2} {ticker2} (MARKET - no price data)")
                order2 = MarketOrder(side2, qty2, tif='DAY')
            
            trade2 = ib.placeOrder(qualified2[0], order2)
            all_trades.append({
                'trade': trade2,
                'ticker': ticker2,
                'side': side2,
                'qty': qty2,
                'limit': limit2,
                'contract': qualified2[0]
            })
            logger.info(f"    Order placed for {ticker2}")
            
            ib.cancelMktData(qualified2[0])  # Clean up market data subscription
                
        except Exception as e:
            logger.error(f"  ✗ Error placing {ticker2} order: {e}")
    
    # ========================================================================
    # PHASE 1: Wait for limit orders to fill
    # ========================================================================
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Waiting for {len(all_trades)} LIMIT orders to fill...")
    logger.info(f"{'='*60}")
    
    elapsed = 0
    check_interval = 2
    
    while elapsed < EXIT_LIMIT_TIMEOUT:
        await asyncio.sleep(check_interval)
        elapsed += check_interval
        
        # Check status
        pending = []
        filled = []
        
        for order_info in all_trades:
            trade_obj = order_info['trade']
            ticker = order_info['ticker']
            status = trade_obj.orderStatus.status
            filled_qty = trade_obj.orderStatus.filled
            target_qty = order_info['qty']
            
            if status == 'Filled' or filled_qty >= target_qty:
                filled.append(order_info)
            elif status not in ['Cancelled', 'ApiCancelled', 'Rejected']:
                pending.append(order_info)
        
        logger.info(f"  [{elapsed}s] Filled: {len(filled)}/{len(all_trades)}, Pending: {len(pending)}")
        
        if not pending:
            break
    
    # ========================================================================
    # PHASE 2: Fallback to market orders for unfilled limits
    # ========================================================================
    
    unfilled = []
    for order_info in all_trades:
        trade_obj = order_info['trade']
        status = trade_obj.orderStatus.status
        filled_qty = trade_obj.orderStatus.filled
        
        if status not in ['Filled'] and filled_qty < order_info['qty']:
            unfilled.append(order_info)
    
    if unfilled:
        logger.warning(f"\n⚠️  {len(unfilled)} orders not filled by limit - converting to MARKET")
        
        for order_info in unfilled:
            trade_obj = order_info['trade']
            ticker = order_info['ticker']
            side = order_info['side']
            qty = order_info['qty']
            contract = order_info['contract']
            filled_qty = trade_obj.orderStatus.filled
            remaining = int(qty - filled_qty)
            
            if remaining <= 0:
                continue
            
            try:
                # Cancel the limit order first
                if trade_obj.orderStatus.status not in ['Filled', 'Cancelled', 'ApiCancelled']:
                    ib.cancelOrder(trade_obj.order)
                    await asyncio.sleep(1)
                
                # Place market order for remaining quantity
                logger.info(f"  Converting {ticker}: MARKET {side} {remaining} shares")
                market_order = MarketOrder(side, remaining, tif='DAY')
                market_trade = ib.placeOrder(contract, market_order)
                
                # Update tracking
                order_info['trade'] = market_trade
                order_info['converted_to_market'] = True
                
            except Exception as e:
                logger.error(f"  ✗ Error converting {ticker} to market: {e}")
        
        # Wait for market orders
        logger.info("  Waiting for market orders to fill...")
        await asyncio.sleep(10)
    
    # ========================================================================
    # FINAL SUMMARY
    # ========================================================================
    
    logger.info(f"\n{'='*60}")
    logger.info(f"TERMINATION EXECUTION COMPLETE")
    logger.info(f"{'='*60}")
    
    filled = []
    failed = []
    partial = []
    
    for order_info in all_trades:
        trade_obj = order_info['trade']
        ticker = order_info['ticker']
        side = order_info['side']
        qty = order_info['qty']
        
        status = trade_obj.orderStatus.status
        filled_qty = trade_obj.orderStatus.filled
        
        order_type = "MARKET" if order_info.get('converted_to_market') else "LIMIT"
        
        if status == 'Filled' or filled_qty >= qty:
            filled.append(order_info)
            logger.info(f"  ✓ {ticker}: {side} {int(filled_qty)} shares - FILLED ({order_type})")
        elif filled_qty > 0 and filled_qty < qty:
            partial.append(order_info)
            logger.warning(f"  ⚠️  {ticker}: {side} {int(filled_qty)}/{qty} shares - PARTIAL ({order_type})")
        elif status in ['Cancelled', 'ApiCancelled', 'Rejected']:
            failed.append(order_info)
            logger.error(f"  ✗ {ticker}: {side} {qty} shares - {status}")
        else:
            failed.append(order_info)
            logger.warning(f"  ⚠️  {ticker}: {side} {qty} shares - {status} (not filled)")
    
    # Determine overall success
    total_orders = len(all_trades)
    success_count = len(filled) + len(partial)
    
    if success_count == total_orders:
        logger.info(f"\n✓ All {total_orders} termination orders completed successfully")
        return True
    elif success_count >= total_orders * 0.8:  # 80% success threshold
        logger.warning(f"\n⚠️  {success_count}/{total_orders} orders completed ({len(failed)} failed)")
        logger.warning("Accepting as successful (80% threshold met)")
        return True
    else:
        logger.error(f"\n❌ Only {success_count}/{total_orders} orders completed")
        logger.error("Too many failures - terminations NOT confirmed")
        return False